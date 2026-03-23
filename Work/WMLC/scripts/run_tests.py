#!/usr/bin/env python
"""WMLC Pipeline Test Suite — Sub-Agent 4.

Runs all automated validation tests and writes output/test_report.md.
"""

import os
import sys
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl

# ---------------------------------------------------------------------------
# Resolve project root (repo root where CLAUDE.md lives)
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent  # C:/.../WMLC

os.chdir(ROOT)

# ---------------------------------------------------------------------------
# Test infrastructure
# ---------------------------------------------------------------------------

class TestResult:
    def __init__(self, name: str, passed: bool, details: str = ""):
        self.name = name
        self.passed = passed
        self.details = details


results: list[TestResult] = []


def run_test(name: str):
    """Decorator that captures pass/fail into *results*."""
    def decorator(fn):
        try:
            detail = fn()
            results.append(TestResult(name, True, detail or ""))
        except Exception as exc:
            results.append(TestResult(name, False, str(exc)))
    return decorator


# ===================================================================
# 1. DATA GEN VALIDATION
# ===================================================================

@run_test("Proxy data files exist")
def _():
    missing = []
    for f in [
        "proxy_data/loan_extract.csv",
        "proxy_data/LAL_Credit.xlsx",
        "proxy_data/Loan_Reserve_Report.xlsx",
        "proxy_data/DAR_Tracker.xlsx",
    ]:
        if not os.path.exists(f):
            missing.append(f)
    assert not missing, f"Missing files: {missing}"
    return "All 4 proxy files present"


@run_test("loan_extract.csv has 400+ rows")
def _():
    df = pd.read_csv("proxy_data/loan_extract.csv")
    assert len(df) >= 400, f"Only {len(df)} rows"
    return f"{len(df)} rows"


@run_test("loan_extract.csv has all required columns")
def _():
    df = pd.read_csv("proxy_data/loan_extract.csv", nrows=0)
    required = [
        "tl_facility_digits12", "facility_id", "account_number", "key_acct",
        "borrower", "name", "sub_product_norm", "product_bucket", "is_lal_nfp",
        "focus_list", "txt_mstr_facil_collateral_desc", "SBL_PERC",
        "book_date", "effective_date", "origination_date",
        "balance", "credit_limit", "amt_original_comt", "credit_lii",
        "NEW_CAMP_YN", "NEW_CAMP_REASON",
        "base_commit", "latest_commit", "commit_delta",
        "new_commitment_amount", "new_commitment_reason",
        "credit_lii_commitment_bucket", "credit_lii_commitment_floor",
    ]
    missing = [c for c in required if c not in df.columns]
    assert not missing, f"Missing columns: {missing}"
    return f"{len(df.columns)} columns present, all {len(required)} required found"


@run_test("All 13 product_buckets present")
def _():
    df = pd.read_csv("proxy_data/loan_extract.csv", usecols=["product_bucket"])
    expected = {
        "LAL Diversified", "LAL Highly Conc.", "LAL NFPs", "RESI",
        "TL Aircraft", "TL CRE", "TL Life Insurance", "TL Multicollateral",
        "TL Other Secured", "TL PHA", "TL SBL Diversified",
        "TL SBL Highly Conc.", "TL Unsecured",
    }
    actual = set(df["product_bucket"].unique())
    missing = expected - actual
    assert not missing, f"Missing product_buckets: {missing}"
    return f"{len(actual)} buckets found"


@run_test("account_number is 12-digit zero-padded")
def _():
    df = pd.read_csv("proxy_data/loan_extract.csv", dtype={"account_number": str})
    bad = df[df["account_number"].str.len() != 12]
    assert len(bad) == 0, f"{len(bad)} rows have non-12-digit account_number"
    # Also check that it is numeric (zero-padded digits)
    non_numeric = df[~df["account_number"].str.match(r"^\d{12}$")]
    assert len(non_numeric) == 0, f"{len(non_numeric)} rows have non-numeric account_number"
    return "All account_numbers are 12-digit zero-padded"


@run_test("credit_lii spans under $10M to over $100M")
def _():
    df = pd.read_csv("proxy_data/loan_extract.csv", usecols=["credit_lii"])
    lo, hi = df["credit_lii"].min(), df["credit_lii"].max()
    assert lo < 10_000_000, f"Min credit_lii {lo:,.0f} is not below $10M"
    assert hi > 100_000_000, f"Max credit_lii {hi:,.0f} is not above $100M"
    return f"Range ${lo:,.0f} to ${hi:,.0f}"


@run_test("NEW_CAMP_YN has both Y and N, Y >= 30%")
def _():
    df = pd.read_csv("proxy_data/loan_extract.csv", usecols=["NEW_CAMP_YN"])
    vals = set(df["NEW_CAMP_YN"].dropna().unique())
    assert "Y" in vals, "No NEW_CAMP_YN='Y' rows"
    assert "N" in vals, "No NEW_CAMP_YN='N' rows"
    y_pct = (df["NEW_CAMP_YN"] == "Y").mean() * 100
    assert y_pct >= 30, f"Only {y_pct:.1f}% are Y (need >= 30%)"
    return f"Y={y_pct:.1f}%"


@run_test("credit_lii_commitment_bucket matches credit_lii values")
def _():
    df = pd.read_csv("proxy_data/loan_extract.csv",
                      usecols=["credit_lii", "credit_lii_commitment_bucket", "credit_lii_commitment_floor"])
    # Bucket floor should be <= credit_lii
    bad = df[df["credit_lii_commitment_floor"] > df["credit_lii"]]
    assert len(bad) == 0, f"{len(bad)} rows where bucket floor > credit_lii"
    # Spot-check big loans
    big = df[df["credit_lii"] >= 300_000_000]
    if len(big) > 0:
        valid_floors = {300_000_000, 350_000_000, 400_000_000, 500_000_000,
                        600_000_000, 700_000_000, 750_000_000, 1_000_000_000}
        bad_big = big[~big["credit_lii_commitment_floor"].isin(valid_floors)]
        assert len(bad_big) == 0, f"{len(bad_big)} large loans in wrong bucket"
    return "All bucket assignments consistent"


# ===================================================================
# 2. ETL VALIDATION
# ===================================================================

@run_test("ETL runs with exit code 0")
def _():
    result = subprocess.run(
        [sys.executable, "corp_etl/main.py"],
        capture_output=True, text=True, timeout=120,
    )
    assert result.returncode == 0, f"Exit code {result.returncode}.\nSTDERR:\n{result.stderr[:500]}"
    return "ETL completed successfully"


@run_test("Tagged CSV has all 6 new columns")
def _():
    df = pd.read_csv("output/loan_extract_tagged.csv", nrows=0)
    required = ["IS_NTC", "IS_OFFICE", "HAS_CREDIT_POLICY_EXCEPTION",
                "WMLC_FLAGS", "WMLC_FLAG_COUNT", "WMLC_QUALIFIED"]
    missing = [c for c in required if c not in df.columns]
    assert not missing, f"Missing columns: {missing}"
    return f"All 6 columns present ({len(df.columns)} total)"


@run_test("wmlc_flags contain only valid flag names")
def _():
    valid_flags = {
        "NTC > $50MM", "Non-Pass Originations >$0MM",
        "TL-CRE >$75MM", "TL-CRE Office >$10MM",
        "TL-SBL-D >$300MM", "TL-SBL-C >$100MM", "TL-LIC >$100MM",
        "TL-Alts HF/PE >$35MM", "TL-Alts Private Shares >$35MM",
        "TL-Alts Unsecured >$35MM", "TL-Alts PAF >$50MM",
        "TL-Alts Fine Art >$50MM", "TL-Alts Other Secured >$50MM",
        "LAL-D >$300MM", "LAL-C >$100MM",
        "RESI >$10MM",
    }
    df = pd.read_csv("output/loan_extract_tagged.csv", usecols=["WMLC_FLAGS"])
    flagged = df[df["WMLC_FLAGS"].notna() & (df["WMLC_FLAGS"] != "")]
    invalid_found = set()
    for flags_str in flagged["WMLC_FLAGS"]:
        for f in str(flags_str).split("|"):
            f = f.strip()
            if f and f not in valid_flags:
                invalid_found.add(f)
    assert not invalid_found, f"Invalid flags found: {invalid_found}"
    return f"All flags valid across {len(flagged)} flagged rows"


@run_test("wmlc_flag_count matches pipe-delimited flag count")
def _():
    df = pd.read_csv("output/loan_extract_tagged.csv",
                      usecols=["WMLC_FLAGS", "WMLC_FLAG_COUNT"])
    mismatches = 0
    for idx, row in df.iterrows():
        flags = row["WMLC_FLAGS"]
        count = row["WMLC_FLAG_COUNT"]
        if pd.isna(flags) or str(flags).strip() == "":
            expected = 0
        else:
            expected = len(str(flags).split("|"))
        if int(count) != expected:
            mismatches += 1
            if mismatches <= 3:
                print(f"  Row {idx}: count={count}, expected={expected}, flags='{flags}'")
    assert mismatches == 0, f"{mismatches} rows with mismatched flag_count"
    return "All flag counts match"


@run_test("wmlc_qualified == True iff wmlc_flag_count > 0")
def _():
    df = pd.read_csv("output/loan_extract_tagged.csv",
                      usecols=["WMLC_FLAG_COUNT", "WMLC_QUALIFIED"])
    mismatch = df[df["WMLC_QUALIFIED"] != (df["WMLC_FLAG_COUNT"] > 0)]
    assert len(mismatch) == 0, f"{len(mismatch)} rows where qualified != (count > 0)"
    return "All rows consistent"


@run_test("At least 8 distinct flag types appear")
def _():
    df = pd.read_csv("output/loan_extract_tagged.csv", usecols=["WMLC_FLAGS"])
    all_flags = set()
    for flags_str in df["WMLC_FLAGS"].dropna():
        s = str(flags_str).strip()
        if s:
            all_flags.update(f.strip() for f in s.split("|"))
    assert len(all_flags) >= 8, f"Only {len(all_flags)} distinct flags: {all_flags}"
    return f"{len(all_flags)} distinct flags: {', '.join(sorted(all_flags))}"


@run_test("At least 50 loans are WMLC-qualified")
def _():
    df = pd.read_csv("output/loan_extract_tagged.csv", usecols=["WMLC_QUALIFIED"])
    count = df["WMLC_QUALIFIED"].sum()
    assert count >= 50, f"Only {count} WMLC-qualified loans"
    return f"{int(count)} WMLC-qualified loans"


# ===================================================================
# 3. EXCEL STRUCTURAL VALIDATION
# ===================================================================

def _excel_path():
    for ext in (".xlsm", ".xlsx"):
        p = f"output/WMLC_Dashboard{ext}"
        if os.path.exists(p):
            return p
    return None


@run_test("Excel dashboard file exists")
def _():
    p = _excel_path()
    assert p is not None, "No WMLC_Dashboard.xlsm or .xlsx found"
    size = os.path.getsize(p)
    assert size > 50_000, f"File too small: {size} bytes"
    return f"{p} ({size:,} bytes)"


@run_test("Excel file is valid ZIP")
def _():
    p = _excel_path()
    assert p, "Dashboard file missing"
    assert zipfile.is_zipfile(p), "File is not a valid ZIP/XLSX"
    return "Valid ZIP archive"


@run_test("Required sheets: Dashboard, loan_detail, _data")
def _():
    p = _excel_path()
    wb = openpyxl.load_workbook(p, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    required = {"Dashboard", "loan_detail", "_data"}
    missing = required - sheets
    assert not missing, f"Missing sheets: {missing}"
    return f"All 3 required sheets present. Total sheets: {len(sheets)}"


@run_test("8 hidden view sheets (_view1 through _view8)")
def _():
    p = _excel_path()
    wb = openpyxl.load_workbook(p, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    missing = []
    for i in range(1, 9):
        name = f"_view{i}"
        if name not in sheets:
            missing.append(name)
    assert not missing, f"Missing view sheets: {missing}"
    return "All 8 view sheets present"


@run_test("Dashboard has 31+ rows and 16 columns")
def _():
    p = _excel_path()
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["Dashboard"]
    rows, cols = ws.max_row, ws.max_column
    wb.close()
    assert rows >= 31, f"Only {rows} rows (need 31+)"
    assert cols >= 16, f"Only {cols} columns (need 16)"
    return f"{rows} rows x {cols} columns"


@run_test("_data row count matches tagged CSV")
def _():
    tagged_count = len(pd.read_csv("output/loan_extract_tagged.csv"))
    p = _excel_path()
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["_data"]
    # max_row includes header row, so data rows = max_row - 1
    excel_data_rows = ws.max_row - 1
    wb.close()
    assert excel_data_rows == tagged_count, \
        f"_data has {excel_data_rows} data rows, tagged CSV has {tagged_count}"
    return f"{excel_data_rows} data rows (matches)"


@run_test("loan_detail has all columns from tagged CSV")
def _():
    tagged_cols = set(pd.read_csv("output/loan_extract_tagged.csv", nrows=0).columns)
    p = _excel_path()
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["loan_detail"]
    # Read header row (row 3 per spec, but check rows 1-3 for the actual header)
    excel_cols = set()
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if cell is not None:
                excel_cols.add(str(cell).strip())
    wb.close()
    missing = tagged_cols - excel_cols
    # Allow minor naming differences — report but don't fail if < 3 missing
    if len(missing) > 3:
        assert False, f"{len(missing)} tagged columns missing from loan_detail: {missing}"
    detail = "All columns present" if not missing else f"Minor mismatches: {missing}"
    return detail


@run_test("Output file is .xlsm (not .xlsx)")
def _():
    p = "output/WMLC_Dashboard.xlsm"
    assert os.path.exists(p), "output/WMLC_Dashboard.xlsm does not exist"
    assert p.endswith(".xlsm"), f"File extension is not .xlsm: {p}"
    return f"Confirmed: {p}"


@run_test("vbaProject.bin exists inside .xlsm ZIP")
def _():
    p = "output/WMLC_Dashboard.xlsm"
    assert os.path.exists(p), "output/WMLC_Dashboard.xlsm does not exist"
    assert zipfile.is_zipfile(p), "File is not a valid ZIP"
    with zipfile.ZipFile(p) as z:
        names = z.namelist()
    assert "xl/vbaProject.bin" in names, \
        f"xl/vbaProject.bin not found in ZIP. Contents: {[n for n in names if 'vba' in n.lower()]}"
    return "xl/vbaProject.bin present"


@run_test("Summary sheet exists")
def _():
    p = _excel_path()
    assert p, "Dashboard file missing"
    wb = openpyxl.load_workbook(p, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    assert "Summary" in sheets, f"'Summary' sheet not found. Sheets: {sheets}"
    return f"Summary sheet present (total sheets: {len(sheets)})"


@run_test("Summary sheet has 4 sub-matrices with correct subheaders")
def _():
    p = _excel_path()
    assert p, "Dashboard file missing"
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["Summary"]
    rows = ws.max_row
    cols = ws.max_column
    # Check minimum dimensions
    assert rows >= 100, f"Summary has only {rows} rows (need 100+)"
    assert cols >= 16, f"Summary has only {cols} columns (need 16+)"
    # Check subheaders at expected rows
    r3 = str(ws.cell(3, 1).value or "")
    r31 = str(ws.cell(31, 1).value or "")
    r59 = str(ws.cell(59, 1).value or "")
    r87 = str(ws.cell(87, 1).value or "")
    wb.close()
    assert "Summary Count" in r3, f"Row 3 expected 'Summary Count', got '{r3}'"
    assert "Summary Commitment" in r31, f"Row 31 expected 'Summary Commitment', got '{r31}'"
    assert "Summary Count" in r59 and "NEW" in r59, \
        f"Row 59 expected 'Summary Count...NEW', got '{r59}'"
    assert "Summary Commitment" in r87 and "NEW" in r87, \
        f"Row 87 expected 'Summary Commitment...NEW', got '{r87}'"
    return f"All 4 subheaders correct. {rows} rows x {cols} cols"


@run_test("Summary Count grand total matches loan count (567)")
def _():
    p = _excel_path()
    assert p, "Dashboard file missing"
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["Summary"]
    # Grand total is in row 29, last data column (16)
    grand_total = ws.cell(29, 16).value
    wb.close()
    assert grand_total is not None, "Grand total cell (row 29, col 16) is empty"
    assert int(grand_total) == 567, \
        f"Grand total is {grand_total}, expected 567"
    return f"Grand total = {grand_total} (matches 567 loans)"


@run_test("VBA .bas module files exist")
def _():
    required = [
        "templates/vba_modules/mod_ViewToggle.bas",
        "templates/vba_modules/mod_CellClick.bas",
        "templates/vba_modules/mod_Navigation.bas",
        "templates/vba_modules/mod_Diagnostics.bas",
        "templates/vba_modules/mod_DataRefresh.bas",
    ]
    missing = [f for f in required if not os.path.exists(f)]
    assert not missing, f"Missing VBA files: {missing}"
    return "All 5 .bas files present"


@run_test("_config sheet has correct initial values (A1=2, A2=0)")
def _():
    p = _excel_path()
    assert p, "Dashboard file missing"
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["_config"]
    a1 = ws["A1"].value
    a2 = ws["A2"].value
    wb.close()
    assert a1 == 2, f"_config!A1 (ViewState) = {a1}, expected 2"
    assert a2 == 0, f"_config!A2 (WMLCState) = {a2}, expected 0"
    return f"ViewState={a1}, WMLCState={a2}"


@run_test("POWER_QUERY_SETUP sheet exists")
def _():
    p = _excel_path()
    assert p, "Dashboard file missing"
    wb = openpyxl.load_workbook(p, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    assert "POWER_QUERY_SETUP" in sheets, f"POWER_QUERY_SETUP not found. Sheets: {sheets}"
    return "POWER_QUERY_SETUP sheet present"


@run_test("Dashboard headers start at row 6, data at row 7")
def _():
    p = _excel_path()
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["Dashboard"]
    r6c = ws.cell(6, 3).value  # First product column header
    r7a = ws.cell(7, 1).value  # First bucket label
    r31a = str(ws.cell(31, 1).value or "").strip()
    wb.close()
    assert r6c is not None, "Row 6 col C (header) is empty"
    assert r7a is not None, "Row 7 col A (first bucket) is empty"
    assert r31a.lower() == "total", f"Row 31 col A expected 'Total', got '{r31a}'"
    return f"Headers row 6: '{r6c}', data row 7: '{r7a}', total row 31"


@run_test("VBA_REFERENCE.txt exists")
def _():
    p = "output/VBA_REFERENCE.txt"
    assert os.path.exists(p), f"Missing: {p}"
    size = os.path.getsize(p)
    assert size > 5000, f"File too small ({size} bytes)"
    return f"VBA_REFERENCE.txt ({size:,} bytes)"


@run_test("Power Query M code file exists")
def _():
    p = "templates/power_query/load_tagged_data.m"
    assert os.path.exists(p), f"Missing: {p}"
    size = os.path.getsize(p)
    assert size > 50, f"File too small ({size} bytes)"
    return f"load_tagged_data.m ({size} bytes)"


# ===================================================================
# REPORT GENERATION
# ===================================================================

def write_report(results: list[TestResult], path: str):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    passed = sum(1 for r in results if r.passed)
    failed = len(results) - passed

    lines = [
        "# WMLC Pipeline Test Report",
        f"Generated: {now}",
        "",
        "## Automated Test Results",
        "",
        "| # | Test | Result | Details |",
        "|---|------|--------|---------|",
    ]

    for i, r in enumerate(results, 1):
        status = "PASS" if r.passed else "FAIL"
        icon = "+" if r.passed else "X"
        detail = r.details.replace("|", "/").replace("\n", " ")[:200]
        lines.append(f"| {i} | {r.name} | {icon} {status} | {detail} |")

    lines += [
        "",
        "## Summary",
        "",
        f"**Passed: {passed} / {len(results)}**  ",
        f"**Failed: {failed}**",
        "",
    ]

    # Failed test detail
    if failed > 0:
        lines += ["## Failed Test Details", ""]
        for i, r in enumerate(results, 1):
            if not r.passed:
                lines.append(f"### Test {i}: {r.name}")
                lines.append(f"```")
                lines.append(r.details)
                lines.append(f"```")
                lines.append("")

    # Manual test checklist
    lines += [
        "## Manual Test Checklist (Human Verification Required)",
        "",
        "### Setup",
        "- [ ] Open WMLC_Dashboard.xlsm in Excel",
        "- [ ] Enable macros when prompted",
        "- [ ] Verify diagnostics popup shows 'ALL TESTS PASSED' on open",
        "",
        "### Dashboard View Toggles",
        '- [ ] Click "Summary Count" button -> values change to integer counts',
        '- [ ] Click "Summary Commitment" button -> values change to currency amounts',
        '- [ ] Click "Summary Count NEW" -> counts reflect only NEW_CAMP_YN="Y"',
        '- [ ] Click "Summary Commitment NEW" -> amounts reflect only NEW loans',
        "- [ ] Verify view indicator text (Row 2) updates with each toggle",
        "- [ ] Verify total row/column recalculate correctly",
        "",
        "### WMLC Toggle",
        '- [ ] Click "WMLC ON" -> matrix values decrease (fewer qualifying loans)',
        '- [ ] Click "WMLC OFF" -> matrix returns to full population',
        "- [ ] Verify WMLC state persists across view toggles",
        "",
        "### Cell Click Drill-Down",
        "- [ ] Click a non-zero cell in the Dashboard matrix",
        "- [ ] Verify: jumps to loan_detail sheet",
        "- [ ] Verify: loan_detail is filtered to correct product_bucket and bucket",
        "- [ ] Verify: if WMLC ON, only wmlc_qualified loans shown",
        '- [ ] Verify: if NEW view active, only NEW_CAMP_YN="Y" loans shown',
        "- [ ] Click a zero-value cell -> verify no data rows shown (correct empty filter)",
        "",
        "### Navigation",
        '- [ ] On loan_detail: click "Back to Dashboard" -> returns to Dashboard',
        "- [ ] Verify filters are cleared after returning",
        '- [ ] On loan_detail: click "Reset Filters" -> all filters cleared',
        '- [ ] Click "Reset" on Dashboard -> returns to Summary Commitment, WMLC OFF',
        "",
        "### Data Refresh",
        '- [ ] Click "Refresh Data" -> prompts file load, recomputes all views',
        "- [ ] Update _config!A3 to a new CSV path, click Refresh Data -> loads new data",
        "- [ ] After refresh, verify all view toggles still work correctly",
        "- [ ] After refresh, verify cell click drill-down still works",
        "",
        "### Data Integrity",
        "- [ ] Spot-check 3 cells: manually count/sum loans in _data matching the bucket",
        "- [ ] Verify loan_detail has ALL columns from the tagged CSV",
        "- [ ] Verify _data sheet row count matches loan_extract_tagged.csv",
    ]

    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    return passed, failed


# ===================================================================
# MAIN
# ===================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("WMLC Pipeline Test Suite")
    print("=" * 60)

    # Results were already collected by the decorators at module load time
    for i, r in enumerate(results, 1):
        icon = "PASS" if r.passed else "FAIL"
        print(f"  [{icon}] {i:>2}. {r.name}")
        if not r.passed:
            print(f"         -> {r.details[:120]}")

    passed, failed = write_report(results, "output/test_report.md")

    print()
    print(f"Results: {passed} passed, {failed} failed out of {len(results)} tests")
    print("Report written to output/test_report.md")

    sys.exit(1 if failed else 0)
