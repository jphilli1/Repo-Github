"""Validate WMLC_Dashboard.xlsm structure and content."""
import openpyxl
import zipfile
import os
import re
import sys

path = "output/WMLC_Dashboard.xlsm"
errors = []

# 1. File exists and is valid ZIP
if not os.path.exists(path):
    path_xlsx = "output/WMLC_Dashboard.xlsx"
    if os.path.exists(path_xlsx):
        print(f"NOTE: .xlsm not found, validating .xlsx instead: {path_xlsx}")
        path = path_xlsx
    else:
        print(f"FAIL: Neither {path} nor {path_xlsx} exists")
        sys.exit(1)

if not zipfile.is_zipfile(path):
    errors.append("Not a valid ZIP/XLSM file")

# 2. VBA present (only for .xlsm)
if path.endswith(".xlsm"):
    with zipfile.ZipFile(path) as z:
        if "xl/vbaProject.bin" not in z.namelist():
            errors.append("No vbaProject.bin -- VBA macros missing")

# 3. All required sheets
wb = openpyxl.load_workbook(path, data_only=True)
required_sheets = {
    "Dashboard", "Loan Detail", "_config", "_chart_data",
    "_view1", "_view2", "_view3", "_view4",
    "_view5", "_view6", "_view7", "_view8",
    "Summary", "POWER_QUERY_SETUP", "Tracker Analytics"
}
actual = set(wb.sheetnames)
missing = required_sheets - actual
if missing:
    errors.append(f"Missing sheets: {missing}")

# 4. Dashboard dimensions
if "Dashboard" in actual:
    ws = wb["Dashboard"]
    if ws.max_row < 30:
        errors.append(f"Dashboard too short: {ws.max_row} rows")
    if ws.max_column < 16:
        errors.append(f"Dashboard too narrow: {ws.max_column} cols")

# 5. Charts on Dashboard
dash_charts = 0
if "Dashboard" in actual:
    ws = wb["Dashboard"]
    dash_charts = len(ws._charts)
    if dash_charts < 1:
        errors.append(f"Dashboard has {dash_charts} charts, expected at least 1")

# 6. Tracker Analytics charts (count only — size checked via XML below)
ta_charts = 0
if "Tracker Analytics" in actual:
    ws_ta = wb["Tracker Analytics"]
    ta_charts = len(ws_ta._charts)
    if ta_charts < 3:
        errors.append(f"Tracker Analytics has {ta_charts} charts, expected at least 3")

# 7. Chart dimensions — read from raw drawing XML (openpyxl doesn't round-trip these)
MIN_WIDTH_CM = 30
MIN_HEIGHT_CM = 10
# Match both <xdr:ext cx="..." cy="..."/> and <ext cx="..." cy="..."/>
# (COM-saved files use xdr: prefix, openpyxl-saved files omit it)
EXT_PATTERN = re.compile(r'<(?:xdr:)?ext\s+cx="(\d+)"\s+cy="(\d+)"')
chart_sizes = []
with zipfile.ZipFile(path) as z:
    for entry in sorted(z.namelist()):
        if "drawings/drawing" in entry and entry.endswith(".xml"):
            content = z.read(entry).decode("utf-8")
            extents = EXT_PATTERN.findall(content)
            for idx, (cx, cy) in enumerate(extents):
                w_cm = int(cx) / 360000
                h_cm = int(cy) / 360000
                chart_label = f"{entry} chart {idx+1}"
                chart_sizes.append((chart_label, w_cm, h_cm))
                if w_cm < MIN_WIDTH_CM:
                    errors.append(f"{chart_label} too narrow: {w_cm:.1f}cm (min {MIN_WIDTH_CM})")
                if h_cm < MIN_HEIGHT_CM:
                    errors.append(f"{chart_label} too short: {h_cm:.1f}cm (min {MIN_HEIGHT_CM})")

# 8. _config defaults
if "_config" in actual:
    cfg = wb["_config"]
    if cfg["A1"].value is None:
        errors.append("_config!A1 (ViewState) is empty")
    if cfg["A2"].value is None:
        errors.append("_config!A2 (WMLCState) is empty")

# 9. _chart_data completeness
if "_chart_data" in actual:
    wsc = wb["_chart_data"]
    checks = {
        "Stacked bar (A1)": wsc["A1"].value,
        "Donut WMLC (A41)": wsc["A41"].value,
        "Donut Non-WMLC (A42)": wsc["A42"].value,
        "Flag breakdown (A45)": wsc["A45"].value,
        "Proximity (A58)": wsc["A58"].value,
    }
    for name, val in checks.items():
        if val is None:
            errors.append(f"_chart_data {name} is empty")

wb.close()

# Report
if errors:
    print(f"VALIDATION FAILED -- {len(errors)} errors:")
    for e in errors:
        print(f"  FAIL: {e}")
    sys.exit(1)
else:
    print("ALL VALIDATION CHECKS PASSED")
    print(f"  Sheets: {len(actual)}")
    print(f"  Dashboard charts: {dash_charts}")
    print(f"  Tracker Analytics charts: {ta_charts}")
    for label, w, h in chart_sizes:
        print(f"  {label}: {w:.1f}x{h:.1f}cm")
    sys.exit(0)
