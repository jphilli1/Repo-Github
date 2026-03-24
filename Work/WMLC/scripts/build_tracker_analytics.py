"""
build_tracker_analytics.py — Tracker Analytics sheet builder for WMLC Dashboard.

Reads proxy_data/WMLC_Tracker.xlsx and builds an openpyxl worksheet named
"Tracker Analytics" with:
  - WMLC Flag Heatmap by Quarter
  - Count by Quarter (grouped bar chart)
  - Gross $ by Quarter (grouped bar chart)
  - Net $ by Quarter (grouped bar chart)
  - Deal Type by Quarter (grouped bar chart)
  - Deal Type by Year (grouped bar chart)

Called from build_dashboard.py before saving the workbook.
"""

import os
import logging
import pandas as pd
from copy import copy

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.text import RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import (
    Paragraph as DrawParagraph,
    ParagraphProperties as DrawParagraphProperties,
    CharacterProperties,
    Font as DrawingFont,
)

logger = logging.getLogger("wmlc_etl.tracker_analytics")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_TRACKER_PATH = os.path.join(REPO, "proxy_data", "WMLC_Tracker.xlsx")

# ── MS Color Palette ──────────────────────────────────────────────────────────
MS_COLORS = {
    'navy':        '002B5C',
    'blue':        '00539B',
    'light_blue':  'D6E4F0',
    'ice_blue':    'F0F4F8',
    'button_gray': 'F4F5F7',
    'border_gray': 'D9DEE3',
    'text_gray':   '6C757D',
    'white':       'FFFFFF',
}

# ── Product area sets ─────────────────────────────────────────────────────────
COUNT_PRODUCT_AREAS = {"LAL", "TL-CRE", "TL-LIQ", "TL-LIC", "TL-ALTS", "DD-MSSB", "FA-MSSB", "PSL"}
DOLLAR_PRODUCT_AREAS = {"LAL", "TL-CRE", "TL-LIQ", "TL-LIC", "TL-ALTS", "PSL"}

# ── Chart bar colors (one per product area, cycling) ──────────────────────────
CHART_COLORS = [
    "002B5C", "00539B", "0077B6", "48CAE4", "D4A017", "C8102E", "6C757D", "2E8B57"
]

DEAL_TYPE_COLORS = [
    "002B5C", "00539B", "0077B6", "48CAE4", "D4A017", "C8102E", "6C757D", "2E8B57"
]

# ── Style constants ───────────────────────────────────────────────────────────
THIN_GRAY = Side(style="thin", color=MS_COLORS['border_gray'])
THIN_GRAY_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
DOUBLE_NAVY_TOP = Border(
    left=THIN_GRAY, right=THIN_GRAY,
    top=Side(style="double", color=MS_COLORS['navy']),
    bottom=THIN_GRAY,
)

NAVY_FILL = PatternFill(start_color=MS_COLORS['navy'], end_color=MS_COLORS['navy'], fill_type="solid")
BLUE_FILL = PatternFill(start_color=MS_COLORS['blue'], end_color=MS_COLORS['blue'], fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color=MS_COLORS['light_blue'], end_color=MS_COLORS['light_blue'], fill_type="solid")
ICE_BLUE_FILL = PatternFill(start_color=MS_COLORS['ice_blue'], end_color=MS_COLORS['ice_blue'], fill_type="solid")
WHITE_FILL = PatternFill(start_color=MS_COLORS['white'], end_color=MS_COLORS['white'], fill_type="solid")

TITLE_FONT = Font(bold=True, size=14, color=MS_COLORS['white'], name="Calibri")
SUBTITLE_FONT = Font(size=11, color=MS_COLORS['white'], name="Calibri")
SECTION_FONT = Font(bold=True, size=12, color=MS_COLORS['white'], name="Calibri")
HEADER_FONT = Font(bold=True, size=10, color=MS_COLORS['white'], name="Calibri")
DATA_FONT = Font(size=10, name="Calibri")
DATA_FONT_BOLD = Font(bold=True, size=10, name="Calibri")
TOTAL_FONT = Font(bold=True, size=10, color=MS_COLORS['navy'], name="Calibri")
HEATMAP_ZERO_FMT = '#,##0;-#,##0;"-"'
DOLLAR_FMT = '$#,##0;-$#,##0;"-"'
COUNT_FMT = '#,##0;-#,##0;"-"'


# ── Helpers ───────────────────────────────────────────────────────────────────

def _heatmap_fill(value, max_val):
    """Return a PatternFill on white→light_blue→navy scale based on value/max_val."""
    if max_val <= 0 or value <= 0:
        return WHITE_FILL
    ratio = min(value / max_val, 1.0)
    # Interpolate from white (255,255,255) through light_blue (214,228,240) to navy (0,43,92)
    if ratio <= 0.5:
        t = ratio / 0.5
        r = int(255 + t * (214 - 255))
        g = int(255 + t * (228 - 255))
        b = int(255 + t * (240 - 255))
    else:
        t = (ratio - 0.5) / 0.5
        r = int(214 + t * (0 - 214))
        g = int(228 + t * (43 - 228))
        b = int(240 + t * (92 - 240))
    hex_color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font_for_heatmap(value, max_val):
    """Dark text for light cells, white text for dark cells."""
    if max_val <= 0 or value <= 0:
        return DATA_FONT
    ratio = min(value / max_val, 1.0)
    if ratio > 0.6:
        return Font(size=10, name="Calibri", color=MS_COLORS['white'])
    return DATA_FONT


def _write_section_header(ws, row, text, max_col):
    """Write a navy-fill section header merged across columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = NAVY_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for ci in range(1, max_col + 1):
        ws.cell(row=row, column=ci).fill = NAVY_FILL
    ws.row_dimensions[row].height = 24


def _write_data_table(ws, start_row, row_labels, col_labels, matrix, num_fmt, show_total=True):
    """Write a data table with headers, alternating rows, and optional total row.

    Returns the row number AFTER the table (next free row).
    """
    max_col = 1 + len(col_labels)

    # Header row
    hr = start_row
    cell_a = ws.cell(row=hr, column=1, value="")
    cell_a.font = HEADER_FONT
    cell_a.fill = BLUE_FILL
    cell_a.border = THIN_GRAY_BORDER
    for ci, label in enumerate(col_labels, 2):
        cell = ws.cell(row=hr, column=ci, value=label)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.border = THIN_GRAY_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    dr = hr + 1
    for ri, row_label in enumerate(row_labels):
        is_alt = (ri % 2 == 1)
        fill = ICE_BLUE_FILL if is_alt else WHITE_FILL

        cell_a = ws.cell(row=dr, column=1, value=row_label)
        cell_a.font = DATA_FONT_BOLD
        cell_a.border = THIN_GRAY_BORDER
        cell_a.fill = fill
        cell_a.alignment = Alignment(horizontal="left")

        for ci, val in enumerate(matrix[ri], 2):
            cell = ws.cell(row=dr, column=ci, value=val)
            cell.number_format = num_fmt
            cell.border = THIN_GRAY_BORDER
            cell.font = DATA_FONT
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right")

        dr += 1

    # Total row
    if show_total:
        cell_t = ws.cell(row=dr, column=1, value="Total")
        cell_t.font = TOTAL_FONT
        cell_t.fill = LIGHT_BLUE_FILL
        cell_t.border = DOUBLE_NAVY_TOP

        for ci in range(2, max_col + 1):
            col_total = sum(
                matrix[ri][ci - 2] for ri in range(len(row_labels))
            )
            cell = ws.cell(row=dr, column=ci, value=col_total)
            cell.font = TOTAL_FONT
            cell.number_format = num_fmt
            cell.fill = LIGHT_BLUE_FILL
            cell.border = DOUBLE_NAVY_TOP
            cell.alignment = Alignment(horizontal="right")
        dr += 1

    return dr


def _write_heatmap_table(ws, start_row, row_labels, col_labels, matrix):
    """Write a heatmap table with color-scaled cells. Returns next free row."""
    max_col = 1 + len(col_labels)
    all_vals = [v for row in matrix for v in row]
    max_val = max(all_vals) if all_vals else 0

    # Header row
    hr = start_row
    cell_a = ws.cell(row=hr, column=1, value="Flag")
    cell_a.font = HEADER_FONT
    cell_a.fill = BLUE_FILL
    cell_a.border = THIN_GRAY_BORDER
    for ci, label in enumerate(col_labels, 2):
        cell = ws.cell(row=hr, column=ci, value=label)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.border = THIN_GRAY_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    dr = hr + 1
    for ri, row_label in enumerate(row_labels):
        cell_a = ws.cell(row=dr, column=1, value=row_label)
        cell_a.font = DATA_FONT_BOLD
        cell_a.border = THIN_GRAY_BORDER
        cell_a.alignment = Alignment(horizontal="left")

        for ci, val in enumerate(matrix[ri], 2):
            cell = ws.cell(row=dr, column=ci, value=val)
            cell.number_format = HEATMAP_ZERO_FMT
            cell.border = THIN_GRAY_BORDER
            cell.fill = _heatmap_fill(val, max_val)
            cell.font = _font_for_heatmap(val, max_val)
            cell.alignment = Alignment(horizontal="right")

        dr += 1

    # Total row
    cell_t = ws.cell(row=dr, column=1, value="Total")
    cell_t.font = TOTAL_FONT
    cell_t.fill = LIGHT_BLUE_FILL
    cell_t.border = DOUBLE_NAVY_TOP

    for ci in range(2, max_col + 1):
        col_total = sum(matrix[ri][ci - 2] for ri in range(len(row_labels)))
        cell = ws.cell(row=dr, column=ci, value=col_total)
        cell.font = TOTAL_FONT
        cell.number_format = HEATMAP_ZERO_FMT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = DOUBLE_NAVY_TOP
        cell.alignment = Alignment(horizontal="right")
    dr += 1

    return dr


def _make_bar_chart(ws, title, data_start_row, data_end_row, n_categories, n_series,
                    series_labels, color_list, anchor_cell):
    """Create a grouped bar chart anchored at anchor_cell.

    data layout: row data_start_row has headers (category labels in cols 2..n_categories+1)
                 rows data_start_row+1..data_end_row-1 have series data
                 col 1 has series names
    """
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.width = 38
    chart.height = 14
    chart.gapWidth = 60

    # Data reference: series in rows (each row = one series)
    data_ref = Reference(ws, min_col=2, min_row=data_start_row,
                         max_col=1 + n_categories, max_row=data_start_row + n_series)
    cats_ref = Reference(ws, min_col=2, min_row=data_start_row,
                         max_col=1 + n_categories)

    chart.add_data(data_ref, from_rows=True, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Apply colors to series
    for idx, s in enumerate(chart.series):
        color = color_list[idx % len(color_list)]
        s.graphicalProperties.solidFill = color

    # No gridlines
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None

    # Transparent plot and chart area
    chart.plot_area.graphicalProperties = GraphicalProperties()
    chart.plot_area.graphicalProperties.noFill = True

    # Axis label styling: Calibri 8pt, MS gray
    axis_font_props = CharacterProperties(
        sz=800, latin=DrawingFont(typeface="Calibri"),
        solidFill=MS_COLORS['text_gray']
    )

    # Y-axis tick label properties
    chart.y_axis.txPr = RichText(
        p=[DrawParagraph(
            pPr=DrawParagraphProperties(defRPr=axis_font_props),
            endParaRPr=axis_font_props
        )]
    )

    # X-axis tick label properties with rotation
    x_axis_font = CharacterProperties(
        sz=800, latin=DrawingFont(typeface="Calibri"),
        solidFill=MS_COLORS['text_gray']
    )
    chart.x_axis.txPr = RichText(
        p=[DrawParagraph(
            pPr=DrawParagraphProperties(defRPr=x_axis_font),
            endParaRPr=x_axis_font
        )]
    )
    chart.x_axis.tickLblPos = "low"

    # Title styling: Calibri 12pt, bold, navy
    title_font = CharacterProperties(
        sz=1200, b=True,
        latin=DrawingFont(typeface="Calibri"),
        solidFill=MS_COLORS['navy']
    )
    chart.title.txPr = RichText(
        p=[DrawParagraph(
            pPr=DrawParagraphProperties(defRPr=title_font),
            endParaRPr=title_font
        )]
    )

    # Legend at bottom
    chart.legend.position = 'b'

    ws.add_chart(chart, anchor_cell)
    return chart


def _divider_row(ws, row, max_col):
    """Write a thin navy divider row."""
    for ci in range(1, max_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = NAVY_FILL
    ws.row_dimensions[row].height = 4


# ── Main builder ──────────────────────────────────────────────────────────────

def build_tracker_analytics_sheet(wb, tracker_path=None):
    """Build the 'Tracker Analytics' sheet into an existing openpyxl Workbook.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The workbook to add the sheet to.
    tracker_path : str or None
        Path to WMLC_Tracker.xlsx. Defaults to proxy_data/WMLC_Tracker.xlsx.
    """
    if tracker_path is None:
        tracker_path = DEFAULT_TRACKER_PATH

    logger.info("Building Tracker Analytics sheet from %s", tracker_path)

    if not os.path.exists(tracker_path):
        logger.warning("Tracker file not found: %s — skipping Tracker Analytics", tracker_path)
        return

    # ── Load tracker data ─────────────────────────────────────────────────
    tracker = pd.read_excel(tracker_path, engine="openpyxl")
    logger.info("Loaded %d rows, %d columns from tracker", len(tracker), len(tracker.columns))

    if tracker.empty:
        logger.warning("Tracker is empty — skipping Tracker Analytics")
        return

    # ── Parse date column ─────────────────────────────────────────────────
    date_col = None
    for col in tracker.columns:
        if "WMLC" in col.upper() and "DATE" in col.upper():
            date_col = col
            break
    if date_col is None:
        # Fallback: first column containing "date"
        for col in tracker.columns:
            if "DATE" in col.upper():
                date_col = col
                break
    if date_col is None:
        logger.warning("No date column found — skipping Tracker Analytics")
        return

    tracker["_DATE"] = pd.to_datetime(tracker[date_col], errors="coerce")
    tracker = tracker.dropna(subset=["_DATE"]).copy()
    tracker["_QUARTER"] = tracker["_DATE"].dt.to_period("Q").astype(str)
    tracker["_YEAR"] = tracker["_DATE"].dt.year.astype(str)

    # ── Find product area column ──────────────────────────────────────────
    prod_area_col = None
    for col in tracker.columns:
        if "PRODUCT" in col.upper() and "AREA" in col.upper():
            prod_area_col = col
            break
    if prod_area_col is None:
        logger.warning("No Product Area column found — skipping Tracker Analytics")
        return

    # ── Find gross / net columns ──────────────────────────────────────────
    gross_col = None
    net_col = None
    for col in tracker.columns:
        cu = col.upper()
        if "GROSS" in cu and ("SIZE" in cu or "TRANSACTION" in cu):
            gross_col = col
        if "NET" in cu and ("SIZE" in cu or "TRANSACTION" in cu):
            net_col = col
    if gross_col is None:
        for col in tracker.columns:
            if "GROSS" in col.upper():
                gross_col = col
                break
    if net_col is None:
        for col in tracker.columns:
            if "NET" in col.upper():
                net_col = col
                break

    if gross_col:
        tracker[gross_col] = pd.to_numeric(tracker[gross_col], errors="coerce").fillna(0)
    if net_col:
        tracker[net_col] = pd.to_numeric(tracker[net_col], errors="coerce").fillna(0)

    # ── Find deal type column ─────────────────────────────────────────────
    deal_type_col = None
    for col in tracker.columns:
        if "DEAL" in col.upper() and "TYPE" in col.upper():
            deal_type_col = col
            break

    # ── Identify flag columns ─────────────────────────────────────────────
    skip_cols = {
        "_DATE", "_QUARTER", "_YEAR", "_BORROWER_KEY", "_BORROWER_KEY_NORM",
        "_FA_MSSB_COMBINED", "_DD_MSSB_COMBINED",
    }
    non_flag_cols = {
        date_col, prod_area_col,
        "Tracker File V", "Relationship Name", "Product Type",
        "Transaction Size (MM) Gross", "Transaction Size Net",
    }
    if gross_col:
        non_flag_cols.add(gross_col)
    if net_col:
        non_flag_cols.add(net_col)
    if deal_type_col:
        non_flag_cols.add(deal_type_col)

    actual_flag_cols = []
    for col in tracker.columns:
        if col in non_flag_cols or col in skip_cols or col.startswith("_"):
            continue
        # Check if column has any "Y" values — it's a flag column
        vals = tracker[col].dropna().astype(str).str.strip().str.upper()
        if (vals == "Y").any():
            actual_flag_cols.append(col)

    logger.info("Found %d flag columns", len(actual_flag_cols))

    # ── A2: Combine FA- and DD- prefixed flags ────────────────────────────
    fa_columns = [col for col in actual_flag_cols if col.upper().startswith("FA")]
    if fa_columns:
        tracker["_FA_MSSB_COMBINED"] = tracker[fa_columns].apply(
            lambda row: "Y" if any(str(v).strip().upper() == "Y" for v in row) else "N",
            axis=1,
        )
        actual_flag_cols = [c for c in actual_flag_cols if not c.upper().startswith("FA")]
        actual_flag_cols.append("_FA_MSSB_COMBINED")
        logger.info("Combined %d FA columns into _FA_MSSB_COMBINED", len(fa_columns))

    dd_columns = [col for col in actual_flag_cols if col.upper().startswith("DD") and col != "_DD_MSSB_COMBINED"]
    if dd_columns:
        tracker["_DD_MSSB_COMBINED"] = tracker[dd_columns].apply(
            lambda row: "Y" if any(str(v).strip().upper() == "Y" for v in row) else "N",
            axis=1,
        )
        actual_flag_cols = [c for c in actual_flag_cols if not (c.upper().startswith("DD") and c != "_DD_MSSB_COMBINED")]
        actual_flag_cols.append("_DD_MSSB_COMBINED")
        logger.info("Combined %d DD columns into _DD_MSSB_COMBINED", len(dd_columns))

    # Display names for combined flags
    flag_display_names = {}
    for fc in actual_flag_cols:
        if fc == "_FA_MSSB_COMBINED":
            flag_display_names[fc] = "FA-MSSB (Combined)"
        elif fc == "_DD_MSSB_COMBINED":
            flag_display_names[fc] = "DD-MSSB (Combined)"
        else:
            flag_display_names[fc] = fc

    # ── A4: Dedup — same borrower in same quarter, keep LATEST ────────────
    tracker.sort_values("_DATE", ascending=False, inplace=True)
    tracker["_BORROWER_KEY"] = tracker.get("Tracker File V", pd.Series(dtype=str)).fillna("")
    mask_empty = tracker["_BORROWER_KEY"].str.strip() == ""
    if "Relationship Name" in tracker.columns:
        tracker.loc[mask_empty, "_BORROWER_KEY"] = tracker.loc[mask_empty, "Relationship Name"].fillna("")
    tracker["_BORROWER_KEY_NORM"] = tracker["_BORROWER_KEY"].str.upper().str.strip()
    n_before = len(tracker)
    tracker_deduped = tracker.drop_duplicates(subset=["_BORROWER_KEY_NORM", "_QUARTER"], keep="first").copy()
    n_after = len(tracker_deduped)
    logger.info("Dedup: %d -> %d rows (removed %d duplicates)", n_before, n_after, n_before - n_after)

    # ── Compute quarters (sorted) ─────────────────────────────────────────
    quarters = sorted(tracker_deduped["_QUARTER"].unique())
    years = sorted(tracker_deduped["_YEAR"].unique())
    logger.info("Quarters: %s", quarters)
    logger.info("Years: %s", years)

    # ── Build worksheet ───────────────────────────────────────────────────
    ws = wb.create_sheet("Tracker Analytics")

    # Determine max columns needed
    max_data_cols = max(len(quarters), len(years), 1)
    max_col = 1 + max_data_cols

    # Column widths
    ws.column_dimensions["A"].width = 32
    for ci in range(2, max_col + 5):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    # ── Row 1: Title ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value="WMLC Tracker Analytics")
    cell.font = TITLE_FONT
    cell.fill = NAVY_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for ci in range(1, max_col + 1):
        ws.cell(row=1, column=ci).fill = NAVY_FILL
    ws.row_dimensions[1].height = 30

    # ── Row 2: Subtitle with date range ───────────────────────────────────
    date_min = tracker_deduped["_DATE"].min().strftime("%b %Y")
    date_max = tracker_deduped["_DATE"].max().strftime("%b %Y")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    cell = ws.cell(row=2, column=1, value=f"Data Range: {date_min} - {date_max}  |  {n_after} deals (deduped from {n_before})")
    cell.font = SUBTITLE_FONT
    cell.fill = BLUE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for ci in range(1, max_col + 1):
        ws.cell(row=2, column=ci).fill = BLUE_FILL
    ws.row_dimensions[2].height = 22

    current_row = 4

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 1: WMLC Flag Heatmap by Quarter
    # ══════════════════════════════════════════════════════════════════════
    hm_max_col = 1 + len(quarters)
    _write_section_header(ws, current_row, "WMLC Flag Heatmap by Quarter", hm_max_col)
    current_row += 1

    # Build heatmap matrix: rows = flags, cols = quarters
    hm_row_labels = [flag_display_names[fc] for fc in actual_flag_cols]
    hm_matrix = []
    for fc in actual_flag_cols:
        row_vals = []
        for q in quarters:
            q_mask = tracker_deduped["_QUARTER"] == q
            is_flagged = tracker_deduped.loc[q_mask, fc].astype(str).str.strip().str.upper() == "Y"
            row_vals.append(int(is_flagged.sum()))
        hm_matrix.append(row_vals)

    current_row = _write_heatmap_table(ws, current_row, hm_row_labels, quarters, hm_matrix)
    current_row += 1

    # Divider
    _divider_row(ws, current_row, hm_max_col)
    current_row += 2

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 2: Count by Quarter (Chart 1)
    # ══════════════════════════════════════════════════════════════════════
    count_areas = sorted([a for a in COUNT_PRODUCT_AREAS if a in tracker_deduped[prod_area_col].unique()])
    ct_max_col = 1 + len(quarters)
    _write_section_header(ws, current_row, "Deal Count by Quarter", ct_max_col)
    current_row += 1

    ct_matrix = []
    for area in count_areas:
        row_vals = []
        for q in quarters:
            mask = (tracker_deduped[prod_area_col] == area) & (tracker_deduped["_QUARTER"] == q)
            row_vals.append(int(mask.sum()))
        ct_matrix.append(row_vals)

    ct_data_start = current_row
    current_row = _write_data_table(ws, current_row, count_areas, quarters, ct_matrix, COUNT_FMT)

    # Chart 1
    chart1_anchor = f"A{current_row + 1}"
    n_series_1 = len(count_areas)
    _make_bar_chart(ws, "Deal Count by Quarter", ct_data_start, ct_data_start + n_series_1,
                    len(quarters), n_series_1, count_areas, CHART_COLORS, chart1_anchor)
    current_row += 30  # space for chart

    # Divider
    _divider_row(ws, current_row, ct_max_col)
    current_row += 2

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 3: Gross $ by Quarter (Chart 2) — DOLLAR_PRODUCT_AREAS only
    # ══════════════════════════════════════════════════════════════════════
    dollar_areas = sorted([a for a in DOLLAR_PRODUCT_AREAS if a in tracker_deduped[prod_area_col].unique()])

    if gross_col:
        _write_section_header(ws, current_row, "Gross Transaction Size ($MM) by Quarter", ct_max_col)
        current_row += 1

        gross_matrix = []
        for area in dollar_areas:
            row_vals = []
            for q in quarters:
                mask = (tracker_deduped[prod_area_col] == area) & (tracker_deduped["_QUARTER"] == q)
                row_vals.append(round(float(tracker_deduped.loc[mask, gross_col].sum()), 1))
            gross_matrix.append(row_vals)

        gross_data_start = current_row
        current_row = _write_data_table(ws, current_row, dollar_areas, quarters, gross_matrix, DOLLAR_FMT)

        chart2_anchor = f"A{current_row + 1}"
        _make_bar_chart(ws, "Gross Transaction Size ($MM) by Quarter",
                        gross_data_start, gross_data_start + len(dollar_areas),
                        len(quarters), len(dollar_areas), dollar_areas, CHART_COLORS, chart2_anchor)
        current_row += 30

        _divider_row(ws, current_row, ct_max_col)
        current_row += 2

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 4: Net $ by Quarter (Chart 3) — DOLLAR_PRODUCT_AREAS only
    # ══════════════════════════════════════════════════════════════════════
    if net_col:
        _write_section_header(ws, current_row, "Net Transaction Size ($MM) by Quarter", ct_max_col)
        current_row += 1

        net_matrix = []
        for area in dollar_areas:
            row_vals = []
            for q in quarters:
                mask = (tracker_deduped[prod_area_col] == area) & (tracker_deduped["_QUARTER"] == q)
                row_vals.append(round(float(tracker_deduped.loc[mask, net_col].sum()), 1))
            net_matrix.append(row_vals)

        net_data_start = current_row
        current_row = _write_data_table(ws, current_row, dollar_areas, quarters, net_matrix, DOLLAR_FMT)

        chart3_anchor = f"A{current_row + 1}"
        _make_bar_chart(ws, "Net Transaction Size ($MM) by Quarter",
                        net_data_start, net_data_start + len(dollar_areas),
                        len(quarters), len(dollar_areas), dollar_areas, CHART_COLORS, chart3_anchor)
        current_row += 30

        _divider_row(ws, current_row, ct_max_col)
        current_row += 2

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 5: Deal Type by Quarter (Chart 4)
    # ══════════════════════════════════════════════════════════════════════
    if deal_type_col and deal_type_col in tracker_deduped.columns:
        deal_types = sorted(tracker_deduped[deal_type_col].dropna().unique().tolist())

        if deal_types:
            dt_max_col = 1 + len(quarters)
            _write_section_header(ws, current_row, "Deal Type Breakdown by Quarter", dt_max_col)
            current_row += 1

            dt_q_matrix = []
            for dt in deal_types:
                row_vals = []
                for q in quarters:
                    mask = (tracker_deduped[deal_type_col] == dt) & (tracker_deduped["_QUARTER"] == q)
                    row_vals.append(int(mask.sum()))
                dt_q_matrix.append(row_vals)

            dt_q_data_start = current_row
            current_row = _write_data_table(ws, current_row, deal_types, quarters, dt_q_matrix, COUNT_FMT)

            chart4_anchor = f"A{current_row + 1}"
            _make_bar_chart(ws, "Deal Type Breakdown by Quarter",
                            dt_q_data_start, dt_q_data_start + len(deal_types),
                            len(quarters), len(deal_types), deal_types, DEAL_TYPE_COLORS, chart4_anchor)
            current_row += 30

            _divider_row(ws, current_row, dt_max_col)
            current_row += 2

            # ══════════════════════════════════════════════════════════════
            # SECTION 6: Deal Type by Year (Chart 5)
            # ══════════════════════════════════════════════════════════════
            dt_y_max_col = 1 + len(years)
            _write_section_header(ws, current_row, "Deal Type Breakdown by Year", dt_y_max_col)
            current_row += 1

            dt_y_matrix = []
            for dt in deal_types:
                row_vals = []
                for y in years:
                    mask = (tracker_deduped[deal_type_col] == dt) & (tracker_deduped["_YEAR"] == y)
                    row_vals.append(int(mask.sum()))
                dt_y_matrix.append(row_vals)

            dt_y_data_start = current_row
            current_row = _write_data_table(ws, current_row, deal_types, years, dt_y_matrix, COUNT_FMT)

            chart5_anchor = f"A{current_row + 1}"
            _make_bar_chart(ws, "Deal Type Breakdown by Year",
                            dt_y_data_start, dt_y_data_start + len(deal_types),
                            len(years), len(deal_types), deal_types, DEAL_TYPE_COLORS, chart5_anchor)
            current_row += 30

    logger.info("Tracker Analytics sheet built — %d charts", len(ws._charts))
    return ws
