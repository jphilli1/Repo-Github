"""
build_dashboard.py — Sub-Agent 3: Excel Dashboard Builder

Reads output/loan_extract_tagged.csv and produces:
  - output/WMLC_Dashboard.xlsm  (data + formatting + named ranges + VBA + buttons)

Two-stage build:
  Stage 1: openpyxl builds the .xlsx with all data, formatting, sheets, named ranges
  Stage 2: COM automation opens the .xlsx, injects VBA modules, creates buttons, saves as .xlsm
"""

import os
import sys
import time
import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference, DoughnutChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.legend import Legend
from openpyxl.chart.text import RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.text import (
    Paragraph as DrawParagraph, ParagraphProperties as DrawParagraphProperties,
    CharacterProperties, Font as DrawingFont,
)

# -- paths ------------------------------------------------------------------
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TAGGED_CSV = os.path.join(REPO, "output", "loan_extract_tagged.csv")
OUTPUT_XLSX = os.path.join(REPO, "output", "WMLC_Dashboard.xlsx")
OUTPUT_XLSM = os.path.join(REPO, "output", "WMLC_Dashboard.xlsm")
VBA_DIR = os.path.join(REPO, "templates", "vba_modules")
PQ_FILE = os.path.join(REPO, "templates", "power_query", "load_tagged_data.m")

# -- MS Color Palette -------------------------------------------------------
MS_COLORS = {
    'navy':        '002B5C',
    'blue':        '00539B',
    'light_blue':  'D6E4F0',
    'ice_blue':    'F0F4F8',
    'button_gray': 'F4F5F7',
    'border_gray': 'D9DEE3',
    'text_gray':   '6C757D',
    'white':       'FFFFFF',
}

# -- constants ---------------------------------------------------------------
PRODUCT_BUCKETS = [
    "LAL Diversified", "LAL Highly Conc.", "LAL NFPs",
    "TL SBL Diversified", "TL SBL Highly Conc.", "TL Life Insurance",
    "TL CRE", "TL Unsecured", "TL Aircraft", "TL PHA",
    "TL Other Secured", "TL Multicollateral", "RESI",
]

# Bucket ladder: list of (label, floor, ceiling) -- descending order
BUCKET_LADDER = [
    ("$1,000,000,000",  1_000_000_000, float("inf")),
    ("$750,000,000",      750_000_000,  999_999_999.99),
    ("$700,000,000",      700_000_000,  749_999_999.99),
    ("$600,000,000",      600_000_000,  699_999_999.99),
    ("$500,000,000",      500_000_000,  599_999_999.99),
    ("$400,000,000",      400_000_000,  499_999_999.99),
    ("$350,000,000",      350_000_000,  399_999_999.99),
    ("$300,000,000",      300_000_000,  349_999_999.99),
    ("$250,000,000",      250_000_000,  299_999_999.99),
    ("$200,000,000",      200_000_000,  249_999_999.99),
    ("$175,000,000",      175_000_000,  199_999_999.99),
    ("$150,000,000",      150_000_000,  174_999_999.99),
    ("$125,000,000",      125_000_000,  149_999_999.99),
    ("$100,000,000",      100_000_000,  124_999_999.99),
    ("$75,000,000",        75_000_000,   99_999_999.99),
    ("$50,000,000",        50_000_000,   74_999_999.99),
    ("$40,000,000",        40_000_000,   49_999_999.99),
    ("$35,000,000",        35_000_000,   39_999_999.99),
    ("$30,000,000",        30_000_000,   34_999_999.99),
    ("$25,000,000",        25_000_000,   29_999_999.99),
    ("$20,000,000",        20_000_000,   24_999_999.99),
    ("$15,000,000",        15_000_000,   19_999_999.99),
    ("$10,000,001",        10_000_001,   14_999_999.99),
    ("$1",                          1,   10_000_000.99),
]

# Defined-range labels
def _range_label(floor, ceil):
    if ceil == float("inf"):
        return f"${floor:,.0f}+"
    return f"${floor:,.0f}-${ceil:,.2f}"


# -- style helpers -----------------------------------------------------------
THIN_GRAY = Side(style="thin", color=MS_COLORS['border_gray'])
THIN_GRAY_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
THIN_WHITE = Side(style="thin", color=MS_COLORS['white'])
THIN_WHITE_BORDER = Border(left=THIN_WHITE, right=THIN_WHITE, top=THIN_WHITE, bottom=THIN_WHITE)
DOUBLE_NAVY_TOP = Border(
    left=THIN_GRAY, right=THIN_GRAY,
    top=Side(style="double", color=MS_COLORS['navy']),
    bottom=THIN_GRAY
)

# Fills
NAVY_FILL = PatternFill(start_color=MS_COLORS['navy'], end_color=MS_COLORS['navy'], fill_type="solid")
BLUE_FILL = PatternFill(start_color=MS_COLORS['blue'], end_color=MS_COLORS['blue'], fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color=MS_COLORS['light_blue'], end_color=MS_COLORS['light_blue'], fill_type="solid")
ICE_BLUE_FILL = PatternFill(start_color=MS_COLORS['ice_blue'], end_color=MS_COLORS['ice_blue'], fill_type="solid")
BUTTON_GRAY_FILL = PatternFill(start_color=MS_COLORS['button_gray'], end_color=MS_COLORS['button_gray'], fill_type="solid")
WHITE_FILL = PatternFill(start_color=MS_COLORS['white'], end_color=MS_COLORS['white'], fill_type="solid")

# Fonts
TITLE_FONT = Font(bold=True, size=18, color=MS_COLORS['white'], name="Calibri")
SUBTITLE_FONT = Font(size=11, color=MS_COLORS['white'], name="Calibri")
INDICATOR_FONT = Font(size=10, color=MS_COLORS['text_gray'], name="Calibri")
HEADER_FONT = Font(bold=True, size=10, color=MS_COLORS['white'], name="Calibri")
DATA_FONT = Font(size=10, name="Calibri")
DATA_FONT_BOLD = Font(bold=True, size=10, name="Calibri")
ITALIC_9_GRAY = Font(italic=True, size=9, color=MS_COLORS['text_gray'], name="Calibri")
TOTAL_FONT = Font(bold=True, size=10, color=MS_COLORS['navy'], name="Calibri")
DETAIL_TITLE_FONT = Font(bold=True, size=14, color=MS_COLORS['white'], name="Calibri")
DETAIL_HEADER_FONT = Font(bold=True, size=10, color=MS_COLORS['white'], name="Calibri")
PQ_TITLE_FONT = Font(bold=True, size=14, color=MS_COLORS['white'], name="Calibri")
PQ_STEP_FONT = Font(size=11, name="Calibri")
PQ_CODE_FONT = Font(size=10, name="Consolas", color="333333")

# Legacy aliases for Summary sheet (keep compatible)
HEADER_FILL = BLUE_FILL
ALT_FILL = ICE_BLUE_FILL
TOTAL_FILL = LIGHT_BLUE_FILL
BOLD_14 = Font(bold=True, size=14, name="Calibri")
BOLD_12 = Font(bold=True, size=12, name="Calibri")
BOLD_11 = Font(bold=True, size=11, name="Calibri")
BOLD_ITALIC_11 = Font(bold=True, italic=True, size=11, name="Calibri")
ITALIC_9 = Font(italic=True, size=9, name="Calibri")
THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DOUBLE_TOP = Border(left=THIN, right=THIN, top=Side(style="double", color="000000"), bottom=THIN)


# -- pivot computation -------------------------------------------------------
def compute_pivot(df, agg, subset_mask=None):
    """Return a 24x13 matrix (list of lists).
    agg: 'count' or 'sum' (of credit_lii).
    subset_mask: boolean Series to pre-filter df.
    """
    sub = df if subset_mask is None else df[subset_mask]
    matrix = []
    for label, floor, ceil in BUCKET_LADDER:
        row_mask = (sub["CREDIT_LII"] >= floor) & (sub["CREDIT_LII"] <= ceil)
        row_data = []
        for pb in PRODUCT_BUCKETS:
            cell_mask = row_mask & (sub["PRODUCT_BUCKET"] == pb)
            if agg == "count":
                row_data.append(int(cell_mask.sum()))
            else:
                row_data.append(float(sub.loc[cell_mask, "CREDIT_LII"].sum()))
        matrix.append(row_data)
    return matrix


# -- Summary sheet builder ---------------------------------------------------
def build_summary_sheet(wb, pivots):
    """Build a static Summary sheet with all 4 view matrices stacked."""
    ws = wb.create_sheet("Summary")

    headers = ["Gross Amount", "Defined Range"] + PRODUCT_BUCKETS + ["Total"]

    # View configs: (title, pivot_key, number_format)
    # Number formats: positive;negative;zero — dash for zero cells
    view_configs = [
        ("Summary Count",          1, '#,##0;-#,##0;"-"'),
        ("Summary Commitment",     2, '$#,##0;-$#,##0;"-"'),
        ("Summary Count - NEW",    3, '#,##0;-#,##0;"-"'),
        ("Summary Commitment - NEW", 4, '$#,##0;-$#,##0;"-"'),
    ]

    # Row 1: title
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = "Portfolio Summary - All Views"
    c.font = BOLD_14
    c.alignment = Alignment(horizontal="center")

    current_row = 3  # start after blank row 2

    for title, pivot_key, num_fmt in view_configs:
        matrix = pivots[pivot_key]

        # Subheader row - navy fill, white text
        cell_sub = ws.cell(row=current_row, column=1, value=title)
        cell_sub.font = Font(bold=True, size=12, color=MS_COLORS['white'], name="Calibri")
        for ci in range(1, 17):
            ws.cell(row=current_row, column=ci).fill = NAVY_FILL
        current_row += 1

        # Column headers - blue fill, white text
        for ci, hdr in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=ci, value=hdr)
            cell.font = Font(bold=True, size=10, color=MS_COLORS['white'], name="Calibri")
            if ci == 2:
                cell.font = Font(bold=True, italic=True, size=10, color=MS_COLORS['white'], name="Calibri")
            cell.fill = BLUE_FILL
            cell.border = THIN_GRAY_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        current_row += 1

        # 24 data rows - alternating white/ice_blue
        for ri, (label, floor, ceil) in enumerate(BUCKET_LADDER):
            is_alt = (ri % 2 == 1)

            # Col A
            cell_a = ws.cell(row=current_row, column=1, value=label)
            cell_a.font = DATA_FONT_BOLD
            cell_a.border = THIN_GRAY_BORDER
            cell_a.alignment = Alignment(horizontal="right")
            cell_a.fill = ICE_BLUE_FILL if is_alt else WHITE_FILL

            # Col B
            cell_b = ws.cell(row=current_row, column=2, value=_range_label(floor, ceil))
            cell_b.font = ITALIC_9_GRAY
            cell_b.border = THIN_GRAY_BORDER
            cell_b.alignment = Alignment(horizontal="left")
            cell_b.fill = ICE_BLUE_FILL if is_alt else WHITE_FILL

            # Cols C-O
            row_total = 0
            for ci, val in enumerate(matrix[ri]):
                cell = ws.cell(row=current_row, column=ci + 3, value=val)
                cell.number_format = num_fmt
                cell.border = THIN_GRAY_BORDER
                cell.alignment = Alignment(horizontal="right")
                cell.font = DATA_FONT
                cell.fill = ICE_BLUE_FILL if is_alt else WHITE_FILL
                row_total += val

            # Col P -- total
            cell_p = ws.cell(row=current_row, column=16, value=row_total)
            cell_p.number_format = num_fmt
            cell_p.border = THIN_GRAY_BORDER
            cell_p.font = TOTAL_FONT
            cell_p.alignment = Alignment(horizontal="right")
            cell_p.fill = ICE_BLUE_FILL if is_alt else WHITE_FILL

            current_row += 1

        # Total row
        ws.cell(row=current_row, column=1, value="Total").font = TOTAL_FONT
        ws.cell(row=current_row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=current_row, column=1).border = DOUBLE_NAVY_TOP
        ws.cell(row=current_row, column=2, value="").fill = LIGHT_BLUE_FILL
        ws.cell(row=current_row, column=2).border = DOUBLE_NAVY_TOP

        data_start_row = current_row - 24  # first data row of this matrix
        for ci in range(3, 17):
            col_total = sum(
                ws.cell(row=r, column=ci).value or 0
                for r in range(data_start_row, current_row)
            )
            cell = ws.cell(row=current_row, column=ci, value=col_total)
            cell.font = TOTAL_FONT
            cell.number_format = num_fmt
            cell.fill = LIGHT_BLUE_FILL
            cell.border = DOUBLE_NAVY_TOP
            cell.alignment = Alignment(horizontal="right")

        current_row += 2  # blank row between matrices

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    for ci in range(3, 17):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    return ws


# -- POWER_QUERY_SETUP sheet builder -----------------------------------------
def build_pq_setup_sheet(wb):
    """Build a visible POWER_QUERY_SETUP sheet with setup instructions."""
    ws = wb.create_sheet("POWER_QUERY_SETUP")

    # Row 1: title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Power Query Setup - One-Time Configuration"
    c.font = PQ_TITLE_FONT
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Instructions
    steps = [
        (3, "Step 1: Go to Data > Get Data > Launch Power Query Editor"),
        (4, "Step 2: Click 'New Source' > 'Blank Query', then open the Advanced Editor"),
        (5, "Step 3: Paste the M code shown below into the Advanced Editor and click 'Done'"),
        (6, "Step 4: Name the query 'tbl_LoanData' and click 'Close & Load To...'"),
        (7, "Step 5: Select 'Existing worksheet' > click the Loan Detail sheet > cell A1 > click OK"),
    ]
    for row_num, text in steps:
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = PQ_STEP_FONT
        cell.alignment = Alignment(wrap_text=True)

    # M code block
    ws.cell(row=9, column=1, value="M Code (copy everything below):").font = Font(
        bold=True, size=11, name="Calibri"
    )

    # Read the M code
    m_code = ""
    if os.path.exists(PQ_FILE):
        with open(PQ_FILE, "r", encoding="utf-8") as f:
            m_code = f.read()

    m_lines = m_code.strip().split("\n")
    for i, line in enumerate(m_lines):
        row_num = 10 + i
        cell = ws.cell(row=row_num, column=1, value=line)
        cell.font = PQ_CODE_FONT
        cell.fill = BUTTON_GRAY_FILL
        cell.alignment = Alignment(wrap_text=False)

    post_code_row = 10 + len(m_lines) + 2

    post_steps = [
        (post_code_row, "Step 6: After loading, the Loan Detail sheet will be populated with loan data"),
        (post_code_row + 1, "Step 7: Click 'Refresh Data' button on Dashboard to recompute all views"),
        (post_code_row + 2, ""),
        (post_code_row + 3, "AFTER SETUP:"),
        (post_code_row + 4, "- To refresh data: click 'Refresh Data' on Dashboard (uses Power Query)"),
        (post_code_row + 5, "- Alternatively, use the corp_etl Python pipeline to produce tagged CSV"),
        (post_code_row + 6, "- Update the CSV path in _config sheet cell A3 if the file location changes"),
        (post_code_row + 7, ""),
        (post_code_row + 8, "NOTE: The DataSourcePath named range in _config!A3 must point to a valid CSV file."),
        (post_code_row + 9, "      Power Query loads into the Loan Detail sheet (not _data)."),
    ]
    for row_num, text in post_steps:
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = PQ_STEP_FONT
        if "AFTER SETUP" in text:
            cell.font = Font(bold=True, size=11, name="Calibri")
        cell.alignment = Alignment(wrap_text=True)

    # Column width
    ws.column_dimensions["A"].width = 120

    return ws


# -- WMLC threshold row mapping (0-indexed into the 24-row matrix) ----------
# Row index in data area (0-23) where WMLC threshold falls for each bucket.
# "Threshold row" = first row AT or ABOVE the WMLC threshold.
# E.g. LAL Diversified threshold is $300MM → row index 7 ($300MM bucket)
WMLC_THRESHOLD_ROW_IDX = {
    "LAL Diversified":    7,   # $300MM
    "LAL Highly Conc.":  13,   # $100MM
    "LAL NFPs":          13,   # $100MM
    "TL SBL Diversified": 7,   # $300MM
    "TL SBL Highly Conc.":13,  # $100MM
    "TL Life Insurance": 13,   # $100MM
    "TL CRE":           14,    # $75MM
    "TL Unsecured":      17,   # $35MM
    "TL Aircraft":       15,   # $50MM
    "TL PHA":            17,   # $35MM
    "TL Other Secured":  15,   # $50MM
    "TL Multicollateral":15,   # $50MM
    "RESI":              22,   # $10MM
}


# VBA-matching absolute Excel row numbers for threshold borders
# (data area rows 7-30; these must match mod_ViewToggle.bas pt() values)
WMLC_THRESHOLD_EXCEL_ROW = {
    "LAL Diversified":    14,
    "LAL Highly Conc.":   20,
    "LAL NFPs":           22,
    "TL SBL Diversified": 14,
    "TL SBL Highly Conc.":20,
    "TL Life Insurance":  20,
    "TL CRE":             21,
    "TL Unsecured":       24,
    "TL Aircraft":        22,
    "TL PHA":             24,
    "TL Other Secured":   22,
    "TL Multicollateral": 22,
    "RESI":               29,
}

# Shade fill for threshold zone (matches VBA RGB(232, 240, 254))
THRESHOLD_SHADE_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
# Medium navy bottom border for threshold row
MEDIUM_NAVY_BOTTOM = Border(
    left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY,
    bottom=Side(style="medium", color=MS_COLORS['navy'])
)


def apply_threshold_formatting(ws_dash):
    """Apply per-column threshold shading and bold borders.

    Replicates VBA ApplyThresholdFormatting so the static workbook
    looks correct before the first macro run.
    """
    # Build header → column mapping from row 6
    header_col = {}
    for ci in range(3, 16):
        hdr = ws_dash.cell(row=6, column=ci).value
        if hdr:
            header_col[str(hdr).strip()] = ci

    for pb, thresh_row in WMLC_THRESHOLD_EXCEL_ROW.items():
        col = header_col.get(pb)
        if col is None:
            continue
        # Shade rows 7 through threshold row
        for r in range(7, thresh_row + 1):
            ws_dash.cell(row=r, column=col).fill = THRESHOLD_SHADE_FILL
        # Bold medium navy bottom border at threshold row
        ws_dash.cell(row=thresh_row, column=col).border = MEDIUM_NAVY_BOTTOM


def build_wmlc_pct_row(ws_dash, pivots):
    """Build row 32: WMLC % concentration for each product bucket.

    Formula: sum(rows from bucket_idx 0..threshold) / sum(all rows) for
    the initial Commitment view (pivot 2).
    """
    ws_dash.cell(row=32, column=1, value="WMLC %").font = TOTAL_FONT
    ws_dash.cell(row=32, column=1).fill = LIGHT_BLUE_FILL
    ws_dash.cell(row=32, column=1).border = THIN_GRAY_BORDER
    ws_dash.cell(row=32, column=1).alignment = Alignment(horizontal="right")
    ws_dash.cell(row=32, column=2, value="").fill = LIGHT_BLUE_FILL
    ws_dash.cell(row=32, column=2).border = THIN_GRAY_BORDER

    matrix = pivots[2]  # Summary Commitment
    weighted_num = 0.0
    weighted_den = 0.0

    for ci, pb in enumerate(PRODUCT_BUCKETS):
        col_idx = ci + 3
        col_total = sum(matrix[ri][ci] for ri in range(24))
        thresh_idx = WMLC_THRESHOLD_ROW_IDX.get(pb, 23)
        above_thresh = sum(matrix[ri][ci] for ri in range(thresh_idx + 1))
        pct = above_thresh / col_total if col_total > 0 else 0.0

        cell = ws_dash.cell(row=32, column=col_idx, value=pct)
        cell.number_format = '0.0%'
        cell.font = TOTAL_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_GRAY_BORDER
        cell.alignment = Alignment(horizontal="right")

        weighted_num += above_thresh
        weighted_den += col_total

    # Col P -- weighted average
    avg_pct = weighted_num / weighted_den if weighted_den > 0 else 0.0
    cell_p = ws_dash.cell(row=32, column=16, value=avg_pct)
    cell_p.number_format = '0.0%'
    cell_p.font = TOTAL_FONT
    cell_p.fill = LIGHT_BLUE_FILL
    cell_p.border = THIN_GRAY_BORDER
    cell_p.alignment = Alignment(horizontal="right")


# -- Threshold $ values per product (used for chart data + distance table) --
WMLC_THRESHOLDS_DOLLARS = {
    "LAL Diversified": 300_000_000, "LAL Highly Conc.": 100_000_000,
    "LAL NFPs": 50_000_000, "TL SBL Diversified": 300_000_000,
    "TL SBL Highly Conc.": 100_000_000, "TL Life Insurance": 100_000_000,
    "TL CRE": 75_000_000, "TL Unsecured": 35_000_000,
    "TL Aircraft": 50_000_000, "TL PHA": 35_000_000,
    "TL Other Secured": 50_000_000, "TL Multicollateral": 50_000_000,
    "RESI": 10_000_000,
}

_CHART_W = 38
_CHART_H = 14


def build_chart_data_sheet(wb, df, pivots):
    """Build _chart_data with all pre-computed visualization data.

    Sections:
      Rows 1-13:   Stacked bar (product × below80/approaching/above)
      Rows 41-42:  Donut split (WMLC qualified $ vs non-WMLC $)
      Rows 45-56:  Flag breakdown by type (flag name, count, commitment sum)
      Rows 58-70:  Threshold proximity counts (product, <80%, 80-99%, >=100%)
    """
    ws = wb.create_sheet("_chart_data")
    credit_lii = df["CREDIT_LII"]
    wmlc_mask = df["WMLC_QUALIFIED"]

    # ── Rows 1-13: Stacked bar data ──────────────────────────────────────
    for i, pb in enumerate(PRODUCT_BUCKETS):
        thresh = WMLC_THRESHOLDS_DOLLARS.get(pb, 999_999_999_999)
        mask = df["PRODUCT_BUCKET"] == pb
        lii = credit_lii[mask]
        below = float(lii[lii < thresh * 0.8].sum())
        approaching = float(lii[(lii >= thresh * 0.8) & (lii < thresh)].sum())
        at_above = float(lii[lii >= thresh].sum())

        row = i + 1
        ws.cell(row=row, column=1, value=pb)
        ws.cell(row=row, column=2, value=below)
        ws.cell(row=row, column=3, value=approaching)
        ws.cell(row=row, column=4, value=at_above)

    # ── Rows 41-42: Donut split ──────────────────────────────────────────
    wmlc_sum = float(credit_lii[wmlc_mask].sum())
    non_wmlc_sum = float(credit_lii[~wmlc_mask].sum())
    ws.cell(row=41, column=1, value="WMLC Qualified")
    ws.cell(row=41, column=2, value=wmlc_sum)
    ws.cell(row=42, column=1, value="Non-WMLC")
    ws.cell(row=42, column=2, value=non_wmlc_sum)

    # ── Rows 45-56: Flag breakdown by type ───────────────────────────────
    flag_col = "WMLC_FLAGS"
    flag_counts = {}
    flag_sums = {}
    for _, row in df.iterrows():
        flags_str = row.get(flag_col, "")
        if not flags_str or str(flags_str).lower() == "nan":
            continue
        lii_val = float(row.get("CREDIT_LII", 0))
        for f in str(flags_str).split("|"):
            f = f.strip()
            if f:
                flag_counts[f] = flag_counts.get(f, 0) + 1
                flag_sums[f] = flag_sums.get(f, 0) + lii_val

    sorted_flags = sorted(flag_counts.keys(), key=lambda x: flag_counts[x], reverse=True)
    for i, flag_name in enumerate(sorted_flags[:12]):
        row_num = 45 + i
        ws.cell(row=row_num, column=1, value=flag_name)
        ws.cell(row=row_num, column=2, value=flag_counts[flag_name])
        ws.cell(row=row_num, column=3, value=flag_sums.get(flag_name, 0))

    # ── Rows 58-70: Threshold proximity counts ───────────────────────────
    for i, pb in enumerate(PRODUCT_BUCKETS):
        thresh = WMLC_THRESHOLDS_DOLLARS.get(pb, 999_999_999_999)
        mask = df["PRODUCT_BUCKET"] == pb
        lii = credit_lii[mask]
        n_below = int((lii < thresh * 0.8).sum())
        n_approach = int(((lii >= thresh * 0.8) & (lii < thresh)).sum())
        n_above = int((lii >= thresh).sum())

        row_num = 58 + i
        ws.cell(row=row_num, column=1, value=pb)
        ws.cell(row=row_num, column=2, value=n_below)
        ws.cell(row=row_num, column=3, value=n_approach)
        ws.cell(row=row_num, column=4, value=n_above)

    ws.sheet_state = "hidden"
    return ws


def _style_chart(chart, title_text):
    """Apply MS styling: navy title, no gridlines, transparent background."""
    chart.title = title_text
    chart.title.txPr = RichText(p=[DrawParagraph(
        pPr=DrawParagraphProperties(defRPr=CharacterProperties(
            sz=1100, b=True,
            solidFill=MS_COLORS['navy'],
            latin=DrawingFont(typeface="Calibri"),
        ))
    )])
    chart.y_axis.majorGridlines = None
    chart.y_axis.minorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.x_axis.minorGridlines = None
    axis_font = CharacterProperties(
        sz=800, solidFill=MS_COLORS['text_gray'],
        latin=DrawingFont(typeface="Calibri"),
    )
    chart.y_axis.txPr = RichText(p=[DrawParagraph(
        pPr=DrawParagraphProperties(defRPr=axis_font))])
    chart.x_axis.txPr = RichText(p=[DrawParagraph(
        pPr=DrawParagraphProperties(defRPr=axis_font))])
    chart.plot_area.graphicalProperties = GraphicalProperties()
    chart.plot_area.graphicalProperties.noFill = True
    chart.plot_area.graphicalProperties.line.noFill = True
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.noFill = True
    chart.graphical_properties.line.noFill = True
    return chart


def build_dashboard_visuals(ws_dash, wb, df):
    """Build all 3 WMLC visualizations + Top 10 below the matrix.

    Layout:
        Row 33:      Navy divider
        Rows 34-50:  VIZ 1 — Threshold Utilization Stacked Bar (chart)
        Row 51:      Navy divider
        Row 52:      Viz 2 header
        Row 53:      Viz 2 column headers
        Rows 54-83:  VIZ 2 — Threshold Distance table (30 rows)
        Row 84:      Navy divider
        Row 85:      Viz 3 header
        Row 86:      Viz 3 column headers (flag names — VBA fills)
        Rows 87-102: VIZ 3 — Flag Overlap Heatmap (VBA fills data)
        Row 103:     Navy divider
        Row 104:     Top 10 header
        Row 105:     Top 10 column headers
        Rows 106-115: Top 10 data
    """
    ws_cd = wb["_chart_data"]
    NUM_PRODUCTS = 13

    # ========== VIZ 1: Threshold Utilization Stacked Bar ==========
    from openpyxl.chart.series import SeriesLabel

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.gapWidth = 60
    chart.width = _CHART_W
    chart.height = _CHART_H

    # Series 1: Below 80% (gray)
    s1 = Reference(ws_cd, min_col=2, min_row=1, max_row=NUM_PRODUCTS)
    chart.add_data(s1, titles_from_data=False)
    chart.series[0].graphicalProperties.solidFill = "D9DEE3"
    chart.series[0].graphicalProperties.line.solidFill = "D9DEE3"
    chart.series[0].title = SeriesLabel(v="Below 80%")

    # Series 2: Approaching 80-99% (amber)
    s2 = Reference(ws_cd, min_col=3, min_row=1, max_row=NUM_PRODUCTS)
    chart.add_data(s2, titles_from_data=False)
    chart.series[1].graphicalProperties.solidFill = "D4A017"
    chart.series[1].graphicalProperties.line.solidFill = "D4A017"
    chart.series[1].title = SeriesLabel(v="Approaching (80-99%)")

    # Series 3: WMLC Qualified >= threshold (navy)
    s3 = Reference(ws_cd, min_col=4, min_row=1, max_row=NUM_PRODUCTS)
    chart.add_data(s3, titles_from_data=False)
    chart.series[2].graphicalProperties.solidFill = "002B5C"
    chart.series[2].graphicalProperties.line.solidFill = "002B5C"
    chart.series[2].title = SeriesLabel(v="WMLC Qualified")

    cats = Reference(ws_cd, min_col=1, min_row=1, max_row=NUM_PRODUCTS)
    chart.set_categories(cats)
    _style_chart(chart, "Threshold Utilization by Product ($)")
    chart.legend = Legend()
    chart.legend.position = "b"
    chart.y_axis.scaling.orientation = "maxMin"
    chart.x_axis.numFmt = '$#,##0,,"M"'

    ws_dash.add_chart(chart, "A34")

    # ========== VIZ 2: Threshold Distance Table (static from proxy) ==========
    # Header
    ws_dash.merge_cells("A52:P52")
    ws_dash["A52"].value = "Threshold Distance \u2014 Top 30 Loans Near WMLC Boundary"
    ws_dash["A52"].font = Font(name="Calibri", bold=True, size=12, color=MS_COLORS['white'])
    ws_dash["A52"].fill = NAVY_FILL
    ws_dash["A52"].alignment = Alignment(vertical="center")
    ws_dash.row_dimensions[52].height = 28

    # Column headers
    dist_headers = ["#", "Borrower", "Product", "Commitment ($)", "Threshold ($)",
                    "Distance ($)", "% of Threshold"]
    for ci, hdr in enumerate(dist_headers, 1):
        cell = ws_dash.cell(row=53, column=ci, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=9, color=MS_COLORS['white'])
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Merged visual bar header
    ws_dash.merge_cells("H53:P53")
    ws_dash["H53"].value = "Threshold Utilization"
    ws_dash["H53"].font = Font(name="Calibri", bold=True, size=9, color=MS_COLORS['white'])
    ws_dash["H53"].fill = BLUE_FILL
    ws_dash["H53"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[53].height = 22

    # Pre-populate from proxy data — loans between 70% and 300% of threshold
    credit_lii = df["CREDIT_LII"]
    candidates = []
    for _, row in df.iterrows():
        pb = row.get("PRODUCT_BUCKET", "")
        thresh = WMLC_THRESHOLDS_DOLLARS.get(pb)
        if thresh is None or thresh == 0:
            continue
        lii = float(row.get("CREDIT_LII", 0))
        pct = lii / thresh
        if 0.7 <= pct <= 3.0:
            candidates.append({
                "borrower": row.get("BORROWER", ""),
                "product": pb,
                "lii": lii,
                "thresh": thresh,
                "pct": pct,
            })
    # Sort by distance from 100% (closest to boundary first)
    candidates.sort(key=lambda x: abs(x["pct"] - 1.0))
    candidates = candidates[:30]

    AMBER = PatternFill("solid", fgColor="D4A017")
    LIGHT_AMBER = PatternFill("solid", fgColor="FFF3CD")
    GREEN_FONT = Font(name="Calibri", size=9, color="007A33")
    AMBER_FONT = Font(name="Calibri", size=9, color="D4A017")

    for i, c in enumerate(candidates):
        r = 54 + i
        ws_dash.row_dimensions[r].height = 18
        is_alt = (i % 2 == 1)
        row_fill = ICE_BLUE_FILL if is_alt else WHITE_FILL

        ws_dash.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center")
        ws_dash.cell(row=r, column=2, value=c["borrower"])
        ws_dash.cell(row=r, column=3, value=c["product"])
        ws_dash.cell(row=r, column=4, value=c["lii"]).number_format = "$#,##0"
        ws_dash.cell(row=r, column=5, value=c["thresh"]).number_format = "$#,##0"
        dist = c["lii"] - c["thresh"]
        dist_cell = ws_dash.cell(row=r, column=6, value=dist)
        dist_cell.number_format = "$#,##0"
        pct_cell = ws_dash.cell(row=r, column=7, value=c["pct"])
        pct_cell.number_format = "0.0%"

        # Color coding
        if dist >= 0:
            dist_cell.font = GREEN_FONT
            pct_cell.fill = NAVY_FILL
            pct_cell.font = Font(name="Calibri", size=9, color=MS_COLORS['white'])
        elif c["pct"] >= 0.9:
            dist_cell.font = AMBER_FONT
            pct_cell.fill = AMBER
            pct_cell.font = Font(name="Calibri", size=9, color=MS_COLORS['white'])
        elif c["pct"] >= 0.8:
            dist_cell.font = AMBER_FONT
            pct_cell.fill = LIGHT_AMBER

        # Base formatting for all cells in row (H:P bars removed — VBA data bar on G)
        for ci in range(1, 17):
            cell = ws_dash.cell(row=r, column=ci)
            if cell.fill == PatternFill():  # unfilled
                cell.fill = row_fill
            cell.border = THIN_GRAY_BORDER
            if cell.font == Font():  # default font
                cell.font = DATA_FONT

    # Fill remaining empty rows (if < 30 candidates)
    for r in range(54 + len(candidates), 84):
        ws_dash.row_dimensions[r].height = 18
        for ci in range(1, 17):
            cell = ws_dash.cell(row=r, column=ci)
            cell.fill = ICE_BLUE_FILL if (r % 2 == 1) else WHITE_FILL
            cell.border = THIN_GRAY_BORDER

    # ========== VIZ 3: Flag Overlap Heatmap (headers only — VBA fills data) ==========
    ws_dash.merge_cells("A85:P85")
    ws_dash["A85"].value = "WMLC Flag Overlap Matrix"
    ws_dash["A85"].font = Font(name="Calibri", bold=True, size=12, color=MS_COLORS['white'])
    ws_dash["A85"].fill = NAVY_FILL
    ws_dash["A85"].alignment = Alignment(vertical="center")
    ws_dash.row_dimensions[85].height = 28
    ws_dash.row_dimensions[86].height = 60  # Tall for rotated text (VBA sets)
    for r in range(87, 103):
        ws_dash.row_dimensions[r].height = 20

    # Pre-populate heatmap from proxy data
    # Collect unique flags
    flag_set = set()
    for flags_str in df["WMLC_FLAGS"].dropna():
        if flags_str and str(flags_str).lower() != "nan":
            for f in str(flags_str).split("|"):
                f = f.strip()
                if f:
                    flag_set.add(f)
    flag_list = sorted(flag_set)[:16]
    num_flags = len(flag_list)
    flag_idx = {f: i for i, f in enumerate(flag_list)}

    # Build co-occurrence matrix
    overlap = [[0] * num_flags for _ in range(num_flags)]
    for flags_str in df["WMLC_FLAGS"].dropna():
        if not flags_str or str(flags_str).lower() == "nan":
            continue
        row_flags = [f.strip() for f in str(flags_str).split("|") if f.strip() in flag_idx]
        for fi in row_flags:
            for fj in row_flags:
                overlap[flag_idx[fi]][flag_idx[fj]] += 1

    max_val = max((overlap[i][j] for i in range(num_flags) for j in range(num_flags)), default=1)
    if max_val == 0:
        max_val = 1

    # Column headers (row 86)
    for fi in range(num_flags):
        cell = ws_dash.cell(row=86, column=fi + 2, value=flag_list[fi][:15])
        cell.font = Font(name="Calibri", size=7, bold=True, color=MS_COLORS['white'])
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(textRotation=90, horizontal="center")

    # Row labels + matrix data
    for fi in range(num_flags):
        r = 87 + fi
        ws_dash.cell(row=r, column=1, value=flag_list[fi][:20]).font = Font(
            name="Calibri", size=8, bold=True)
        for fj in range(num_flags):
            val = overlap[fi][fj]
            cell = ws_dash.cell(row=r, column=fj + 2, value=val)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Calibri", size=8)
            # Color scale: white → navy
            intensity = val / max_val
            rr = int(255 - intensity * 255)
            gg = int(255 - intensity * (255 - 43))
            bb = int(255 - intensity * (255 - 92))
            cell.fill = PatternFill("solid", fgColor=f"{rr:02X}{gg:02X}{bb:02X}")
            if intensity > 0.5:
                cell.font = Font(name="Calibri", size=8, color=MS_COLORS['white'])
            else:
                cell.font = Font(name="Calibri", size=8, color=MS_COLORS['navy'])
            if fi == fj:
                cell.font = Font(name="Calibri", size=8, bold=True,
                                 color=MS_COLORS['white'] if intensity > 0.5 else MS_COLORS['navy'])
            cell.border = Border(
                left=Side(style="hair", color="C8C8C8"),
                right=Side(style="hair", color="C8C8C8"),
                top=Side(style="hair", color="C8C8C8"),
                bottom=Side(style="hair", color="C8C8C8"),
            )

    # ========== TOP 10 BORROWERS ==========
    TITLE_ROW = 104
    HDR_ROW = 105
    DATA_START = 106

    ws_dash.merge_cells(f"A{TITLE_ROW}:E{TITLE_ROW}")
    cell = ws_dash.cell(row=TITLE_ROW, column=1, value="Top 10 Exposures")
    cell.font = Font(name="Calibri", bold=True, size=13, color=MS_COLORS['white'])
    cell.fill = NAVY_FILL
    cell.alignment = Alignment(vertical="center")
    ws_dash.row_dimensions[TITLE_ROW].height = 28

    top10_headers = ["Rank", "Borrower", "Product", "Commitment", "WMLC Flags"]
    for ci, hdr in enumerate(top10_headers, 1):
        cell = ws_dash.cell(row=HDR_ROW, column=ci, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=10, color=MS_COLORS['white'])
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[HDR_ROW].height = 22

    top10 = df.nlargest(10, "CREDIT_LII")
    for rank, (_, row) in enumerate(top10.iterrows(), 1):
        data_row = DATA_START + rank - 1
        is_alt = (rank % 2 == 0)
        row_fill = ICE_BLUE_FILL if is_alt else WHITE_FILL
        cells_data = [
            rank,
            row.get("BORROWER", ""),
            row.get("PRODUCT_BUCKET", ""),
            row.get("CREDIT_LII", 0),
            row.get("WMLC_FLAGS", ""),
        ]
        for ci, val in enumerate(cells_data, 1):
            cell = ws_dash.cell(row=data_row, column=ci, value=val)
            cell.fill = row_fill
            cell.border = THIN_GRAY_BORDER
            cell.font = DATA_FONT
            if ci == 4:
                cell.number_format = '$#,##0'
            if ci == 1:
                cell.alignment = Alignment(horizontal="center")


# -- COM post-processing: inject VBA + create buttons -----------------------
def inject_vba_and_buttons(xlsx_path, xlsm_path, vba_dir):
    """Open .xlsx via COM, inject VBA, create buttons, save as .xlsm"""
    import win32com.client
    import winreg

    # Ensure VBA project access is enabled in registry
    try:
        key_path = r"Software\Microsoft\Office\16.0\Excel\Security"
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0,
                            winreg.KEY_READ | winreg.KEY_SET_VALUE)
        winreg.SetValueEx(key, "AccessVBOM", 0, winreg.REG_DWORD, 1)
        winreg.CloseKey(key)
        print("  Enabled VBA project access via registry.")
    except Exception as e:
        print(f"  Warning: Could not set registry key for VBA access: {e}")

    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AutomationSecurity = 1  # msoAutomationSecurityLow

        abs_xlsx = os.path.abspath(xlsx_path)
        abs_xlsm = os.path.abspath(xlsm_path)

        print(f"  Opening {abs_xlsx} via COM ...")
        wb = excel.Workbooks.Open(abs_xlsx)

        vb_proj = wb.VBProject

        # Test access
        try:
            _ = vb_proj.VBComponents.Count
            print(f"  VBProject access OK ({vb_proj.VBComponents.Count} existing components)")
        except Exception as e:
            print(f"  ERROR: Cannot access VBProject: {e}")
            print("  Please enable 'Trust access to the VBA project object model' in")
            print("  File > Options > Trust Center > Trust Center Settings > Macro Settings")
            raise

        # --- Import standard modules ---
        def read_bas_code(filename):
            """Read .bas file and strip Attribute VB_Name line."""
            path = os.path.join(vba_dir, filename)
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            # Strip Attribute VB_Name line
            lines = [l for l in lines if not l.strip().startswith("Attribute VB_Name")]
            return "".join(lines)

        # mod_ViewToggle
        print("  Injecting mod_ViewToggle ...")
        comp = vb_proj.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
        comp.Name = "mod_ViewToggle"
        code = read_bas_code("mod_ViewToggle.bas")
        comp.CodeModule.AddFromString(code)

        # mod_Navigation
        print("  Injecting mod_Navigation ...")
        comp = vb_proj.VBComponents.Add(1)
        comp.Name = "mod_Navigation"
        code = read_bas_code("mod_Navigation.bas")
        comp.CodeModule.AddFromString(code)

        # mod_Diagnostics
        print("  Injecting mod_Diagnostics ...")
        comp = vb_proj.VBComponents.Add(1)
        comp.Name = "mod_Diagnostics"
        code = read_bas_code("mod_Diagnostics.bas")
        comp.CodeModule.AddFromString(code)

        # mod_DataRefresh
        print("  Injecting mod_DataRefresh ...")
        comp = vb_proj.VBComponents.Add(1)
        comp.Name = "mod_DataRefresh"
        code = read_bas_code("mod_DataRefresh.bas")
        comp.CodeModule.AddFromString(code)

        # mod_CellClick -- goes in Dashboard sheet module
        print("  Injecting mod_CellClick into Dashboard sheet module ...")
        cell_click_code = read_bas_code("mod_CellClick.bas")
        # Strip comment header lines about where to paste
        filtered_lines = []
        for line in cell_click_code.split("\n"):
            stripped = line.strip()
            if stripped.startswith("' ====") or stripped.startswith("' IMPORTANT"):
                continue
            filtered_lines.append(line)
        cell_click_code = "\n".join(filtered_lines)

        # Find Dashboard sheet's code module
        dash_sheet = wb.Sheets("Dashboard")
        dash_codename = dash_sheet.CodeName
        print(f"  Dashboard CodeName: {dash_codename}")
        for comp in vb_proj.VBComponents:
            if comp.Name == dash_codename:
                comp.CodeModule.AddFromString(cell_click_code)
                print("  Injected Worksheet_SelectionChange into Dashboard module.")
                break

        # --- Navy RGB for buttons ---
        # #002B5C = R=0, G=43, B=92 -> COM RGB = R + G*256 + B*65536
        NAVY_RGB = 0 + (43 * 256) + (92 * 65536)  # 6037248
        WHITE_RGB = 255 + (255 * 256) + (255 * 65536)

        # --- Create buttons on Dashboard ---
        print("  Creating Dashboard buttons ...")
        dash = wb.Sheets("Dashboard")

        # Button layout: 110w, 22h, 12px gap between each
        # Row 4 starts at left=10, top calculated from row position
        BTN_W = 110
        BTN_H = 22
        GAP = 12
        ROW4_TOP = 60   # approximate pixel top for row 4
        ROW5_TOP = 92   # approximate pixel top for row 5

        def _btn_positions(names_macros, top):
            """Generate (name, text, left, top, w, h, macro) tuples with gaps."""
            result = []
            left = 10
            for name, text, macro in names_macros:
                # Wider for "Summary Commitment"
                w = 130 if len(text) > 13 else BTN_W
                result.append((name, text, left, top, w, BTN_H, macro))
                left += w + GAP
            return result

        row4_buttons = _btn_positions([
            ("btnSummaryCount",      "Summary Count",       "ShowSummaryCount"),
            ("btnSummaryCommitment", "Summary Commitment",  "ShowSummaryCommitment"),
            ("btnCountNew",          "Count NEW",           "ShowSummaryCountNew"),
            ("btnCommitmentNew",     "Commitment NEW",      "ShowSummaryCommitmentNew"),
        ], ROW4_TOP)

        row5_buttons = _btn_positions([
            ("btnWMLCOn",            "WMLC ON",             "WMLCOn"),
            ("btnWMLCOff",           "WMLC OFF",            "WMLCOff"),
            ("btnReset",             "Reset",               "ResetDashboard"),
            ("btnRefreshData",       "Refresh Data",        "RefreshData"),
            ("btnDiagnostics",       "Run Diagnostics",     "TestRefreshCycle"),
        ], ROW5_TOP)

        all_dash_buttons = row4_buttons + row5_buttons
        for name, text, left, top, width, height, macro in all_dash_buttons:
            shp = dash.Shapes.AddShape(5, left, top, width, height)  # 5 = msoShapeRoundedRectangle
            shp.Name = name
            shp.TextFrame2.TextRange.Text = text
            shp.TextFrame2.TextRange.Font.Size = 9
            shp.TextFrame2.TextRange.Font.Bold = True
            shp.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = WHITE_RGB
            shp.TextFrame2.TextRange.ParagraphFormat.Alignment = 2  # msoAlignCenter
            shp.Fill.ForeColor.RGB = NAVY_RGB
            shp.Line.Visible = 0  # False - flat design, no border
            shp.OnAction = macro

        # Diagnostics button gets a distinct dark teal color (#003366)
        try:
            diag_btn = dash.Shapes("btnDiagnostics")
            diag_btn.Fill.ForeColor.RGB = 0x00 + (0x33 * 256) + (0x66 * 65536)  # #003366
        except Exception:
            pass

        # --- Create buttons on Loan Detail ---
        print("  Creating Loan Detail buttons ...")
        detail = wb.Sheets("Loan Detail")

        for name, text, left, top, width, height, macro in [
            ("btnBack",         "\u2190 Back to Dashboard", 5, 18, 160, 24, "BackToDashboard"),
            ("btnResetFilters", "Reset Filters",          175, 18, 110, 24, "ResetFilters"),
        ]:
            shp = detail.Shapes.AddShape(5, left, top, width, height)
            shp.Name = name
            shp.TextFrame2.TextRange.Text = text
            shp.TextFrame2.TextRange.Font.Size = 9
            shp.TextFrame2.TextRange.Font.Bold = True
            shp.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = WHITE_RGB
            shp.TextFrame2.TextRange.ParagraphFormat.Alignment = 2
            shp.Fill.ForeColor.RGB = NAVY_RGB
            shp.Line.Visible = 0
            shp.OnAction = macro

        # --- Save as .xlsm ---
        print(f"  Saving as {abs_xlsm} ...")
        wb.SaveAs(abs_xlsm, FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
        print("  Saved .xlsm successfully.")

    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        # Release COM objects
        wb = None
        excel = None


# -- VBA_REFERENCE.txt generator --------------------------------------------
def generate_vba_reference():
    """Read all .bas files and PQ M code, assemble into output/VBA_REFERENCE.txt"""
    ref_path = os.path.join(REPO, "output", "VBA_REFERENCE.txt")
    lines = []
    lines.append("=" * 80)
    lines.append("WMLC Dashboard - VBA & Power Query Reference")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 80)
    lines.append("")
    lines.append("This file contains all VBA module source code and Power Query M code")
    lines.append("embedded in the WMLC_Dashboard.xlsm workbook.")
    lines.append("")
    lines.append("LAYOUT NOTE: Dashboard grid starts at row 7 (title=1, subtitle=2,")
    lines.append("indicator=3, buttons=4-5, headers=6, data=7-30, total=31).")
    lines.append("")

    # VBA modules
    bas_files = sorted([f for f in os.listdir(VBA_DIR) if f.endswith('.bas')])
    for bas_file in bas_files:
        bas_path = os.path.join(VBA_DIR, bas_file)
        with open(bas_path, "r", encoding="utf-8") as f:
            content = f.read()
        lines.append("-" * 80)
        lines.append(f"MODULE: {bas_file}")
        lines.append("-" * 80)
        lines.append(content)
        lines.append("")

    # Power Query
    if os.path.exists(PQ_FILE):
        with open(PQ_FILE, "r", encoding="utf-8") as f:
            pq_content = f.read()
        lines.append("-" * 80)
        lines.append("POWER QUERY: load_tagged_data.m")
        lines.append("-" * 80)
        lines.append(pq_content)
        lines.append("")

    lines.append("=" * 80)
    lines.append("END OF VBA REFERENCE")
    lines.append("=" * 80)

    with open(ref_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  VBA_REFERENCE.txt written to {ref_path}")
    return ref_path


# -- main build --------------------------------------------------------------
def main():
    # Kill any running Excel before we start
    print("Killing any running EXCEL.EXE processes ...")
    subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], capture_output=True)
    time.sleep(2)

    # Delete old output files if they exist
    for f in [OUTPUT_XLSX, OUTPUT_XLSM]:
        if os.path.exists(f):
            try:
                os.remove(f)
                print(f"  Deleted old {os.path.basename(f)}")
            except Exception as e:
                print(f"  Warning: could not delete {f}: {e}")

    print("Reading tagged CSV ...")
    df = pd.read_csv(TAGGED_CSV)
    print(f"  {len(df)} rows, {len(df.columns)} columns")

    # Normalize column names to uppercase for consistent access
    df.columns = [c.upper() for c in df.columns]

    # Ensure boolean columns
    df["WMLC_QUALIFIED"] = df["WMLC_QUALIFIED"].astype(str).str.strip().str.lower() == "true"
    df["CREDIT_LII"] = pd.to_numeric(df["CREDIT_LII"], errors="coerce").fillna(0)

    # Masks
    mask_new = df["NEW_CAMP_YN"] == "Y"
    mask_wmlc = df["WMLC_QUALIFIED"]
    mask_new_wmlc = mask_new & mask_wmlc

    # Compute 8 pivots
    print("Computing 8 view pivots ...")
    pivots = {
        1: compute_pivot(df, "count"),
        2: compute_pivot(df, "sum"),
        3: compute_pivot(df, "count", mask_new),
        4: compute_pivot(df, "sum",   mask_new),
        5: compute_pivot(df, "count", mask_wmlc),
        6: compute_pivot(df, "sum",   mask_wmlc),
        7: compute_pivot(df, "count", mask_new_wmlc),
        8: compute_pivot(df, "sum",   mask_new_wmlc),
    }

    wb = Workbook()

    # == "Loan Detail" sheet — BLANK landing zone for Power Query ===========
    # Power Query loads data here. No pre-populated data, no Table object.
    # VBA MasterRefresh reads from this sheet once PQ populates it.
    print("Building Loan Detail sheet (blank — PQ landing zone) ...")
    ws_data = wb.active
    ws_data.title = "Loan Detail"
    ws_data["A1"].value = "\u2190 Power Query loads data here. See POWER_QUERY_SETUP sheet."
    ws_data["A1"].font = Font(italic=True, size=9, color="999999", name="Calibri")
    ws_data.freeze_panes = "A2"

    # == Dashboard sheet =====================================================
    print("Building Dashboard sheet ...")
    ws_dash = wb.create_sheet("Dashboard")

    # Row 1 -- title (merged A1:P1, navy fill, white 18pt bold)
    ws_dash.merge_cells("A1:P1")
    c = ws_dash["A1"]
    c.value = "Wealth Management Lending Committee \u2014 Portfolio Dashboard"
    c.font = TITLE_FONT
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 36

    # Row 2 -- subtitle (merged A2:P2, blue fill, white 11pt)
    ws_dash.merge_cells("A2:P2")
    c2 = ws_dash["A2"]
    c2.value = "Credit Risk | 1st Line of Defense"
    c2.font = SUBTITLE_FONT
    c2.fill = BLUE_FILL
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[2].height = 24

    # Row 3 -- date/view indicator (merged A3:P3, white bg, gray text)
    ws_dash.merge_cells("A3:P3")
    c3 = ws_dash["A3"]
    c3.value = f"As of: {datetime.now().strftime('%m/%d/%Y')} | View: Summary Commitment | WMLC: OFF"
    c3.font = INDICATOR_FONT
    c3.fill = WHITE_FILL
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[3].height = 22

    # Rows 4-5 -- button area (light gray fill)
    for row_num in [4, 5]:
        ws_dash.row_dimensions[row_num].height = 32
        for ci in range(1, 17):
            cell = ws_dash.cell(row=row_num, column=ci)
            cell.fill = BUTTON_GRAY_FILL

    # Row 6 -- column headers (navy fill, white text, bold 10pt, center)
    headers = ["Gross Amount", "Defined Range"] + PRODUCT_BUCKETS + ["Total"]
    ws_dash.row_dimensions[6].height = 28
    for ci, hdr in enumerate(headers, 1):
        cell = ws_dash.cell(row=6, column=ci, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.border = THIN_WHITE_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    # Rows 7-30 -- 24 data rows (alternating white/ice_blue)
    initial_matrix = pivots[2]  # Summary Commitment
    for ri, (label, floor, ceil) in enumerate(BUCKET_LADDER):
        row_num = ri + 7
        is_alt = (ri % 2 == 1)
        row_fill = ICE_BLUE_FILL if is_alt else WHITE_FILL

        # Col A -- Gross Amount
        cell_a = ws_dash.cell(row=row_num, column=1, value=label)
        cell_a.font = DATA_FONT_BOLD
        cell_a.border = THIN_GRAY_BORDER
        cell_a.alignment = Alignment(horizontal="right")
        cell_a.fill = row_fill

        # Col B -- Defined Range
        cell_b = ws_dash.cell(row=row_num, column=2, value=_range_label(floor, ceil))
        cell_b.font = ITALIC_9_GRAY
        cell_b.border = THIN_GRAY_BORDER
        cell_b.alignment = Alignment(horizontal="left")
        cell_b.fill = row_fill

        # Cols C-O -- data
        row_total = 0
        for ci, val in enumerate(initial_matrix[ri]):
            cell = ws_dash.cell(row=row_num, column=ci + 3, value=val)
            cell.number_format = '$#,##0;-$#,##0;"-"'
            cell.font = DATA_FONT
            cell.border = THIN_GRAY_BORDER
            cell.alignment = Alignment(horizontal="right")
            cell.fill = row_fill
            row_total += val

        # Col P -- total
        cell_p = ws_dash.cell(row=row_num, column=16, value=row_total)
        cell_p.number_format = '$#,##0'
        cell_p.border = THIN_GRAY_BORDER
        cell_p.font = TOTAL_FONT
        cell_p.alignment = Alignment(horizontal="right")
        cell_p.fill = LIGHT_BLUE_FILL

    # Row 31 -- Total row (bold navy text, light blue fill, double-line top)
    ws_dash.cell(row=31, column=1, value="Total").font = TOTAL_FONT
    ws_dash.cell(row=31, column=1).fill = LIGHT_BLUE_FILL
    ws_dash.cell(row=31, column=1).border = DOUBLE_NAVY_TOP
    ws_dash.cell(row=31, column=2, value="").fill = LIGHT_BLUE_FILL
    ws_dash.cell(row=31, column=2).border = DOUBLE_NAVY_TOP

    for ci in range(3, 17):
        col_total = sum(
            ws_dash.cell(row=r, column=ci).value or 0 for r in range(7, 31)
        )
        cell = ws_dash.cell(row=31, column=ci, value=col_total)
        cell.font = TOTAL_FONT
        cell.number_format = '$#,##0;-$#,##0;"-"'
        cell.fill = LIGHT_BLUE_FILL
        cell.border = DOUBLE_NAVY_TOP
        cell.alignment = Alignment(horizontal="right")

    # Row 32 -- WMLC % concentration
    print("Building WMLC % concentration row ...")
    build_wmlc_pct_row(ws_dash, pivots)

    # Threshold shading + bold borders (match VBA ApplyThresholdFormatting)
    print("Applying threshold formatting ...")
    apply_threshold_formatting(ws_dash)

    # Column widths
    ws_dash.column_dimensions["A"].width = 18
    ws_dash.column_dimensions["B"].width = 28
    for ci in range(3, 16):
        ws_dash.column_dimensions[get_column_letter(ci)].width = 14
    ws_dash.column_dimensions["P"].width = 14

    # Data row heights
    for r in range(7, 31):
        ws_dash.row_dimensions[r].height = 16
    ws_dash.row_dimensions[31].height = 20  # Total
    ws_dash.row_dimensions[32].height = 20  # WMLC %

    # Freeze panes at A7
    ws_dash.freeze_panes = "A7"

    # == _chart_data sheet (hidden) ==========================================
    print("Building _chart_data sheet ...")
    build_chart_data_sheet(wb, df, pivots)

    # == Layout for visualization zones ======================================
    print("Setting visualization layout ...")
    DIVIDER_FILL = PatternFill("solid", fgColor=MS_COLORS['navy'])

    # Navy divider rows
    for dr in [33, 51, 84, 103]:
        ws_dash.row_dimensions[dr].height = 4
        for ci in range(1, 17):
            ws_dash.cell(row=dr, column=ci).fill = DIVIDER_FILL

    # Viz 1 chart zone: rows 34-50 (25px each for chart room)
    for r in range(34, 51):
        ws_dash.row_dimensions[r].height = 25

    # == Build all 3 visualizations + Top 10 =================================
    print("Building WMLC visualizations ...")
    build_dashboard_visuals(ws_dash, wb, df)

    # == 8 hidden view sheets ================================================
    print("Building 8 hidden view sheets ...")
    for view_num in range(1, 9):
        ws_v = wb.create_sheet(f"_view{view_num}")
        matrix = pivots[view_num]
        for ri, row_vals in enumerate(matrix):
            for ci, val in enumerate(row_vals):
                ws_v.cell(row=ri + 1, column=ci + 1, value=val)
        ws_v.sheet_state = "hidden"

    # == loan_detail removed — "Loan Detail" sheet IS the data sheet now ======
    # No 50K row copy needed. Drill-down filters "Loan Detail" directly.

    # == Summary sheet =======================================================
    print("Building Summary sheet ...")
    build_summary_sheet(wb, pivots)

    # == POWER_QUERY_SETUP sheet =============================================
    print("Building POWER_QUERY_SETUP sheet ...")
    build_pq_setup_sheet(wb)

    # == Tracker Analytics sheet ================================================
    print("Building Tracker Analytics sheet ...")
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from build_tracker_analytics import build_tracker_analytics_sheet
    build_tracker_analytics_sheet(wb)

    # == _config sheet (hidden) ==============================================
    print("Building _config sheet ...")
    ws_cfg = wb.create_sheet("_config")
    ws_cfg["A1"].value = 2   # ViewState (Summary Commitment)
    ws_cfg["A2"].value = 0   # WMLCState (OFF)
    ws_cfg["A3"].value = TAGGED_CSV  # DataSourcePath
    ws_cfg.sheet_state = "hidden"

    # == Named ranges ========================================================
    from openpyxl.workbook.defined_name import DefinedName

    dn_view = DefinedName("ViewState", attr_text="_config!$A$1")
    wb.defined_names.add(dn_view)

    dn_wmlc = DefinedName("WMLCState", attr_text="_config!$A$2")
    wb.defined_names.add(dn_wmlc)

    dn_path = DefinedName("DataSourcePath", attr_text="_config!$A$3")
    wb.defined_names.add(dn_path)

    # == Move Dashboard to first position ====================================
    wb.move_sheet("Dashboard", offset=-1)

    # == Save .xlsx (Stage 1) ================================================
    print(f"Saving .xlsx to {OUTPUT_XLSX} ...")
    wb.save(OUTPUT_XLSX)
    size = os.path.getsize(OUTPUT_XLSX)
    print(f"  Saved! Size: {size:,} bytes ({size/1024:.1f} KB)")

    # == Validation ==========================================================
    sheets = wb.sheetnames
    print(f"  Sheets: {sheets}")
    assert "Dashboard" in sheets
    assert "Loan Detail" in sheets
    assert "_config" in sheets
    assert "_chart_data" in sheets
    assert "Summary" in sheets
    assert "POWER_QUERY_SETUP" in sheets
    assert "Tracker Analytics" in sheets
    for i in range(1, 9):
        assert f"_view{i}" in sheets

    # Verify named ranges
    dn_names = list(wb.defined_names)
    assert "ViewState" in dn_names, f"ViewState not in {dn_names}"
    assert "WMLCState" in dn_names, f"WMLCState not in {dn_names}"
    assert "DataSourcePath" in dn_names, f"DataSourcePath not in {dn_names}"

    print("\n=== Stage 1 complete (openpyxl .xlsx) ===")

    # Quick pivot sanity check
    for v in range(1, 9):
        total = sum(sum(r) for r in pivots[v])
        print(f"  _view{v} grand total: {total:,.0f}")

    # == Generate VBA_REFERENCE.txt ==========================================
    print("\nGenerating VBA_REFERENCE.txt ...")
    generate_vba_reference()

    # == Stage 2: COM post-processing ========================================
    print("\n=== Stage 2: COM automation (VBA + buttons) ===")
    try:
        inject_vba_and_buttons(OUTPUT_XLSX, OUTPUT_XLSM, VBA_DIR)

        # Verify .xlsm
        xlsm_size = os.path.getsize(OUTPUT_XLSM)
        print(f"\n=== Dashboard build complete ===")
        print(f"  Output: {OUTPUT_XLSM}")
        print(f"  Size:   {xlsm_size:,} bytes ({xlsm_size/1024:.1f} KB)")
        print(f"  Sheets: {len(sheets)} ({', '.join(s for s in sheets if not s.startswith('_'))})")
        print(f"  Hidden: {', '.join(s for s in sheets if s.startswith('_'))}")

        # Verify ZIP contains vbaProject.bin
        import zipfile
        with zipfile.ZipFile(OUTPUT_XLSM, 'r') as zf:
            names = zf.namelist()
            has_vba = any("vbaProject" in n for n in names)
            print(f"  vbaProject.bin in ZIP: {has_vba}")
            if not has_vba:
                print("  WARNING: VBA project not found in .xlsm!")

    except Exception as e:
        print(f"\n  COM automation failed: {e}")
        print("  The .xlsx file was created successfully at:")
        print(f"    {OUTPUT_XLSX}")
        print("  You will need to manually import VBA modules.")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
