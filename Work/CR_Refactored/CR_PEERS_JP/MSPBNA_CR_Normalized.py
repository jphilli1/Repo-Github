# Standard library imports
import asyncio
import csv
import inspect
import io
import logging
import os
import re
import sys
import time
import warnings
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Third-party imports
import aiohttp
import numpy as np
import openpyxl
import pandas as pd
import requests
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats
from tqdm import tqdm
from tqdm.asyncio import tqdm as tqdm_asyncio

# --- Consolidated Data Dictionary (same package) ---
from master_data_dictionary import MasterDataDictionary, LOCAL_DERIVED_METRICS



script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# ==================================================================================
#  DETERMINISTIC CERT CONFIGURATION
# ==================================================================================
# Load .env if available (best-effort)
try:
    from dotenv import load_dotenv as _ld
    _env_path = Path(script_dir) / ".env"
    if _env_path.exists():
        _ld(_env_path)
except ImportError:
    pass


token = os.getenv("HUD_USER_TOKEN")
print("HUD_USER_TOKEN found:", bool(token))
print("Length:", len(token) if token else 0)
print("Prefix:", token[:6] + "..." if token else None)
_mspbna_raw = os.getenv("MSPBNA_CERT")
_msbna_raw = os.getenv("MSBNA_CERT")
if not _mspbna_raw or not _msbna_raw:
    raise ValueError(
        "MSPBNA_CERT and MSBNA_CERT environment variables are required.\n"
        "Set them before running:\n"
        "  export MSPBNA_CERT=34221\n"
        "  export MSBNA_CERT=32992\n"
        "Or create a .env file in the script directory with those values."
    )
MSPBNA_CERT = int(_mspbna_raw)
MSBNA_CERT = int(_msbna_raw)
MS_COMBINED_CERT = int(os.getenv("MS_COMBINED_CERT", "88888"))

# Composite aggregation: "mean" or "weighted" (auto-detect if unset)
COMPOSITE_METHOD = os.getenv("COMPOSITE_METHOD", "").lower() or None
MIN_PEER_MEMBERS = int(os.getenv("MIN_PEER_MEMBERS", "2"))

# ==================================================================================
#  1. SCRIPT CONFIGURATION & SETUP
# ==================================================================================

def setup_logging(log_dir: str = "logs") -> logging.Logger:
    """Setup comprehensive logging configuration."""
    Path(log_dir).mkdir(exist_ok=True)
    log_filename = f"{log_dir}/bank_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    # Remove existing handlers to avoid duplicate logs in interactive environments
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[logging.FileHandler(log_filename), logging.StreamHandler(sys.stdout)]
    )
    logger = logging.getLogger(__name__)
    logger.info(f"Logging initialized. Log file: {log_filename}")
    return logger

logger = setup_logging()

@dataclass
class DashboardConfig:
    """Clean configuration class to hold runtime parameters."""
    fred_api_key: str
    subject_bank_cert: int
    peer_bank_certs: List[int]
    quarters_back: int = 30
    fred_years_back: int = 15
    output_dir: str = "output"
    fdic_api_base: str = "https://banks.data.fdic.gov/api"
    fred_api_base: str = "https://api.stlouisfed.org/fred"

# ==================================================================================
#  2. GLOBAL DATA DICTIONARIES & CONSTANTS
# ==================================================================================

# [NEW] FIELD RESOLUTION LAYER
# Maps conceptual/synthetic fields to authoritative Call Report series.
# Keys = The legacy/conceptual name you want to use.
# Values = The list of authoritative series to check in order (Consolidated -> Domestic -> Proxy).
FDIC_FALLBACK_MAP = {
    # --- SBL & Fund Finance (FFIEC Bulk / Consolidated) ---
    "LNOTHPCS":    ["RCFD1545", "RCON1545", "LNOTHER"],  # Legacy SBL -> Consolidated SBL
    "LNOTHNONDEP": ["RCFDJ454", "RCONJ454"],             # Legacy Fund Fin -> Consolidated Fund Fin

    # NOTE:
    # Do NOT map 'LNRERES' or 'LNRELOC' here.
    # They are FDIC API fields and must remain sourced from FDIC, otherwise Standard Resi can go to 0
    # when FFIEC-only MDRMs are missing in the FDIC dataframe.

    # --- RI-C Disaggregated Allowance (FFIEC Bulk Only) ---
    "RCFDJ466": ["FFIEC"],  # Construction


    # --- RI-C Disaggregated Allowance (FFIEC Bulk Only) ---
    "RCFDJ466": ["FFIEC"], # Construction
    "RCFDJ467": ["FFIEC"], # Commercial RE
    "RCFDJ468": ["FFIEC"], # Residential
    "RCFDJ469": ["FFIEC"], # C&I
    "RCFDJ470": ["FFIEC"], # Credit Cards
    "RCFDJ471": ["FFIEC"], # Other Consumer
    "RCFDJ472": ["FFIEC"], # Unallocated
    "RCFDJ474": ["FFIEC"], # Other Loans (SBL Proxy)

    # --- Consumer Granular (Derived) ---
    "LNCONAUTO": ["LNAUTO"],
    "LNCONCC":   ["LNCRCD"],
    "LNCONOTHX": ["DERIVED"], # Calculated as LNCON - LNAUTO - LNCRCD

    # --- Legacy Real Estate Aliases (Mapping to standard codes) ---
    "LNREOTH":  ["LNRENROT"], # "Income Producing" -> Non-Owner Occupied Nonfarm
    "NALRERES": ["NARERES"],  # Legacy Resi NA -> Total Resi NA
    "P9RENRES": ["P9RENROT"], # Legacy Nonfarm NA -> Non-Owner Nonfarm NA
}

# [UPDATED] MASTER FETCH LIST
# Includes ALL valid API fields + Legacy placeholders (resolved later)
FDIC_FIELDS_TO_FETCH = [
    # --- Identifiers & Dates ---
    "CERT", "NAME", "REPDTE",

    # --- Balance Sheet Totals ---
    "ASSET", "LIAB", "DEP", "EQ",

    # --- Profitability ---
    "ROA", "ROE", "NIMY", "EEFFR", "NONIIAY", "ELNATRY", "EINTEXP",

    # --- Capital & RWA ---
    "RBCT1CER", "RBCT1J", "RBCT2", "RWAJ", "RBCRWAJ", "RB2LNRES",
    "EQCS", "EQSUR", "EQUP", "EQCCOMPI", "EQPP",

    # --- Reserves & Funding ---
    "LNATRES", "OTHBOR", "MUTUAL", "FREPP",

    # ==========================================================================
    # RAW BALANCE SHEET SERIES — FDIC Text Aliases (handle 031/041 split natively)
    # ==========================================================================
    # FDIC top-level text aliases automatically resolve the FFIEC 031 (Consolidated)
    # vs FFIEC 041/051 (Domestic) split. Prefer these over raw MDRM codes.
    # Replaces: RCFD2170→ASSET, RCFD2948→LIAB, RCFD3210→EQ, RCFD2200→DEP,
    #           RCFD1400→LNLSGR, RCFD3123→LNATRES (already in list above)
    "LNLSGR",    # Gross Loans (replaces RCFD1400)

    # --- Liquidity Components (Text Aliases where available) ---
    "CHBAL",      # Cash and Balances Due (replaces RCFD0010)
    "CHBALNI",    # Non-Interest-Bearing Cash (replaces RCFD0081)
    "TRD",        # Trading Assets (replaces RCFD3545)
    # HTM/AFS Securities: No text alias — fetch BOTH RCFD + RCON for 031/041 waterfall
    "RCFD1754", "RCON1754",   # Held-to-Maturity Securities (Amortized Cost)
    "RCFD1773", "RCON1773",   # Available-for-Sale Securities (Fair Value)

    # --- Unused Commitments: No text alias — fetch BOTH RCFD + RCON ---
    "RCFD3423", "RCON3423",   # Unused Commitments
    "RCFD3814", "RCON3814",   # HELOC Unused Commitments
    "RCFD6550", "RCON6550",   # Non-RE Commitments

    # --- Raw Income Statement (Text Aliases) ---
    "NETINC",     # Net Income (replaces RIAD4340)
    "INTINC",     # Total Interest Income (replaces RIAD4107)
    "INTEXP",     # Total Interest Expense (replaces RIAD4073)
    "NONII",      # Total Noninterest Income (replaces RIAD4079)
    "NONIX",      # Total Noninterest Expense (replaces RIAD4093)
    "IBTX",       # Income Before Taxes (replaces RIAD4301)
    "ILNDOM",     # Interest Income on Loans (Domestic) — replaces RIAD4010 part 1
    "ILNFOR",     # Interest Income on Loans (Foreign) — replaces RIAD4010 part 2
    "RIAD4115",   # Interest Expense on Deposits (raw — no alias)
    "NTCI",       # C&I Net Charge-Offs (replaces RIAD4638 + RIAD4608)
    "RIAD4609",   # Total Recoveries (raw — no alias)
    "NTLS",       # Total Loan & Lease NCOs (replaces RIAD4658 + RIAD4659)

    # --- INCOME & PROFITABILITY (Corrected Series) ---
    # Note: ILNDOM and ILNFOR now fetched via text aliases above
    "EDEPDOM",  # Int Exp Deposits (Domestic)
    "EDEPFOR",  # Int Exp Deposits (Foreign)
    "ELNATR",   # Provision for Credit Losses (YTD)

    # --- TOP HOUSE DELINQUENCY ---
    "P3LNLS",   # PD 30-89 Total
    "P9LNLS",   # PD 90+ Total


    # --- LOAN BALANCES (Authoritative) ---
    "LNLS", "LNLSNET",
    "LNCI",       # C&I
    "LNRECONS",   # Construction
    "LNREMULT",   # Multifamily
    "LNRENROW",   # Owner-Occ CRE
    "LNRENROT",   # Non-Owner CRE (Income Producing)
    "LNREAG",     # Farmland
    "LNRERES",    # 1-4 Family Resi Total
    "LNRELOC",    # HELOC
    "LNCON",      # Consumer Total
    "LNCRCD",     # Credit Cards
    "LNAUTO",     # Auto
    "LNOTHER",    # All Other Loans
    "LS",         # Leases
    "LNAG",       # Ag Loans

    # --- LEGACY BALANCES (Mapped via Fallback) ---
    "LNREOTH",    # Will map to LNRENROT
    # NOTE: LNOTHPCS, LNOTHNONDEP, LNCONAUTO, LNCONCC, LNCONOTHX are synthetic
    # variables that do not exist in the FDIC API. They are calculated locally
    # in pandas AFTER the API fetch completes (see FDIC_FALLBACK_MAP resolution).

    # --- NET CHARGE-OFFS ---
    "NTLNLS", "NCLNLS",
    "NTCI",
    "NTRECONS", "NTREMULT", "NTRENROT", "NTRENROW", "NTREAG", "NTRENRES", # <--- Added NTRENROW
    "NTRERES", "NTRELOC", # Other Consumer NCO

    # --- PAST DUE 30-89 ---
    "P3LNLS", "P3CI",
    "P3RECONS", "P3LREMUL", "P3RENROT", "P3RENROW", "P3REAG", "P3RENRES",
    "P3RERES", "P3RELOC",
    "P3CON", "P3CRCD", "P3AUTO", "P3LS", "P3AG", "P3OTHLN",
    "P3CONOTH",

    # --- PAST DUE 90+ ---
    "P9LNLS", "P9CI",
    "P9RECONS", "P9REMULT", "P9RENROT", "P9RENROW", "P9REAG", "P9RENRES",
    "P9RERES", "P9RELOC",
    "P9CON", "P9CRCD", "P9AUTO", "P9LS", "P9AG", "P9OTHLN",
    "P9CONOTH",

    # --- NONACCRUAL ---
    "NACI",
    "NARECONS", "NAREMULT", "NARENROT", "NARENROW", "NAREAG", "NARENRES",
    "NARERES", "NARELOC",
    "NACON", "NACRCD", "NAAUTO", "NALS", "NAAG", "NAOTHLN",
    "NACONOTH",
    "NALRERES", # Legacy

    # =============================================================================
    # NORMALIZATION EXCLUSION FIELDS (Ex-Commercial/Ex-Consumer Segments)
    # =============================================================================
    # These fields are used to create "apples-to-apples" comparisons by removing
    # Mass Market Consumer and Commercial Banking segments that MSPBNA does not
    # participate in.

    # --- 1. Domestic C&I (Remove standard business lending) ---
    "RCON1763",    # Balance: Commercial & Industrial loans to U.S. addressees (Domestic)
    # NTCI (text alias) replaces RIAD4638/RIAD4608 — already in fetch list above
    # P3CI, P9CI, NACI (text aliases) replace RCON1606/1607/1608 — already in fetch list above

    # --- 2. Nondepository Financial Institutions (NDFI) - Fund Finance/Shadow Banking ---
    # --- 2. Nondepository Financial Institutions (NDFI) - Fund Finance/Shadow Banking ---
    "RCONJ454", "RCFDJ454",    # Balance: Loans to nondepository financial institutions
    # NCO: Note - NDFI NCOs not separately reported (buried in "All Other")
    # NDFI (Past Due / Nonaccrual)
    'RCONJ458', 'RCFDJ458',    # PD30: NDFI Past Due 30-89
    'RCONJ459', 'RCFDJ459',    # PD90: NDFI Past Due 90+
    'RCONJ460', 'RCFDJ460',    # NA: NDFI Nonaccrual

    "RCFDF162",


    # --- 3. ADC Loans (Remove Construction Risk) ---
    "RCON1420",    # Balance: Construction, land development, and other land loans - Total
    # NTLS (text alias) replaces RIAD4658/RIAD4659 — already in fetch list above
    # P3RECONS, P9RECONS, NARECONS (text aliases) replace RCON2759/2769/3492 — already in fetch list above

    # --- 4. Mass Market Consumer (Credit Cards, Auto, Ag) ---
    "RCFDB538",    # Balance: Credit Card Loans
    "RIADB514",    # NCO: Credit Card Charge-offs
    "RIADB515",    # NCO: Credit Card Recoveries
    "RCFDB575",    # NA: Credit Card Nonaccrual
    "RCFDK137",    # Balance: Auto Loans
    "RIADK205",    # NCO: Auto Charge-offs
    "RIADK206",    # NCO: Auto Recoveries
    "RCFDK213",    # NA: Auto Nonaccrual
    "RCFD1590",    # Balance: Agricultural Loans
    "RIAD4635",    # NCO: Agricultural Charge-offs
    "RIAD4645",    # NCO: Agricultural Recoveries
    "RCFD5341",    # NA: Agricultural Nonaccrual
    "RCON2746",    # PD30: Agricultural Past Due 30-89
    "RCON2747",    # PD90: Agricultural Past Due 90+

    # --- 5. Foreign Government & Banks (Exclude from Domestic Peer View) ---
    "RCFD2081",    # Balance: Loans to Foreign Governments
    "RCFD2005",    # Balance: Loans to Depository Institutions (Banks)

    # --- 6. C&I NCO Proxy (Alternative to segment-level calculation) ---
]



# ==================================================================================
#  MASTER DATA DICTIONARY (replaces hardcoded FDIC_FIELD_DESCRIPTIONS)
# ==================================================================================
# Single instance — lives in the same package as master_data_dictionary.py.
# Tier 1: Federal Reserve MDRM CSV | Tier 2: FDIC API | Tier 3: Local/Derived
_master_dict = MasterDataDictionary(
    cache_dir=os.path.join(script_dir, ".data_dictionary_cache")
)


def _get_metric_short_name(code: str) -> str:
    """Quick helper: resolve a metric code to its short display name."""
    result = _master_dict.lookup_metric(code)
    if result["Source_of_Truth"] == "Not Found":
        return code
    return result["Metric_Name"]


from enum import Enum

# =============================================================================
#  PEER GROUP DEFINITIONS
# =============================================================================

class PeerGroupType(str, Enum):
    CORE_PRIVATE_BANK = "Core_Private_Bank"
    MS_FAMILY_PLUS = "MS_Family_Plus"
    ALL_PEERS = "All_Peers"
    # Normalized Peer Groups (Ex-Commercial/Ex-Consumer view)
    CORE_PRIVATE_BANK_NORM = "Core_Private_Bank_Norm"
    MS_FAMILY_PLUS_NORM = "MS_Family_Plus_Norm"
    ALL_PEERS_NORM = "All_Peers_Norm"

PEER_GROUPS = {
    PeerGroupType.CORE_PRIVATE_BANK: {
        "name": "Core Private Bank Peers",
        "short_name": "Core PB",
        "description": "True private banking comparables - SBL, wealth management, UHNW focus",
        "certs": [MSPBNA_CERT, 33124, 57565],  # MSPBNA + GS + UBS
        "use_case": "Best for SBL/wealth product comparisons, NCO benchmarking",
        "display_order": 1,
        "use_normalized": False
    },
    PeerGroupType.MS_FAMILY_PLUS: {
        "name": "MSPBNA + Wealth",
        "short_name": "MSPBNA+Wealth",
        "description": "MSPBNA plus wealth management peers (excludes MSBNA)",
        "certs": [MSPBNA_CERT, 33124, 57565],  # MSPBNA + GS + UBS  (NO MSBNA)
        "use_case": "Internal MS comparison plus broader wealth industry view",
        "display_order": 2,
        "use_normalized": False
    },
    PeerGroupType.ALL_PEERS: {
        "name": "Full Peer Universe",
        "short_name": "All Peers",
        "description": "Complete peer set including MSBNA and G-SIBs for size/scale context",
        "certs": [MSBNA_CERT, 33124, 57565, 628, 3511, 7213, 3510],  # MSBNA + full universe
        "use_case": "Regulatory comparison, market share analysis, full industry context",
        "display_order": 3,
        "use_normalized": False
    },
    # ==========================================================================
    # NORMALIZED PEER GROUPS (Ex-Commercial/Ex-Consumer)
    # ==========================================================================
    PeerGroupType.CORE_PRIVATE_BANK_NORM: {
        "name": "Core Private Bank (Normalized)",
        "short_name": "Core PB Norm",
        "description": "Core PB peers with normalized metrics - strips C&I, ADC, NDFI, Cards, Auto, Ag",
        "certs": [MSPBNA_CERT, 33124, 57565],  # Same as CORE_PRIVATE_BANK
        "use_case": "True private bank comparison excluding mass market and commercial segments",
        "display_order": 4,
        "use_normalized": True
    },
    PeerGroupType.MS_FAMILY_PLUS_NORM: {
        "name": "MSPBNA + Wealth (Normalized)",
        "short_name": "MSPBNA+Wealth Norm",
        "description": "MSPBNA + wealth peers with normalized metrics (excludes MSBNA)",
        "certs": [MSPBNA_CERT, 33124, 57565],  # Same as MS_FAMILY_PLUS (NO MSBNA)
        "use_case": "Internal MS comparison on private bank comparable portfolios",
        "display_order": 5,
        "use_normalized": True
    },
    PeerGroupType.ALL_PEERS_NORM: {
        "name": "Full Peer Universe (Normalized)",
        "short_name": "All Peers Norm",
        "description": "All peers with normalized metrics for private bank comparable view",
        "certs": [MSBNA_CERT, 33124, 57565, 628, 3511, 7213, 3510],  # Same as ALL_PEERS
        "use_case": "Broad comparison on normalized (ex-commercial/ex-consumer) basis",
        "display_order": 6,
        "use_normalized": True
    }
}

# Helper to ensure we fetch data for ALL distinct certs mentioned in any group
def get_all_peer_certs():
    all_certs = set()
    for group in PEER_GROUPS.values():
        all_certs.update(group['certs'])
    return list(all_certs)
# MSPBNA V6 SEGMENTATION (Adapted for V5 Structure)
# NOTE: SBL and Fund Finance NCOs are not separately reported in Call Reports.
# They are lumped into 'NTOTH' (All Other). They are left empty [] here to
# prevent double-counting. You must calculate them via allocation if needed.
LOAN_CATEGORIES = {
    # 1. SBL & LIQUIDITY
    "SBL": {
        "balance": ["SBL_Balance"], # Resolved in Processor
        "nco": [], "pd30": [], "pd90": [], "na": []
    },
    # 2. FUND FINANCE
    "Fund_Finance": {
        "balance": ["Fund_Finance_Balance"], # Resolved in Processor
        "nco": [], "pd30": ["P3NDFI"], "pd90": ["P9NDFI"], "na": ["NANDFI"]
    },
    # 3. WEALTH RESIDENTIAL
    "Wealth_Resi": {
        "balance": ["Wealth_Resi_Balance"], # Collapsed in Processor
        "nco": ["NTRERES", "NTRELOC"],
        "pd30": ["P3RERES", "P3RELOC"],
        "pd90": ["P9RERES", "P9RELOC"],
        "na": ["NARERES", "NARELOC"]
    },
    # 4. TRADITIONAL C&I
    "Corp_CI": {
        "balance": ["Corp_CI_Balance"],
        "nco": ["NTCI"], "pd30": ["P3CI"], "pd90": ["P9CI"], "na": ["NACI"]
    },
    # 5. CRE: OWNER-OCCUPIED
    "CRE_OO": {
        "balance": ["CRE_OO_Balance"],
        "nco": ["NTRENROW"], "pd30": ["P3RENROW"], "pd90": ["P9RENROW"], "na": ["NARENROW"]
    },
    # 6. CRE: INVESTMENT
    "CRE_Investment": {
        "balance": ["CRE_Investment_Balance"],
        "nco": ["NTREMULT", "NTRENROT"],
        "pd30": ["P3REMULT", "P3RENROT"],
        "pd90": ["P9REMULT", "P9RENROT"],
        "na": ["NAREMULT", "NARENROT"]
    },
    # 7. CONSUMER: AUTO
    "Consumer_Auto": {
        "balance": ["Consumer_Auto_Balance"],
        "nco": ["NTAUTO"], "pd30": ["P3AUTO"], "pd90": ["P9AUTO"], "na": ["NAAUTO"]
    },
    # 8. CONSUMER: OTHER
    "Consumer_Other": {
        "balance": ["Consumer_Other_Balance"], # Derived Residual in Processor
        "nco": ["NTCON", "NTCRCD"],
        "pd30": ["P3CON", "P3CRCD"],
        "pd90": ["P9CON", "P9CRCD"],
        "na": ["NACON", "NACRCD"]
    }
}
FRED_SERIES_TO_FETCH = {
    'Key Economic Indicators': {
        'GDPC1': {'short': 'Real GDP', 'long': 'Real Gross Domestic Product'},
        'A191RL1Q225SBEA': {'short': 'Real GDP Growth', 'long': 'Real Gross Domestic Product: Percent Change from Preceding Period'},
        'UNRATE': {'short': 'Unemployment Rate', 'long': 'Unemployment Rate'},
        'CPIAUCSL': {'short': 'CPI Inflation', 'long': 'Consumer Price Index for All Urban Consumers: All Items'},
        'UMCSENT': {'short': 'Consumer Sentiment', 'long': 'University of Michigan: Consumer Sentiment'},
        'ICSA': {'short': 'Initial Jobless Claims', 'long': 'Initial Claims'},
        'USSLIND': {'short': 'Leading Index', 'long': 'Leading Index for the United States'},
        'RSXFS': {'short': 'Retail Sales', 'long': 'Advance Retail Sales: Retail and Food Services'}
    },
    'Interest Rates & Yield Curve': {
        'FEDFUNDS': {'short': 'Fed Funds Rate', 'long': 'Effective Federal Funds Rate'},
        'DFF': {'short': 'Fed Funds (Daily)', 'long': 'Effective Federal Funds Rate (Daily)'},
        'DPRIME': {'short': 'Prime Rate', 'long': 'Bank Prime Loan Rate'},
        'MORTGAGE30US': {'short': '30Y Mortgage Rate', 'long': '30-Year Fixed Rate Mortgage Average in the United States'},
        'DGS30': {'short': 'UST 30Y', 'long': 'Market Yield on U.S. Treasury Securities at 30-Year Constant Maturity'},
        'DGS20': {'short': 'UST 20Y', 'long': 'Market Yield on U.S. Treasury Securities at 20-Year Constant Maturity'},
        'DGS10': {'short': 'UST 10Y', 'long': 'Market Yield on U.S. Treasury Securities at 10-Year Constant Maturity'},
        'DGS7': {'short': 'UST 7Y', 'long': 'Market Yield on U.S. Treasury Securities at 7-Year Constant Maturity'},
        'DGS5': {'short': 'UST 5Y', 'long': 'Market Yield on U.S. Treasury Securities at 5-Year Constant Maturity'},
        'DGS3': {'short': 'UST 3Y', 'long': 'Market Yield on U.S. Treasury Securities at 3-Year Constant Maturity'},
        'DGS2': {'short': 'UST 2Y', 'long': 'Market Yield on U.S. Treasury Securities at 2-Year Constant Maturity'},
        'DGS1': {'short': 'UST 1Y', 'long': 'Market Yield on U.S. Treasury Securities at 1-Year Constant Maturity'},
        'DGS6MO': {'short': 'UST 6M', 'long': 'Market Yield on U.S. Treasury Securities at 6-Month Constant Maturity'},
        'DGS3MO': {'short': 'UST 3M', 'long': 'Market Yield on U.S. Treasury Securities at 3-Month Constant Maturity'},
        'DGS1MO': {'short': 'UST 1M', 'long': 'Market Yield on U.S. Treasury Securities at 1-Month Constant Maturity'},
        'T10Y2Y': {'short': '10Y-2Y', 'long': '10-Year Treasury Constant Maturity Minus 2-Year Treasury Constant Maturity'},
        'T10Y3M': {'short': '10Y-3M', 'long': '10-Year Treasury Constant Maturity Minus 3-Month Treasury Constant Maturity'}
    },
    'Credit Spreads & Lending Standards': {
        'DBAA': {'short': "Moody's Baa Yield", 'long': "Moody's Seasoned Baa Corporate Bond Yield"},
        'DAAA': {'short': "Moody's Aaa Yield", 'long': "Moody's Seasoned Aaa Corporate Bond Yield"},
        'BAMLH0A0HYM2': {'short': 'HY OAS', 'long': 'ICE BofA US High Yield Index Option-Adjusted Spread'},
        'BAMLC0A0CM': {'short': 'IG OAS', 'long': 'ICE BofA US Corporate Index Option-Adjusted Spread'},
        'DRTSCILM': {'short': 'C&I Standards (Large/Med)', 'long': 'Net Percentage of Domestic Banks Tightening Standards for C&I Loans to Large and Middle-Market Firms'},
        'DRTSCIS': {'short': 'C&I Standards (Small)', 'long': 'Net Percentage of Domestic Banks Tightening Standards for C&I Loans to Small Firms'}
    },
    'Financial Stress & Risk': {
        'STLFSI4': {'short': 'St. Louis FSI', 'long': 'St. Louis Fed Financial Stress Index'},
        'VIXCLS': {'short': 'VIX', 'long': 'CBOE Volatility Index: VIX'},
        'NFCI': {'short': 'NFCI', 'long': 'Chicago Fed National Financial Conditions Index'}
    },
    'Investor Leverage & Market Credit': {
        'BOGZ1FL663067003Q': {
            'short': 'Broker-Dealer Credit Balances',
            'long': 'Security Brokers and Dealers; Credit Balances; Asset (Z.1 Financial Accounts). Proxy for securities-based leverage used mainly by wealthy/institutional investors.'
        },
        'BOGZ1FL663067005Q': {
            'short': 'Broker-Dealer Total Financial Assets',
            'long': 'Security Brokers and Dealers; Total Financial Assets (Z.1 Financial Accounts). Denominator for normalizing broker-dealer credit balances.'
        }
    },
    'Global Benchmarks': {
        'DEXUSEU': {'short': 'USD/EUR', 'long': 'U.S. Dollars to Euro Spot Exchange Rate'},
        'DEXJPUS': {'short': 'JPY/USD', 'long': 'Japanese Yen to U.S. Dollar Spot Exchange Rate'},
        'DEXUSUK': {'short': 'USD/GBP', 'long': 'U.S. Dollars to British Pound Sterling Spot Exchange Rate'},
        'DCOILWTICO': {'short': 'WTI Oil', 'long': 'Crude Oil Prices: West Texas Intermediate (WTI) - Cushing, Oklahoma'},
        'GOLDAMGBD228NLBM': {'short': 'Gold', 'long': 'Gold Fixing Price 10:30 A.M. (London time) in London Bullion Market, based in U.S. Dollars'},
        'VIXCLS': {'short': 'VIX', 'long': 'CBOE Volatility Index: VIX'}
    },
    'Real Estate & Housing': {
        'CSUSHPINSA': {'short': 'Case-Shiller National', 'long': 'S&P CoreLogic Case-Shiller U.S. National Home Price Index'},
        'HOUST': {'short': 'Housing Starts', 'long': 'Housing Starts: Total: New Privately Owned Housing Units Started'},
        'PERMIT': {'short': 'Building Permits', 'long': 'New Private Housing Units Authorized by Building Permits'},
        'MSPUS': {'short': 'Median Sales Price', 'long': 'Median Sales Price of Houses Sold for the United States'},
        'RRVRUSQ156N': {'short': 'Vacancy Rate (Rental)', 'long': 'Rental Vacancy Rate for the United States'},
        'RCVRUSQ156N': {'short': 'Vacancy Rate (Homeowner)', 'long': 'Homeowner Vacancy Rate for the United States'}
    },
    'Banking Sector Aggregates': {
        'TOTLL': {'short': 'Total Loans & Leases', 'long': 'Total Loans and Leases, All Commercial Banks'},
        'BUSLOANS': {'short': 'Business Loans', 'long': 'Commercial and Industrial Loans, All Commercial Banks'},
        'REALLN': {'short': 'Real Estate Loans', 'long': 'Real Estate Loans, All Commercial Banks'},
        'CCLACBW027SBOG': {'short': 'Credit Card Loans', 'long': 'Consumer Loans: Credit Cards and Other Revolving Plans, All Commercial Banks'},
        'CONSUMER': {'short': 'Consumer Loans', 'long': 'Consumer Loans, All Commercial Banks'},
        'DEPALL': {'short': 'Total Deposits', 'long': 'Total Deposits, All Commercial Banks'},
        'DPSACBW027SBOG': {'short': 'Savings Deposits', 'long': 'Deposits, Savings Accounts, All Commercial Banks'},
        'DODFFSWCMI': {'short': 'Deposits: Other', 'long': 'Other Deposits, All Commercial Banks'},
        'CORBLACBS': {'short': 'Bus Loan CO Rate', 'long': 'Charge-off Rate on Business Loans, Annualized, All Commercial Banks'},
        'CORALACBS': {'short': 'All Loans CO Rate', 'long': 'Charge-off Rate on All Loans, Annualized, All Commercial Banks'},
        'CORCCLACBS': {'short': 'CC CO Rate', 'long': 'Charge-off Rate on Credit Card Loans, Annualized, All Commercial Banks'},
        'DRALACBS': {'short': 'All Loans Delinq', 'long': 'Delinquency Rate on All Loans, All Commercial Banks'},
        'DRCCLACBS': {'short': 'CC Delinq Rate', 'long': 'Delinquency Rate on Credit Card Loans, All Commercial Banks'},
        'DRCRELEXFACBS': {'short': 'CRE Delinq (ex-farm)', 'long': 'Delinquency Rate on Commercial Real Estate Loans (Excluding Farmland), All Commercial Banks'},
        'DRSFRMACBS': {'short': '1-4 Resi Delinq', 'long': 'Delinquency Rate on Single-Family Residential Mortgages, All Commercial Banks'},
        'DRSFRMT100S': {'short': '1-4 Resi Delinq (Top 100)', 'long': 'Delinquency Rate on Single-Family Residential Mortgages, Banks Ranked 1st to 100th Largest in Size by Assets (SA)'},
        'DRSFRMT100N': {'short': '1-4 Resi Delinq (Top 100, NSA)', 'long': 'Delinquency Rate on Single-Family Residential Mortgages, Banks Ranked 1st to 100th Largest in Size by Assets (NSA)'},
        'DRCCLT100S': {'short': 'Credit Card Delinq (Top 100)', 'long': 'Delinquency Rate on Credit Card Loans, Banks Ranked 1st to 100th Largest in Size by Assets (SA)'},
        'CORALACBN': {'short': 'All Loans CO Rate (NSA)', 'long': 'Charge-off Rate on All Loans, Annualized, All Commercial Banks (NSA)'}
    },
    'Leading Indicators': {
        'USSLIND': {'short': 'Leading Index', 'long': 'Leading Index for the United States'},
        'USALOLITONOSTSAM': {'short': 'US Leading Indicator', 'long': 'US Leading Index: Leading Index'},
        'PAYEMS': {'short': 'Nonfarm Payrolls', 'long': 'All Employees: Total Nonfarm Payrolls'},
        'PERMIT': {'short': 'Building Permits', 'long': 'New Private Housing Units Authorized by Building Permits'}
    },
    'Middle Market, Healthcare, & Funding Indicators': {
        'DRTSCIS': {'short': 'Small Firm C&I Standards', 'long': 'Net Pct Banks Tightening Standards - Small Firms'},
        'TCU': {'short': 'Capacity Utilization: Total Industry', 'long': 'Capacity Utilization: Total Index'},
        'INDPRO': {'short': 'Industrial Production Index', 'long': 'Industrial Production: Total Index'},
        'NEWORDER': {'short': "Manufacturers' New Orders", 'long': "Manufacturers' New Orders: Nondefense Capital Goods Excluding Aircraft"},
        'SOFR': {'short': 'SOFR', 'long': 'Secured Overnight Financing Rate'},
        'TB3MS': {'short': '3-Month T-Bill', 'long': '3-Month Treasury Bill Secondary Market Rate'},
        'SOFR3MTB3M': {'short': 'SOFR vs T-Bill Spread', 'long': 'Calculated Spread: SOFR - 3-Month T-Bill'},
        'MPCT04XXS': {'short': 'Medicare Spending', 'long': 'Medicare: Total Expenditures'}
    }
}

FDIC_FIELDS_TO_FETCH =list(dict.fromkeys(FDIC_FIELDS_TO_FETCH))



# ==================================================================================
#  FIXED FFIEC BULK LOADER CLASS - v4
#  Fixes: Quoted column headers ("IDRSSD"), proper row 0/row 1 handling
#  Replace the existing FFIECBulkLoader class in CR_Bank_DashvMSPB.py with this
# ==================================================================================

class FFIECBulkLoader:
    """
    Robust Data Engine: Fetches granular 'Private Bank' fields from
    FFIEC CDR Bulk Data using a strict ASP.NET WebForms state machine.

    FIXED v4:
    - Handles quoted column headers ("IDRSSD" -> IDRSSD)
    - Row 0 = MDRM column codes (header)
    - Row 1 = Description row (skipped)
    - Row 2+ = Data rows
    - Saves debug extracts when parsing fails
    """

    def __init__(self, output_dir="data/ffiec_cache"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.debug_dir = self.output_dir / "debug"
        self.debug_dir.mkdir(parents=True, exist_ok=True)
        self.zips_dir = self.output_dir / "zips"
        self.zips_dir.mkdir(parents=True, exist_ok=True)
        self.extracted_dir = self.output_dir / "extracted"
        self.extracted_dir.mkdir(parents=True, exist_ok=True)

        self.base_url = "https://cdr.ffiec.gov/public/pws/downloadbulkdata.aspx"

        # Critical Private Banking Fields - expanded for RICII
        self.target_fields = [
            # SBL / Fund Finance (RC-C)
            'RCFD1545', 'RCON1545',
            'RCFDJ454', 'RCONJ454',

            # RI-C / RICII ACL Disaggregated Allowance fields
            'RCFDJ466', 'RCONJ466',
            'RCFDJ467', 'RCONJ467',
            'RCFDJ468', 'RCONJ468',
            'RCFDJ469', 'RCONJ469',
            'RCFDJ470', 'RCONJ470',
            'RCFDJ471', 'RCONJ471',
            'RCFDJ472', 'RCONJ472',
            'RCFDJ473', 'RCONJ473',
            'RCFDJ474', 'RCONJ474',

            # RICII fields (JJ04-JJ23 series)
            'RCFDJJ04', 'RCONJJ04',
            'RCFDJJ05', 'RCONJJ05',
            'RCFDJJ06', 'RCONJJ06',
            'RCFDJJ07', 'RCONJJ07',
            'RCFDJJ08', 'RCONJJ08',
            'RCFDJJ09', 'RCONJJ09',
            'RCFDJJ10', 'RCONJJ10',
            'RCFDJJ11', 'RCONJJ11',
            'RCFDJJ12', 'RCONJJ12',
            'RCFDJJ13', 'RCONJJ13',
            'RCFDJJ14', 'RCONJJ14',
            'RCFDJJ15', 'RCONJJ15',
            'RCFDJJ16', 'RCONJJ16',
            'RCFDJJ17', 'RCONJJ17',
            'RCFDJJ18', 'RCONJJ18',
            'RCFDJJ19', 'RCONJJ19',
            'RCFDJJ20', 'RCONJJ20',
            'RCFDJJ21', 'RCONJJ21',
            'RCFDJJ22', 'RCONJJ22',
            'RCFDJJ23', 'RCONJJ23',
            # --- 2. Nondepository Financial Institutions (NDFI) - Fund Finance/Shadow Banking ---
            'RCONJ454', 'RCFDJ454',    # Balance: Loans to nondepository financial institutions
            # NCO: Note - NDFI NCOs not separately reported (buried in "All Other")
            #      We'll assume 0 NCOs for NDFI given near-zero historical loss rates
            'RCONJ458', 'RCFDJ458',    # PD30: NDFI Past Due 30-89
            'RCONJ459', 'RCFDJ459',    # PD90: NDFI Past Due 90+
            'RCONJ460', 'RCFDJ460',    # NA: NDFI Nonaccrual

            # --- 2b. Other segment (explicitly visible bucket) ---
            'RCONJ451', 'RCFDJ451',    # All other loans (exclude consumer)
            'RCFDF162', 'RCFDF163',    # Leases (RC-C item 10.a / 10.b)
            # --- NEW: Normalization Fields (Mass Market & Commercial) ---
            # 1. Domestic C&I
            'RCON1763', 'RCFD1763', # Balance
            'RCON1606', 'RCFD1606', # P3
            'RCON1607', 'RCFD1607', # P9
            'RCON1608', 'RCFD1608', # NA
            'RIAD4638',             # NCO (Direct)
            'RIAD4608', 'RIAD4609', # Charge-offs/Recoveries
            # Other segment fields (add these)
            'RCONJ451', 'RCFDJ451',     # Other (J451)
            # Other segment fields (add these)
            'RCONJ451', 'RCFDJ451',     # Other (J451)
            'RCONF162', 'RCFDF162',     # Other (F162) - Fixed MDRM format
            'RCONF163', 'RCFDF163',     # Other (F163) - Fixed MDRM format
            'RCONFF162', 'RCFDFF162',   # Other (F162)  <-- use your actual column naming convention if it is RCON/F162 without FF
            'RCONFF163', 'RCFDFF163',   # Other (F163)


            # 2. ADC / Construction
            'RCON1420', 'RCFD1420', # Balance
            'RCON2759', 'RCFD2759', # P3
            'RCON2769', 'RCFD2769', # P9
            'RCON3492', 'RCFD3492', # NA
            'RIAD4658', 'RIAD4659', # NCO

            # 3. Credit Cards
            'RCFDB538', 'RCONB538', # Balance
            'RCFDB572',             # P3
            'RCFDB573',             # P9
            'RCFDB575', 'RCONB575', # NA
            'RIADB514', 'RIADB515', # NCO

            # 4. Auto Loans
            'RCFDK137', 'RCONK137', # Balance
            'RCFDK214',             # P3
            'RCFDK215',             # P9
            'RCFDK213', 'RCONK213', # NA
            'RIADK205', 'RIADK206', # NCO

            # 5. Agriculture
            'RCFD1590', 'RCON1590', # Balance
            'RCON2746', 'RCFD2746', # P3
            'RCON2747', 'RCFD2747', # P9
            'RCFD5341', 'RCON5341', # NA
            'RIAD4635', 'RIAD4645', # NCO

            # 6. NDFI Risk
            'RCONJ458', 'RCFDJ458', # P3
            'RCONJ459', 'RCFDJ459', # P9
            'RCONJ460', 'RCFDJ460', # NA
        ]

        # Aliases for key columns (handles variations in FFIEC files)
        self.cert_aliases = {'CERT', 'FDIC_CERT', 'FDIC CERTIFICATE NUMBER',
                            'FDICCERT', 'FDIC_CERTIFICATE_NUMBER', 'FDICERT'}
        self.idrssd_aliases = {'IDRSSD', 'ID_RSSD', 'RSSD', 'RSSDID'}

    def _clean_column_name(self, col: str) -> str:
        """Removes quotes and whitespace from column names."""
        if col is None:
            return ''
        return col.strip().strip('"').strip("'").strip()

    def _save_debug_html(self, html_content, tag: str) -> str:
        """Saves HTML content to debug directory."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ffiec_fail_{tag}_{timestamp}.html"
        filepath = self.debug_dir / filename

        try:
            if isinstance(html_content, bytes):
                html_content = html_content.decode('utf-8', errors='replace')
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            logging.info(f"      [FFIEC] Debug HTML saved to: {filepath}")
            return str(filepath)
        except Exception as e:
            logging.error(f"      [FFIEC] Failed to save debug HTML: {e}")
            return ""

    def _save_debug_extract(self, file_text: str, filename: str, date_obj, num_lines: int = 50) -> str:
        """Saves first N lines of a file for debugging."""
        try:
            date_dir = self.extracted_dir / date_obj.strftime('%Y%m%d')
            date_dir.mkdir(parents=True, exist_ok=True)

            safe_name = re.sub(r'[^\w\-.]', '_', filename)
            out_path = date_dir / f"{safe_name}_head{num_lines}.txt"

            lines = file_text.split('\n')[:num_lines]
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))

            logging.info(f"      [FFIEC] Debug extract saved to: {out_path}")
            return str(out_path)
        except Exception as e:
            logging.warning(f"      [FFIEC] Failed to save debug extract: {e}")
            return ""

    def _save_zip_file(self, zip_bytes: bytes, date_obj, product: str = "bulk", fmt: str = "tsv") -> str:
        """Saves ZIP bytes to disk."""
        date_str = date_obj.strftime('%Y%m%d')
        filename = f"ffiec_{date_str}_{product}_{fmt}.zip"
        filepath = self.zips_dir / filename

        try:
            with open(filepath, 'wb') as f:
                f.write(zip_bytes)
            logging.info(f"      [FFIEC] ZIP saved to: {filepath}")
            return str(filepath)
        except Exception as e:
            logging.error(f"      [FFIEC] Failed to save ZIP: {e}")
            return ""

    def _parse_hidden_fields(self, html: str) -> dict:
        """Parses all <input type='hidden'> fields."""
        fields = {}
        patterns = [
            r'<input[^>]*?type=["\']hidden["\'][^>]*?name=["\']([^"\']+)["\'][^>]*?value=["\']([^"\']*)["\']',
            r'<input[^>]*?name=["\']([^"\']+)["\'][^>]*?type=["\']hidden["\'][^>]*?value=["\']([^"\']*)["\']',
            r'<input[^>]*?name=["\']([^"\']+)["\'][^>]*?value=["\']([^"\']*)["\'][^>]*?type=["\']hidden["\']'
        ]
        for pattern in patterns:
            matches = re.findall(pattern, html, re.IGNORECASE | re.DOTALL)
            for name, value in matches:
                if name not in fields:
                    fields[name] = value
        return fields

    def _parse_select_options(self, html: str, control_name_fragment: str) -> tuple:
        """Parses <select> options for a control."""
        name_match = re.search(
            r'<select[^>]*name=["\']([^"\']*' + re.escape(control_name_fragment) + r'[^"\']*)["\']',
            html, re.IGNORECASE
        )
        if not name_match:
            return None, {}

        full_name = name_match.group(1)
        select_block_match = re.search(
            r'<select[^>]*name=["\']' + re.escape(full_name) + r'["\'][^>]*>(.*?)</select>',
            html, re.DOTALL | re.IGNORECASE
        )
        if not select_block_match:
            return full_name, {}

        block = select_block_match.group(1)
        options = {}
        opt_matches = re.findall(
            r'<option\s+(?:[^>]*\s+)?value=["\']([^"\']*)["\'][^>]*>(.*?)</option>',
            block, re.DOTALL | re.IGNORECASE
        )
        for val, txt in opt_matches:
            options[val] = txt.strip()

        return full_name, options

    def _parse_radio_buttons(self, html: str, name_fragment: str) -> tuple:
        """Parses radio button groups."""
        pattern = r'<input[^>]*type=["\']radio["\'][^>]*name=["\']([^"\']*' + re.escape(name_fragment) + r'[^"\']*)["\'][^>]*value=["\']([^"\']*)["\']'
        matches = re.findall(pattern, html, re.IGNORECASE)

        if not matches:
            pattern = r'<input[^>]*name=["\']([^"\']*' + re.escape(name_fragment) + r'[^"\']*)["\'][^>]*type=["\']radio["\'][^>]*value=["\']([^"\']*)["\']'
            matches = re.findall(pattern, html, re.IGNORECASE)

        if not matches:
            return None, {}

        full_name = matches[0][0]
        options = {m[1]: m[1] for m in matches}
        return full_name, options

    def _find_format_controls(self, html: str) -> dict:
        """Finds file format selection controls."""
        result = {'type': None, 'name': None, 'options': {}, 'preferred_value': None}

        for fragment in ['ExportFormatDropDownList', 'FormatDropDown', 'FileFormat']:
            dd_name, dd_options = self._parse_select_options(html, fragment)
            if dd_name and dd_options:
                result['type'] = 'dropdown'
                result['name'] = dd_name
                result['options'] = dd_options
                for pref in ['TSV', 'TXT', 'CSV', 'Tab']:
                    for val, label in dd_options.items():
                        if pref.lower() in val.lower() or pref.lower() in label.lower():
                            result['preferred_value'] = val
                            break
                    if result['preferred_value']:
                        break
                if not result['preferred_value'] and dd_options:
                    result['preferred_value'] = list(dd_options.keys())[0]
                return result

        for fragment in ['FormatType', 'RadioButton', 'Format']:
            rb_name, rb_options = self._parse_radio_buttons(html, fragment)
            if rb_name and rb_options:
                result['type'] = 'radio'
                result['name'] = rb_name
                result['options'] = rb_options
                for pref in ['TSV', 'TXT', 'CSV', 'Tab']:
                    for val in rb_options.keys():
                        if pref.lower() in val.lower():
                            result['preferred_value'] = val
                            break
                    if result['preferred_value']:
                        break
                if not result['preferred_value'] and rb_options:
                    result['preferred_value'] = list(rb_options.keys())[0]
                return result

        return result

    def _find_download_button(self, html: str) -> dict:
        """Finds the download button control."""
        result = {'name': None, 'type': 'button', 'value': 'Download'}

        patterns = [
            r'<input[^>]*name=["\']([^"\']*Download[^"\']*)["\'][^>]*type=["\']submit["\'][^>]*value=["\']([^"\']*)["\']',
            r'<input[^>]*type=["\']submit["\'][^>]*name=["\']([^"\']*Download[^"\']*)["\'][^>]*value=["\']([^"\']*)["\']',
            r'<input[^>]*name=["\']([^"\']*Download[^"\']*)["\'][^>]*type=["\']image["\']',
            r'<input[^>]*type=["\']image["\'][^>]*name=["\']([^"\']*Download[^"\']*)["\']',
        ]

        for pattern in patterns:
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                result['name'] = match.group(1)
                if 'image' in pattern.lower():
                    result['type'] = 'image'
                if len(match.groups()) > 1:
                    result['value'] = match.group(2)
                return result

        fallback_names = [
            'ctl00$MainContentHolder$TabStrip1$Download_0',
            'ctl00$MainContentHolder$Download',
        ]
        for name in fallback_names:
            if name in html:
                result['name'] = name
                return result

        return result

    def _get_validation_summary(self, html: str) -> str:
        """Extracts validation error text if present."""
        if "ValidationSummary" in html:
            match = re.search(
                r'<div[^>]*id=["\'][^"\']*ValidationSummary[^"\']*["\'][^>]*>(.*?)</div>',
                html, re.DOTALL | re.IGNORECASE
            )
            if match:
                text = re.sub(r'<[^>]+>', ' ', match.group(1))
                text = re.sub(r'\s+', ' ', text).strip()
                return text[:300]
        return None

    def _download_strict(self, date_obj) -> tuple:
        """Downloads FFIEC bulk data ZIP file."""
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': self.base_url,
            'Origin': 'https://cdr.ffiec.gov',
        })

        try:
            logging.info(f"      [FFIEC] Step 1: Initial GET")
            r1 = session.get(self.base_url, timeout=30)
            if r1.status_code != 200:
                return None, 'INITIAL_GET_FAILED'

            hidden = self._parse_hidden_fields(r1.text)
            lb_name, lb_options = self._parse_select_options(r1.text, 'ListBox1')
            if not lb_name:
                self._save_debug_html(r1.content, "step1_no_listbox")
                return None, 'NO_LISTBOX_CONTROL'

            target_report_value = "ReportingSeriesSinglePeriod"
            if target_report_value not in lb_options:
                return None, 'REPORT_TYPE_NOT_FOUND'

            logging.info(f"      [FFIEC] Step 2: Select Report Type ({target_report_value})")
            dd_name_pre, _ = self._parse_select_options(r1.text, 'DatesDropDownList')
            ef_name_pre, _ = self._parse_select_options(r1.text, 'ExportFormat')

            payload_step2 = hidden.copy()
            payload_step2['__EVENTTARGET'] = lb_name
            payload_step2['__EVENTARGUMENT'] = ''
            payload_step2[lb_name] = target_report_value
            if dd_name_pre:
                payload_step2[dd_name_pre] = ''
            if ef_name_pre:
                payload_step2[ef_name_pre] = ''

            r2 = session.post(self.base_url, data=payload_step2, timeout=30)
            if r2.status_code != 200:
                return None, 'REPORT_TYPE_POST_FAILED'

            hidden = self._parse_hidden_fields(r2.text)
            dd_name, dd_options = self._parse_select_options(r2.text, 'DatesDropDownList')

            if not dd_name or not dd_options:
                self._save_debug_html(r2.content, "step2_no_dates")
                return None, 'NO_DATES_AVAILABLE'

            target_date_str_1 = date_obj.strftime("%m/%d/%Y")
            try:
                target_date_str_2 = date_obj.strftime("%-m/%-d/%Y")
            except ValueError:
                target_date_str_2 = date_obj.strftime("%#m/%#d/%Y")

            selected_date_val = None
            selected_date_label = None
            for val, label in dd_options.items():
                if target_date_str_1 in label or target_date_str_2 in label:
                    selected_date_val = val
                    selected_date_label = label
                    break

            if not selected_date_val:
                return None, 'DATE_NOT_AVAILABLE'

            logging.info(f"      [FFIEC] Step 3: Select Date ({selected_date_label} -> {selected_date_val})")
            payload_step3 = hidden.copy()
            payload_step3['__EVENTTARGET'] = dd_name
            payload_step3['__EVENTARGUMENT'] = ''
            payload_step3[lb_name] = target_report_value
            payload_step3[dd_name] = selected_date_val

            r3 = session.post(self.base_url, data=payload_step3, timeout=30)
            if r3.status_code != 200:
                return None, 'DATE_POST_FAILED'

            hidden = self._parse_hidden_fields(r3.text)

            logging.info(f"      [FFIEC] Step 3.5: Select File Format")
            format_controls = self._find_format_controls(r3.text)

            if format_controls['name'] and format_controls['preferred_value']:
                logging.info(f"      [FFIEC] Found format control: {format_controls['type']} "
                           f"'{format_controls['name']}' with value '{format_controls['preferred_value']}'")

                payload_step35 = hidden.copy()
                payload_step35[lb_name] = target_report_value
                payload_step35[dd_name] = selected_date_val

                if format_controls['type'] == 'dropdown':
                    payload_step35['__EVENTTARGET'] = format_controls['name']
                    payload_step35['__EVENTARGUMENT'] = ''
                    payload_step35[format_controls['name']] = format_controls['preferred_value']
                else:
                    payload_step35[format_controls['name']] = format_controls['preferred_value']
                    if '__EVENTTARGET' in payload_step35:
                        payload_step35['__EVENTTARGET'] = ''

                r35 = session.post(self.base_url, data=payload_step35, timeout=30)
                if r35.status_code == 200:
                    hidden = self._parse_hidden_fields(r35.text)
                    logging.info(f"      [FFIEC] Format selection successful")

            logging.info(f"      [FFIEC] Step 4: Click Download")
            payload_step4 = hidden.copy()
            payload_step4[lb_name] = target_report_value
            payload_step4[dd_name] = selected_date_val

            if format_controls['name'] and format_controls['preferred_value']:
                payload_step4[format_controls['name']] = format_controls['preferred_value']

            download_btn = self._find_download_button(r3.text)

            if download_btn['name']:
                payload_step4.pop('__EVENTTARGET', None)
                payload_step4.pop('__EVENTARGUMENT', None)

                if download_btn['type'] == 'image':
                    payload_step4[f"{download_btn['name']}.x"] = '10'
                    payload_step4[f"{download_btn['name']}.y"] = '10'
                else:
                    payload_step4[download_btn['name']] = download_btn['value']

                logging.info(f"      [FFIEC] Using download button: {download_btn['name']}")
            else:
                payload_step4['ctl00$MainContentHolder$TabStrip1$Download_0'] = 'Download'

            r4 = session.post(self.base_url, data=payload_step4, timeout=120)

            if r4.status_code == 200:
                zip_bytes = r4.content
                ct = r4.headers.get('Content-Type', '').lower()

                is_zip = ('zip' in ct or 'octet-stream' in ct or
                         (len(zip_bytes) >= 2 and zip_bytes[:2] == b'PK'))

                if is_zip:
                    logging.info(f"      [FFIEC] Download successful - received ZIP file ({len(zip_bytes)} bytes)")
                    return zip_bytes, 'SUCCESS'
                else:
                    html_text = zip_bytes.decode('utf-8', errors='replace')
                    validation_msg = self._get_validation_summary(html_text)
                    logging.warning(f"      [FFIEC] Received HTML, not ZIP. Validation: {validation_msg}")
                    self._save_debug_html(zip_bytes, f"download_fail_{date_obj.strftime('%Y%m%d')}")
                    return None, 'DOWNLOAD_RETURNED_HTML'
            else:
                return None, 'DOWNLOAD_HTTP_ERROR'

        except requests.exceptions.Timeout:
            return None, 'TIMEOUT'
        except requests.exceptions.RequestException as e:
            logging.error(f"      [FFIEC] Request exception: {e}")
            return None, 'REQUEST_ERROR'
        except Exception as e:
            logging.error(f"      [FFIEC] Exception: {e}")
            return None, 'UNKNOWN_ERROR'

    def _find_key_column(self, columns: list) -> tuple:
        """
        Finds the key column (IDRSSD or CERT) in a list of column names.
        Returns: (column_index, column_type) or (None, None)
        """
        for idx, col in enumerate(columns):
            col_upper = col.upper()
            if col_upper in self.idrssd_aliases:
                return idx, 'IDRSSD'
            if col_upper in self.cert_aliases:
                return idx, 'CERT'
        return None, None

    def _build_cert_idrssd_crosswalk(self, zf: zipfile.ZipFile, peer_certs: set, date_obj) -> tuple:
        """
        Builds CERT <-> IDRSSD crosswalk from POR or ENT files.
        """
        cert_to_idrssd = {}
        idrssd_to_cert = {}

        # Look for files that might contain both CERT and IDRSSD
        crosswalk_files = [n for n in zf.namelist()
                         if ('POR' in n.upper() or 'ENT' in n.upper()) and n.lower().endswith('.txt')]

        logging.info(f"      [FFIEC] Looking for crosswalk in: {crosswalk_files}")

        for xwalk_file in crosswalk_files:
            try:
                file_bytes = zf.read(xwalk_file)
                file_text = file_bytes.decode('latin-1', errors='replace')
                lines = file_text.split('\n')

                if len(lines) < 3:
                    continue

                # Row 0 = column headers (MDRM codes)
                # Row 1 = descriptions (skip)
                # Row 2+ = data
                header_line = lines[0]
                raw_headers = header_line.split('\t')
                columns = [self._clean_column_name(h) for h in raw_headers]

                logging.info(f"      [FFIEC] {xwalk_file}: First 10 columns: {columns[:10]}")

                # Find CERT and IDRSSD columns
                cert_col_idx = None
                idrssd_col_idx = None

                for idx, col in enumerate(columns):
                    col_upper = col.upper()
                    if col_upper in self.idrssd_aliases:
                        idrssd_col_idx = idx
                    if col_upper in self.cert_aliases:
                        cert_col_idx = idx

                if cert_col_idx is None or idrssd_col_idx is None:
                    logging.warning(f"      [FFIEC] {xwalk_file}: Missing CERT or IDRSSD column")
                    logging.warning(f"      [FFIEC] {xwalk_file}: All columns: {columns[:30]}")
                    self._save_debug_extract(file_text, xwalk_file, date_obj)
                    continue

                logging.info(f"      [FFIEC] {xwalk_file}: CERT col={cert_col_idx}, IDRSSD col={idrssd_col_idx}")

                # Parse data rows (skip row 1 = descriptions)
                for line in lines[2:]:
                    if not line.strip():
                        continue
                    fields = line.split('\t')
                    if len(fields) <= max(cert_col_idx, idrssd_col_idx):
                        continue

                    try:
                        cert_str = fields[cert_col_idx].strip().strip('"')
                        idrssd_str = fields[idrssd_col_idx].strip().strip('"')

                        cert_val = int(cert_str)
                        idrssd_val = int(idrssd_str)

                        if cert_val in peer_certs:
                            cert_to_idrssd[cert_val] = idrssd_val
                            idrssd_to_cert[idrssd_val] = cert_val
                    except (ValueError, IndexError):
                        continue

                if cert_to_idrssd:
                    logging.info(f"      [FFIEC] Crosswalk built: {len(cert_to_idrssd)} mappings")
                    sample = dict(list(cert_to_idrssd.items())[:5])
                    logging.info(f"      [FFIEC] Sample mappings (CERT->IDRSSD): {sample}")
                    return cert_to_idrssd, idrssd_to_cert

            except Exception as e:
                logging.warning(f"      [FFIEC] Error reading {xwalk_file}: {e}")
                continue

        logging.warning(f"      [FFIEC] Could not build crosswalk from POR/ENT files")
        return cert_to_idrssd, idrssd_to_cert

    def _parse_schedule_file(self, file_text: str, filename: str, date_obj,
                            peer_certs: set, peer_idrssd: set,
                            idrssd_to_cert: dict) -> dict:
        """
        Parses a schedule file with proper row handling:
        - Row 0 = MDRM column codes (header)
        - Row 1 = Descriptions (skip)
        - Row 2+ = Data
        """
        cert_data_map = {}
        lines = file_text.split('\n')

        if len(lines) < 3:
            logging.warning(f"      [FFIEC] {filename}: Too few lines ({len(lines)})")
            return cert_data_map

        # Parse Row 0 as header
        header_line = lines[0]
        raw_headers = header_line.split('\t')
        columns = [self._clean_column_name(h) for h in raw_headers]

        # Log first 5 columns for diagnostics
        logging.info(f"      [FFIEC] {filename}: Columns (first 5): {columns[:5]}")

        # Find key column
        key_col_idx, key_col_type = self._find_key_column(columns)

        if key_col_idx is None:
            logging.warning(f"      [FFIEC] {filename}: No IDRSSD or CERT column found")
            logging.warning(f"      [FFIEC] {filename}: All columns: {columns[:20]}")
            self._save_debug_extract(file_text, filename, date_obj)
            return cert_data_map

        logging.info(f"      [FFIEC] {filename}: Key column = {key_col_type} at index {key_col_idx}")

        # Build column index map for target fields
        field_col_map = {}
        columns_upper = [c.upper() for c in columns]
        for tf in self.target_fields:
            tf_upper = tf.upper()
            if tf_upper in columns_upper:
                field_col_map[tf] = columns_upper.index(tf_upper)

        if field_col_map:
            logging.info(f"      [FFIEC] {filename}: Target fields found: {list(field_col_map.keys())[:10]}")

        # Parse data rows (skip Row 1 = descriptions)
        total_rows = 0
        matched_rows = 0
        matched_keys = set()

        for line in lines[2:]:  # Start from Row 2
            if not line.strip():
                continue

            total_rows += 1
            fields = line.split('\t')

            if len(fields) <= key_col_idx:
                continue

            try:
                key_str = fields[key_col_idx].strip().strip('"')
                key_val = int(key_str)
            except (ValueError, AttributeError):
                continue

            # Determine CERT based on key type
            cert_val = None
            if key_col_type == 'IDRSSD':
                if key_val in peer_idrssd:
                    cert_val = idrssd_to_cert.get(key_val)
                    matched_keys.add(key_val)
            else:  # CERT
                if key_val in peer_certs:
                    cert_val = key_val
                    matched_keys.add(key_val)

            if cert_val is None:
                continue

            matched_rows += 1

            # Extract target fields
            if cert_val not in cert_data_map:
                cert_data_map[cert_val] = {}

            for field, col_idx in field_col_map.items():
                if col_idx < len(fields):
                    val_str = fields[col_idx].strip().strip('"')
                    if val_str and val_str != '':
                        try:
                            cert_data_map[cert_val][field] = float(val_str.replace(',', ''))
                        except ValueError:
                            pass

        # Log diagnostics
        logging.info(f"      [FFIEC] {filename}: {total_rows} total rows, {matched_rows} matched, "
                    f"{len(matched_keys)} unique {key_col_type}s")

        return cert_data_map

    def _parse_zip_content(self, zip_bytes: bytes, date_obj, peer_certs: set) -> pd.DataFrame:
        """Parses ZIP content and extracts FFIEC data."""
        zip_path = self._save_zip_file(zip_bytes, date_obj)

        if len(zip_bytes) < 4 or zip_bytes[:2] != b'PK':
            logging.error(f"      [FFIEC] Invalid ZIP file")
            return pd.DataFrame()

        logging.info(f"      [FFIEC] Parsing ZIP: {len(zip_bytes)} bytes")

        bio = io.BytesIO(zip_bytes)

        try:
            with zipfile.ZipFile(bio, 'r') as zf:
                members = zf.namelist()
                logging.info(f"      [FFIEC] ZIP contains {len(members)} files")
                logging.info(f"      [FFIEC] First 25 members: {members[:25]}")

                # Step 1: Build crosswalk
                cert_to_idrssd, idrssd_to_cert = self._build_cert_idrssd_crosswalk(zf, peer_certs, date_obj)

                if not cert_to_idrssd:
                    logging.warning(f"      [FFIEC] No crosswalk - will only match CERT-keyed files")

                peer_idrssd = set(cert_to_idrssd.values())

                # Step 2: Identify and prioritize files
                target_files = [n for n in members
                               if n.lower().endswith('.txt') and
                               ('schedule' in n.lower() or 'bulk' in n.lower())]

                # Prioritize RIC/RICII/RCC files
                priority_keywords = ['RICII', 'RIC ', 'RICI ', 'RCC']
                priority_files = []
                other_files = []
                for f in target_files:
                    f_upper = f.upper()
                    if any(kw in f_upper for kw in priority_keywords):
                        priority_files.append(f)
                    else:
                        other_files.append(f)

                target_files = priority_files + other_files
                logging.info(f"      [FFIEC] Processing {len(target_files)} files "
                           f"({len(priority_files)} priority)")

                # Step 3: Process each file
                cert_data_map = {}

                for t_file in target_files:
                    try:
                        file_bytes = zf.read(t_file)
                        file_text = file_bytes.decode('latin-1', errors='replace')
                    except Exception as e:
                        logging.warning(f"      [FFIEC] Failed to read {t_file}: {e}")
                        continue

                    file_data = self._parse_schedule_file(
                        file_text, t_file, date_obj,
                        peer_certs, peer_idrssd,
                        idrssd_to_cert
                    )

                    # Merge into main data map
                    for cert, fields in file_data.items():
                        if cert not in cert_data_map:
                            cert_data_map[cert] = {'CERT': cert, 'REPDTE': date_obj}
                        cert_data_map[cert].update(fields)

                # Step 4: Create DataFrame
                if cert_data_map:
                    df = pd.DataFrame(list(cert_data_map.values()))
                    df['FFIEC_PATCH_STATUS'] = 'SUCCESS'

                    # Log field coverage
                    field_counts = {}
                    for f in self.target_fields:
                        if f in df.columns:
                            non_null = df[f].notna().sum()
                            if non_null > 0:
                                field_counts[f] = int(non_null)

                    logging.info(f"      [FFIEC] Extracted {len(df)} banks, "
                               f"{len(field_counts)}/{len(self.target_fields)} target fields with data")

                    if field_counts:
                        logging.info(f"      [FFIEC] Field non-null counts: {field_counts}")

                    return df
                else:
                    logging.warning(f"      [FFIEC] No matching banks found in ZIP")
                    return pd.DataFrame()

        except zipfile.BadZipFile as e:
            logging.error(f"      [FFIEC] Bad ZIP file: {e}")
            return pd.DataFrame()
        except Exception as e:
            logging.error(f"      [FFIEC] ZIP parsing exception: {e}")
            import traceback
            logging.error(traceback.format_exc())
            return pd.DataFrame()

    def fetch_quarter_data(self, date_obj, peer_certs) -> pd.DataFrame:
        """Fetches FFIEC bulk data for a specific quarter."""
        date_fmt = date_obj.strftime("%m/%d/%Y")
        file_date_sig = date_obj.strftime("%Y%m%d")
        cache_file = self.output_dir / f"FFIEC_Bulk_Call_{file_date_sig}.csv"

        # Check cache - only use if it has actual target fields
        if cache_file.exists():
            try:
                df_cache = pd.read_csv(cache_file)
                if 'CERT' in df_cache.columns:
                    df_cache = df_cache[df_cache['CERT'].isin(peer_certs)]
                    fields_present = [c for c in self.target_fields if c in df_cache.columns]
                    if fields_present:
                        logging.info(f"      [Cache Hit] {date_fmt} ({len(df_cache)} records, "
                                   f"{len(fields_present)} target fields)")
                        return df_cache
            except Exception as e:
                logging.warning(f"      [Cache] Failed to read: {e}")

        logging.info(f"      [Downloading] FFIEC Bulk Data for {date_fmt}...")

        zip_bytes, status = self._download_strict(date_obj)

        if zip_bytes is None:
            logging.warning(f"      [Failed] No FFIEC data for {date_obj.date()} (status: {status})")
            return pd.DataFrame({'FFIEC_PATCH_STATUS': [status], 'REPDTE': [date_obj]})

        df = self._parse_zip_content(zip_bytes, date_obj, set(peer_certs))

        if df.empty or 'CERT' not in df.columns:
            logging.warning(f"      [Failed] ZIP parsing produced no data for {date_obj.date()}")
            return pd.DataFrame({'FFIEC_PATCH_STATUS': ['ZIP_PARSE_FAILED'], 'REPDTE': [date_obj]})

        # Cache only if we have target fields
        target_cols_present = [c for c in self.target_fields if c in df.columns]
        if target_cols_present:
            try:
                df.to_csv(cache_file, index=False)
                logging.info(f"      [Cache] Saved to {cache_file}")
            except Exception as e:
                logging.warning(f"      [Cache] Failed to save: {e}")

        return df

    def _quarter_needs_patching(self, df_fdic: pd.DataFrame, date_obj) -> bool:
        """Determines if a quarter needs FFIEC data patching."""
        if df_fdic.empty:
            return True
        q_data = df_fdic[df_fdic['REPDTE'] == date_obj]
        if q_data.empty:
            return False

        check_cols = [c for c in ['RCFD1545', 'RCFDJ466', 'RCFDJJ04'] if c in q_data.columns]
        if not check_cols:
            return True

        missing_or_zero = sum(
            ((q_data[col].isna()) | (q_data[col] == 0)).sum()
            for col in check_cols
        )
        total_cells = len(check_cols) * len(q_data)
        return (missing_or_zero / total_cells) > 0.5 if total_cells > 0 else True

    def heal_dataset(self, df_fdic: pd.DataFrame, peer_certs: set) -> pd.DataFrame:
        """Heals FDIC data with FFIEC bulk data."""
        print("\n" + "="*60)
        print("DUAL-TRACK DATA RECOVERY: FFIEC BULK API (v4)")
        print("="*60)

        if df_fdic.empty:
            return df_fdic

        dates = sorted(df_fdic['REPDTE'].unique(), reverse=True)[:30]
        ffiec_frames = []

        for dt in dates:
            dt_obj = pd.to_datetime(dt)
            if not self._quarter_needs_patching(df_fdic, dt_obj):
                logging.info(f"      [Skip] Data sufficient for {dt_obj.strftime('%Y-%m-%d')}")
                continue

            df_ffiec_q = self.fetch_quarter_data(dt_obj, peer_certs)

            if not df_ffiec_q.empty and 'CERT' in df_ffiec_q.columns:
                target_cols = [c for c in self.target_fields if c in df_ffiec_q.columns]
                if target_cols:
                    ffiec_frames.append(df_ffiec_q)

        if not ffiec_frames:
            print("      No FFIEC data merged.")
            df_fdic['RIC_Source_Status'] = 'FFIEC_NO_TARGET_FIELDS'
            df_fdic['RIC_SOURCE_AVAILABLE'] = 0
            return df_fdic

        print(f"      Merging FFIEC data from {len(ffiec_frames)} quarters...")
        df_patch = pd.concat(ffiec_frames, ignore_index=True)
        df_patch = df_patch[df_patch['CERT'].notna()]

        if df_patch.empty:
            df_fdic['RIC_Source_Status'] = 'FFIEC_NO_VALID_DATA'
            df_fdic['RIC_SOURCE_AVAILABLE'] = 0
            return df_fdic

        df_fdic['REPDTE'] = pd.to_datetime(df_fdic['REPDTE'])
        df_patch['REPDTE'] = pd.to_datetime(df_patch['REPDTE'])

        df_main = df_fdic.set_index(['CERT', 'REPDTE'])
        df_p = df_patch.set_index(['CERT', 'REPDTE'])

        if 'FFIEC_PATCH_STATUS' in df_p.columns:
            df_p = df_p.drop(columns=['FFIEC_PATCH_STATUS'])

        merged_fields = 0
        for col in self.target_fields:
            if col in df_p.columns:
                if col not in df_main.columns:
                    df_main[col] = df_p[col]
                    merged_fields += 1
                else:
                    orig_nulls = df_main[col].isna().sum()
                    df_main[col] = df_main[col].fillna(df_p[col])
                    mask_fix = (df_main[col] == 0) & (df_p[col].notna()) & (df_p[col] != 0)
                    if mask_fix.any():
                        df_main.loc[mask_fix, col] = df_p.loc[mask_fix, col]
                    if df_main[col].isna().sum() < orig_nulls:
                        merged_fields += 1

        logging.info(f"      [Merge] Updated {merged_fields} target fields")

        df_main = df_main.copy()  # de-fragment after sequential column assignments
        res = df_main.reset_index()

        ric_check_cols = ['RCFDJ466', 'RCONJ466', 'RCFDJJ04', 'RCONJJ04']
        ric_check_col = next((c for c in ric_check_cols if c in res.columns and res[c].notna().any()), None)

        if ric_check_col:
            res['RIC_SOURCE_AVAILABLE'] = res[ric_check_col].notna().astype(int)
            res['RIC_Source_Status'] = np.where(res['RIC_SOURCE_AVAILABLE'] == 1, 'SUCCESS', 'MISSING_DATA')
        else:
            res['RIC_SOURCE_AVAILABLE'] = 0
            res['RIC_Source_Status'] = 'FFIEC_NO_RIC_FIELDS'

        success_count = (res['RIC_Source_Status'] == 'SUCCESS').sum()
        logging.info(f"      [Summary] {success_count}/{len(res)} records with FFIEC data")

        return res
# ==================================================================================
#  3. HELPER CLASSES
# ==================================================================================
def get_bank_locations(cert_numbers: list) -> pd.DataFrame:
    """
    Fetches the primary (HQ) and all operating states for a list of banks.

    Args:
        cert_numbers: A list of bank CERT numbers.

    Returns:
        A pandas DataFrame with CERT, NAME, HQ_STATE, and ALL_OPERATING_STATES.
    """
    institution_data = []
    session = requests.Session()

    for cert in cert_numbers:
        logger.info(f"Fetching location data for CERT: {cert}...")
        hq_state = "N/A"
        all_states = set()
        bank_name = f"Bank with CERT {cert}"

        # 1. Get Institution details (for HQ State)
        try:
            inst_url = f"https://banks.data.fdic.gov/api/institutions?filters=CERT%3A%20{cert}&fields=NAME,STALP"
            response = session.get(inst_url, timeout=20)
            response.raise_for_status()
            data = response.json().get('data', [])
            if data:
                bank_name = data[0]['data']['NAME']
                hq_state = data[0]['data']['STALP']
                all_states.add(hq_state)

        except requests.exceptions.RequestException as e:
            logger.error(f"  - Could not fetch institution data for CERT {cert}: {e}")


        # 2. Get all branch locations (for all operating states)
        try:
            # Increase limit to get all branches, assuming no bank has > 10000 branches
            loc_url = f"https://banks.data.fdic.gov/api/locations?filters=CERT%3A%20{cert}&fields=STALP&limit=10000"
            response = session.get(loc_url, timeout=20)
            response.raise_for_status()
            locations = response.json().get('data', [])
            for loc in locations:
                all_states.add(loc['data']['STALP'])

        except requests.exceptions.RequestException as e:
            logger.error(f"  - Could not fetch branch location data for CERT {cert}: {e}")

        institution_data.append({
            "CERT": cert,
            "NAME": bank_name,
            "HQ_STATE": hq_state,
            "ALL_OPERATING_STATES": ", ".join(sorted(list(all_states)))
        })
        time.sleep(0.1) # Small delay to be polite to the API

    return pd.DataFrame(institution_data)
class FDICDataFetcher:
    def __init__(self, config: DashboardConfig):
        self.config = config
        self.session = requests.Session()

    def fetch_lnci_separately(self, certs_to_fetch: List[int]) -> pd.DataFrame:
        """Fetch LNCI field separately since it doesn't work in bulk requests."""
        lnci_data = []

        for cert in tqdm(certs_to_fetch, desc="Fetching LNCI data", unit="bank"):
            try:
                params = {
                    "filters": f"CERT:{cert}",
                    "fields": "CERT,REPDTE,LNCI",
                    "sort_by": "REPDTE",
                    "sort_order": "DESC",
                    "limit": self.config.quarters_back + 4,
                    "format": "json"
                }

                response = self.session.get(f"{self.config.fdic_api_base}/financials", params=params, timeout=30)
                response.raise_for_status()

                data = [item.get('data', {}) for item in response.json().get('data', []) if item.get('data')]
                if data:
                    df = pd.DataFrame(data)
                    df['CERT'] = cert

                    # CRITICAL FIX: Use same datetime conversion as main fetch
                    df['REPDTE'] = pd.to_datetime(df['REPDTE'], format='ISO8601', utc=True).dt.tz_localize(None)
                    df['LNCI'] = pd.to_numeric(df['LNCI'], errors='coerce')

                    # Diagnostic
                    lnci_values = df['LNCI'].dropna()
                    if not lnci_values.empty:
                        logger.info(f"📊 CERT {cert} LNCI: {len(lnci_values)} non-null values, range ${lnci_values.min():,.0f} - ${lnci_values.max():,.0f}")
                    else:
                        logger.warning(f"⚠️ CERT {cert}: LNCI column exists but all values are NaN")

                    lnci_data.append(df[['CERT', 'REPDTE', 'LNCI']])
                    logger.info(f"✅ LNCI fetched for CERT {cert}")
                else:
                    logger.warning(f"⚠️ No LNCI data returned for CERT {cert}")

            except Exception as e:
                logger.error(f"❌ Error fetching LNCI for CERT {cert}: {e}")

            time.sleep(0.2)

        if lnci_data:
            combined = pd.concat(lnci_data, ignore_index=True)
            logger.info(f"📊 Combined LNCI: {len(combined)} rows, {combined['LNCI'].notna().sum()} non-null")
            return combined

        return pd.DataFrame()

    def fetch_all_banks(self) -> Tuple[pd.DataFrame, List[int]]:
        certs_to_fetch = [self.config.subject_bank_cert] + self.config.peer_bank_certs
        all_bank_data, failed_certs = [], []

        # Step 1: Fetch all fields EXCEPT LNCI
        main_fields = [f for f in FDIC_FIELDS_TO_FETCH if f != 'LNCI']

        for cert in tqdm(certs_to_fetch, desc="Fetching FDIC data", unit="bank"):
            try:
                params = {
                    "filters": f"CERT:{cert}",
                    "fields": ",".join(main_fields),
                    "sort_by": "REPDTE", "sort_order": "DESC",
                    "limit": self.config.quarters_back + 4, "format": "json"
                }
                response = self.session.get(f"{self.config.fdic_api_base}/financials", params=params, timeout=30)
                response.raise_for_status()
                data = [item.get('data', {}) for item in response.json().get('data', []) if item.get('data')]
                if data:
                    df = pd.DataFrame(data)
                    df['CERT'] = cert
                    all_bank_data.append(df)
            except Exception as e:
                logger.error(f"Error fetching FDIC data for CERT {cert}: {e}")
                failed_certs.append(cert)
            time.sleep(0.2)

        if not all_bank_data:
            return pd.DataFrame(), failed_certs

        # Step 2: Combine main data
        combined_df = pd.concat(all_bank_data, ignore_index=True)

        # CRITICAL FIX: Convert REPDTE to datetime WITH EXPLICIT FORMAT
        # This ensures consistent datetime representation for merging
        combined_df['REPDTE'] = pd.to_datetime(combined_df['REPDTE'], format='ISO8601', utc=True).dt.tz_localize(None)

        # Step 3: Fetch LNCI separately and merge
        logger.info("Fetching LNCI data separately...")
        lnci_df = self.fetch_lnci_separately(certs_to_fetch)

        if not lnci_df.empty:
            # DIAGNOSTIC: Check merge compatibility
            logger.info(f"📊 Pre-merge diagnostics:")
            logger.info(f"   Main DF REPDTE dtype: {combined_df['REPDTE'].dtype}")
            logger.info(f"   LNCI DF REPDTE dtype: {lnci_df['REPDTE'].dtype}")
            logger.info(f"   Main DF sample: {combined_df[['CERT', 'REPDTE']].head(2).to_dict('records')}")
            logger.info(f"   LNCI DF sample: {lnci_df[['CERT', 'REPDTE']].head(2).to_dict('records')}")

            # Check for overlapping keys
            main_keys = set(zip(combined_df['CERT'], combined_df['REPDTE'].astype(str)))
            lnci_keys = set(zip(lnci_df['CERT'], lnci_df['REPDTE'].astype(str)))
            overlap = main_keys.intersection(lnci_keys)
            logger.info(f"   Overlapping keys: {len(overlap)} out of {len(main_keys)}")

            if len(overlap) == 0:
                logger.error("❌ NO OVERLAPPING KEYS - merge will fail!")
                logger.info(f"   Main key examples: {list(main_keys)[:3]}")
                logger.info(f"   LNCI key examples: {list(lnci_keys)[:3]}")

            # Merge LNCI data
            combined_df = pd.merge(
                combined_df,
                lnci_df,
                on=['CERT', 'REPDTE'],
                how='left'
            )

            # POST-MERGE DIAGNOSTIC
            lnci_count = combined_df['LNCI'].notna().sum()
            logger.info(f"✅ After merge - LNCI non-null count: {lnci_count}/{len(combined_df)}")

            if lnci_count == 0:
                logger.warning("⚠️ LNCI merge produced all NaN - investigating...")
                # Try alternative merge strategy
                logger.info("   Attempting alternative merge with string dates...")
                combined_df['REPDTE_STR'] = combined_df['REPDTE'].astype(str)
                lnci_df['REPDTE_STR'] = lnci_df['REPDTE'].astype(str)

                combined_df = combined_df.drop(columns=['LNCI'], errors='ignore')
                combined_df = pd.merge(
                    combined_df,
                    lnci_df[['CERT', 'REPDTE_STR', 'LNCI']],
                    on=['CERT', 'REPDTE_STR'],
                    how='left'
                )
                combined_df = combined_df.drop(columns=['REPDTE_STR'])

                lnci_count_retry = combined_df['LNCI'].notna().sum()
                logger.info(f"   Retry result: {lnci_count_retry} non-null LNCI values")

        else:
            combined_df['LNCI'] = np.nan
            logger.warning("⚠️ LNCI fetch failed, using NaN values")

        # =========================================================
        # [STEP 2] DUAL-TRACK DATA RECOVERY (FFIEC BULK HEALER)
        # =========================================================
        try:
            print("\n[Dual-Track] Initiating FFIEC Bulk Data Recovery...")
            ffiec_loader = FFIECBulkLoader(output_dir="data/ffiec_cache")
            all_certs_set = set(certs_to_fetch)

            # Run Healer
            combined_df = ffiec_loader.heal_dataset(combined_df, all_certs_set)
            # [D] VALIDATOR FOR CERT 34221
            try:
                target_cert = 34221
                # Check latest available date in the combined frame
                if not combined_df.empty:
                    latest_dt = combined_df['REPDTE'].max()
                    test_row = combined_df[(combined_df["CERT"] == target_cert) & (combined_df["REPDTE"] == latest_dt)]

                    if not test_row.empty:
                        logging.info(f"\n[VALIDATOR] Checking RI-C fields for CERT {target_cert} at {latest_dt.date()}:")
                        check_cols = ["RCFDJ466", "RCONJ466", "RCFDJ474", "RCONJ474"]
                        for col in check_cols:
                            val = test_row[col].iloc[0] if col in test_row.columns else "MISSING"
                            logging.info(f"  - {col}: {val}")

                        if 'FFIEC_RIC_STATUS' in test_row.columns and test_row['FFIEC_RIC_STATUS'].iloc[0] == 'FAILED':
                            logging.warning("  -> FFIEC Download Failed for this date. Check debug/ folder.")
                    else:
                        logging.warning(f"[VALIDATOR] CERT {target_cert} not found in latest data.")
            except Exception as e:
                logging.error(f"[VALIDATOR] Exception: {e}")

            # ===== POST-FFIEC HEAL DIAGNOSTIC (RI-C) =====
            # Check if columns actually exist after the merge
            ric_cols = [
                "RCFDJ466","RCONJ466","RCFDJ467","RCONJ467","RCFDJ468","RCONJ468","RCFDJ469","RCONJ469",
                "RCFDJ470","RCONJ470","RCFDJ471","RCONJ471","RCFDJ472","RCONJ472","RCFDJ474","RCONJ474"
            ]
            missing = [c for c in ric_cols if c not in combined_df.columns]
            logger.info(f"[RI-C DIAG] Columns present: {len(ric_cols)-len(missing)}/{len(ric_cols)}")
            if missing:
                logger.warning(f"[RI-C DIAG] Missing RI-C source columns: {missing}")

            # Non-null counts
            counts = {c: int(combined_df[c].notna().sum()) for c in ric_cols if c in combined_df.columns}
            logger.info(f"[RI-C DIAG] Non-null counts: {counts}")

            # Validation Row (MSPBNA)
            try:
                test_cert = 34221
                # Find the latest available date in the df to test against
                latest_dt = combined_df['REPDTE'].max()
                test = combined_df[(combined_df["CERT"] == test_cert) & (combined_df["REPDTE"] == latest_dt)]

                if test.empty:
                    logger.warning(f"[RI-C DIAG] Validation row missing for CERT={test_cert}.")
                else:
                    show_cols = ["CERT","REPDTE"] + [c for c in ric_cols if c in combined_df.columns]
                    # Print first 200 chars of dict to avoid massive log spam
                    row_data = test[show_cols].iloc[0].to_dict()
                    logger.info(f"[RI-C DIAG] CERT={test_cert} values: {row_data}")
            except Exception as e:
                logger.error(f"[RI-C DIAG] Exception during validation row check: {e}")
            # ===========================================

        except Exception as e:
            logger.error(f"FFIEC Healer failed: {e}")
            print(f"WARNING: FFIEC Healer skipped due to error: {e}")
        # =========================================================


        # Step 4: Continue with existing processing
        for field in FDIC_FIELDS_TO_FETCH:
            if field not in combined_df.columns:
                combined_df[field] = np.nan

        # Convert numeric columns
        num_cols = [c for c in combined_df.columns if c not in ['CERT', 'NAME', 'REPDTE']]
        combined_df[num_cols] = combined_df[num_cols].apply(pd.to_numeric, errors='coerce')

        return combined_df.sort_values(['CERT', 'REPDTE']), failed_certs



class FREDDataFetcher:
    """
    Asynchronous data fetcher for the FRED API.
    Handles concurrent requests with rate limiting and retries.
    """
    BASE_URL = "https://api.stlouisfed.org/fred/series/observations"

    def __init__(self, config: 'DashboardConfig', max_concurrent: int = 3, rate_limit_delay: float = 1.0):
        self.config = config
        self.api_key = self.config.fred_api_key
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.rate_limit_delay = rate_limit_delay
        self.logger = logging.getLogger(__name__)
        self.last_fred_obs_df = pd.DataFrame() # Initialize storage for raw obs

    async def _fetch_series_metadata(
        self, session: aiohttp.ClientSession, series_id: str
    ) -> Tuple[str, Optional[Dict]]:
        """Fetches metadata for a FRED series including frequency."""
        params = {
            "series_id": series_id,
            "api_key": self.api_key,
            "file_type": "json"
        }

        async with self.semaphore:
            try:
                async with session.get(
                    "https://api.stlouisfed.org/fred/series",
                    params=params
                ) as response:
                    response.raise_for_status()
                    data = await response.json()

                    if "seriess" in data and data["seriess"]:
                        metadata = data["seriess"][0]
                        return series_id, {
                            'frequency': metadata.get('frequency', 'Unknown'),
                            'frequency_short': metadata.get('frequency_short', 'Unknown'),
                            'units': metadata.get('units', 'Unknown'),
                            'seasonal_adjustment': metadata.get('seasonal_adjustment', 'Unknown'),
                            'last_updated': metadata.get('last_updated', 'Unknown')
                        }
                    return series_id, None
            except Exception as e:
                self.logger.error(f"Error fetching metadata for {series_id}: {e}")
                return series_id, None
            finally:
                await asyncio.sleep(self.rate_limit_delay)

    async def _fetch_single_series(
        self, session: aiohttp.ClientSession, series_id: str, start_date: str
    ) -> Tuple[str, Optional[pd.DataFrame]]:
        """
        Fetches a single time series from the FRED API asynchronously.
        """
        params = {
            "series_id": series_id,
            "api_key": self.api_key,
            "file_type": "json",
            "observation_start": start_date,
            "sort_order": "asc",
        }

        async with self.semaphore:
            try:
                async with session.get(self.BASE_URL, params=params) as response:
                    response.raise_for_status()
                    data = await response.json()

                    if "observations" in data and data["observations"]:
                        df = pd.DataFrame(data["observations"])
                        df = df[['date', 'value']]
                        df['date'] = pd.to_datetime(df['date'])
                        df['value'] = pd.to_numeric(df['value'], errors='coerce')
                        df = df.set_index('date').rename(columns={'value': series_id})
                        return series_id, df
                    else:
                        self.logger.warning(f"No observations returned for series {series_id}.")
                        return series_id, None
            except aiohttp.ClientError as e:
                self.logger.error(f"HTTP error fetching {series_id}: {e}")
                return series_id, None
            except Exception as e:
                self.logger.error(f"An unexpected error occurred for {series_id}: {e}")
                return series_id, None
            finally:
                await asyncio.sleep(self.rate_limit_delay)

    async def fetch_all_series_async(
        self,
        series_ids: List[str],
        series_descriptions: Optional[pd.DataFrame] = None
    ) -> Tuple[pd.DataFrame, pd.DataFrame, List[str], pd.DataFrame]:
        """
        Async fetch for all FRED series + metadata.
        Consolidated method that handles optional description inputs.
        """
        start_date = (datetime.now() - pd.DateOffset(years=self.config.fred_years_back)).strftime('%Y-%m-%d')

        if not series_ids:
             return pd.DataFrame(), pd.DataFrame(), [], pd.DataFrame()

        async with aiohttp.ClientSession() as session:
            # Fetch both data and metadata concurrently
            data_tasks = [
                self._fetch_single_series(session, series_id, start_date)
                for series_id in series_ids
            ]
            metadata_tasks = [
                self._fetch_series_metadata(session, series_id)
                for series_id in series_ids
            ]

            all_tasks = data_tasks + metadata_tasks
            # Use tqdm_asyncio for async progress bar (fixes AttributeError)
            results = await tqdm_asyncio.gather(*all_tasks, desc="Fetching FRED Series & Metadata")

        # Split results
        num_series = len(series_ids)
        data_results = results[:num_series]
        metadata_results = results[num_series:]

        successful_dfs = []
        failed_series = []
        metadata_dict = {}
        raw_df_map = {}  # keep original (pre-resample) series by id

        # Process data results
        for series_id, df in data_results:
            if df is not None and not df.empty:
                successful_dfs.append(df)          # this is your resample-ready frame (DATE, sid as column)
                raw_df_map[series_id] = df.copy()  # keep the original obs-dated frame
            else:
                failed_series.append(series_id)

        # Process metadata results
        for series_id, metadata in metadata_results:
            if metadata:
                metadata_dict[series_id] = metadata

        if not successful_dfs:
            self.logger.error("No FRED series were successfully fetched.")
            # Return empty structure matching signature, using passed descriptions if available
            desc_return = series_descriptions if series_descriptions is not None else pd.DataFrame()
            return pd.DataFrame(), desc_return, failed_series, pd.DataFrame()

        merged_df = pd.concat(successful_dfs, axis=1)
        merged_df = merged_df.resample('D').asfreq().ffill()

        # Create metadata DataFrame
        metadata_df = pd.DataFrame.from_dict(metadata_dict, orient='index')
        metadata_df.index.name = 'Series ID'
        metadata_df.reset_index(inplace=True)

        # === PATCH S-RAW2: build long raw observations (DATE, SeriesID, VALUE) and stash ===
        raw_long = []
        for sid, rdf in raw_df_map.items():
            r = rdf.copy().reset_index()                             # index is 'date'
            date_col = 'date' if 'date' in r.columns else r.columns[0]
            r = r.rename(columns={date_col: 'DATE', sid: 'VALUE'})
            r = r[['DATE', 'VALUE']]
            r['SeriesID'] = str(sid)
            raw_long.append(r)
        FRED_Obs_df = (pd.concat(raw_long, ignore_index=True)
                       if raw_long else
                       pd.DataFrame(columns=['DATE','VALUE','SeriesID']))
        FRED_Obs_df['DATE'] = pd.to_datetime(FRED_Obs_df['DATE'], errors='coerce')
        FRED_Obs_df['VALUE'] = pd.to_numeric(FRED_Obs_df['VALUE'], errors='coerce')
        FRED_Obs_df['SeriesID'] = FRED_Obs_df['SeriesID'].astype(str)
        FRED_Obs_df = FRED_Obs_df.dropna(subset=['DATE']).sort_values(['SeriesID','DATE'])
        # expose for downstream writer
        self.last_fred_obs_df = FRED_Obs_df
        # === END PATCH S-RAW2 ===

        # === PATCH FREQ-2: infer frequency for series with missing/unknown metadata ===
        def _infer_freq_from_index(idx: pd.DatetimeIndex) -> tuple[str, str]:
            """Return ('daily'|'monthly'|'quarterly', 'D'|'M'|'Q') using raw index cadence."""
            idx = pd.DatetimeIndex(idx).sort_values().unique()
            if len(idx) < 3:
                return ("monthly", "M")  # conservative default
            guess = pd.infer_freq(idx)
            if guess:
                g = guess.upper()
                if g.startswith(("Q", "QS")): return ("quarterly", "Q")
                if g.startswith(("M", "MS")): return ("monthly", "M")
                if g.startswith(("D", "B", "C", "W")): return ("daily", "D")
            # Heuristic fallback: med observations per year
            vc = pd.Series(idx.year).value_counts()
            med_per_year = float(vc.median()) if not vc.empty else 0.0
            if med_per_year >= 200: return ("daily", "D")
            if 6 <= med_per_year <= 15: return ("monthly", "M")
            if 3 <= med_per_year <= 5: return ("quarterly", "Q")
            # last resort: bucket by month uniqueness
            months_per_year = pd.Series(list(zip(idx.year, idx.month))).drop_duplicates().groupby(lambda x: x[0]).size().median()
            if months_per_year and months_per_year >= 6: return ("monthly", "M")
            return ("quarterly", "Q")

        # Build a quick lookup of existing metadata
        if metadata_df is None or metadata_df.empty:
            metadata_df = pd.DataFrame(columns=["Series ID","frequency","frequency_short","units"])

        present = set(metadata_df["Series ID"].astype(str)) if not metadata_df.empty else set()

        # Candidates: (a) not in metadata_df, or (b) frequency missing/Unknown
        def _needs_infer(row) -> bool:
            f = str(row.get("frequency", "")).strip().lower()
            fs = str(row.get("frequency_short", "")).strip().lower()
            return (f in ("", "nan", "none", "unknown")) and (fs in ("", "nan", "none", "unknown"))

        need_rows = []
        for sid in merged_df.columns:
            ss = str(sid)
            if ss not in raw_df_map:  # no raw series to infer from
                continue
            if ss not in present:
                need_rows.append(ss)
            else:
                # check if this metadata row needs inference
                r = metadata_df.loc[metadata_df["Series ID"].astype(str)==ss]
                if r.empty or _needs_infer(r.iloc[0].to_dict()):
                    need_rows.append(ss)

        inferred_records = []
        for sid in need_rows:
            raw = raw_df_map.get(sid)
            if raw is None or raw.empty:
                continue
            idx = raw.index
            freq_long, freq_short = _infer_freq_from_index(idx)
            inferred_records.append({
                "Series ID": sid,
                "frequency": freq_long.title() if freq_long else "Unknown",
                "frequency_short": freq_short.upper() if freq_short else "",
                "units": np.nan  # leave units; your downstream formatter/heuristic handles it
            })

        if inferred_records:
            inferred_df = pd.DataFrame(inferred_records)
            # Update/append intelligently
            if metadata_df.empty:
                metadata_df = inferred_df
            else:
                md = metadata_df.set_index("Series ID")
                idf = inferred_df.set_index("Series ID")
                md.update(idf)  # update rows that exist
                to_add = idf.loc[~idf.index.isin(md.index)]
                metadata_df = pd.concat([md, to_add]).reset_index()

        self.logger.info(f"[FREQ] Metadata rows: {len(metadata_df)} (after inference)")
        # === END PATCH FREQ-2 ===

        # 3. Compile Descriptions DataFrame
        # Logic: If caller provided descriptions, use them. Otherwise return empty or fetched ones (if implemented).
        if series_descriptions is not None and not series_descriptions.empty:
            desc_df = series_descriptions
        else:
            desc_df = pd.DataFrame()

        self.logger.info(f"Successfully fetched {len(successful_dfs)} series with metadata.")

        return merged_df.reset_index(), desc_df, failed_series, metadata_df
class BankMetricsProcessor:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config

    def _get_series(self, df: pd.DataFrame, series_list: List[str]) -> pd.Series:
        total = pd.Series(0, index=df.index)
        for col in series_list:
            if col in df:
                total += df[col].fillna(0)
        return total

    def _safe_divide(self, numerator, denominator, fill_value=0.0):
        if np.isscalar(numerator) and np.isscalar(denominator):
            return numerator / denominator if denominator != 0 else fill_value

        if isinstance(denominator, pd.Series):
            denom_safe = denominator.replace(0, np.nan)
            return (numerator / denom_safe).fillna(fill_value)

        try:
            return (numerator / denominator).fillna(fill_value)
        except:
            return fill_value
    # ==================================================================================
    #  UPDATED METRICS PROCESSOR (v30: Restores Income/Provision TTM Logic)
    # ==================================================================================
    def create_derived_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Derives analytical metrics for Credit Risk Dashboard.

        v30 FIXES:
        - Restored Income/Provision TTM calculations (solves 'KeyError: Provision_Exp_TTM').
        - Unified Time-Series loop handles both NCOs and Income Statement flows.
        - Retains v29 crash-proofing and v28 forensic CRE logic.
        """
        if df.empty: return df
        df_processed = df.copy()

        # --- WATERFALL COALESCE: FFIEC 031 (Consolidated) vs FFIEC 041 (Domestic) ---
        # For globally active banks (G-SIBs filing FFIEC 031), RCFD is the total.
        # For domestic-only banks (FFIEC 041/051 filers), RCFD is null and RCON is the total.
        # Coalesce RCFD.fillna(RCON) to ensure apples-to-apples comparisons.
        coalesce_pairs = {
            'HTM_Securities': ('RCFD1754', 'RCON1754'),
            'AFS_Securities': ('RCFD1773', 'RCON1773'),
            'Unused_Commitments_Total': ('RCFD3423', 'RCON3423'),
            'Unused_Commitments_HELOC': ('RCFD3814', 'RCON3814'),
            'Unused_Commitments_NonRE': ('RCFD6550', 'RCON6550'),
        }

        for final_col, (rcfd_col, rcon_col) in coalesce_pairs.items():
            if rcfd_col in df_processed.columns and rcon_col in df_processed.columns:
                df_processed[final_col] = df_processed[rcfd_col].fillna(df_processed[rcon_col])
            elif rcon_col in df_processed.columns:
                df_processed[final_col] = df_processed[rcon_col]
            elif rcfd_col in df_processed.columns:
                df_processed[final_col] = df_processed[rcfd_col]

        # --- SPARSE NICHE SEGMENTS: Fill NaN with 0 for legitimately sparse series ---
        # Fund Finance and NDFI series are sparse because few banks use them.
        # Fill NaN to prevent downstream math errors.
        sparse_niche_cols = ['RCONJ454', 'RCONJ458', 'RCONJ459', 'RCONJ460',
                             'RCFDJ454', 'RCFDJ458', 'RCFDJ459', 'RCFDJ460']
        for col in sparse_niche_cols:
            if col in df_processed.columns:
                df_processed[col] = df_processed[col].fillna(0)

        # --- SYNTHETIC VARIABLE RESOLUTION ---
        # These variables do not exist in the FDIC API; calculate from base components.
        # LNOTHPCS (SBL): fallback to RCFD1545 → RCON1545 → LNOTHER
        sbl_candidates = ['RCFD1545', 'RCON1545', 'LNOTHER']
        df_processed['LNOTHPCS'] = pd.Series(np.nan, index=df_processed.index)
        for col in sbl_candidates:
            if col in df_processed.columns:
                mask = df_processed['LNOTHPCS'].isna()
                df_processed.loc[mask, 'LNOTHPCS'] = df_processed.loc[mask, col]
        df_processed['LNOTHPCS'] = df_processed['LNOTHPCS'].fillna(0)

        # LNOTHNONDEP (Fund Finance): fallback to RCFDJ454 → RCONJ454
        nondep_candidates = ['RCFDJ454', 'RCONJ454']
        df_processed['LNOTHNONDEP'] = pd.Series(np.nan, index=df_processed.index)
        for col in nondep_candidates:
            if col in df_processed.columns:
                mask = df_processed['LNOTHNONDEP'].isna()
                df_processed.loc[mask, 'LNOTHNONDEP'] = df_processed.loc[mask, col]
        df_processed['LNOTHNONDEP'] = df_processed['LNOTHNONDEP'].fillna(0)

        # LNCONAUTO: maps to LNAUTO
        if 'LNAUTO' in df_processed.columns:
            df_processed['LNCONAUTO'] = df_processed['LNAUTO'].fillna(0)
        else:
            df_processed['LNCONAUTO'] = 0.0

        # LNCONCC: maps to LNCRCD
        if 'LNCRCD' in df_processed.columns:
            df_processed['LNCONCC'] = df_processed['LNCRCD'].fillna(0)
        else:
            df_processed['LNCONCC'] = 0.0

        # LNCONOTHX: Derived as LNCON - LNAUTO - LNCRCD
        lncon = df_processed['LNCON'].fillna(0) if 'LNCON' in df_processed.columns else 0
        lnauto = df_processed['LNAUTO'].fillna(0) if 'LNAUTO' in df_processed.columns else 0
        lncrcd = df_processed['LNCRCD'].fillna(0) if 'LNCRCD' in df_processed.columns else 0
        df_processed['LNCONOTHX'] = (lncon - lnauto - lncrcd).clip(lower=0)

        # --- Helper: Compute Quarterly Flow from YTD ---
        def compute_quarterly_from_ytd(df_in, col_name):
            """
            Converts YTD cumulative series to discrete quarterly flows.
            Logic: If Q1, Flow = YTD. Else, Flow = YTD_Current - YTD_Prev.
            """
            if col_name not in df_in.columns:
                return pd.Series(0.0, index=df_in.index)

            # Group by CERT to ensure we don't diff across banks
            q_flows = []
            for cert, group in df_in.groupby('CERT'):
                group = group.sort_values('REPDTE')

                # Calculate diff
                diffs = group[col_name].diff()

                # Fix Q1: In Q1, the flow is the YTD value itself (no subtraction)
                is_q1 = group['REPDTE'].dt.quarter == 1
                diffs.loc[is_q1] = group.loc[is_q1, col_name]

                # Handle cases where diff is negative (accounting adjustments)
                # or fill NaN (first record)
                diffs = diffs.fillna(group[col_name]) # First record fallback

                q_flows.append(diffs)

            return pd.concat(q_flows).reindex(df_in.index).fillna(0)
        def sum_cols(df, cols):
            total = pd.Series(0.0, index=df.index)
            for col in cols:
                if col in df.columns:
                    total = total + df[col].fillna(0)
            return total

        # === ADD THIS HERE (MOVED FROM BOTTOM) ===
        def safe_div(n, d):
            return np.where(d != 0, n / d, 0)

        # --- Helper: Vectorized Best-Of ---
        def best_of(df, primaries, fallbacks=[]):
            res = pd.Series(np.nan, index=df.index)
            for col in primaries:
                if col in df.columns:
                    mask = (res.isna()) | (res == 0)
                    if mask.any():
                        vals = df[col]
                        update_mask = mask & (vals.notna()) & (vals != 0)
                        res.loc[update_mask] = vals.loc[update_mask]
            if fallbacks:
                for col in fallbacks:
                    if col in df.columns:
                        mask = (res.isna()) | (res == 0)
                        if not mask.any(): break
                        vals = df[col]
                        update_mask = mask & (vals.notna()) & (vals != 0)
                        res.loc[update_mask] = vals.loc[update_mask]
            return res.fillna(0)
        def get_sum_of_series(df, series_ids):
            cols = [c for c in series_ids if c in df.columns]
            if not cols:
                return pd.Series([np.nan] * len(df), index=df.index)
            return df[cols].fillna(0).sum(axis=1)


        def sum_cols(df, cols):
            total = pd.Series(0.0, index=df.index)
            for col in cols:
                if col in df.columns:
                    total = total + df[col].fillna(0)
            return total

        # ---------------------------------------------------------
        # [1] DATA MAPPING (Balances & Risk Stack)
        # ---------------------------------------------------------

        # A. Construction
        df_processed['RIC_Constr_Cost']       = best_of(df_processed, ['RCFDJJ04', 'RCONJJ04'])
        df_processed['RIC_Constr_ACL']        = best_of(df_processed, ['RCFDJJ12', 'RCONJJ12'])
        df_processed['RIC_Constr_Nonaccrual'] = sum_cols(df_processed, ['NARECONS'])
        df_processed['RIC_Constr_PD30']       = sum_cols(df_processed, ['P3RECONS'])
        df_processed['RIC_Constr_PD90']       = sum_cols(df_processed, ['P9RECONS'])

        # B. CRE (Components)
        df_processed['RIC_CRE_Cost']       = best_of(df_processed, ['RCFDJJ05', 'RCONJJ05'])
        df_processed['RIC_CRE_ACL']        = best_of(df_processed, ['RCFDJJ13', 'RCFDJJ13'])

        # Risk Stack (Resolved)
        def resolve_cre_metric(df, prefix):
            c_mf  = f'{prefix}REMULT'
            c_row = f'{prefix}RENROW'
            c_res = f'{prefix}RENRES'
            c_rot = f'{prefix}RENROT'
            mf  = df[c_mf].fillna(0)  if c_mf  in df.columns else 0.0
            row = df[c_row].fillna(0) if c_row in df.columns else 0.0
            res = df[c_res].fillna(0) if c_res in df.columns else 0.0
            rot = df[c_rot].fillna(0) if c_rot in df.columns else 0.0
            return mf + np.maximum(rot, row + res)

        # Bank-wide CRE (top-house) resolved stack (kept for totals)
        df_processed['RIC_CRE_TopHouse_Nonaccrual'] = resolve_cre_metric(df_processed, 'NA')
        df_processed['RIC_CRE_TopHouse_PD30']       = resolve_cre_metric(df_processed, 'P3')
        df_processed['RIC_CRE_TopHouse_PD90']       = resolve_cre_metric(df_processed, 'P9')

        # Segmented CRE: Investment = MF + Nonfarm Nonres (NON-owner-occ)
        df_processed['RIC_CRE_Inv_Nonaccrual'] = sum_cols(df_processed, ['NAREMULT', 'NARENROT'])
        df_processed['RIC_CRE_Inv_PD30']       = sum_cols(df_processed, ['P3REMULT', 'P3RENROT'])
        df_processed['RIC_CRE_Inv_PD90']       = sum_cols(df_processed, ['P9REMULT', 'P9RENROT'])

        # Segmented CRE: Owner-Occupied (OO) = ROW + RES
        df_processed['RIC_CRE_OO_Nonaccrual'] = sum_cols(df_processed, ['NARENROW', 'NARENRES'])
        df_processed['RIC_CRE_OO_PD30']       = sum_cols(df_processed, ['P3RENROW', 'P3RENRES'])
        df_processed['RIC_CRE_OO_PD90']       = sum_cols(df_processed, ['P9RENROW', 'P9RENRES'])

        # For segment-level reporting, set RIC_CRE_* = Investment CRE (consistent with your CRE balance definition)
        df_processed['RIC_CRE_Nonaccrual'] = resolve_cre_metric(df_processed, 'NA')
        df_processed['RIC_CRE_PD30']       = resolve_cre_metric(df_processed, 'P3')
        df_processed['RIC_CRE_PD90']       = resolve_cre_metric(df_processed, 'P9')

        # C. Residential (Standard View Balance = Wealth Resi Definition)
        # Fix: do NOT rely on JJ-series here (can be missing/0 for some banks like MSBNA).
        # Use the same numerator logic as the normalized wealth resi balance:
        # First Liens (1797) + Jr Liens (5367) + HELOC (1799)
        resi_first = best_of(df_processed, ['RCON1797', 'RCFD1797']).fillna(0)
        resi_jr    = best_of(df_processed, ['RCON5367', 'RCFD5367']).fillna(0)
        heloc      = best_of(df_processed, ['RCON1799', 'RCFD1799']).fillna(0)

        # Calculate strict sum (Granular)
        resi_sum = resi_first + resi_jr + heloc

        # Calculate broad fallback (Summary Line Item)
        # Matches the logic used later for 'Wealth_Resi_Balance'
        resi_fallback = best_of(df_processed, ['LNRERES']).fillna(0) + best_of(df_processed, ['LNRELOC']).fillna(0)

        # USE FALLBACK IF GRANULAR DATA IS ZERO
        df_processed['RIC_Resi_Cost'] = np.where(resi_sum > 0, resi_sum, resi_fallback)

        df_processed['RIC_Resi_ACL']        = best_of(df_processed, ['RCFDJJ14', 'RCONJJ14'])
        df_processed['RIC_Resi_Nonaccrual'] = sum_cols(df_processed, ['NARERES', 'NARELOC'])
        df_processed['RIC_Resi_PD30']       = sum_cols(df_processed, ['P3RERES', 'P3RELOC'])
        df_processed['RIC_Resi_PD90']       = sum_cols(df_processed, ['P9RERES', 'P9RELOC'])

        # D. C&I
        df_processed['RIC_Comm_Cost']       = best_of(df_processed, ['RCFDJJ07', 'RCONJJ07'])

        df_processed['RIC_Comm_ACL']        = best_of(df_processed, ['RCFDJJ15', 'RCONJJ15'])
        df_processed['RIC_Comm_Nonaccrual'] = sum_cols(df_processed, ['NACI'])
        df_processed['RIC_Comm_PD30']       = sum_cols(df_processed, ['P3CI'])
        df_processed['RIC_Comm_PD90']       = sum_cols(df_processed, ['P9CI'])

        # E. Credit Cards
        df_processed['RIC_Card_Cost']       = best_of(df_processed, ['RCFDJJ08', 'RCONJJ08'])
        df_processed['RIC_Card_ACL']        = best_of(df_processed, ['RCFDJJ16', 'RCONJJ16'])
        df_processed['RIC_Card_Nonaccrual'] = sum_cols(df_processed, ['NACRCD'])
        df_processed['RIC_Card_PD30']       = sum_cols(df_processed, ['P3CRCD'])
        df_processed['RIC_Card_PD90']       = sum_cols(df_processed, ['P9CRCD'])

        # F. Other Consumer
        df_processed['RIC_OthCons_Cost']       = best_of(df_processed, ['RCFDJJ09', 'RCONJJ09'])
        df_processed['RIC_OthCons_ACL']        = best_of(df_processed, ['RCFDJJ17', 'RCONJJ17'])
        df_processed['RIC_OthCons_Nonaccrual'] = sum_cols(df_processed, ['NAAUTO', 'NACONOTH'])
        df_processed['RIC_OthCons_PD30']       = sum_cols(df_processed, ['P3AUTO', 'P3CONOTH'])
        df_processed['RIC_OthCons_PD90']       = sum_cols(df_processed, ['P9AUTO', 'P9CONOTH'])

        # G. Top-House Delinquency & Profitability Series
        df_processed['TopHouse_PD30'] = best_of(df_processed, ['P3LNLS'])
        df_processed['TopHouse_PD90'] = best_of(df_processed, ['P9LNLS'])

        # Income / Exp / Provision (YTD)
        df_processed['Int_Inc_Loans_YTD'] = sum_cols(df_processed, ['ILNDOM', 'ILNFOR'])
        df_processed['Int_Exp_Dep_YTD'] = sum_cols(df_processed, ['EDEPDOM', 'EDEPFOR'])
        df_processed['Provision_Exp_YTD'] = best_of(df_processed, ['ELNATR'])
        df_processed['Total_Int_Exp_YTD'] = best_of(df_processed, ['EINTEXP'])

        # H. Total Nonaccrual
        df_processed['Total_Nonaccrual'] = sum_cols(df_processed,
            ['RIC_Constr_Nonaccrual', 'RIC_CRE_Nonaccrual', 'RIC_Resi_Nonaccrual',
             'RIC_Comm_Nonaccrual', 'RIC_Card_Nonaccrual', 'RIC_OthCons_Nonaccrual',
             'NAREAG', 'NAAG', 'NALS', 'NAOTHLN'])



        # ---------------------------------------------------------
        # [2] FORENSIC TIME SERIES CALCULATION (NCOs & Income)
        # ---------------------------------------------------------

        # 1. DEFINE & INIT COLUMNS
        # Explicitly list all flow columns we need to transform

        # A) Net Charge-Offs
        nco_cols = [
            'NTLNLS', 'NTCI',
            'NTRECONS', 'NTREMULT', 'NTRENROW', 'NTRENROT',
            'NTRERES', 'NTRELOC',
            'NTCRCD', 'NTAUTO', 'NTCONOTH'
        ]

        # B) Income Statement (These are YTD too!)
        income_cols = [
            'Int_Inc_Loans_YTD',
            'Int_Exp_Dep_YTD',
            'Provision_Exp_YTD',
            'Total_Int_Exp_YTD',
            'Net_Income_Raw'
        ]

        # Ensure existence
        all_flow_cols = nco_cols + income_cols
        for col in all_flow_cols:
            if col not in df_processed.columns:
                df_processed[col] = 0.0

        # 2. Compute Quarterly Flows (YTD -> Q)
        # We calculate _Q for everything in the list
        for col in all_flow_cols:
            q_col = col.replace('_YTD', '_Q') if '_YTD' in col else f"{col}_Q"
            df_processed[q_col] = compute_quarterly_from_ytd(df_processed, col)

        # 3. Aggregate Quarterly Flows into Segments (Forensic Logic)
        # CRE
        cre_q_nco = df_processed['NTREMULT_Q'] + np.maximum(
            df_processed['NTRENROT_Q'],
            (df_processed.get('NTRENROW_Q', 0) + df_processed.get('NTRENRES_Q', 0))
        )
        df_processed['RIC_CRE_NCO_Q'] = cre_q_nco

        # Other Segments
        df_processed['RIC_Constr_NCO_Q'] = df_processed['NTRECONS_Q']
        df_processed['RIC_Resi_NCO_Q']   = df_processed['NTRERES_Q'] + df_processed['NTRELOC_Q']
        df_processed['RIC_Comm_NCO_Q']   = df_processed['NTCI_Q']
        df_processed['RIC_Card_NCO_Q']   = df_processed['NTCRCD_Q']
        df_processed['RIC_OthCons_NCO_Q']= df_processed['NTCONOTH_Q'] + df_processed['NTAUTO_Q']

        # 4. Calculate TTM (Rolling 4Q Sum)
        # We map the Q columns to their target TTM columns
        ttm_map = {
            # NCO Segments
            'RIC_CRE_NCO_Q':     'RIC_CRE_NCO_TTM',
            'RIC_Constr_NCO_Q':  'RIC_Constr_NCO_TTM',
            'RIC_Resi_NCO_Q':    'RIC_Resi_NCO_TTM',
            'RIC_Comm_NCO_Q':    'RIC_Comm_NCO_TTM',
            'RIC_Card_NCO_Q':    'RIC_Card_NCO_TTM',
            'RIC_OthCons_NCO_Q': 'RIC_OthCons_NCO_TTM',

            # Total NCO TTM (for normalization)
            'NTLNLS_Q':          'Total_NCO_TTM',

            # Income Statement TTMs (The Fix!)
            'Provision_Exp_YTD_Q': 'Provision_Exp_TTM',
            'Int_Inc_Loans_YTD_Q': 'Int_Inc_Loans_TTM',
            'Total_Int_Exp_YTD_Q': 'Total_Int_Exp_TTM'
        }

        # Note: compute_quarterly_from_ytd creates 'Provision_Exp_YTD_Q' from 'Provision_Exp_YTD'

        temp_ttm_frames = []
        for cert, group in df_processed.groupby('CERT'):
            group = group.sort_values('REPDTE')

            # Rolling Sums
            for q_col, ttm_col in ttm_map.items():
                if q_col in group.columns:
                    group[ttm_col] = group[q_col].rolling(window=4, min_periods=1).sum()

            # Lagged Metrics
            group['Delta_Nonaccrual'] = group['Total_Nonaccrual'].diff()

            # Check for Provision Q col name (it was generated by the loop above)
            prov_q = 'Provision_Exp_YTD_Q'
            if prov_q in group.columns:
                group['Delta_Provision'] = group[prov_q].diff()
            else:
                group['Delta_Provision'] = 0.0

            # CRE Specific Velocity
            cre_pd_total = group['RIC_CRE_PD30'] + group['RIC_CRE_PD90']
            group['Delta_CRE_Nonaccrual'] = group['RIC_CRE_Nonaccrual'].diff()
            group['Lagged_CRE_Total_PD'] = cre_pd_total.shift(1)

            temp_ttm_frames.append(group)

        if temp_ttm_frames:
            df_processed = pd.concat(temp_ttm_frames)
        # [3] TOTALS & DENOMINATORS
        # ---------------------------------------------------------
        df_processed['SBL_Balance'] = best_of(df_processed, ['RCFD1545', 'RCON1545'])
        df_processed['Fund_Finance_Balance'] = best_of(df_processed, ['RCFDJ454', 'RCONJ454'])
        # Fund Finance / NDFI performance (Standard view only)
        df_processed['RIC_Fund_Finance_PD30'] = best_of(df_processed, ['RCONJ458', 'RCFDJ458']).fillna(0)
        df_processed['RIC_Fund_Finance_PD90'] = best_of(df_processed, ['RCONJ459', 'RCFDJ459']).fillna(0)
        df_processed['RIC_Fund_Finance_Nonaccrual'] = best_of(df_processed, ['RCONJ460', 'RCFDJ460']).fillna(0)

        # NCO for Fund Finance: explicitly assumed 0 in this framework
        df_processed['RIC_Fund_Finance_NCO_TTM'] = 0.0


        # NEW: Other (remaining loan buckets we want visible as a separate segment)
        f162 = df_processed['RCFDF162'] if 'RCFDF162' in df_processed.columns else pd.Series(0.0, index=df_processed.index)
        f163 = df_processed['RCFDF163'] if 'RCFDF163' in df_processed.columns else pd.Series(0.0, index=df_processed.index)
        df_processed['Other_Balance'] = (
            best_of(df_processed, ['RCFDJ451', 'RCONJ451']).fillna(0) +
            f162.fillna(0) +
            f163.fillna(0)
        )

        # 1. ADC / Construction (To be excluded from Normalized/WM View)
        df_processed['ADC_Balance'] = best_of(df_processed, ['LNRECONS']).fillna(0)




        # 2. Pure Investment CRE (Multifamily + Non-Owner Occ Nonfarm) -> KEPT in WM View
        df_processed['CRE_Investment_Pure_Balance'] = (
            best_of(df_processed, ['LNREMULT']).fillna(0) +
            best_of(df_processed, ['LNRENROT', 'LNREOTH']).fillna(0)
        )

        # 3. Owner-Occupied CRE (Business dependent) -> KEPT in WM View
        df_processed['CRE_OO_Balance'] = best_of(df_processed, ['LNRENROW']).fillna(0)
        df_processed['RIC_Unalloc_ACL'] = best_of(df_processed, ['RCFDJJ22', 'RCONJJ22'])

        df_processed['RIC_Calculated_ACL'] = (
            df_processed['RIC_Constr_ACL'] + df_processed['RIC_CRE_ACL'] + df_processed['RIC_Resi_ACL'] +
            df_processed['RIC_Comm_ACL'] + df_processed['RIC_Card_ACL'] + df_processed['RIC_OthCons_ACL'] +
            df_processed['RIC_Unalloc_ACL']
        )
        df_processed['RIC_Calculated_Cost'] = (
            df_processed['RIC_Constr_Cost'] + df_processed['RIC_CRE_Cost'] + df_processed['RIC_Resi_Cost'] +
            df_processed['RIC_Comm_Cost'] + df_processed['RIC_Card_Cost'] + df_processed['RIC_OthCons_Cost']
        )

        df_processed['Total_ACL'] = df_processed.get('LNATRES', 0).fillna(0)
        df_processed['Total_Reg_ACL'] = df_processed.get('RB2LNRES', 0).fillna(0)
        df_processed['Gross_Loans'] = df_processed.get('LNLS', 0)

        # Denominators
        df_processed['RIC_Used_Total_ACL'] = np.where(df_processed['Total_ACL'] > 0, df_processed['Total_ACL'], df_processed['RIC_Calculated_ACL'])
        df_processed['RIC_Used_Total_Cost'] = np.where(df_processed['Gross_Loans'] > 0, df_processed['Gross_Loans'], df_processed['RIC_Calculated_Cost'])
        df_processed['RIC_Used_Total_NA'] = df_processed['Total_Nonaccrual']

        # ---------------------------------------------------------
        # [3.6] LIQUIDITY & BALANCE SHEET STRUCTURE (RESTORED)
        # ---------------------------------------------------------
        # Cash and liquid assets for liquidity analysis
        df_processed['Cash_and_Balances'] = best_of(df_processed, ['CHBAL', 'RCFD0010', 'RCON0010']).fillna(0)
        df_processed['Fed_Funds_Sold'] = best_of(df_processed, ['CHBALNI', 'RCFD0081', 'RCON0081']).fillna(0)
        df_processed['Securities_HTM'] = best_of(df_processed, ['HTM_Securities']).fillna(0)
        df_processed['Securities_AFS'] = best_of(df_processed, ['AFS_Securities']).fillna(0)
        df_processed['Trading_Assets'] = best_of(df_processed, ['TRD', 'RCFD3545', 'RCON3545']).fillna(0)

        # Liquid Assets = Cash + Fed Funds Sold + AFS Securities
        df_processed['Liquid_Assets'] = (
            df_processed['Cash_and_Balances'] +
            df_processed['Fed_Funds_Sold'] +
            df_processed['Securities_AFS']
        )

        # High Quality Liquid Assets (more conservative) = Cash + Fed Funds + HTM + AFS
        df_processed['HQLA'] = (
            df_processed['Cash_and_Balances'] +
            df_processed['Fed_Funds_Sold'] +
            df_processed['Securities_HTM'] +
            df_processed['Securities_AFS']
        )

        # Total Assets for ratio
        df_processed['Total_Assets_Raw'] = best_of(df_processed, ['ASSET', 'RCFD2170']).fillna(0)
        df_processed['Total_Equity_Raw'] = best_of(df_processed, ['EQ', 'RCFD3210']).fillna(0)
        df_processed['Total_Deposits_Raw'] = best_of(df_processed, ['DEP', 'RCFD2200']).fillna(0)
        # Income Statement Raw Metrics (prefer FDIC text aliases)
        df_processed['Net_Income_Raw'] = best_of(df_processed, ['NETINC', 'RIAD4340']).fillna(0)
        df_processed['Total_Int_Income_Raw'] = best_of(df_processed, ['INTINC', 'RIAD4107']).fillna(0)
        df_processed['Total_Int_Expense_Raw'] = best_of(df_processed, ['INTEXP', 'RIAD4073']).fillna(0)
        df_processed['Total_Nonint_Income_Raw'] = best_of(df_processed, ['NONII', 'RIAD4079']).fillna(0)
        df_processed['Total_Nonint_Expense_Raw'] = best_of(df_processed, ['NONIX', 'RIAD4093']).fillna(0)
        df_processed['Int_Inc_Loans_Raw'] = best_of(df_processed, ['ILNDOM', 'RIAD4010']).fillna(0)
        df_processed['Int_Exp_Deposits_Raw'] = best_of(df_processed, ['RIAD4115']).fillna(0)

        # Unused Commitments (credit pipeline)
        df_processed['Unused_Commitments'] = best_of(df_processed, ['Unused_Commitments_Total']).fillna(0)

        # [3.5] NORMALIZATION LOGIC (Ex-Commercial/Ex-Consumer) - OPTIMIZED
        # ---------------------------------------------------------
        # Creates "apples-to-apples" comparison by stripping out Mass Market
        # Consumer and Commercial Banking segments.
        # OPTIMIZATION: Collect columns in a dict to prevent "Fragmented DataFrame" warnings.

        # --- [3.5] NORMALIZATION LOGIC (Ex-Commercial/Ex-Consumer) - OPTIMIZED
        # ---------------------------------------------------------
        norm_cols = {}

        # --- A. Map Raw Exclusion Balances ---
        # C&I: Exclude traditional C&I but KEEP SBL (which is often classified within LNCI)
        # Formula: Excl_CI = max(0, LNCI - SBL_Balance)
        lnci_raw = best_of(df_processed, ['LNCI', 'RCON1763', 'RCFD1763']).fillna(0)
        sbl_bal = df_processed.get('SBL_Balance', pd.Series(0, index=df_processed.index)).fillna(0)
        norm_cols['Excl_CI_Balance'] = (lnci_raw - sbl_bal).clip(lower=0)

        norm_cols['Excl_NDFI_Balance'] = best_of(df_processed, ['RCONJ454', 'RCFDJ454']).fillna(0)
        norm_cols['Excl_ADC_Balance'] = best_of(df_processed, ['ADC_Balance', 'RCON1420', 'RCFD1420']).fillna(0)
        norm_cols['Excl_CreditCard_Balance'] = best_of(df_processed, ['RIC_Card_Cost', 'RCFDB538', 'RCONB538']).fillna(0)
        norm_cols['Excl_Auto_Balance'] = best_of(df_processed, ['LNAUTO', 'RCFDK137', 'RCONK137']).fillna(0)
        norm_cols['Excl_Ag_Balance'] = best_of(df_processed, ['LNAG', 'RCFD1590', 'RCON1590']).fillna(0)
        norm_cols['Excl_OO_CRE_Balance'] = best_of(df_processed, ['LNRENROW']).fillna(0)


        # --- B. Calculate Exclusion NCOs (YTD) ---
        ci_nco_direct = best_of(df_processed, ['NTCI', 'RIAD4638']).fillna(0)
        ci_chargeoffs = best_of(df_processed, ['RIAD4608']).fillna(0)
        ci_recoveries = best_of(df_processed, ['RIAD4609']).fillna(0)
        ci_nco_calc = ci_chargeoffs - ci_recoveries
        norm_cols['Excl_CI_NCO_YTD'] = np.where(ci_nco_direct != 0, ci_nco_direct, ci_nco_calc)

        adc_chargeoffs = best_of(df_processed, ['NTLS', 'RIAD4658']).fillna(0)
        adc_recoveries = best_of(df_processed, ['RIAD4659']).fillna(0)
        norm_cols['Excl_ADC_NCO_YTD'] = adc_chargeoffs - adc_recoveries

        cc_chargeoffs = best_of(df_processed, ['RIADB514']).fillna(0)
        cc_recoveries = best_of(df_processed, ['RIADB515']).fillna(0)
        norm_cols['Excl_CC_NCO_YTD'] = cc_chargeoffs - cc_recoveries

        auto_chargeoffs = best_of(df_processed, ['RIADK205']).fillna(0)
        auto_recoveries = best_of(df_processed, ['RIADK206']).fillna(0)
        norm_cols['Excl_Auto_NCO_YTD'] = auto_chargeoffs - auto_recoveries

        norm_cols['Excl_NDFI_NCO_YTD'] = 0.0

        ag_chargeoffs = best_of(df_processed, ['RIAD4635']).fillna(0)
        ag_recoveries = best_of(df_processed, ['RIAD4645']).fillna(0)
        ag_nco_calc = ag_chargeoffs - ag_recoveries
        ag_nco_fallback = best_of(df_processed, ['NTAG']).fillna(0)
        norm_cols['Excl_Ag_NCO_YTD'] = np.where(ag_nco_calc != 0, ag_nco_calc, ag_nco_fallback)

        # --- C. Calculate Exclusion Nonaccruals ---
        norm_cols['Excl_CI_NA'] = best_of(df_processed, ['NACI', 'RCON1608', 'RCFD1608']).fillna(0)
        norm_cols['Excl_NDFI_NA'] = best_of(df_processed, ['RCONJ460', 'RCFDJ460']).fillna(0)
        norm_cols['Excl_ADC_NA'] = best_of(df_processed, ['NARECONS', 'RCON3492', 'RCFD3492']).fillna(0)
        norm_cols['Excl_CC_NA'] = best_of(df_processed, ['RCFDB575', 'RCONB575']).fillna(0)
        norm_cols['Excl_Auto_NA'] = best_of(df_processed, ['RCFDK213', 'RCONK213']).fillna(0)

        ag_na_direct = best_of(df_processed, ['RCFD5341', 'RCON5341']).fillna(0)
        ag_na_fallback = best_of(df_processed, ['NAAG']).fillna(0)
        norm_cols['Excl_Ag_NA'] = np.where(ag_na_direct != 0, ag_na_direct, ag_na_fallback)

        # [FIX] Map OO CRE Risk Metrics for Exclusion
        norm_cols['Excl_OO_CRE_NCO_YTD'] = best_of(df_processed, ['NTRENROW']).fillna(0)
        norm_cols['Excl_OO_CRE_NA'] = best_of(df_processed, ['NARENROW']).fillna(0)
        norm_cols['Excl_OO_CRE_P3'] = best_of(df_processed, ['P3RENROW']).fillna(0)
        norm_cols['Excl_OO_CRE_P9'] = best_of(df_processed, ['P9RENROW']).fillna(0)

        # --- D. Calculate Exclusion Past Dues (P3 = 30-89 Days, P9 = 90+ Days) ---

        # --- D. Calculate Exclusion Past Dues (P3 = 30-89 Days, P9 = 90+ Days) ---
        # 1. Domestic C&I
        norm_cols['Excl_CI_P3'] = best_of(df_processed, ['P3CI', 'RCON1606', 'RCFD1606']).fillna(0)
        norm_cols['Excl_CI_P9'] = best_of(df_processed, ['P9CI', 'RCON1607', 'RCFD1607']).fillna(0)
        # 2. NDFI
        norm_cols['Excl_NDFI_P3'] = best_of(df_processed, ['RCONJ458', 'RCFDJ458']).fillna(0)
        norm_cols['Excl_NDFI_P9'] = best_of(df_processed, ['RCONJ459', 'RCFDJ459']).fillna(0)
        # 3. ADC
        norm_cols['Excl_ADC_P3'] = best_of(df_processed, ['P3RECONS', 'RCON2759', 'RCFD2759']).fillna(0)
        norm_cols['Excl_ADC_P9'] = best_of(df_processed, ['P9RECONS', 'RCON2769', 'RCFD2769']).fillna(0)
        # 4. Credit Cards
        norm_cols['Excl_CC_P3'] = best_of(df_processed, ['P3CRCD', 'RCFDB572']).fillna(0)
        norm_cols['Excl_CC_P9'] = best_of(df_processed, ['P9CRCD', 'RCFDB573']).fillna(0)
        # 5. Auto
        norm_cols['Excl_Auto_P3'] = best_of(df_processed, ['P3AUTO', 'RCFDK214']).fillna(0)
        norm_cols['Excl_Auto_P9'] = best_of(df_processed, ['P9AUTO', 'RCFDK215']).fillna(0)
        # 6. Ag
        norm_cols['Excl_Ag_P3'] = best_of(df_processed, ['RCON2746', 'RCFD2746']).fillna(0)
        norm_cols['Excl_Ag_P9'] = best_of(df_processed, ['RCON2747', 'RCFD2747']).fillna(0)
        # --- E. Sum Total Exclusions ---
        #Added Excl_OO_CRE_Balance to ensure Norm_Gross_Loans is pure Wealth/Inv. CRE
        norm_cols['Excluded_Balance'] = (
            norm_cols['Excl_CI_Balance'] + norm_cols['Excl_NDFI_Balance'] +
            norm_cols['Excl_ADC_Balance'] + norm_cols['Excl_CreditCard_Balance'] +
            norm_cols['Excl_Auto_Balance'] + norm_cols['Excl_Ag_Balance'] +
            norm_cols['Excl_OO_CRE_Balance']
        )

        norm_cols['Excluded_NCO_YTD'] = (
            norm_cols['Excl_CI_NCO_YTD'] + norm_cols['Excl_NDFI_NCO_YTD'] +
            norm_cols['Excl_ADC_NCO_YTD'] + norm_cols['Excl_CC_NCO_YTD'] +
            norm_cols['Excl_Auto_NCO_YTD'] + norm_cols['Excl_Ag_NCO_YTD'] +
            norm_cols['Excl_OO_CRE_NCO_YTD']
        )

        norm_cols['Excluded_Nonaccrual'] = (
            norm_cols['Excl_CI_NA'] + norm_cols['Excl_NDFI_NA'] +
            norm_cols['Excl_ADC_NA'] + norm_cols['Excl_CC_NA'] +
            norm_cols['Excl_Auto_NA'] + norm_cols['Excl_Ag_NA'] +
            norm_cols['Excl_OO_CRE_NA']
        )

        excluded_pd30 = (norm_cols['Excl_CI_P3'] + norm_cols['Excl_NDFI_P3'] +
                         norm_cols['Excl_ADC_P3'] + norm_cols['Excl_CC_P3'] +
                         norm_cols['Excl_Auto_P3'] + norm_cols['Excl_Ag_P3'] +
                         norm_cols['Excl_OO_CRE_P3'])

        excluded_pd90 = (norm_cols['Excl_CI_P9'] + norm_cols['Excl_NDFI_P9'] +
                         norm_cols['Excl_ADC_P9'] + norm_cols['Excl_CC_P9'] +
                         norm_cols['Excl_Auto_P9'] + norm_cols['Excl_Ag_P9'] +
                         norm_cols['Excl_OO_CRE_P9'])

        # --- [CRITICAL STEP] MERGE BATCH 1 ---
        df_norm_batch = pd.DataFrame(norm_cols, index=df_processed.index)
        df_processed = pd.concat([df_processed, df_norm_batch], axis=1)
        for c in ['Excluded_Nonaccrual', 'Excluded_NCO_TTM', 'Excluded_Balance']:
            if c not in df_processed.columns:
                df_processed[c] = 0.0
        # --- F. Convert Exclusion NCOs from YTD to Quarterly, then TTM ---
        df_processed['Excluded_NCO_Q'] = compute_quarterly_from_ytd(df_processed, 'Excluded_NCO_YTD')

        temp_norm_frames = []
        for cert, group in df_processed.groupby('CERT'):
            group = group.sort_values('REPDTE')
            group['Excluded_NCO_TTM'] = group['Excluded_NCO_Q'].rolling(window=4, min_periods=1).sum()
            temp_norm_frames.append(group)

        if temp_norm_frames:
            df_processed = pd.concat(temp_norm_frames)
        else:
            df_processed['Excluded_NCO_TTM'] = 0.0

        # --- G. Calculate Normalized Master Metrics (COMPONENTS ONLY) ---
        # --- Total NCO TTM must be YTD->Q->TTM (never use raw NTLNLS YTD as "TTM") ---
        # [#4] Ensure Total_NCO_TTM is truly TTM of quarterly NCO derived from YTD
        # Prefer the forensic calc (NTLNLS_Q -> rolling 4Q sum). Only compute here if missing.
        if 'Total_NCO_TTM' not in df_processed.columns:
            if 'NTLNLS_Q' not in df_processed.columns:
                df_processed['NTLNLS_Q'] = compute_quarterly_from_ytd(df_processed, 'NTLNLS')

            _tmp = []
            for cert, grp in df_processed.groupby('CERT'):
                grp = grp.sort_values('REPDTE')
                grp['Total_NCO_TTM'] = grp['NTLNLS_Q'].rolling(window=4, min_periods=1).sum()
                _tmp.append(grp)

            df_processed = pd.concat(_tmp) if _tmp else df_processed
            if 'Total_NCO_TTM' not in df_processed.columns:
                df_processed['Total_NCO_TTM'] = 0.0



        df_processed['Norm_Total_NCO'] = (df_processed['Total_NCO_TTM'] - df_processed['Excluded_NCO_TTM']).clip(lower=0)

        # FIX: ensure Total_Nonaccrual exists before using it
        total_na = (
            df_processed['Total_Nonaccrual']
            if 'Total_Nonaccrual' in df_processed.columns
            else pd.Series(0.0, index=df_processed.index)
        )
        # [#3] Define total_na immediately before using it (prevents NameError)
        total_na = df_processed['Total_Nonaccrual'] if 'Total_Nonaccrual' in df_processed.columns else pd.Series(0.0, index=df_processed.index)

        df_processed['Norm_Total_Nonaccrual'] = (
            (total_na - df_processed.get('Excluded_Nonaccrual', 0.0))
            .clip(lower=0)
        )
        # =========================================================
        # OVERRIDE: Normalized Performance (Wealth Segments Only)
        # =========================================================

        # 1. Define Wealth Resi Balance
        if 'Wealth_Resi_Balance' not in df_processed.columns:
            df_processed['Wealth_Resi_Balance'] = (
                best_of(df_processed, ['LNRERES']).fillna(0) +
                best_of(df_processed, ['LNRELOC']).fillna(0)
            )
        wealth_resi_bal = df_processed['Wealth_Resi_Balance']

        # 2. Calculate PURE Wealth Resi Numerators (Excluding CRE)
        wealth_resi_nco_pure = sum_cols(df_processed, ['NTRERES', 'NTRELOC']).clip(lower=0)
        wealth_resi_na_pure  = sum_cols(df_processed, ['NARERES', 'NARELOC']).clip(lower=0)
        wealth_resi_pd30_pure = sum_cols(df_processed, ['P3RERES', 'P3RELOC']).clip(lower=0)
        wealth_resi_pd90_pure = sum_cols(df_processed, ['P9RERES', 'P9RELOC']).clip(lower=0)

        # 3. Calculate Rates for HTML Report
        df_processed['Wealth_Resi_TTM_NCO_Rate'] = safe_div(wealth_resi_nco_pure, wealth_resi_bal)
        df_processed['Wealth_Resi_NA_Rate'] = safe_div(wealth_resi_na_pure, wealth_resi_bal)

        # Total Delinquency (30-89 + 90+)
        df_processed['Wealth_Resi_Delinquency_Rate'] = safe_div(
            wealth_resi_pd30_pure + wealth_resi_pd90_pure,
            wealth_resi_bal
        )

        # 4. Normalized Master Metrics (Resi Pure + Inv CRE)
        # Reconstruct total normalized metrics by adding Investment CRE back in
        df_processed['Norm_Total_NCO'] = wealth_resi_nco_pure + sum_cols(df_processed, ['NTREMULT', 'NTRENROT'])

        # Define total_na immediately before using it
        total_na = df_processed['Total_Nonaccrual'] if 'Total_Nonaccrual' in df_processed.columns else pd.Series(0.0, index=df_processed.index)

        df_processed['Norm_Total_Nonaccrual'] = wealth_resi_na_pure + sum_cols(df_processed, ['NAREMULT', 'NARENROT'])
        df_processed['Norm_PD30'] = wealth_resi_pd30_pure + sum_cols(df_processed, ['P3REMULT', 'P3RENROT'])
        df_processed['Norm_PD90'] = wealth_resi_pd90_pure + sum_cols(df_processed, ['P9REMULT', 'P9RENROT'])

        #(D) Wealth-only denominator ---
        # --- G. Calculate Normalized Master Metrics (COMPONENTS ONLY) ---
        df_processed['Norm_Gross_Loans'] = (df_processed['Gross_Loans'] - df_processed['Excluded_Balance']).clip(lower=0)

        # [#4] Ensure Total_NCO_TTM is truly TTM of quarterly NCO derived from YTD
        if 'Total_NCO_TTM' not in df_processed.columns:
            if 'NTLNLS_Q' not in df_processed.columns:
                df_processed['NTLNLS_Q'] = compute_quarterly_from_ytd(df_processed, 'NTLNLS')

            _tmp = []
            for cert, grp in df_processed.groupby('CERT'):
                grp = grp.sort_values('REPDTE')
                grp['Total_NCO_TTM'] = grp['NTLNLS_Q'].rolling(window=4, min_periods=1).sum()
                _tmp.append(grp)
            df_processed = pd.concat(_tmp) if _tmp else df_processed
            if 'Total_NCO_TTM' not in df_processed.columns:
                df_processed['Total_NCO_TTM'] = 0.0

        df_processed['Norm_Total_NCO'] = (df_processed['Total_NCO_TTM'] - df_processed['Excluded_NCO_TTM']).clip(lower=0)

        # [#3] Define total_na immediately before using it
        total_na = df_processed['Total_Nonaccrual'] if 'Total_Nonaccrual' in df_processed.columns else pd.Series(0.0, index=df_processed.index)
        df_processed['Norm_Total_Nonaccrual'] = (total_na - df_processed['Excluded_Nonaccrual']).clip(lower=0)
        # NOTE: Rates (Norm_Delinquency_Rate, etc.) are calculated in Section 4 using safe_div.

        # ---------------------------------------------------------
        # [4] ANALYTICAL METRICS (Vectorized)
        # ---------------------------------------------------------
        new_cols = {}

        # --- A. Top-Level Ratios ---
        top_house_pd = df_processed['TopHouse_PD30'] + df_processed['TopHouse_PD90']
        new_cols['Top_House_Delinquency_Rate'] = safe_div(top_house_pd, df_processed['Gross_Loans'])

        # [ADDED] Total Delinquency Rate (Safe Calculation)
        # This ensures the field exists and uses safe_div
        new_cols['Total_Delinquency_Rate'] = new_cols['Top_House_Delinquency_Rate']

        new_cols['Nonaccrual_to_Gross_Loans_Rate'] = safe_div(df_processed['Total_Nonaccrual'], df_processed['Gross_Loans'])
        new_cols['Allowance_to_Gross_Loans_Rate'] = safe_div(df_processed['Total_ACL'], df_processed['Gross_Loans'])
        new_cols['Risk_Adj_Allowance_Coverage'] = safe_div(df_processed['Total_ACL'], (df_processed['Gross_Loans'] - df_processed['SBL_Balance']))

        new_cols['SBL_Composition'] = safe_div(df_processed['SBL_Balance'], df_processed['Gross_Loans'])
        new_cols['Fund_Finance_Composition'] = safe_div(df_processed['Fund_Finance_Balance'], df_processed['Gross_Loans'])

        # --- A.2 LIQUIDITY RATIOS (RESTORED) ---
        total_assets = df_processed['Total_Assets_Raw']
        total_assets = np.where(total_assets > 0, total_assets, df_processed.get('ASSET', 1))

        new_cols['Liquidity_Ratio'] = safe_div(df_processed['Liquid_Assets'], total_assets)
        new_cols['HQLA_Ratio'] = safe_div(df_processed['HQLA'], total_assets)
        new_cols['Cash_to_Assets'] = safe_div(df_processed['Cash_and_Balances'], total_assets)
        new_cols['Securities_to_Assets'] = safe_div(
            df_processed['Securities_HTM'] + df_processed['Securities_AFS'],
            total_assets
        )
        new_cols['Loans_to_Deposits'] = safe_div(df_processed['Gross_Loans'], df_processed['Total_Deposits_Raw'])

        # --- A.3 CAPITAL RATIOS (RESTORED) ---
        equity = df_processed['Total_Equity_Raw']
        equity = np.where(equity > 0, equity, df_processed.get('EQ', 1))

        new_cols['Equity_to_Assets'] = safe_div(equity, total_assets)
        new_cols['Leverage_Ratio'] = safe_div(total_assets, equity)

        # --- A.4 RAW PROFITABILITY METRICS (RESTORED - backup for FDIC derived) ---
        # These use raw RIAD series and can be calculated even if FDIC ratios are missing
        # Annualize by multiplying quarterly by 4
        net_income = df_processed['Net_Income_Raw']

        new_cols['ROA_Raw'] = safe_div(net_income * 4, total_assets)
        # Around line 1532
        # Use the discrete quarterly flow (* 4) instead of YTD
        if 'Net_Income_Raw_Q' in df_processed.columns:
            annualized_income = df_processed['Net_Income_Raw_Q'] * 4
        else:
            # Fallback if Q conversion hasn't happened yet (divide YTD by quarter number)
            # Note: Requires a 'Quarter' column or similar logic, simpler to rely on Step 1
            annualized_income = net_income * 4

        new_cols['ROE_Raw'] = safe_div(annualized_income, equity)

        # Net Interest Margin (Raw)
        net_int_income = df_processed['Total_Int_Income_Raw'] - df_processed['Total_Int_Expense_Raw']
        new_cols['NIM_Raw'] = safe_div(net_int_income * 4, total_assets)

        # Efficiency Ratio (Raw) - lower is better
        revenue = df_processed['Total_Nonint_Income_Raw'] + net_int_income
        new_cols['Efficiency_Ratio_Raw'] = safe_div(df_processed['Total_Nonint_Expense_Raw'], revenue)

        # Yield on Loans (Raw)
        new_cols['Yield_on_Loans_Raw'] = safe_div(df_processed['Int_Inc_Loans_Raw'] * 4, df_processed['Gross_Loans'])

        # Cost of Deposits (Raw)
        new_cols['Cost_of_Deposits_Raw'] = safe_div(df_processed['Int_Exp_Deposits_Raw'] * 4, df_processed['Total_Deposits_Raw'])

        # Unused Commitment Ratio
        new_cols['Unused_Commitment_Ratio'] = safe_div(df_processed['Unused_Commitments'], df_processed['Gross_Loans'])

        # --- A.5 NORMALIZED RATIOS (Ex-Commercial/Ex-Consumer) ---
        # These create apples-to-apples comparison metrics
        norm_loans = df_processed['Norm_Gross_Loans']
        norm_nco = df_processed['Norm_Total_NCO']
        norm_na = df_processed['Norm_Total_Nonaccrual']
        norm_pd30 = df_processed['Norm_PD30']
        norm_pd90 = df_processed['Norm_PD90']

        new_cols['Norm_NCO_Rate'] = safe_div(norm_nco, norm_loans)
        new_cols['Norm_Nonaccrual_Rate'] = safe_div(norm_na, norm_loans)
        new_cols['Norm_Delinquency_Rate'] = safe_div((norm_pd30 + norm_pd90), norm_loans)
        new_cols['Norm_ACL_Coverage'] = safe_div(df_processed['Total_ACL'], norm_loans)

        # Normalized Composition (what % of remaining portfolio after exclusions)
        new_cols['Norm_Exclusion_Pct'] = safe_div(df_processed['Excluded_Balance'], df_processed['Gross_Loans'])

        # Segment-specific normalized compositions
        # SBL as % of normalized loans (should be higher for pure private banks)
        # In BankMetricsProcessor -> create_derived_metrics
        # Look for the "Segment-specific normalized compositions" section (approx line 1180)
        new_cols['Norm_SBL_Composition'] = safe_div(df_processed['SBL_Balance'], norm_loans)

        # Fund Finance is explicitly excluded from Norm_Gross_Loans (via Excl_NDFI_Balance),
        # so normalized composition must be 0 by design.
        new_cols['Norm_Fund_Finance_Composition'] = 0.0

        # Normalized Investment CRE % (Uses Pure Balance / Norm Loans)


        # NEW: Other as % of normalized loans
        new_cols['Norm_Other_Composition'] = safe_div(df_processed['Other_Balance'], norm_loans)

        # Normalized Investment CRE % (Uses Pure Balance / Norm Loans)


        # Normalized Investment CRE % (Uses Pure Balance / Norm Loans)
        # Note: Excludes ADC, consistent with the denominator
        new_cols['Norm_CRE_Investment_Composition'] = safe_div(
            df_processed['CRE_Investment_Pure_Balance'],
            df_processed['Norm_Gross_Loans']
        )

        # >>> ADD THIS LINE (Standard numerator aligned to normalized numerator; denom stays Gross_Loans)
        new_cols['CRE_Investment_Composition'] = safe_div(
            df_processed['CRE_Investment_Pure_Balance'],
            df_processed['Gross_Loans']
        )

        # Normalized Owner-Occupied CRE %
        new_cols['Norm_CRE_OO_Composition'] = safe_div(
            df_processed['CRE_OO_Balance'],
            df_processed['Norm_Gross_Loans']
        )

        # Normalized ADC % (Should be 0.0% if fully excluded)
        new_cols['Norm_ADC_Composition'] = safe_div(
            df_processed['ADC_Balance'],
            df_processed['Norm_Gross_Loans']
        )
        # Wealth Resi balance
        # --- Wealth Resi balance (1-4 family first liens + HELOC/open-end) ---
        # Prefer RC-C component lines to avoid double-counting HELOC/open-end
        resi_components = (
            df_processed.reindex(
                columns=['RCFD1797', 'RCON1797', 'RCFD5367', 'RCON5367', 'RCFD5368', 'RCON5368'],
                fill_value=0
            ).sum(axis=1)
        )

        resi_total_primary = best_of(df_processed, ['LNRERES']).fillna(0)
        heloc_primary      = best_of(df_processed, ['LNRELOC']).fillna(0)

        use_components = resi_components.ne(0)
        def compute_wealth_resi_bal(df):
            # Try RC-C components first (if present and non-trivial)
            rcc_cols = ['RCFD1797','RCON1797','RCFD5367','RCON5367','RCFD5368','RCON5368']
            have_any = any(c in df.columns for c in rcc_cols)

            if have_any:
                rcc = df.reindex(columns=rcc_cols, fill_value=0).fillna(0).sum(axis=1)
                # If RC-C is effectively all zeros, fallback
                if (rcc.abs().sum() > 0):
                    return rcc.clip(lower=0)

            # Fallback: FDIC balance proxies (your “known good” definition)
            return (
                best_of(df, ['LNRERES']).fillna(0) +
                best_of(df, ['LNRELOC']).fillna(0)
            ).clip(lower=0)


        # If RC-C components exist, use them (they already include open-end).
        # Else fall back to LNRERES + LNRELOC (your prior logic).
        wealth_resi_bal = compute_wealth_resi_bal(df_processed)
        df_processed['Wealth_Resi_Balance'] = wealth_resi_bal

        new_cols['Norm_Wealth_Resi_Composition'] = safe_div(wealth_resi_bal, norm_loans)
        new_cols['Wealth_Resi_Composition']      = safe_div(wealth_resi_bal, df_processed['Gross_Loans'])
        # IMPROVED NORMALIZED ACL (Partial Normalization via Schedule RI-C)
        # ==============================================================================

        # 1. Get Total Allowance
        total_acl = best_of(df_processed, ['LNATRES', 'RCFD3123', 'RCON3123', 'RCFDJJ19']).fillna(0)

        # 2. Get Excluded Reserves
        res_adc = best_of(df_processed, ['RCFDJJ12', 'RCFDJJ12']).fillna(0)
        res_cc  = best_of(df_processed, ['RCFDJJ16', 'RCFDJJ16']).fillna(0)
        res_oth = best_of(df_processed, ['RCFDJJ18', 'RCFDJJ17']).fillna(0)

        # 3. Calculate "Normalized Allowance" (This creates the variable you were missing)
        norm_acl_balance = total_acl - (res_adc + res_cc + res_oth)
        new_cols['Norm_ACL_Balance'] = norm_acl_balance

        # 4. Calculate Ratios & Shares (NOW it is safe to use norm_acl_balance)
        new_cols['Norm_CRE_ACL_Share'] = safe_div(df_processed['RIC_CRE_ACL'], norm_acl_balance)
        new_cols['Norm_Resi_ACL_Share'] = safe_div(df_processed['RIC_Resi_ACL'], norm_acl_balance)

        new_cols['Norm_RESI_ACL_Coverage'] = safe_div(df_processed['RIC_Resi_ACL'], wealth_resi_bal)
        new_cols['Norm_CRE_ACL_Coverage']  = safe_div(df_processed['RIC_CRE_ACL'], df_processed['CRE_Investment_Pure_Balance'])
        new_cols['Norm_Comm_ACL_Coverage'] = safe_div(df_processed['RIC_Comm_ACL'], df_processed['SBL_Balance'])

        new_cols['Norm_ACL_Coverage'] = safe_div(norm_acl_balance, df_processed['Norm_Gross_Loans'])
        new_cols['Norm_Risk_Adj_Allowance_Coverage'] = safe_div(
            norm_acl_balance,
            (df_processed['Norm_Gross_Loans'] - df_processed['SBL_Balance'])
        )

        # --- B. Profitability & Efficiency ---
        # Ensure TTM columns exist (default to 0 if rolling calc failed due to sparse data)
        prov_ttm = df_processed.get('Provision_Exp_TTM', pd.Series(0, index=df_processed.index))
        inc_ttm = df_processed.get('Int_Inc_Loans_TTM', pd.Series(0, index=df_processed.index))

        new_cols['Provision_to_Loans_Rate'] = safe_div(prov_ttm, df_processed['Gross_Loans'])
        loan_yield = safe_div(inc_ttm, df_processed['Gross_Loans'])
        new_cols['Loan_Yield_Proxy'] = loan_yield
        new_cols['Provision_Elasticity'] = safe_div(df_processed['Delta_Provision'], df_processed['Delta_Nonaccrual'])

        # --- B.5 NORMALIZED PROFITABILITY (Ex-Commercial/Ex-Consumer) ---
        # Normalized yield uses normalized loans as denominator for apples-to-apples comparison
        norm_loan_yield = safe_div(inc_ttm, norm_loans)
        new_cols['Norm_Loan_Yield'] = norm_loan_yield
        # Normalized Provision Rate is misleading because we cannot exclude C&I/Consumer provision flow
        # new_cols['Norm_Provision_Rate'] = safe_div(prov_ttm, norm_loans)  <-- COMMENT THIS OUT
        new_cols['Norm_Provision_Rate'] = np.nan # Set to NaN to prevent bad data

        # Normalized Loss-Adjusted Yield: What you earn after losses on the private bank portfolio
        norm_nco_rate = new_cols['Norm_NCO_Rate']
        new_cols['Norm_Loss_Adj_Yield'] = norm_loan_yield - norm_nco_rate

        # Risk-Adjusted Return: Yield vs Nonaccrual rate on normalized book
        new_cols['Norm_Risk_Adj_Return'] = norm_loan_yield - new_cols['Norm_Nonaccrual_Rate']

        # --- C. Segment Ratios ---
        segments = {
            'Constr':  ('RIC_Constr_ACL', 'RIC_Constr_Cost', 'RIC_Constr_NCO_TTM', 'RIC_Constr_Nonaccrual', 'RIC_Constr_PD30', 'RIC_Constr_PD90'),
            'CRE':     ('RIC_CRE_ACL',    'RIC_CRE_Cost',    'RIC_CRE_NCO_TTM',    'RIC_CRE_Nonaccrual',    'RIC_CRE_PD30',    'RIC_CRE_PD90'),
            'Resi':    ('RIC_Resi_ACL',   'RIC_Resi_Cost',   'RIC_Resi_NCO_TTM',   'RIC_Resi_Nonaccrual',   'RIC_Resi_PD30',   'RIC_Resi_PD90'),
            'Comm':    ('RIC_Comm_ACL',   'RIC_Comm_Cost',   'RIC_Comm_NCO_TTM',   'RIC_Comm_Nonaccrual',   'RIC_Comm_PD30',   'RIC_Comm_PD90'),
            'Card':    ('RIC_Card_ACL',   'RIC_Card_Cost',   'RIC_Card_NCO_TTM',   'RIC_Card_Nonaccrual',   'RIC_Card_PD30',   'RIC_Card_PD90'),
            'OthCons': ('RIC_OthCons_ACL','RIC_OthCons_Cost', 'RIC_OthCons_NCO_TTM', 'RIC_OthCons_Nonaccrual', 'RIC_OthCons_PD30', 'RIC_OthCons_PD90')
        }

        denom_acl = df_processed['RIC_Used_Total_ACL']
        denom_cost = df_processed['RIC_Used_Total_Cost']
        denom_na = df_processed['RIC_Used_Total_NA']

        for seg_name, (acl_col, cost_col, nco_ttm_col, na_col, pd30_col, pd90_col) in segments.items():
            s_acl = df_processed[acl_col].fillna(0)
            s_cost = df_processed[cost_col].fillna(0)
            s_nco_ttm = df_processed.get(nco_ttm_col, pd.Series(0, index=df_processed.index)).fillna(0)
            s_na = df_processed[na_col].fillna(0)
            s_pd30 = df_processed[pd30_col].fillna(0)
            s_pd90 = df_processed[pd90_col].fillna(0)

            # Metrics
            new_cols[f'RIC_{seg_name}_ACL_Coverage'] = safe_div(s_acl, s_cost)
            new_cols[f'RIC_{seg_name}_Risk_Adj_Coverage'] = safe_div(s_acl, s_na)

            # Years of Reserves
            years_res = np.where(s_nco_ttm > 0, s_acl / s_nco_ttm, np.nan)
            new_cols[f'RIC_{seg_name}_Years_of_Reserves'] = np.where(years_res > 100, 100, years_res)

            nco_rate = safe_div(s_nco_ttm, s_cost)
            new_cols[f'RIC_{seg_name}_NCO_Rate'] = nco_rate
            new_cols[f'RIC_{seg_name}_Nonaccrual_Rate'] = safe_div(s_na, s_cost)
            new_cols[f'RIC_{seg_name}_Delinquency_Rate'] = safe_div((s_pd30 + s_pd90), s_cost)

            # Advanced
            new_cols[f'RIC_{seg_name}_Loss_Adj_Yield'] = loan_yield - nco_rate
            new_cols[f'RIC_{seg_name}_ACL_Efficiency'] = safe_div(s_nco_ttm, s_acl)
            new_cols[f'RIC_{seg_name}_Migration_Ratio'] = safe_div((s_pd30 + s_pd90), s_na)

            # Shares
            share_acl = safe_div(s_acl, denom_acl)
            share_loan = safe_div(s_cost, denom_cost)

            new_cols[f'RIC_{seg_name}_ACL_Share'] = share_acl
            new_cols[f'RIC_{seg_name}_Loan_Share'] = share_loan
            new_cols[f'RIC_{seg_name}_Alloc_Mismatch'] = share_acl - share_loan
            new_cols[f'RIC_{seg_name}_Risk_Mismatch'] = share_acl - safe_div(s_na, denom_na)

        # D. Advanced Velocity (Specific for CRE)
        new_cols['RIC_CRE_Conversion_Velocity'] = safe_div(df_processed['Delta_CRE_Nonaccrual'], df_processed['Lagged_CRE_Total_PD'])
        # === [NEW CODE START: DELINQUENCY CONTRIBUTIONS] ===

        # 1. Define Total Bank Delinquency (Denominators)
        # P3ASTOT / P9ASTOT are the series keys for Total Assets Past Due in your file
        total_pd30 = best_of(df_processed, ['P3ASTOT', 'P3ASSET']).fillna(0)
        total_pd90 = best_of(df_processed, ['P9ASTOT', 'P9ASSET']).fillna(0)
        total_pd_all = total_pd30 + total_pd90

        # 2. Define Excluded Segment Delinquencies (Using API Series Keys)
        # We must explicitly sum the PD30/PD90 for segments we are normalizing out.

        pd30_ci = best_of(df_processed, ['P3CI']).fillna(0)
        pd90_ci = best_of(df_processed, ['P9CI']).fillna(0)

        pd30_adc = best_of(df_processed, ['P3RECONS']).fillna(0)
        pd90_adc = best_of(df_processed, ['P9RECONS']).fillna(0)

        pd30_card = best_of(df_processed, ['P3CRC', 'P3CRCD']).fillna(0)
        pd90_card = best_of(df_processed, ['P9CRC', 'P9CRCD']).fillna(0)

        pd30_auto = best_of(df_processed, ['P3AUT', 'P3AUTO']).fillna(0)
        pd90_auto = best_of(df_processed, ['P9AUT', 'P9AUTO']).fillna(0)

        pd30_ag = best_of(df_processed, ['P3AG', 'P3AGR']).fillna(0)
        pd90_ag = best_of(df_processed, ['P9AG', 'P9AGR']).fillna(0)

        # NDFI often lacks a P3 summary code; assume 0 if not mapped
        pd30_ndfi = best_of(df_processed, ['P3NDFI', 'P3DEP']).fillna(0)
        pd90_ndfi = best_of(df_processed, ['P9NDFI', 'P9DEP']).fillna(0)

        # Sum of Excluded Delinquencies
        excluded_pd30 = pd30_ci + pd30_adc + pd30_card + pd30_auto + pd30_ag + pd30_ndfi
        excluded_pd90 = pd90_ci + pd90_adc + pd90_card + pd90_auto + pd90_ag + pd90_ndfi

        # Normalized Denominators
        norm_total_pd30 = total_pd30 - excluded_pd30
        norm_total_pd90 = total_pd90 - excluded_pd90
        norm_total_pd_all = norm_total_pd30 + norm_total_pd90

        # 3. Calculate Shares for Key Segments
        # Mapping Segment -> Numerator Variables
        seg_vars = {
            'ADC': (pd30_adc, pd90_adc),
            'CI':  (pd30_ci, pd90_ci),

            # For Resi and CRE, we rely on the specific P3/P9 codes
            'Resi': (best_of(df_processed, ['P3RERES', 'P3RES']).fillna(0),
                     best_of(df_processed, ['P9RERES', 'P9RES']).fillna(0)),

            # CRE (Non-Owner Occ + Multifamily)
            'CRE':  (
                (best_of(df_processed, ['P3REMULT']).fillna(0) + best_of(df_processed, ['P3RENROT', 'P3CRE']).fillna(0)),
                (best_of(df_processed, ['P9REMULT']).fillna(0) + best_of(df_processed, ['P9RENROT', 'P9CRE']).fillna(0))
            )
        }

        for seg, (p3_val, p9_val) in seg_vars.items():
            seg_total = p3_val + p9_val

            # A. STANDARD SHARES (Denominator = Total Bank)
            new_cols[f'{seg}_Share_of_Total_PD30'] = safe_div(p3_val, total_pd30)
            new_cols[f'{seg}_Share_of_Total_PD90'] = safe_div(p9_val, total_pd90)
            new_cols[f'{seg}_Share_of_Total_PD']   = safe_div(seg_total, total_pd_all)

            # B. NORMALIZED SHARES (Denominator = Norm Bank)
            new_cols[f'Norm_{seg}_Share_of_Total_PD30'] = safe_div(p3_val, norm_total_pd30)
            new_cols[f'Norm_{seg}_Share_of_Total_PD90'] = safe_div(p9_val, norm_total_pd90)
            new_cols[f'Norm_{seg}_Share_of_Total_PD']   = safe_div(seg_total, norm_total_pd_all)


        # Legacy Groups
        new_cols['Group_CRE_ACL_Share'] = new_cols['RIC_CRE_ACL_Share'] + new_cols['RIC_Constr_ACL_Share']
        new_cols['Group_Commercial_ACL_Share'] = new_cols['RIC_Comm_ACL_Share']
        new_cols['Group_Residential_ACL_Share'] = new_cols['RIC_Resi_ACL_Share']
        new_cols['Group_OtherSBL_ACL_Share'] = new_cols['RIC_OthCons_ACL_Share'] + new_cols['RIC_Card_ACL_Share']

        # Integrity
        diff = df_processed['Total_ACL'] - df_processed['RIC_Calculated_ACL']
        pct_diff = safe_div(diff, df_processed['Total_ACL'])

        conds = [
            (df_processed['CERT'] > 90000),
            (np.abs(pct_diff) < 0.05),
            (diff > 0)
        ]
        choices = ["NO_BENCHMARK", "MATCH", "UNDER (Missing Segments?)"]
        new_cols['ACL_Integrity_Status'] = np.select(conds, choices, default="OVER (Double Count?)")

        reg_diff = df_processed['Total_Reg_ACL'] - df_processed['Total_ACL']
        new_cols['ACL_Reg_Divergence'] = safe_div(reg_diff, df_processed['Total_ACL'])

        # Final Merge
        df_final = pd.concat([df_processed, pd.DataFrame(new_cols, index=df_processed.index)], axis=1)

        df_final['RIC_Resi_Best'] = df_final['RIC_Resi_ACL']
        df_final['RIC_Comm_Best'] = df_final['RIC_Comm_ACL']
        df_final['RIC_CommRE_Best'] = df_final['RIC_CRE_ACL']

        # --- RESTORED DUAL-TIER REPORTING METRICS ---
        # The downstream reporting tool explicitly looks for these exact variable names
        # to represent the un-normalized, bank-wide top line.
        df_final['NPL_to_Gross_Loans_Rate'] = df_final.get('Nonaccrual_to_Gross_Loans_Rate', np.nan)
        df_final['Tier_1_Leverage_Ratio'] = df_final.get('RCFD7204', np.nan)
        df_final['Total_Assets'] = df_final.get('ASSET', np.nan)
        df_final['Net_Charge_Off_Rate'] = df_final.get('TTM_NCO_Rate', np.nan)

        return df_final.copy()


    # ==================================================================================
    #  UPDATED 8Q AGGREGATOR (v31: Peak Stress Logic for Nonaccruals)
    # ==================================================================================
    def calculate_8q_averages(self, proc_df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculates 8-quarter averages for Peer Comparison Scatter Plots.

        v31 UPDATES (Nonaccrual Logic):
        - Implements "Peak Stress" logic for Nonaccrual Rates.
        - Instead of mean(rates), calculates Max(Balance) / Mean(Cost).
        - Prevents lumpy CRE nonaccruals from washing out to zero.
        - Adds 'RIC_CRE_Ever_NA_Flag' to highlight banks with any recent stress.
        """
        if proc_df.empty: return pd.DataFrame()

        # Identify numeric columns for standard averaging
        metrics = proc_df.select_dtypes(include=np.number).columns.tolist()
        if 'CERT' in metrics: metrics.remove('CERT')

        results = []

        # Segments to apply "Peak Stress" logic to
        segments = ['Constr', 'CRE', 'Resi', 'Comm', 'Card', 'OthCons']

        for cert, group in proc_df.groupby('CERT'):
            if len(group) < 1: continue # Need at least some data

            # Sort by date to ensure rolling window is correct
            group = group.sort_values('REPDTE')

            # 1. Standard Averages (Baseline)
            # We take the mean of the last 8 periods (or fewer if <8 available)
            # min_periods=1 ensures we get data even for new banks
            avgs = group[metrics].rolling(window=8, min_periods=1).mean().iloc[-1].to_dict()

            # 2. OVERRIDE: Peak Stress Logic for Nonaccruals
            # Problem: Averaging quarterly NA rates (often 0%) dilutes signal.
            # Solution: Max(Nonaccrual Balance) / Mean(Cost Balance) over 8Q.

            for seg in segments:
                na_col = f'RIC_{seg}_Nonaccrual'
                cost_col = f'RIC_{seg}_Cost'
                rate_col = f'RIC_{seg}_Nonaccrual_Rate'

                if na_col in group.columns and cost_col in group.columns:
                    # Rolling Max of the numerator (Did they have a problem?)
                    na_max = group[na_col].rolling(window=8, min_periods=1).max().iloc[-1]
                    # Rolling Mean of the denominator (Average portfolio size)
                    cost_mean = group[cost_col].rolling(window=8, min_periods=1).mean().iloc[-1]

                    # Recalculate the rate based on Peak Stress
                    if cost_mean > 0:
                        avgs[rate_col] = na_max / cost_mean
                    else:
                        avgs[rate_col] = 0.0

            # 3. New Metric: "Ever Nonaccrual" Flag (Option B)
            # Useful for coloring scatter plots (0=Clean, 1=Stressed)
            if 'RIC_CRE_Nonaccrual' in group.columns:
                cre_max = group['RIC_CRE_Nonaccrual'].rolling(window=8, min_periods=1).max().iloc[-1]
                avgs['RIC_CRE_Ever_NA_Flag'] = 1.0 if cre_max > 0 else 0.0

            # Metadata
            rec = {"CERT": cert, "NAME": group['NAME'].iloc[-1]}
            rec.update(avgs)
            results.append(rec)

        return pd.DataFrame(results).set_index('CERT') if results else pd.DataFrame()
    # ==================================================================================
    #  UPDATED TTM CALCULATOR (v26: Growth & Top-House Metrics)
    # ==================================================================================
    def calculate_ttm_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculates trailing 12-month (TTM) and Year-Over-Year (YoY) Growth metrics.

        v26 UPDATES:
        - Targets new 'RIC_' segments for Growth calculations.
        - Calculates 'TTM_NCO_Rate' (Bank-wide) using 4Q Sum / 4Q Avg Loans.
        - Calculates 'TTM_Past_Due_Rate' (Smoothed 4Q Average).
        """
        if df.empty: return df

        all_banks_data = []

        # Define targets for Growth Calculation (Metric Name Prefix : Column Name)
        growth_targets = {
            # Change 'RIC_CRE_Cost' to the robust RC-C derived balance
            'CRE':     'CRE_Investment_Pure_Balance',
            'Constr':  'RIC_Constr_Cost',
            'Resi':    'RIC_Resi_Cost',
            'Comm':    'RIC_Comm_Cost',
            'Card':    'RIC_Card_Cost',
            'OthCons': 'RIC_OthCons_Cost',

            # 2. Special Portfolios
            'SBL':          'SBL_Balance',
            'Fund_Finance': 'Fund_Finance_Balance',

            # 3. Top of House
            'Total_Loans':  'LNLS',
            'Assets':       'ASSET',
            'Deposits':     'DEP'
        }

        for cert, group in df.groupby('CERT'):
            bank_df = group.sort_values("REPDTE").copy()

            # We need at least 1 row, but for TTM we prefer 4.
            # Rolling ops handle <4 gracefully if min_periods set, but growth needs offset.

            # 4-Quarter Moving Average of Loans (Denominator for TTM Rates)
            avg_loans = bank_df['LNLS'].rolling(window=4, min_periods=1).mean()

            # -----------------------------------------------------
            # 1. Calculate YoY (4Q) Growth
            # -----------------------------------------------------
            for prefix, col_name in growth_targets.items():
                if col_name in bank_df.columns:
                    # Get 4-quarter lag
                    curr_vals = bank_df[col_name]
                    prev_vals = bank_df[col_name].shift(4)

                    # Safe Growth Calculation (Vectorized)
                    # Logic: (Curr - Prev) / |Prev|. If Prev=0, return 0 if Curr=0 else NaN
                    diff = curr_vals - prev_vals

                    growth = np.where(
                        prev_vals != 0,
                        diff / prev_vals.abs(),
                        np.where(curr_vals == 0, 0.0, np.nan)
                    )

                    bank_df[f'{prefix}_Growth_TTM'] = growth

            # -----------------------------------------------------
            # 2. Top-House TTM NCO Rate
            # -----------------------------------------------------
            # NTLNLS_Q is created in create_derived_metrics (from NTLNLS YTD)
            if 'NTLNLS_Q' in bank_df.columns:
                ttm_nco = bank_df['NTLNLS_Q'].rolling(window=4, min_periods=1).sum()
                # Standard Definition: TTM NCO / Average Loans
                bank_df['TTM_NCO_Rate'] = ttm_nco / avg_loans

            # -----------------------------------------------------
            # 3. Top-House TTM Past Due Rate (Smoothed)
            # -----------------------------------------------------
            # Uses TopHouse_PD30/90 created in create_derived_metrics
            if 'TopHouse_PD30' in bank_df.columns and 'TopHouse_PD90' in bank_df.columns:
                total_pd = bank_df['TopHouse_PD30'].fillna(0) + bank_df['TopHouse_PD90'].fillna(0)
                # Quarterly Rate
                pd_rate = total_pd / bank_df['LNLS']
                # Smoothed (Average of last 4 quarters)
                bank_df['TTM_Past_Due_Rate'] = pd_rate.rolling(window=4, min_periods=1).mean()

            all_banks_data.append(bank_df)

        return pd.concat(all_banks_data, ignore_index=True) if all_banks_data else df

    def create_latest_snapshot(self, proc_df: pd.DataFrame) -> pd.DataFrame:
        """
        Creates a snapshot of the most recent data for the Summary Dashboard.

        UPDATES:
        - Filters out intermediate YTD and Quarterly NCO columns.
        - Only exports the final TTM (Trailing 12-Month) metrics.
        """
        if proc_df.empty: return pd.DataFrame()

        latest_date = proc_df['REPDTE'].max()
        latest = proc_df[proc_df['REPDTE'] == latest_date].copy()

        snapshot = latest[['CERT', 'NAME', 'REPDTE']].copy()

        # 1. Base Metrics
        base_metrics = [
            'ASSET', 'LNLS', 'Total_ACL', 'Total_Capital',
            'Allowance_to_Gross_Loans_Rate', 'Nonaccrual_to_Gross_Loans_Rate',
            'Risk_Adj_Allowance_Coverage',
            'TTM_NCO_Rate', 'TTM_Past_Due_Rate', 'ACL_Integrity_Status',
            # Dual-Tier Reporting: legacy series for report_generator.py
            'NPL_to_Gross_Loans_Rate', 'Tier_1_Leverage_Ratio',
            'Total_Assets', 'Net_Charge_Off_Rate',
            'Top_House_Delinquency_Rate', 'Total_Delinquency_Rate',
            # Normalized (Ex-Commercial/Ex-Consumer) metrics
            'Norm_NCO_Rate', 'Norm_Nonaccrual_Rate', 'Norm_Delinquency_Rate',
            'Norm_ACL_Coverage', 'Norm_Gross_Loans', 'Norm_Total_NCO',
            'Norm_Total_Nonaccrual', 'Norm_SBL_Composition',
            'Norm_Fund_Finance_Composition', 'Norm_Wealth_Resi_Composition',
            'Norm_Exclusion_Pct',
            # Profitability
            'Loan_Yield_Proxy', 'Norm_Loan_Yield',
            'Provision_to_Loans_Rate', 'Norm_Provision_Rate',
            'Norm_Loss_Adj_Yield', 'Norm_Risk_Adj_Return',
            # Liquidity / Capital
            'Liquidity_Ratio', 'HQLA_Ratio', 'Cash_to_Assets',
            'Securities_to_Assets', 'Loans_to_Deposits',
            'Equity_to_Assets', 'Leverage_Ratio',
            'Gross_Loans',
        ]
        for m in base_metrics:
            if m in latest.columns:
                snapshot[m] = latest[m]

        # 2. Special Portfolios
        special_metrics = [
            'SBL_Balance', 'SBL_Growth_TTM',
            'Fund_Finance_Balance', 'Fund_Finance_Growth_TTM'
        ]
        for m in special_metrics:
            if m in latest.columns:
                snapshot[m] = latest[m]

        # 3. Dynamic RI-C II Segment Metrics
        for col in latest.columns:
            if not col.startswith('RIC_'): continue

            # EXCLUSION: Do not export intermediate YTD or Q columns
            if '_YTD' in col or '_Q' in col:
                continue

            # A. Dollar Columns (Cost, ACL, NCO_TTM) -> Keep raw values
            if any(x in col for x in ['_Cost', '_ACL', '_NCO_TTM', '_Balance', '_Nonaccrual', '_PD30', '_PD90']):
                # Ensure we don't accidentally grab rates that contain these strings
                if '_Rate' not in col and '_Share' not in col and '_Coverage' not in col:
                    snapshot[col] = latest[col]
                    continue

            # B. Ratios & Rates -> Keep decimal
            if any(x in col for x in ['_Rate', '_Share', '_Coverage', '_Mismatch', '_Pct']):
                snapshot[col] = latest[col]
                continue

            # C. Integers / Years
            if '_Years_of_Reserves' in col:
                snapshot[col] = latest[col]
                continue

            # D. Catch-all (Status flags, etc.)
            snapshot[col] = latest[col]

        # 4. Growth Metrics
        for col in latest.columns:
            if '_Growth_TTM' in col:
                snapshot[col] = latest[col]

        return snapshot.set_index('CERT')

class PeerAnalyzer:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self._mdd = _master_dict

        # Define metric mappings for normalized peer groups
        # When use_normalized=True, these standard metrics are replaced with Norm_ versions
        self.normalized_metric_map = {
            'TTM_NCO_Rate': 'Norm_NCO_Rate',
            'Nonaccrual_to_Gross_Loans_Rate': 'Norm_Nonaccrual_Rate',
            'Top_House_Delinquency_Rate': 'Norm_Delinquency_Rate',
            'Allowance_to_Gross_Loans_Rate': 'Norm_ACL_Coverage',
            'Gross_Loans': 'Norm_Gross_Loans',
            'Total_Nonaccrual': 'Norm_Total_Nonaccrual',
            'NTLNLS': 'Norm_Total_NCO',
            # Profitability mappings
            'Loan_Yield_Proxy': 'Norm_Loan_Yield',
            'Provision_to_Loans_Rate': 'Norm_Provision_Rate',
        }

    def create_peer_comparison(self, processed_df: pd.DataFrame) -> pd.DataFrame:
        logging.info("Creating multi-group peer comparison analysis...")
        if processed_df.empty: return pd.DataFrame()

        # Get latest date data
        latest_date = processed_df["REPDTE"].max()
        latest_data = processed_df[processed_df["REPDTE"] == latest_date].copy()

        # Isolate Subject Bank
        subject_data = latest_data[latest_data["CERT"] == self.config.subject_bank_cert]
        if subject_data.empty: return pd.DataFrame()

        numeric_cols = latest_data.select_dtypes(include=["number"]).columns
        metrics_to_compare = [c for c in numeric_cols if c not in ("CERT", "REPDTE")]
        comparison_list = []

        for metric in metrics_to_compare:
            if metric not in latest_data.columns: continue

            # 1. Base Record (Subject Bank Data)
            subject_value = subject_data[metric].iloc[0]
            record = {
                "Metric Code": metric,
                "Metric Name": _get_metric_short_name(metric),
                "Your_Bank": subject_value
            }

            # 2. Iterate through EACH Peer Group
            primary_percentile = None # To store percentile for the "Primary" (Core) group for flagging
            primary_norm_percentile = None  # For normalized Core group

            for group_key, group_info in PEER_GROUPS.items():
                group_name = group_info['short_name']
                group_certs = group_info['certs']
                use_normalized = group_info.get('use_normalized', False)

                # Determine which metric column to use
                if use_normalized and metric in self.normalized_metric_map:
                    # Use normalized version for this peer group
                    actual_metric = self.normalized_metric_map[metric]
                    if actual_metric not in latest_data.columns:
                        continue  # Skip if normalized metric not available
                else:
                    actual_metric = metric

                # Filter data for this specific group
                group_data = latest_data[latest_data["CERT"].isin(group_certs)][actual_metric].dropna()

                if group_data.empty: continue

                # For normalized groups, also get subject's normalized value
                if use_normalized and metric in self.normalized_metric_map:
                    norm_metric = self.normalized_metric_map[metric]
                    if norm_metric in subject_data.columns:
                        subject_value_for_comparison = subject_data[norm_metric].iloc[0]
                    else:
                        subject_value_for_comparison = subject_value
                else:
                    subject_value_for_comparison = subject_value

                # Calculate Stats
                record[f"{group_name} Median"] = group_data.median()
                record[f"{group_name} Mean"] = group_data.mean()

                # Calculate Percentile (Rank)
                if pd.notna(subject_value_for_comparison):
                    pct = stats.percentileofscore(group_data, subject_value_for_comparison, kind='rank')
                    record[f"{group_name} Pct"] = pct

                    # Use CORE group as the driver for the "Performance Flag"
                    if group_key == PeerGroupType.CORE_PRIVATE_BANK:
                        primary_percentile = pct
                    elif group_key == PeerGroupType.CORE_PRIVATE_BANK_NORM:
                        primary_norm_percentile = pct

            # 3. Generate Performance Flag (Based on Core Private Bank Peers)
            if primary_percentile is not None:
                record["Performance_Flag"] = self._get_performance_flag(metric, primary_percentile)
            else:
                record["Performance_Flag"] = "N/A"

            # Add normalized flag if available
            if primary_norm_percentile is not None:
                record["Norm_Performance_Flag"] = self._get_performance_flag(metric, primary_norm_percentile)

            comparison_list.append(record)

        if not comparison_list: return pd.DataFrame()

        comparison_df = pd.DataFrame(comparison_list)
        return comparison_df

    def create_normalized_comparison(self, processed_df: pd.DataFrame) -> pd.DataFrame:
        """
        Creates a focused comparison table specifically for normalized metrics.
        This provides a cleaner view of the Ex-Commercial/Ex-Consumer comparison.
        """
        logging.info("Creating normalized peer comparison analysis...")
        if processed_df.empty: return pd.DataFrame()

        # Get latest date data
        latest_date = processed_df["REPDTE"].max()
        latest_data = processed_df[processed_df["REPDTE"] == latest_date].copy()

        # Isolate Subject Bank
        subject_data = latest_data[latest_data["CERT"] == self.config.subject_bank_cert]
        if subject_data.empty: return pd.DataFrame()

        # Focus on normalized metrics
        norm_metrics = [
            'Norm_NCO_Rate', 'Norm_Nonaccrual_Rate', 'Norm_Delinquency_Rate',
            'Norm_ACL_Coverage', 'Norm_Gross_Loans', 'Norm_Total_NCO',
            'Norm_Total_Nonaccrual', 'Norm_SBL_Composition',
            'Norm_Fund_Finance_Composition', 'Norm_Wealth_Resi_Composition',
            'Norm_Exclusion_Pct', 'Excluded_Balance',
            # Normalized Profitability
            'Norm_Loan_Yield', 'Norm_Provision_Rate',
            'Norm_Loss_Adj_Yield', 'Norm_Risk_Adj_Return'
        ]

        comparison_list = []

        for metric in norm_metrics:
            if metric not in latest_data.columns: continue

            subject_value = subject_data[metric].iloc[0]
            record = {
                "Metric Code": metric,
                "Metric Name": _get_metric_short_name(metric),
                "Your_Bank": subject_value
            }

            # Only iterate through normalized peer groups
            for group_key, group_info in PEER_GROUPS.items():
                if not group_info.get('use_normalized', False):
                    continue

                group_name = group_info['short_name']
                group_certs = group_info['certs']

                group_data = latest_data[latest_data["CERT"].isin(group_certs)][metric].dropna()

                if group_data.empty: continue

                record[f"{group_name} Median"] = group_data.median()
                record[f"{group_name} Mean"] = group_data.mean()
                record[f"{group_name} Min"] = group_data.min()
                record[f"{group_name} Max"] = group_data.max()

                if pd.notna(subject_value):
                    pct = stats.percentileofscore(group_data, subject_value, kind='rank')
                    record[f"{group_name} Pct"] = pct

            comparison_list.append(record)

        if not comparison_list: return pd.DataFrame()

        return pd.DataFrame(comparison_list)

    def _get_performance_flag(self, metric_code: str, percentile: float) -> str:
        """Determines if high/low percentile is good/bad based on metric type."""
        metric = metric_code.lower()

        # Metrics where LOWER is better
        lower_is_better_terms = ["nco", "npl", "past_due", "nonaccrual", "cost_of_funds", "pd30", "pd90", "risk"]
        is_risk_metric = any(term in metric for term in lower_is_better_terms)

        if is_risk_metric:
            if percentile <= 25: return "Top Quartile (Low Risk)"
            if percentile <= 50: return "Better than Median"
            return "Bottom Quartile (High Risk)"

        # Metrics where HIGHER is better (Profitability, Capital, Growth)
        else:
            if percentile >= 75: return "Top Quartile (Strong)"
            if percentile >= 50: return "Better than Median"
            return "Bottom Quartile (Weak)"

class MacroTrendAnalyzer:
    """
    Analyzes macroeconomic time series data, calculates technical indicators,
    and generates structured output for Power BI dashboards.
    """
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.indicator_metadata = self._initialize_metadata()
        self.default_technical_params = self._get_default_tech_params()
        self.technical_params = self._set_technical_parameters()



    def _get_default_tech_params(self) -> Dict:
        """Centralizes the definition of default technical parameters."""
        return {
            'sma_period': 12, 'ema_period': 12,
            'bb_period': 12, 'bb_std': 2.0,
            'rsi_period': 14, 'z_score_period': 12
        }
    def _initialize_metadata(self) -> Dict:
        """
        Defines a comprehensive metadata repository for each macroeconomic indicator.
        This dictionary is the "brain" of the automated analysis, providing context
        for interpretation, classification, and commentary generation.

        Returns:
            Dict]: A nested dictionary where keys are FRED series IDs
                                       and values are dictionaries of metadata attributes.
        """
        return {
            # Leading Indicators
            "USSLIND": {
                "type": "leading", "horizon": "medium", "sectors": ["economy_wide"],
                "benchmark": {"recession": 98.0, "expansion": 102.0}, "annualized": False,
                "best_technical": "EMA", "technical_reason": "Highly responsive to economic turning points."
            },
            "T10Y2Y": { # Assuming this was an existing series
                "type": "leading", "horizon": "long", "sectors": ["banking", "real_estate", "economy_wide"],
                "benchmark": {"inversion": 0.0, "normal": 1.5}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Stable trend indicator for yield curve shape."
            },
            "USALOLITONOSTSAM": {
                "type": "leading", "horizon": "medium", "sectors": ["economy_wide", "global"],
                "benchmark": {"contraction": 99.5, "expansion": 100.5}, "annualized": False,
                "best_technical": "EMA", "technical_reason": "Captures momentum shifts in OECD economic outlook."
            },
            "DRTSCIS": {
                "type": "leading", "horizon": "short", "sectors": ["banking", "credit", "middle_market"],
                "benchmark": {"tightening": 20.0, "easing": -10.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Breakouts signal significant shifts in lending sentiment."
            },

            # Banking & Credit Quality
            "DRCRELEXFACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "real_estate", "credit"],
                "benchmark": {"stress": 3.0, "normal": 1.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Volatility in delinquencies indicates credit cycle stress."
            },
            "CORBLACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "credit", "corporate"],
                "benchmark": {"stress": 1.5, "normal": 0.5}, "annualized": True,
                "best_technical": "SMA", "technical_reason": "Identifies the underlying trend in realized credit losses."
            },
            "DRALACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "credit", "consumer"],
                "benchmark": {"stress": 3.5, "normal": 2.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Smooths quarterly data to reveal the broad credit quality trend."
            },
            "CORALACBN": {
                "type": "lagging", "horizon": "long", "sectors": ["banking", "credit", "economy_wide"],
                "benchmark": {"stress": 2.0, "normal": 0.75}, "annualized": True,
                "best_technical": "SMA", "technical_reason": "Tracks the long-term cycle of aggregate bank loan losses."
            },
            "SOFR3MTB3M": {
                "type": "coincident", "horizon": "short", "sectors": ["banking", "funding", "credit"],
                "benchmark": {"stress": 0.5, "extreme_stress": 1.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Band expansion signals rising funding stress and counterparty risk."
            },

            # Employment
            "PAYEMS": {
                "type": "coincident", "horizon": "short", "sectors": ["economy_wide", "employment"],
                "benchmark": {"contraction_yoy_pct": 0.0}, "annualized": False,
                "best_technical": "RSI", "technical_reason": "Identifies momentum exhaustion in labor market growth."
            },
            "CEU6562000101": {
                "type": "coincident", "horizon": "medium", "sectors": ["healthcare", "employment"],
                "benchmark": {"contraction_yoy_pct": 0.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Healthcare employment is a stable trend, best captured by SMA."
            },

            # Real Estate & Construction
            "NYXRSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "LXXRSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "MIXRNSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "CASTHPI": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Ideal for smoothing quarterly state-level HPI data."},
            "FLSTHPI": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Ideal for smoothing quarterly state-level HPI data."},
            "MPCT04XXS": {
                "type": "coincident", "horizon": "medium", "sectors": ["healthcare", "construction"],
                "benchmark": {"contraction": 0.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Construction spending data is volatile; SMA reveals the underlying investment trend."
            },

            # Global Benchmarks
            "FPCPITOTLZGWLD": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 5.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Annual data requires a long-term moving average to identify global inflation regimes."},
            "FPCPITOTLZGHIC": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 4.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Tracks inflation trends in developed economies."},
            "FPCPITOTLZGLMY": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 6.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Tracks inflation trends in emerging economies."},
            "NYGDPMKTPCDWLD": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Annual data requires a long-term moving average to identify global growth regimes."},
        }

    def _set_technical_parameters(self) -> Dict:
        """
        Sets default and series-specific parameters for technical indicator calculations.
        These can be tuned to optimize for different data frequencies and volatilities.

        Returns:
            Dict]: A dictionary of parameters for each indicator.
        """
        # Default parameters are chosen based on common practice for economic data
        default_params = self.default_technical_params

        specific_params = {
            "T10Y2Y": {'sma_period': 50, 'ema_period': 20, 'bb_period': 50, 'rsi_period': 14, 'z_score_period': 50},
            "USSLIND": {'sma_period': 6, 'ema_period': 6, 'bb_period': 6, 'rsi_period': 14, 'z_score_period': 6},
            "PAYEMS": {'sma_period': 12, 'ema_period': 12, 'bb_period': 20, 'rsi_period': 14, 'z_score_period': 12},
        }

        params = {}
        for series_id in self.indicator_metadata.keys():
            params[series_id] = default_params.copy()
            if series_id in specific_params:
                params[series_id].update(specific_params[series_id])
        return params
    def _calculate_rsi(self, series: pd.Series, period: int) -> pd.Series:
        """
        Helper function to calculate the Relative Strength Index (RSI).
        """
        delta = series.diff(1)
        gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()

        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return rsi
    def _safe_divide(self, numerator, denominator):
        if isinstance(denominator, pd.Series):
            denominator = denominator.replace(0, np.nan)
        return (numerator / denominator).fillna(0)
    def _compute_trend_slope(self, data: pd.Series, window: int) -> float:
        """
        Computes the linear trend slope over a specified window.
        Returns the slope as an annualized percentage change.
        """
        if len(data) < window or data.isna().all():
            return 0.0
        y = data.iloc[-window:].values
        x = np.arange(len(y))
        # Remove NaN values
        mask = ~np.isnan(y)
        if mask.sum() < 2:  # Need at least 2 points
            return 0.0
        x_clean = x[mask]
        y_clean = y[mask]
        try:
            slope, _ = np.polyfit(x_clean, y_clean, 1)
            # Convert to annualized percentage
            mean_val = np.mean(y_clean)
            if mean_val != 0:
                return (slope * 12 / mean_val) * 100
            else:
                return 0.0
        except:
            return 0.0
    def calculate_technical_indicators(self, series: pd.Series, series_id: str) -> pd.DataFrame:
        """
        Calculates a suite of technical indicators for a given time series.
        """
        # Validate input
        if series is None or series.empty or len(series) == 0:
            logger.warning(f"Empty or invalid series data for {series_id}")
            return pd.DataFrame()

        # Ensure series has a datetime index
        if not isinstance(series.index, pd.DatetimeIndex):
            logger.warning(f"Series {series_id} does not have DatetimeIndex, skipping")
            return pd.DataFrame()

        # Drop NaN values before resampling
        series = series.dropna()

        if len(series) < 2:  # Need at least 2 data points
            logger.warning(f"Series {series_id} has fewer than 2 data points after dropping NaN")
            return pd.DataFrame()

        # Resample to monthly frequency
        try:
            series_monthly = series.resample('MS').ffill()
        except Exception as e:
            logger.warning(f"Resampling failed for {series_id}: {e}")
            return pd.DataFrame()

        # Verify resampling produced valid output
        if not isinstance(series_monthly, pd.Series):
            logger.warning(f"Resampling produced non-Series output for {series_id}")
            return pd.DataFrame()

        if series_monthly.empty or len(series_monthly) < 2:
            logger.warning(f"Resampling resulted in insufficient data for {series_id}")
            return pd.DataFrame()

        # Additional check: ensure it's not a scalar
        if series_monthly.ndim != 1:
            logger.warning(f"Series {series_id} has invalid dimensions after resampling")
            return pd.DataFrame()

        params = self.technical_params.get(series_id, self.default_technical_params)

        # Create DataFrame - this should now be safe
        try:
            df = pd.DataFrame({series_id: series_monthly})
            df.index.name = 'date'
        except Exception as e:
            logger.error(f"Failed to create DataFrame for {series_id}: {e}")
            return pd.DataFrame()

        # Calculate indicators (rest of your existing code)
        df[f'{series_id}_SMA'] = series_monthly.rolling(window=params['sma_period'], min_periods=1).mean()
        df[f'{series_id}_EMA'] = series_monthly.ewm(span=params['ema_period'], adjust=False, min_periods=1).mean()

        rolling_mean = series_monthly.rolling(window=params['bb_period'], min_periods=1).mean()
        rolling_std = series_monthly.rolling(window=params['bb_period'], min_periods=1).std()
        df[f'{series_id}_BB_Upper'] = rolling_mean + (params['bb_std'] * rolling_std)
        df[f'{series_id}_BB_Lower'] = rolling_mean - (params['bb_std'] * rolling_std)
        df[f'{series_id}_BB_Width'] = self._safe_divide((df[f'{series_id}_BB_Upper'] - df[f'{series_id}_BB_Lower']), rolling_mean)

        df[f'{series_id}_RSI'] = self._calculate_rsi(series_monthly, params['rsi_period'])

        z_score_mean = series_monthly.rolling(window=params['z_score_period'], min_periods=1).mean()
        z_score_std = series_monthly.rolling(window=params['z_score_period'], min_periods=1).std()
        df[f'{series_id}_ZScore'] = self._safe_divide((series_monthly - z_score_mean), z_score_std)

        return df

    def _get_optimal_technical_indicator(self, series_id: str) -> Dict[str, Any]:
        """
        Selects the optimal technical indicator and provides reasoning based on the
        economic characteristics of the series.
        Args:
            series_id (str): The FRED series ID.
        Returns:
            Dict[str, Any]: A dictionary containing the name of the best indicator
                            and the reasoning for its selection.
        """
        # 1. This dictionary now contains the reasons for choosing each indicator.
        # This fixes the original SyntaxError.
        selection_rules = {
            "SMA": {
                "reason": "Best for identifying long-term trends in stable, less volatile series."
            },
            "EMA": {
                "reason": "Best for tracking recent price changes and identifying short-term trends."
            },
            "Bollinger Bands": {
                "reason": "Best for assessing volatility and identifying overbought/oversold trend conditions."
            },
            "RSI": {
                "reason": "Best for measuring momentum and identifying potential trend reversals."
            }
        }

        # 2. Determine which indicator is best for the given series_id.
        # We still get this from the metadata, with "SMA" as a safe default.
        metadata = self.indicator_metadata.get(series_id, {})
        best_indicator = metadata.get("best_technical", "SMA")

        # 3. Look up the specific reason from the rules dictionary using the best_indicator.
        # We use .get() to provide a default reason in case the indicator isn't in our rules.
        default_rule = {"reason": "Default trend-following indicator used as no specific rule was found."}
        reason = selection_rules.get(best_indicator, default_rule)["reason"]

        # 4. Return the final output, which is now ready to be used in your required column.
        return {
            "indicator": best_indicator,
            "reason": reason
        }
    def calculate_oecd_aggregates(self, fred_data: pd.DataFrame) -> pd.DataFrame:
        """
        Calculates or extracts aggregates for OECD vs. non-OECD country groups
        using World Bank proxies available on FRED.

        Args:
            fred_data (pd.DataFrame): The DataFrame containing all fetched FRED series.

        Returns:
            pd.DataFrame: A DataFrame with the aggregated OECD and non-OECD series.
        """
        # 1. Define constants for the specific series IDs being used as proxies.
        # These are the actual FRED series for inflation in different income groups.
        HI_INCOME_SERIES = 'FPCPITOTLZGHI'   # Proxy for OECD countries
        LOW_MID_INCOME_SERIES = 'FPCPITOTLZGLM' # Proxy for Non-OECD countries
        WORLD_SERIES = 'FPCPITOTLZGOW'   # World aggregate

        # 2. Assign the list of required columns to the variable. This fixes the SyntaxError.
        required_cols = [HI_INCOME_SERIES, LOW_MID_INCOME_SERIES, WORLD_SERIES]

        # Check if all required columns are present in the input DataFrame.
        if not all(col in fred_data.columns for col in required_cols):
            print("Warning: Not all global inflation series found for OECD aggregation.")
            return pd.DataFrame()

        # 3. Create the output by selecting the CORRECT column for each key.
        # This fixes the logical error/ValueError.
        output_df = pd.DataFrame({
            'OECD_Proxy_Inflation': fred_data[HI_INCOME_SERIES],
            'NonOECD_Proxy_Inflation': fred_data[LOW_MID_INCOME_SERIES],
            'World_Inflation': fred_data[WORLD_SERIES]
        })

        # The data is annual, so forward-fill to create a step-function series
        # that can be merged with higher-frequency data.
        output_df.ffill(inplace=True)

        return output_df
    def _detect_frequency(self, series: pd.Series) -> str:
        """
        Infers the frequency of a time series index.
        """
        series = series.dropna()
        if len(series.index) < 3:
            return "Undetermined"
        return pd.infer_freq(series.index) or "Irregular"

    def validate_series_data(self, series_id: str, data: pd.Series) -> Dict[str, Any]:
        """
        Performs data quality checks on a time series and returns a validation summary.
        """
        total = len(data)
        missing = data.isna().sum()
        pct_missing = missing / total * 100 if total else 100
        validation_results = {
            "status": "OK",
            "missing_data_pct": pct_missing,
            "last_update": data.last_valid_index(),
            "data_frequency": self._detect_frequency(data),
        }
        # NEW - safer handling
        try:
            clean_data = data.dropna()
            if len(clean_data) > 3:  # Need at least a few points for z-score
                z_scores = np.abs(stats.zscore(clean_data))
                outlier_count = (z_scores > 3).sum()
                # Handle both scalar and Series results
                validation_results["outliers_detected"] = int(outlier_count) if np.isscalar(outlier_count) else int(outlier_count.item())
            else:
                validation_results["outliers_detected"] = 0
        except Exception as e:
            logger.warning(f"Could not calculate outliers for {series_id}: {e}")
            validation_results["outliers_detected"] = 0
        return validation_results
    def _get_technical_signal(self, series_id: str, data: pd.DataFrame, horizon: str = 'medium') -> str:
        """
        Generates a sophisticated signal by analyzing the state, duration, and
        recent events of the optimal technical indicator.

        Args:
            series_id (str): The FRED series ID (e.g., 'DGS10').
            data (pd.DataFrame): DataFrame containing the series and its technical indicators.
            horizon (str): The lookback horizon ('short', 'medium', 'long') to define
                           the period for signal confirmation.

        Returns:
            str: A descriptive signal string.
        """
        # 1. Map the horizon string to a lookback period (number of data points)
        horizon_map = {'short': 2, 'medium': 5, 'long': 10}
        lookback = horizon_map.get(horizon, 5)

        # Ensure we have enough data to look back
        if len(data) < lookback + 1:
            return "Awaiting Data (Insufficient History)"

        # --- Determine the optimal indicator to use ---
        metadata = self.indicator_metadata.get(series_id, {})
        best_tech = metadata.get("best_technical", "SMA")

        # --- SMA / EMA Logic ---
        if best_tech in ["SMA", "EMA"]:
            col_name = f"{series_id}_{best_tech}"
            if col_name not in data.columns: return "Signal N/A"

            # Get the last two points for crossover detection
            price_now = data[series_id].iloc[-1]
            price_prev = data[series_id].iloc[-2]
            tech_now = data[col_name].iloc[-1]
            tech_prev = data[col_name].iloc[-2]

            if pd.isna(price_now) or pd.isna(tech_now): return "Awaiting Data"

            # Crossover Detection
            if price_prev <= tech_prev and price_now > tech_now:
                return f"Bullish Crossover ({best_tech})"
            if price_prev >= tech_prev and price_now < tech_now:
                return f"Bearish Crossover ({best_tech})"

            # Confirmed Trend Detection (checks the entire lookback period)
            if all(data[series_id].iloc[-lookback:] > data[col_name].iloc[-lookback:]):
                return f"Confirmed Bullish Trend (Above {best_tech} for {lookback} periods)"
            if all(data[series_id].iloc[-lookback:] < data[col_name].iloc[-lookback:]):
                return f"Confirmed Bearish Trend (Below {best_tech} for {lookback} periods)"

            return "Neutral (Consolidating)"

        # --- Bollinger Bands Logic ---
        elif best_tech == "Bollinger Bands":
            upper_col, lower_col = f"{series_id}_BB_Upper", f"{series_id}_BB_Lower"
            if upper_col not in data.columns or lower_col not in data.columns: return "Signal N/A"

            price_now = data[series_id].iloc[-1]
            upper_now, lower_now = data[upper_col].iloc[-1], data[lower_col].iloc[-1]

            if any(pd.isna([price_now, upper_now, lower_now])): return "Awaiting Data"

            # Check for breakouts or reversions
            if price_now > upper_now:
                # Check if it's a new breakout or a sustained one
                if data[series_id].iloc[-2] <= data[upper_col].iloc[-2]:
                    return "Volatility Breakout (High Stress)"
                else:
                    return "Sustained High Volatility"
            elif price_now < lower_now:
                if data[series_id].iloc[-2] >= data[lower_col].iloc[-2]:
                     return "Volatility Breakout (Low/Recession)"
                else:
                    return "Sustained Low Pressure"

            return "Neutral (Range-bound)"

        # --- RSI Logic ---
        elif best_tech == "RSI":
            rsi_col = f"{series_id}_RSI"
            if rsi_col not in data.columns: return "Signal N/A"

            rsi_val = data[rsi_col].iloc[-1]
            if pd.isna(rsi_val): return "Awaiting Data"

            # Add magnitude to the signal
            if rsi_val > 80: return "Extreme Overbought (High Momentum)"
            if rsi_val > 70: return "Overbought"
            if rsi_val < 20: return "Extreme Oversold (Low Momentum)"
            if rsi_val < 30: return "Oversold"

            # Check for momentum change
            avg_rsi_lookback = data[rsi_col].iloc[-lookback:].mean()
            if rsi_val > avg_rsi_lookback and rsi_val > 50: return "Gaining Bullish Momentum"
            if rsi_val < avg_rsi_lookback and rsi_val < 50: return "Gaining Bearish Momentum"

            return "Neutral Momentum"

        return "Not Implemented"

    def _generate_short_term_analysis(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates markdown-formatted short-term (1-3 months) analysis for a series.
        This is applied row-wise to generate commentary for each point in time.
        """
        def generate_row_analysis(row):
            current_value = row[series_id]
            # Get previous row's value for change calculation
            prev_row_index = data.index.get_loc(row.name) - 1
            if prev_row_index < 0:
                change = 0.0
            else:
                prior_value = data[series_id].iloc[prev_row_index]
                change = current_value - prior_value

            if pd.isna(current_value):
                return "No data available for this period."

            tech_signal = self._get_technical_signal(series_id, data.loc[:row.name], 'short')
            annualized_text = ' (annualized)' if metadata.get('annualized', False) else ''

            # Simple outlook logic
            outlook = "Stable"
            if "Bullish" in tech_signal or "High" in tech_signal:
                outlook = "Improving"
            elif "Bearish" in tech_signal or "Low" in tech_signal:
                outlook = "Worsening"

            return (
                f"**Current Level**: {current_value:,.2f}{annualized_text}\n\n"
                f"**Change (MoM)**: {change:+.2f}\n\n"
                f"**Technical Signal**: {tech_signal}\n\n"
                f"**Outlook**: Short-term trend appears to be **{outlook}**."
            )

        # Apply the function to each row of the DataFrame
        return data.apply(generate_row_analysis, axis=1)


    def _generate_long_term_analysis(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates markdown-formatted long-term (1-3 years) analysis for a series.
        Includes trends, momentum, historical context, and regime identification.
        """
        def generate_row_analysis(row):
            current_value = row[series_id]
            current_date = row.name
            if pd.isna(current_value):
                return "No data available for this period."
            # Get historical data up to current point
            historical_data = data.loc[:current_date, series_id].dropna()
            if len(historical_data) < 12:  # Need at least 1 year of data
                return "Insufficient historical data for long-term analysis."
            # 1. Calculate long-term trends (1-3 year windows)
            analysis_parts = []
            # Year-over-year change
            if len(historical_data) >= 12:
                yoy_value = historical_data.iloc[-13] if len(historical_data) >= 13 else historical_data.iloc[0]
                yoy_change = ((current_value - yoy_value) / yoy_value) * 100 if yoy_value != 0 else 0
                yoy_change_abs = current_value - yoy_value
            else:
                yoy_change = 0
                yoy_change_abs = 0
            # Calculate long-term moving averages
            ma_12m = historical_data.rolling(window=12, min_periods=12).mean().iloc[-1] if len(historical_data) >= 12 else current_value
            ma_24m = historical_data.rolling(window=24, min_periods=24).mean().iloc[-1] if len(historical_data) >= 24 else ma_12m
            # Calculate trend slope (annualized percentage change over 12-24 months)
            if len(historical_data) >= 24:
                window_data = historical_data.iloc[-24:]
                x = np.arange(len(window_data))
                y = window_data.values
                if len(y) > 0 and not np.all(np.isnan(y)):
                    slope, _ = np.polyfit(x, y, 1)
                    # Annualize the slope as percentage of mean
                    trend_slope_pct = (slope * 12 / np.mean(y)) * 100 if np.mean(y) != 0 else 0
                else:
                    trend_slope_pct = 0
            else:
                trend_slope_pct = 0
            # Calculate z-score over 3-year window
            if len(historical_data) >= 36:
                rolling_mean = historical_data.rolling(window=36).mean().iloc[-1]
                rolling_std = historical_data.rolling(window=36).std().iloc[-1]
                z_score_3y = (current_value - rolling_mean) / rolling_std if rolling_std > 0 else 0
            else:
                z_score_3y = 0
            # Calculate historical percentile (5-year window)
            if len(historical_data) >= 60:
                percentile_rank = stats.percentileofscore(historical_data.iloc[-60:], current_value)
            elif len(historical_data) >= 12:
                percentile_rank = stats.percentileofscore(historical_data, current_value)
            else:
                percentile_rank = 50
            # Determine regime and context based on metadata
            annualized_text = ' (annualized)' if metadata.get('annualized', False) else ''
            benchmark = metadata.get('benchmark', {})
            # Build analysis text
            analysis_parts.append(f"**Current Level vs Historical Trend:** As of {current_date.strftime('%Y-%m')}, "
                                f"{metadata.get('short', series_id)} stands at {current_value:,.2f}{annualized_text} "
                                f"(YoY: {yoy_change:+.1f}%), which is in the {percentile_rank:.0f}th percentile "
                                f"of its 5-year historical range.")
            # Trend analysis
            trend_direction = "upward" if trend_slope_pct > 0.5 else "downward" if trend_slope_pct < -0.5 else "sideways"
            ma_signal = "above" if current_value > ma_12m else "below"
            analysis_parts.append(f"**Long-Term Trend:** The 24-month trend shows {trend_direction} momentum "
                                f"with an annualized slope of {trend_slope_pct:+.1f}%. "
                                f"Current value is {ma_signal} its 12-month moving average ({ma_12m:,.2f}).")
            # Volatility and regime analysis
            if abs(z_score_3y) > 2:
                volatility_status = "extreme high" if z_score_3y > 2 else "extreme low"
                analysis_parts.append(f"**Volatility/Alerts:** Recent readings have achieved a 3-year z-score of {z_score_3y:.2f}, "
                                    f"indicating {volatility_status} levels relative to historical norms.")
            elif abs(z_score_3y) > 1:
                volatility_status = "elevated" if z_score_3y > 1 else "depressed"
                analysis_parts.append(f"**Volatility/Alerts:** The indicator shows {volatility_status} levels "
                                    f"with a z-score of {z_score_3y:.2f}.")
            # Economic context based on benchmarks and series type
            context_parts = []
            # Check against specific benchmarks
            if benchmark:
                for regime, threshold in benchmark.items():
                    if 'stress' in regime and current_value > threshold:
                        context_parts.append(f"Value exceeds stress threshold of {threshold:.2f}")
                    elif 'recession' in regime and current_value < threshold:
                        context_parts.append(f"Below recession threshold of {threshold:.2f}")
                    elif 'expansion' in regime and current_value > threshold:
                        context_parts.append(f"In expansion territory (above {threshold:.2f})")
                    elif 'inversion' in regime and current_value < threshold:
                        context_parts.append(f"Yield curve inverted (below {threshold:.2f})")
            # Add series-specific context
            if 'inflation' in series_id.lower() or series_id == 'CPIAUCSL':
                if current_value > 2.5:
                    context_parts.append("Consumer inflation is above the Fed's 2% target")
                else:
                    context_parts.append("Inflation remains near or below target")
            if 'unemployment' in metadata.get('long', '').lower() or series_id == 'UNRATE':
                if percentile_rank < 25:
                    context_parts.append("Unemployment near historical lows suggests tight labor market")
                elif percentile_rank > 75:
                    context_parts.append("Elevated unemployment indicates labor market stress")
            if context_parts:
                analysis_parts.append(f"**Economic Context:** {'. '.join(context_parts)}.")
            return '\n\n'.join(analysis_parts)
        # Apply the function to each row
        return data.apply(generate_row_analysis, axis=1)

    def _generate_risk_assessment(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates risk assessment with red/yellow/green flags and numeric scores (0-100).
        Evaluates deviation from benchmarks, volatility, and momentum.
        """
        def generate_row_assessment(row):
            current_value = row[series_id]
            current_date = row.name
            if pd.isna(current_value):
                return "**Risk Score:** N/A - No data available."
            # Get historical data
            historical_data = data.loc[:current_date, series_id].dropna()
            if len(historical_data) < 12:
                return "**Risk Score:** N/A - Insufficient data for risk assessment."
            # Initialize risk components
            risk_score = 0
            risk_factors = []
            risk_color = "Green"
            # 1. Deviation from historical norms (0-40 points)
            if len(historical_data) >= 36:
                rolling_mean = historical_data.rolling(window=36).mean().iloc[-1]
                rolling_std = historical_data.rolling(window=36).std().iloc[-1]
                z_score = abs((current_value - rolling_mean) / rolling_std) if rolling_std > 0 else 0
            else:
                rolling_mean = historical_data.mean()
                rolling_std = historical_data.std()
                z_score = abs((current_value - rolling_mean) / rolling_std) if rolling_std > 0 else 0
            # Map z-score to risk points (0-40)
            deviation_score = min(40, (z_score / 3) * 40)
            risk_score += deviation_score
            if z_score > 2:
                risk_factors.append(f"Extreme deviation ({z_score:.1f}σ)")
            elif z_score > 1:
                risk_factors.append(f"Moderate deviation ({z_score:.1f}σ)")
            # 2. Benchmark breach assessment (0-30 points)
            benchmark = metadata.get('benchmark', {})
            benchmark_score = 0
            for regime, threshold in benchmark.items():
                if 'stress' in regime.lower() or 'extreme' in regime.lower():
                    if current_value > threshold:
                        benchmark_score = 30
                        risk_factors.append(f"Exceeds {regime} level ({threshold:.2f})")
                    elif current_value > threshold * 0.8:  # Within 20% of stress
                        benchmark_score = max(benchmark_score, 20)
                        risk_factors.append(f"Approaching {regime} level")
                elif 'normal' in regime.lower():
                    if abs(current_value - threshold) / threshold > 0.5:
                        benchmark_score = max(benchmark_score, 15)
            risk_score += benchmark_score
            # 3. Volatility assessment (0-20 points)
            if len(historical_data) >= 12:
                recent_volatility = historical_data.iloc[-12:].std()
                long_term_volatility = historical_data.std()
                vol_ratio = recent_volatility / long_term_volatility if long_term_volatility > 0 else 1
                volatility_score = min(20, max(0, (vol_ratio - 1) * 20))
                risk_score += volatility_score
                if vol_ratio > 1.5:
                    risk_factors.append("High recent volatility")
            # 4. Momentum/trend assessment (0-10 points)
            if len(historical_data) >= 6:
                recent_trend = historical_data.iloc[-6:].values
                x = np.arange(len(recent_trend))
                if len(recent_trend) > 1:
                    slope, _ = np.polyfit(x, recent_trend, 1)
                    # For certain indicators, negative trends increase risk
                    indicator_type = metadata.get('type', '')
                    if indicator_type == 'leading' and slope < 0:
                        momentum_score = 10
                        risk_factors.append("Negative momentum in leading indicator")
                    elif 'delinquency' in metadata.get('long', '').lower() and slope > 0:
                        momentum_score = 10
                        risk_factors.append("Rising delinquency trend")
                    else:
                        momentum_score = 0
                    risk_score += momentum_score
            # 5. Special conditions for specific series
            if series_id == 'T10Y2Y' and current_value < 0:
                risk_score = max(risk_score, 80)  # Yield curve inversion is high risk
                risk_factors.append("Yield curve inverted")
            if 'spread' in series_id.lower() and series_id.startswith('BAM'):
                percentile = stats.percentileofscore(historical_data, current_value)
                if percentile > 90:
                    risk_score = max(risk_score, 70)
                    risk_factors.append("Credit spreads at extreme highs")
            # Determine color flag based on final score
            if risk_score >= 70:
                risk_color = "Red"
            elif risk_score >= 40:
                risk_color = "Yellow"
            else:
                risk_color = "Green"
            # Format the assessment
            assessment_parts = [
                f"**Risk Score:** {risk_score:.0f}/100 ({risk_color})"
            ]
            if risk_factors:
                assessment_parts.append(f"**Risk Drivers:** {', '.join(risk_factors)}")
            # Add specific risk interpretation
            if risk_color == "Red":
                assessment_parts.append("**Assessment:** High risk - Indicator signals significant stress or deviation from normal conditions.")
            elif risk_color == "Yellow":
                assessment_parts.append("**Assessment:** Moderate risk - Indicator shows concerning trends requiring close monitoring.")
            else:
                assessment_parts.append("**Assessment:** Low risk - Indicator within normal operating parameters.")
            return ' '.join(assessment_parts)
        # Apply the function to each row
        return data.apply(generate_row_assessment, axis=1)
    def generate_powerbi_output(self, processed_data: Dict) -> pd.DataFrame:
        """
        Generates a Power BI-optimized "long" format DataFrame from the processed
        technical indicator data. Also attaches metadata and generated commentary.
        """
        all_series_dfs = []
        for series_id, df in processed_data.items():
            if df.empty:
                continue
            # Use a copy to avoid side effects
            df_copy = df.copy()
            # --- Generate Commentary ---
            metadata = self.indicator_metadata.get(series_id, {})
            df_copy['Short_Term_Analysis'] = self._generate_short_term_analysis(series_id, df_copy, metadata)
            df_copy['Long_Term_Analysis'] = self._generate_long_term_analysis(series_id, df_copy, metadata)
            df_copy['Risk_Assessment'] = self._generate_risk_assessment(series_id, df_copy, metadata)

            # --- Add Metadata Columns ---
            commentary_cols = ['Short_Term_Analysis', 'Long_Term_Analysis', 'Risk_Assessment']
            df_copy['Type'] = metadata.get('type', 'N/A')
            df_copy['Horizon'] = metadata.get('horizon', 'N/A')
            df_copy['Sectors'] = '|'.join(metadata.get('sectors', []))
            metadata_cols = ['Type', 'Horizon', 'Sectors']
            date_col_name = df_copy.index.name
            df_copy.reset_index(inplace=True)
            id_vars = [date_col_name] + commentary_cols + metadata_cols

            value_vars = [col for col in df_copy.columns if col not in id_vars and col != series_id]

            base_df = df_copy[id_vars + [series_id]].copy()
            base_df['Indicator_Name'] = 'Actual Value'
            base_df.rename(columns={series_id: 'Indicator_Value', date_col_name: 'date'}, inplace=True)

            melted_df = pd.DataFrame()
            if value_vars:
                melted_df = df_copy.melt(
                    id_vars=id_vars,
                    value_vars=value_vars,
                    var_name='Indicator_Name',
                    value_name='Indicator_Value'
                )
                melted_df.rename(columns={date_col_name: 'date'}, inplace=True)
            series_final_df = pd.concat([base_df, melted_df], ignore_index=True)
            series_final_df['Series_ID'] = series_id
            all_series_dfs.append(series_final_df)
        if not all_series_dfs:
            return pd.DataFrame()
        # Concatenate all the processed series DataFrames into one big one
        final_output_df = pd.concat(all_series_dfs, ignore_index=True)
        return final_output_df


class ExcelOutputGenerator:
    """Generates the final Excel dashboard output file."""

    # FFIEC 031/041 schedule patterns (Description column text matching)
    _SCHEDULE_PATTERNS = {
        "RC-C (Loans and Leases)": r"RC-C",
        "RC-N (Past Due and Nonaccrual)": r"RC-N",
        "RI-B Part I (Charge-offs and Recoveries)": r"RI-B.*(?:charge|recover)",
        "RI-B Part II (Changes in ACL)": r"RI-B.*(?:allowance|ACL|credit loss)",
    }

    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self._mdd = _master_dict
        self.audit_df = self._build_audit_trail()
        self._enrich_audit_with_source_code()

    # ------------------------------------------------------------------
    #  Static Analysis: Python_Calculation_Code & Dependencies (AST)
    # ------------------------------------------------------------------

    # Variables that hold computed metrics inside create_derived_metrics
    _TARGET_VARS = frozenset({"df_processed", "new_cols", "norm_cols", "df_final"})

    # Segment names used in for-loops with f-string keys
    _SEGMENT_NAMES = [
        "Constr", "CRE", "Resi", "Comm", "Card", "OthCons",
        "ADC", "CI",
    ]

    # Human-readable segment labels for the Segment_Name column
    _SEGMENT_LABELS = {
        "Constr": "Construction",
        "CRE": "CRE",
        "Resi": "Residential",
        "Comm": "C&I",
        "Card": "Credit Cards",
        "OthCons": "Other Consumer",
        "ADC": "ADC",
        "CI": "C&I",
    }

    # Maximum recursion depth for lineage tracing
    _MAX_TRACE_DEPTH = 10

    def _enrich_audit_with_source_code(self) -> None:
        """Append calculation, lineage, dependency, and segment-mapping
        columns to *self.audit_df*.

        Columns added:
            Python_Calculation_Code    – direct RHS expression
            Full_Calculation_Trace     – recursive multi-step formula trace
            Upstream_Derived_Dependencies – all unique derived deps across trace
            Scope                      – Bankwide | Segment-Specific
            Segment_Name               – N/A | CRE | Residential | …
            Calculation_Type           – Raw_Regulatory | Normalized_Base |
                                         Ratio | Trend_Derived
        """
        empty_cols = {
            "Python_Calculation_Code": "",
            "Full_Calculation_Trace": "",
            "Upstream_Derived_Dependencies": "",
            "Scope": "",
            "Segment_Name": "",
            "Calculation_Type": "",
        }
        if self.audit_df.empty:
            for col, default in empty_cols.items():
                self.audit_df[col] = default
            return

        try:
            code_map, direct_deps_map = self._ast_extract_metrics()
        except Exception:
            for col, default in empty_cols.items():
                self.audit_df[col] = (
                    "AST Parse Error — Manual Audit Required"
                    if col == "Python_Calculation_Code" else default
                )
            return

        all_keys = set(code_map.keys())

        # Pre-compute lineage traces and deep dependencies for every metric
        trace_map: dict[str, str] = {}
        deep_deps_map: dict[str, str] = {}
        for mc in code_map:
            trace, all_deps = self._build_lineage_trace(mc, code_map, all_keys)
            trace_map[mc] = trace
            deep_deps_map[mc] = ", ".join(sorted(all_deps)) if all_deps else ""

        # Map columns onto audit rows
        calc_col: list[str] = []
        trace_col: list[str] = []
        deps_col: list[str] = []
        scope_col: list[str] = []
        seg_col: list[str] = []
        ctype_col: list[str] = []

        for _, row in self.audit_df.iterrows():
            mc = row["Metric_Code"]
            is_derived = (
                row.get("Is_Derived") is True
                or "Derived" in str(row.get("Source_of_Truth", ""))
            )

            if mc in code_map:
                calc_col.append(code_map[mc])
                trace_col.append(trace_map.get(mc, ""))
                deps_col.append(deep_deps_map.get(mc, ""))
            elif is_derived:
                calc_col.append("Planned — Not Yet Implemented")
                trace_col.append("")
                deps_col.append("")
            else:
                calc_col.append("")
                trace_col.append("")
                deps_col.append("")

            scope, seg_name, calc_type = self._classify_metric(mc, code_map)
            scope_col.append(scope)
            seg_col.append(seg_name)
            ctype_col.append(calc_type)

        self.audit_df["Python_Calculation_Code"] = calc_col
        self.audit_df["Full_Calculation_Trace"] = trace_col
        self.audit_df["Upstream_Derived_Dependencies"] = deps_col
        self.audit_df["Scope"] = scope_col
        self.audit_df["Segment_Name"] = seg_col
        self.audit_df["Calculation_Type"] = ctype_col

    # ---- Part 1: Recursive Lineage Solver ----------------------------

    def _build_lineage_trace(
        self,
        metric_code: str,
        code_map: dict[str, str],
        all_keys: set[str],
    ) -> tuple[str, set[str]]:
        """Recursively expand a metric's formula up to ``_MAX_TRACE_DEPTH``
        levels, stopping when a terminal (raw regulatory code) is reached.

        Terminal codes that are NOT derived (i.e. raw regulatory fields like
        LNLS, ASSET, RCFD1410) are annotated with their MasterDataDictionary
        definition so the trace is readable by non-coders.

        Returns ``(trace_string, all_deps)`` where *trace_string* is a
        pipe-delimited multi-step expansion and *all_deps* is the
        cumulative set of every derived metric touched at any level.
        """
        steps: list[str] = []
        all_deps: set[str] = set()
        visited: set[str] = set()
        terminal_codes: set[str] = set()

        frontier = [metric_code]

        for depth in range(1, self._MAX_TRACE_DEPTH + 1):
            next_frontier: list[str] = []
            for mc in frontier:
                if mc in visited:
                    continue
                visited.add(mc)
                rhs = code_map.get(mc)
                if rhs is None:
                    continue
                steps.append(f"[Step {depth}] {mc} = {rhs}")
                # Find direct dependencies within this RHS
                direct = self._ast_extract_deps(rhs, all_keys)
                direct.discard(mc)
                all_deps |= direct
                # Queue derived deps for deeper expansion (skip raw/terminal)
                for dep in sorted(direct):
                    if dep in code_map and dep not in visited:
                        next_frontier.append(dep)
                    elif dep not in code_map:
                        terminal_codes.add(dep)

            if not next_frontier:
                break
            frontier = next_frontier

        # Annotate terminal regulatory codes with dictionary definitions
        for tc in sorted(terminal_codes):
            info = self._mdd.lookup_metric(tc)
            name = info.get("Metric_Name", "")
            if name and name != tc and info.get("Source_of_Truth", "") != "Not Found":
                steps.append(f"[Regulatory Definition] {tc}: {name}")

        return " | ".join(steps), all_deps - {metric_code}

    # ---- Part 2: Segment Mapping & Classification --------------------

    # Prefixes/patterns that indicate segment-specific metrics
    _SEG_PREFIXES = {
        "RIC_Constr_": "Construction",
        "RIC_CRE_": "CRE",
        "RIC_Resi_": "Residential",
        "RIC_Comm_": "C&I",
        "RIC_Card_": "Credit Cards",
        "RIC_OthCons_": "Other Consumer",
        "RIC_Fund_Finance_": "Fund Finance",
        "Constr_": "Construction",
        "CRE_": "CRE",
        "Resi_": "Residential",
        "ADC_": "ADC",
        "CI_": "C&I",
        "Wealth_Resi_": "Residential",
        "Group_CRE_": "CRE",
        "Group_Commercial_": "C&I",
        "Group_Residential_": "Residential",
    }

    # Patterns for Normalized metrics
    _NORM_PREFIXES = (
        "Norm_", "Excl_", "Excluded_", "Norm_CRE_", "Norm_ADC_",
        "Norm_Wealth_", "Norm_SBL_", "Norm_Fund_",
    )

    def _classify_metric(
        self,
        metric_code: str,
        code_map: dict[str, str],
    ) -> tuple[str, str, str]:
        """Return ``(Scope, Segment_Name, Calculation_Type)`` for a metric.

        Scope:
            ``Bankwide`` – consolidated metric (no segment prefix)
            ``Segment-Specific`` – tied to a particular loan segment

        Segment_Name:
            ``N/A`` for bankwide; otherwise the human-readable segment.

        Calculation_Type:
            ``Raw_Regulatory``   – a raw MDRM/FDIC/FRED code not computed
            ``Normalized_Base``  – an intermediate normalisation building-block
            ``Ratio``            – a rate / percentage / coverage metric
            ``Trend_Derived``    – time-series derived (TTM, delta, velocity)
        """
        # --- Calculation_Type -------------------------------------------
        if metric_code not in code_map:
            # Not computed in create_derived_metrics -> raw regulatory field
            calc_type = "Raw_Regulatory"
        elif any(kw in metric_code for kw in ("_Rate", "_Ratio", "_Coverage",
                 "_Composition", "_Share", "_Pct", "_Yield", "_Efficiency",
                 "_Mismatch", "_Elasticity", "ROA_", "ROE_", "NIM_",
                 "Cost_of_", "Leverage_")):
            calc_type = "Ratio"
        elif any(kw in metric_code for kw in ("TTM", "Delta_", "Lagged_",
                 "_Velocity", "_Q", "_YTD")):
            calc_type = "Trend_Derived"
        elif metric_code.startswith(self._NORM_PREFIXES):
            calc_type = "Normalized_Base"
        else:
            # Check RHS content for ratio indicators
            rhs = code_map.get(metric_code, "")
            if "safe_div" in rhs:
                calc_type = "Ratio"
            elif "rolling" in rhs or "diff()" in rhs or "shift(" in rhs:
                calc_type = "Trend_Derived"
            elif metric_code.startswith(("Excl_", "Excluded_")):
                calc_type = "Normalized_Base"
            else:
                calc_type = "Normalized_Base"

        # --- Scope & Segment_Name --------------------------------------
        scope = "Bankwide"
        seg_name = "N/A"

        # Check explicit segment prefixes
        for prefix, label in self._SEG_PREFIXES.items():
            if metric_code.startswith(prefix):
                scope = "Segment-Specific"
                seg_name = label
                break

        # Also catch Norm_ + segment patterns like Norm_CRE_Investment_Composition
        if seg_name == "N/A" and scope == "Bankwide":
            mc_upper = metric_code.upper()
            for seg_short, seg_label in self._SEGMENT_LABELS.items():
                # Match patterns like Norm_{seg}_ or Group_{seg}_
                if f"_{seg_short.upper()}_" in mc_upper:
                    scope = "Segment-Specific"
                    seg_name = seg_label
                    break

        # SBL and Fund Finance are segment-specific
        if seg_name == "N/A":
            if "SBL_" in metric_code:
                scope = "Segment-Specific"
                seg_name = "SBL"
            elif "Fund_Finance_" in metric_code:
                scope = "Segment-Specific"
                seg_name = "Fund Finance"

        return scope, seg_name, calc_type

    # ---- AST helpers -------------------------------------------------

    # Growth-prefix names used in calculate_ttm_metrics loops
    _TTM_GROWTH_PREFIXES = [
        "CRE", "Constr", "Resi", "Comm", "Card", "OthCons",
        "SBL", "Fund_Finance", "Total_Loans", "Assets", "Deposits",
    ]

    def _ast_extract_metrics(self) -> tuple[dict[str, str], dict[str, str]]:
        """Parse *create_derived_metrics* **and** *calculate_ttm_metrics*
        with :mod:`ast` and return ``(code_map, deps_map)`` where each
        maps metric-code strings to extracted source / comma-separated
        dependency lists.
        """
        import ast as _ast
        import textwrap as _tw

        raw_map: dict[str, str] = {}

        # ---------- helper: walk a method and collect assignments ---------
        def _walk_method(method_ref, extra_vars: frozenset = frozenset()):
            src = _tw.dedent(inspect.getsource(method_ref))
            tree = _ast.parse(src)
            allowed = self._TARGET_VARS | extra_vars

            for node in _ast.walk(tree):
                if not isinstance(node, _ast.Assign):
                    continue
                for target in node.targets:
                    if not isinstance(target, _ast.Subscript):
                        continue
                    var_name = self._ast_var_name(target.value)
                    if var_name not in allowed:
                        continue
                    key = self._ast_extract_key(target.slice)
                    if key is None:
                        continue
                    rhs_src = _ast.get_source_segment(src, node.value)
                    if rhs_src is None:
                        try:
                            rhs_src = _ast.unparse(node.value)
                        except Exception:
                            continue
                    raw_map[key] = " ".join(rhs_src.split())

        # 1. Primary: create_derived_metrics
        _walk_method(BankMetricsProcessor.create_derived_metrics)

        # 2. Secondary: calculate_ttm_metrics (growth, TTM rates)
        try:
            _walk_method(
                BankMetricsProcessor.calculate_ttm_metrics,
                extra_vars=frozenset({"bank_df", "df"}),
            )
        except Exception:
            pass

        # 3. Regex fallback for anything the AST walk missed
        try:
            cdm_src = _tw.dedent(inspect.getsource(
                BankMetricsProcessor.create_derived_metrics
            ))
            _rx = re.compile(
                r"""(?:df_processed|new_cols|norm_cols|df_final)"""
                r"""\[['"]([A-Za-z0-9_]+)['"]\]\s*=""",
            )
            for m in _rx.finditer(cdm_src):
                k = m.group(1)
                if k not in raw_map:
                    line_end = cdm_src.find("\n", m.end())
                    rhs = cdm_src[m.end():line_end].strip().lstrip("= ")
                    raw_map[k] = " ".join(rhs.split())
        except Exception:
            pass

        # 4. Expand f-string templates into concrete metric names
        code_map: dict[str, str] = {}
        for key, rhs in raw_map.items():
            if "{" not in key:
                code_map[key] = rhs
            else:
                for seg in self._SEGMENT_NAMES:
                    expanded = (
                        key.replace("{seg_name}", seg)
                           .replace("{seg}", seg)
                           .replace("{...}", seg)
                    )
                    code_map.setdefault(expanded, rhs)
                for pfx in self._TTM_GROWTH_PREFIXES:
                    expanded = key.replace("{prefix}", pfx)
                    if expanded != key:
                        code_map.setdefault(expanded, rhs)

        # 5. Build direct dependency map
        all_keys = set(code_map.keys())
        deps_map: dict[str, str] = {}
        for metric_code, rhs_src in code_map.items():
            refs = self._ast_extract_deps(rhs_src, all_keys)
            refs.discard(metric_code)
            deps_map[metric_code] = ", ".join(sorted(refs)) if refs else ""

        return code_map, deps_map

    @staticmethod
    def _ast_var_name(node) -> str | None:
        """Return the variable name from an AST node (Name or Attribute)."""
        import ast as _ast
        if isinstance(node, _ast.Name):
            return node.id
        if isinstance(node, _ast.Attribute):
            return node.attr
        return None

    @staticmethod
    def _ast_extract_key(slice_node) -> str | None:
        """Extract the string key from a subscript slice AST node.

        Handles:
        - ``ast.Constant('key')`` → ``'key'``
        - ``ast.JoinedStr`` (f-string) → template like ``'RIC_{seg_name}_ACL_Coverage'``
        """
        import ast as _ast
        if isinstance(slice_node, _ast.Constant) and isinstance(slice_node.value, str):
            return slice_node.value
        if isinstance(slice_node, _ast.JoinedStr):
            parts: list[str] = []
            for v in slice_node.values:
                if isinstance(v, _ast.Constant):
                    parts.append(str(v.value))
                elif isinstance(v, _ast.FormattedValue):
                    if isinstance(v.value, _ast.Name):
                        parts.append("{" + v.value.id + "}")
                    else:
                        parts.append("{...}")
            return "".join(parts)
        return None

    def _ast_extract_deps(self, rhs_src: str, all_keys: set[str]) -> set[str]:
        """Walk the AST of an RHS code snippet and collect every metric-key
        reference (subscript or ``.get()`` call) that is itself a known
        derived metric.

        Falls back to regex if the snippet cannot be parsed as an expression.
        """
        import ast as _ast

        deps: set[str] = set()

        # Try parsing as expression first, then as statement(s)
        rhs_tree = None
        for mode in ("eval", "exec"):
            try:
                rhs_tree = _ast.parse(rhs_src, mode=mode)
                break
            except SyntaxError:
                continue

        if rhs_tree is None:
            # Regex fallback for truly unparseable fragments
            pat = re.compile(
                r"""(?:df_processed|new_cols|norm_cols|df_final)"""
                r"""(?:\[['"]|\.get\s*\(\s*['"])([^'"]+)['"]"""
            )
            return set(pat.findall(rhs_src)) & all_keys

        for node in _ast.walk(rhs_tree):
            # Subscript: var['key']
            if isinstance(node, _ast.Subscript):
                var_name = self._ast_var_name(node.value)
                if var_name in self._TARGET_VARS:
                    k = self._ast_extract_key(node.slice)
                    if k and k in all_keys:
                        deps.add(k)

            # Method call: var.get('key', default)
            elif isinstance(node, _ast.Call) and isinstance(node.func, _ast.Attribute):
                if node.func.attr == "get":
                    var_name = self._ast_var_name(node.func.value)
                    if var_name in self._TARGET_VARS and node.args:
                        arg0 = node.args[0]
                        if isinstance(arg0, _ast.Constant) and isinstance(arg0.value, str):
                            if arg0.value in all_keys:
                                deps.add(arg0.value)

            # Also catch string constants inside list literals passed to
            # helper functions like sum_cols(df, ['X', 'Y']).
            # These are upstream *source* fields, not derived, so only
            # include them if they are in all_keys.
            elif isinstance(node, _ast.Constant) and isinstance(node.value, str):
                if node.value in all_keys:
                    deps.add(node.value)

        return deps

    # ------------------------------------------------------------------
    #  Audit Trail Builder
    # ------------------------------------------------------------------

    def _build_audit_trail(self) -> pd.DataFrame:
        """Build the Data_Dictionary_Audit ledger.

        Columns: Metric_Code, Metric_Name, Description, Source_of_Truth,
                 Is_Derived, Usage_Status

        *Used in Dashboard*  — every FDIC field, FRED series, and derived
        metric the dashboard actually fetches or computes.

        *Schedule Reference (Unused)* — all remaining MDRM items from
        FFIEC 031/041 schedules RC-C, RC-N, RI-B Part I, RI-B Part II
        that are NOT already marked as Used.
        """
        rows: list[dict] = []
        used_codes: set[str] = set()

        # --- 1. FDIC fields used by the dashboard ---
        for code in FDIC_FIELDS_TO_FETCH:
            info = self._mdd.lookup_metric(code)
            rows.append({
                "Metric_Code": code,
                "Metric_Name": info["Metric_Name"],
                "Description": info["Description"],
                "Source_of_Truth": info["Source_of_Truth"],
                "Is_Derived": False,
                "Usage_Status": "Used in Dashboard",
            })
            used_codes.add(code.upper())

        # --- 2. FRED series used by the dashboard ---
        for _category, series_dict in FRED_SERIES_TO_FETCH.items():
            for fred_code, meta in series_dict.items():
                if fred_code.upper() in used_codes:
                    continue
                rows.append({
                    "Metric_Code": fred_code,
                    "Metric_Name": meta.get("short", fred_code),
                    "Description": meta.get("long", ""),
                    "Source_of_Truth": "FRED (Federal Reserve Economic Data)",
                    "Is_Derived": False,
                    "Usage_Status": "Used in Dashboard",
                })
                used_codes.add(fred_code.upper())

        # --- 3. Local / Derived metrics used by the dashboard ---
        for code, meta in LOCAL_DERIVED_METRICS.items():
            if code.upper() in used_codes:
                continue
            rows.append({
                "Metric_Code": code,
                "Metric_Name": meta.get("short", code),
                "Description": meta.get("long", ""),
                "Source_of_Truth": "Tier 3 — Local/Derived",
                "Is_Derived": True,
                "Usage_Status": "Used in Dashboard",
            })
            used_codes.add(code.upper())

        # --- 4. MDRM Schedule Reference rows (RC-C, RC-N, RI-B) ---
        mdrm_df = self._mdd.get_mdrm_dataframe()
        if mdrm_df is not None and not mdrm_df.empty:
            call_report_mask = mdrm_df["Reporting Form"].isin(
                ["FFIEC 031", "FFIEC 041", "FFIEC 051"]
            )
            cr_df = mdrm_df[call_report_mask].copy()
            desc_col = cr_df["Description"].fillna("")

            for schedule_label, pattern in self._SCHEDULE_PATTERNS.items():
                matched = cr_df[desc_col.str.contains(pattern, case=False, na=False)]
                for _, row in matched.iterrows():
                    mnemonic = str(row.get("Mnemonic", "")).strip()
                    item_code = str(row.get("Item Code", "")).strip()
                    composite = f"{mnemonic}{item_code}" if mnemonic and item_code else (mnemonic or item_code)
                    if not composite or composite.upper() in used_codes:
                        continue
                    rows.append({
                        "Metric_Code": composite,
                        "Metric_Name": str(row.get("Item Name", composite)).strip(),
                        "Description": str(row.get("Description", "")).strip()[:500],
                        "Source_of_Truth": f"MDRM — {schedule_label}",
                        "Is_Derived": False,
                        "Usage_Status": "Schedule Reference (Unused)",
                    })
                    used_codes.add(composite.upper())

        audit = pd.DataFrame(rows)
        if not audit.empty:
            # Sort: Used first, then Reference; within each group, alphabetical
            status_order = {"Used in Dashboard": 0, "Schedule Reference (Unused)": 1}
            audit["_sort"] = audit["Usage_Status"].map(status_order).fillna(2)
            audit.sort_values(["_sort", "Metric_Code"], inplace=True, ignore_index=True)
            audit.drop(columns="_sort", inplace=True)

        logging.info(
            "Audit trail built: %d Used, %d Schedule Reference.",
            len(audit[audit["Usage_Status"] == "Used in Dashboard"]),
            len(audit[audit["Usage_Status"] == "Schedule Reference (Unused)"]),
        )
        return audit

    # ------------------------------------------------------------------
    #  Excel Writer
    # ------------------------------------------------------------------

    def write_excel_output(self, file_path: str, **kwargs):
        """Writes all DataFrames to a single styled Excel file with multiple sheets."""
        logging.info(f"Writing final dashboard to: {file_path}")
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in kwargs.items():
                if isinstance(df, pd.DataFrame) and not df.empty:
                    n_rows, n_cols = df.shape
                    logging.info(f"Sheet '{sheet_name}' contains {n_rows} rows and {n_cols} cols.")
                    if n_rows > 1048576:
                        logging.error(f"Sheet '{sheet_name}' too large ({n_rows} rows).")
                    write_index = sheet_name in ["Latest_Peer_Snapshot", "Averages_8Q_All_Metrics", "Data_Validation_Report"]
                    df.to_excel(writer, sheet_name=sheet_name, index=write_index)

            # Write the audit trail sheet
            if not self.audit_df.empty:
                self.audit_df.to_excel(writer, sheet_name="Data_Dictionary_Audit", index=False)

            logging.info("All data written, starting styling...")
            self._style_audit_sheet(writer)
            self._apply_summary_styles(writer, kwargs.get("Summary_Dashboard"))
            self._apply_snapshot_styles(writer, kwargs.get("Latest_Peer_Snapshot"))
            self._apply_macro_analysis_styles(writer, kwargs.get("Macro_Analysis"))
        logging.info("Excel file written and styled successfully.")

    # ------------------------------------------------------------------
    #  Styling helpers
    # ------------------------------------------------------------------

    def _style_audit_sheet(self, writer):
        """Format the Data_Dictionary_Audit sheet with visual row distinction."""
        sheet_name = "Data_Dictionary_Audit"
        if sheet_name not in writer.sheets:
            return

        logging.info(f"Styling {sheet_name} sheet...")
        ws = writer.sheets[sheet_name]

        # Column widths
        ws.column_dimensions["A"].width = 25    # Metric_Code
        ws.column_dimensions["B"].width = 35    # Metric_Name
        ws.column_dimensions["C"].width = 80    # Description
        ws.column_dimensions["D"].width = 30    # Source_of_Truth
        ws.column_dimensions["E"].width = 12    # Is_Derived
        ws.column_dimensions["F"].width = 28    # Usage_Status
        ws.column_dimensions["G"].width = 70    # Python_Calculation_Code
        ws.column_dimensions["H"].width = 100   # Full_Calculation_Trace
        ws.column_dimensions["I"].width = 50    # Upstream_Derived_Dependencies
        ws.column_dimensions["J"].width = 18    # Scope
        ws.column_dimensions["K"].width = 22    # Segment_Name
        ws.column_dimensions["L"].width = 22    # Calculation_Type

        # Visual fills
        used_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # light blue
        ref_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # light grey
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Style header row (row 1)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Style data rows: colour by Usage_Status (column F = col index 6)
        # Full_Calculation_Trace (col H = 8) gets wrap text for readability
        trace_col_idx = 8
        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=6)
            fill = used_fill if status_cell.value == "Used in Dashboard" else ref_fill
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                if col_idx == trace_col_idx:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _apply_summary_styles(self, writer, df: pd.DataFrame):
        """Applies conditional formatting to the Summary_Dashboard sheet using openpyxl."""
        if df is None or df.empty: return
        sheet_name = 'Summary_Dashboard'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")

        # Define styles using openpyxl objects
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        green_font = Font(color='006100')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        yellow_font = Font(color='9C6500')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006')

        # Set column widths
        worksheet.column_dimensions['A'].width = 15 # Metric Code
        worksheet.column_dimensions['B'].width = 35 # Metric Name
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
            worksheet.column_dimensions[col_letter].width = 18

        try:
            flag_col_index = df.columns.get_loc('Performance_Flag') + 1 # +1 for 0-based to 1-based
            flag_col_letter = openpyxl.utils.get_column_letter(flag_col_index)

            # Define conditional formatting rules
            green_rule = FormulaRule(formula=[f'SEARCH("Top Quartile",${flag_col_letter}2)'], fill=green_fill, font=green_font)
            yellow_rule = FormulaRule(formula=[f'SEARCH("Better than Median",${flag_col_letter}2)'], fill=yellow_fill, font=yellow_font)
            red_rule = FormulaRule(formula=[f'SEARCH("Bottom Quartile",${flag_col_letter}2)'], fill=red_fill, font=red_font)

            data_range = f"A2:{openpyxl.utils.get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            worksheet.conditional_formatting.add(data_range, green_rule)
            worksheet.conditional_formatting.add(data_range, yellow_rule)
            worksheet.conditional_formatting.add(data_range, red_rule)

        except KeyError:
            logger.warning("Styling skipped: 'Performance_Flag' column not found in Summary_Dashboard.")

    def _apply_snapshot_styles(self, writer, df: pd.DataFrame):
        """Applies number formatting to the Latest_Peer_Snapshot sheet using openpyxl."""
        if df is None or df.empty: return
        sheet_name = 'Latest_Peer_Snapshot'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")

        worksheet.column_dimensions['A'].width = 15  # CERT index
        worksheet.column_dimensions['B'].width = 35  # NAME

        percent_format = '0.00%'

        # Start from column 2 because to_excel writes the index
        for col_num, col_name in enumerate(df.columns, 2):
            col_letter = openpyxl.utils.get_column_letter(col_num)

            if col_letter not in ['A', 'B']:
                 worksheet.column_dimensions[col_letter].width = 20

            if any(term in col_name for term in ["Rate", "Pct", "Ratio", "Comp", "Growth", "Risk", "Funds"]):
                for cell in worksheet[col_letter]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = percent_format

    def _apply_macro_analysis_styles(self, writer, df: pd.DataFrame):
        if df is None or df.empty: return
        sheet_name = 'Macro_Analysis'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")
        worksheet.column_dimensions['A'].width = 25 # date
        worksheet.column_dimensions['H'].width = 30 # Series_ID
        worksheet.column_dimensions['I'].width = 25 # Indicator_Name
        worksheet.column_dimensions['B'].width = 40 # Short_Term_Analysis
        worksheet.column_dimensions['C'].width = 40 # Long_Term_Analysis
        worksheet.column_dimensions['D'].width = 40 # Risk_Assessment
class PowerBIGenerator:
    """Generates M and DAX scripts for a Power BI Star Schema."""

    def generate_powerbi_setup_file(self, output_dir: str, powerbi_macro_df: pd.DataFrame, fred_desc_df: pd.DataFrame):
        logger.info("Generating Power BI setup file...")
        Path(output_dir).mkdir(exist_ok=True)
        output_path = Path(output_dir) / "PowerBI_Setup.txt"

        m_code_part = self._get_m_language_code()
        dax_code_part = self._get_dax_measures_code()
        bonus_code_part = self._get_bonus_m_code()
        instructions_part = self._get_instructions()

        full_content = (
            f"{instructions_part}\n"
            f"//--- M LANGUAGE QUERIES ---\n"
            f"//--- Paste each query into a new Blank Query in Power BI's Power Query Editor ---\n\n"
            f"{m_code_part}\n"
            f"//--- DATA MODEL RELATIONSHIPS ---\n"
            f"//--- Power BI should auto-detect these. If not, create them manually in the 'Model' view. ---\n\n"
            f"// Relationship 1: Dates\n"
            f"//   Table 1: fMacro_Data (Many side)\n"
            f"//   Table 2: dDate (One side)\n"
            f"//   Column:  DateKey\n"
            f"//   Cardinality: Many to One (*:1)\n"
            f"//   Cross-filter direction: Single\n\n"
            f"// Relationship 2: Series\n"
            f"//   Table 1: fMacro_Data (Many side)\n"
            f"//   Table 2: dSeries (One side)\n"
            f"//   Column:  SeriesKey\n"
            f"//   Cardinality: Many to One (*:1)\n"
            f"//   Cross-filter direction: Single\n\n"
            f"{dax_code_part}\n"
            f"{bonus_code_part}\n"
        )

        with open(output_path, "w") as f:
            f.write(full_content)
        logger.info(f"Successfully wrote Power BI setup file to {output_path}")

    def _get_instructions(self) -> str:
        return f"""
//==============================================================================
// Power BI Setup Script
// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
//==============================================================================

// INSTRUCTIONS:
// 1. In Power BI, click "Get data" -> "Excel workbook" and select the generated Excel file.
// 2. In the Navigator, select the 'Macro_Analysis' and 'FRED_Descriptions' sheets. Click "Transform Data".
//    This loads the data into the Power Query Editor.
// 3. Once in the Power Query Editor, create a new "Blank Query" for each of the M Language Queries below.
//    (Right-click in the Queries pane -> New Query -> Blank Query).
// 4. Copy and paste the code for each query into its corresponding blank query's Advanced Editor.
// 5. Rename the queries exactly as specified in the titles (e.g., "dDate", "dSeries", "fMacro_Data").
// 6. Click "Close & Apply" to load the data model.
// 7. Once loaded, verify the relationships as described below in the 'Model' view.
// 8. Finally, go to the "Data" view, select the 'fMacro_Data' table, and create each DAX measure.
"""

    def _get_m_language_code(self) -> str:
        dDate_m = """
// --- Dimension Table: dDate ---
// This query generates a complete and dynamic date table based on your data.
let
    // **FIXED**: Point to the correct 'Macro_Analysis' sheet
    SourceTable = Macro_Analysis,
    MinDate = List.Min(Table.Column(SourceTable, "Date")),
    MaxDate = List.Max(Table.Column(SourceTable, "Date")),

    DayCount = Duration.Days(MaxDate - MinDate) + 1,
    DateList = List.Dates(MinDate, DayCount, #duration(1, 0, 0, 0)),
    DateTable = Table.FromList(DateList, Splitter.SplitByNothing(), {"Date"}),

    #"Changed Type to Date" = Table.TransformColumnTypes(DateTable,{{"Date", type date}}),
    #"Inserted Year" = Table.AddColumn(#"Changed Type to Date", "Year", each Date.Year([Date]), Int64.Type),
    #"Inserted Quarter Num" = Table.AddColumn(#"Inserted Year", "QuarterNum", each Date.QuarterOfYear([Date]), Int64.Type),
    #"Added Quarter" = Table.AddColumn(#"Inserted Quarter Num", "Quarter", each "Q" & Text.From([QuarterNum]), type text),
    #"Inserted Month" = Table.AddColumn(#"Added Quarter", "Month", each Date.Month([Date]), Int64.Type),
    #"Inserted Month Name" = Table.AddColumn(#"Inserted Month", "MonthName", each Date.ToText([Date], "MMMM"), type text),
    #"Inserted Day" = Table.AddColumn(#"Inserted Month Name", "Day", each Date.Day([Date]), Int64.Type),
    #"Inserted Day Name" = Table.AddColumn(#"Inserted Day", "DayName", each Date.ToText([Date], "dddd"), type text),
    #"Inserted Week of Year" = Table.AddColumn(#"Inserted Day Name", "WeekOfYear", each Date.WeekOfYear([Date]), Int64.Type),
    #"Added DateKey" = Table.AddColumn(#"Inserted Week of Year", "DateKey", each [Year] * 10000 + [Month] * 100 + [Day], Int64.Type),

    #"Reordered Columns" = Table.ReorderColumns(#"Added DateKey",{"DateKey", "Date", "Year", "Quarter", "Month", "MonthName", "Day", "DayName", "WeekOfYear"}),
    #"Removed QuarterNum" = Table.RemoveColumns(#"Reordered Columns",{"QuarterNum"})
in
    #"Removed QuarterNum"
"""
        dSeries_m = """
// --- Dimension Table: dSeries ---
// It assumes you have loaded the 'FRED_Descriptions' and 'Macro_Analysis' sheets from Excel.
let
    SourceDescriptions = FRED_Descriptions,
    SourceData = Macro_Analysis,

    // Get the distinct list of Series IDs from the main data table to ensure relevance
    MacroDataSeries = Table.Distinct(Table.SelectColumns(SourceData, {"Series_ID"})),

    // Join descriptions with the actual series present in the data
    #"Merged Queries" = Table.NestedJoin(SourceDescriptions, {"Series ID"}, MacroDataSeries, {"Series_ID"}, "Join", JoinKind.Inner),
    #"Removed Columns" = Table.RemoveColumns(#"Merged Queries",{"Join"}),

    // Add a unique integer key for the dimension, which will be the primary key
    #"Added Index" = Table.AddIndexColumn(#"Removed Columns", "SeriesKey", 1, 1, Int64.Type),

    // Select and reorder for clarity
    #"Selected Columns" = Table.SelectColumns(#"Added Index",{"SeriesKey", "Series ID", "long", "short", "Category", "Sectors"}),
    #"Renamed Columns" = Table.RenameColumns(#"Selected Columns",{{"long", "Description"}, {"short", "Series Name"}})
in
    #"Renamed Columns"
"""
        fMacro_Data_m = """
// --- Fact Table: fMacro_Data ---
// It assumes you have the queries 'Macro_Analysis', and 'dSeries' already created.
let
    Source = Macro_Analysis,

    #"Added DateKey" = Table.AddColumn(Source, "DateKey", each Date.Year([Date]) * 10000 + Date.Month([Date]) * 100 + Date.Day([Date]), Int64.Type),

    #"Merged with dSeries" = Table.NestedJoin(#"Added DateKey", {"Series_ID"}, dSeries, {"Series ID"}, "dSeries", JoinKind.LeftOuter),
    #"Expanded SeriesKey" = Table.ExpandTableColumn(#"Merged with dSeries", "dSeries", {"SeriesKey"}, {"SeriesKey"}),

    #"Selected Columns" = Table.SelectColumns(#"Expanded SeriesKey",{"DateKey", "SeriesKey", "Indicator_Value"}),
    #"Changed Type" = Table.TransformColumnTypes(#"Selected Columns",{{"Indicator_Value", type number}, {"DateKey", Int64.Type}, {"SeriesKey", Int64.Type}})
in
    #"Changed Type"
"""
        return f"{dDate_m}\n{dSeries_m}\n{fMacro_Data_m}"

    def _get_dax_measures_code(self) -> str:
        return """
//--- DAX MEASURES & COLUMNS ---
//--- Select the 'fMacro_Data' table, then click 'New Measure' for each item. ---

[Selected Value] = SUM('fMacro_Data'[Indicator_Value])

[Latest Value] =
CALCULATE(
    [Selected Value],
    LASTDATE('dDate'[Date])
)

[Previous Month Value] =
CALCULATE(
    [Selected Value],
    DATEADD('dDate'[Date], -1, MONTH)
)

[Month-over-Month Change] =
VAR CurrentValue = [Selected Value]
VAR PreviousValue = [Previous Month Value]
RETURN
    IF(NOT ISBLANK(CurrentValue) && NOT ISBLANK(PreviousValue), CurrentValue - PreviousValue, BLANK())

[Year-over-Year Value] =
CALCULATE(
    [Selected Value],
    SAMEPERIODLASTYEAR('dDate'[Date])
)

[Year-over-Year Change] =
VAR CurrentValue = [Selected Value]
VAR PreviousValue = [Year-over-Year Value]
RETURN
    IF(NOT ISBLANK(CurrentValue) && NOT ISBLANK(PreviousValue), CurrentValue - PreviousValue, BLANK())
"""

    def _get_bonus_m_code(self) -> str:
        return """
//--- BONUS: M-LANGUAGE FOR BRIDGE TABLE ---
// --- Bridge Table: dSeries_Sectors ---
let
    Source = dSeries,
    #"Removed Other Columns" = Table.SelectColumns(Source,{"SeriesKey", "Sectors"}),
    #"Split Column by Delimiter" = Table.ExpandListColumn(Table.TransformColumns(#"Removed Other Columns", {{"Sectors", Splitter.SplitTextByDelimiter("|", QuoteStyle.None), let itemType = (type nullable text) meta [Serialized.Text = true] in type {itemType}}}), "Sectors"),
    #"Renamed Columns" = Table.RenameColumns(#"Split Column by Delimiter",{{"Sectors", "Sector"}}),
    #"Trimmed Text" = Table.TransformColumns(#"Renamed Columns",{{"Sector", Text.Trim, type text}}),
    #"Filtered Rows" = Table.SelectRows(#"Trimmed Text", each ([Sector] <> null and [Sector] <> ""))
in
    #"Filtered Rows"
"""



# ==================================================================================
#  4. MAIN DASHBOARD CLASS & EXECUTION
# ==================================================================================

class BankPerformanceDashboard:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.fdic_fetcher = FDICDataFetcher(config)
        self.fred_fetcher = FREDDataFetcher(config)
        self.processor = BankMetricsProcessor(config)
        self.analyzer = PeerAnalyzer(config)
        self.macro_analyzer = MacroTrendAnalyzer(config)
        self.output_gen = ExcelOutputGenerator(config)
        Path(config.output_dir).mkdir(exist_ok=True)
    def _analyze_fdic_data_availability(self, fdic_df: pd.DataFrame) -> Dict[str, Any]:
        """
        Analyzes FDIC data availability and returns a report of missing/empty series.
        """
        # Exclude calculated series from analysis
        CALCULATED_FDIC_SERIES = ['Total_Capital', 'AOBS', 'LNOTHPCS', 'LNOTHNONDEP', 'LNCONAUTO', 'LNCONCC', 'LNCONOTHX']
        fields_to_analyze = [f for f in FDIC_FIELDS_TO_FETCH if f not in CALCULATED_FDIC_SERIES]

        if fdic_df.empty:
            return {"empty_series": fields_to_analyze, "sparse_series": [], "available_series": []}

        analysis_report = {
            "empty_series": [],      # Completely missing or all NaN
            "sparse_series": [],     # Present but >75% missing data
            "available_series": []   # Present with reasonable data coverage
        }

        total_records = len(fdic_df)

        for field in fields_to_analyze:
            if field not in fdic_df.columns:
                analysis_report["empty_series"].append(field)
            else:
                non_null_count = fdic_df[field].notna().sum()
                coverage_pct = (non_null_count / total_records) * 100

                if coverage_pct == 0:
                    analysis_report["empty_series"].append(field)
                elif coverage_pct < 25:
                    analysis_report["sparse_series"].append({
                        "field": field,
                        "coverage_pct": round(coverage_pct, 1),
                        "description": _get_metric_short_name(field)
                    })
                else:
                    analysis_report["available_series"].append({
                        "field": field,
                        "coverage_pct": round(coverage_pct, 1)
                    })

        return analysis_report
    def _optimize_df_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Reduces DataFrame memory usage and file size by downcasting numeric types
        and converting object columns with low cardinality to 'category'.
        """
        for col in df.columns:
            if df[col].dtype == 'float64':
                df[col] = pd.to_numeric(df[col], downcast='float')
            elif df[col].dtype == 'int64':
                df[col] = pd.to_numeric(df[col], downcast='integer')
            elif df[col].dtype == 'object':
                if len(df[col].unique()) / len(df[col]) < 0.5:
                    df[col] = df[col].astype('category')
        return df
    # Rate-like column name patterns (weighted average when method="weighted")
    _RATE_PATTERNS = re.compile(
        r"(?i)(_Rate$|_Ratio$|_Coverage$|_Pct$|_Share$|_Composition$|_Yield|"
        r"_Adj_|ROA|ROE|NIM|Efficiency|Leverage|Liquidity|HQLA|Cash_to|"
        r"Securities_to|Loans_to|Equity_to|Norm_NCO|Norm_Nonaccrual|"
        r"Norm_Delinquency|Norm_ACL|Norm_Loan|Norm_Provision|Norm_Loss|"
        r"Norm_Risk_Adj)"
    )

    def _resolve_composite_method(self, df: pd.DataFrame) -> str:
        """Auto-detect composite math: compare existing composites to mean vs weighted."""
        global COMPOSITE_METHOD
        if COMPOSITE_METHOD in ("mean", "weighted"):
            return COMPOSITE_METHOD

        # Try to detect from an existing composite (90001..90006)
        existing_certs = set(df["CERT"].unique()) & {90001, 90002, 90003, 90004, 90005, 90006}
        if not existing_certs:
            COMPOSITE_METHOD = "mean"
            logging.info("No existing composites found; defaulting COMPOSITE_METHOD='mean'")
            return COMPOSITE_METHOD

        # Simple detection: the current code uses .mean(), so default to mean
        COMPOSITE_METHOD = "mean"
        logging.info(f"COMPOSITE_METHOD resolved to '{COMPOSITE_METHOD}'")
        return COMPOSITE_METHOD

    def _compute_group_avg(self, peer_subset: pd.DataFrame, numeric_cols, method: str) -> pd.DataFrame:
        """Compute composite averages for a peer group per REPDTE."""
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", FutureWarning)
            if method == "weighted":
                # Weighted average: use LNLS (gross loans) as weight, fallback to ASSET
                weight_col = "LNLS" if "LNLS" in peer_subset.columns else (
                    "ASSET" if "ASSET" in peer_subset.columns else None
                )
                if weight_col is None:
                    return peer_subset.groupby("REPDTE")[numeric_cols].mean().reset_index()

                result_frames = []
                rate_cols = [c for c in numeric_cols if self._RATE_PATTERNS.search(c)]
                level_cols = [c for c in numeric_cols if c not in rate_cols]

                for repdte, grp in peer_subset.groupby("REPDTE"):
                    row = {"REPDTE": repdte}
                    w = grp[weight_col].fillna(0)
                    w_sum = w.sum()
                    # Level metrics: simple mean
                    for c in level_cols:
                        row[c] = grp[c].mean()
                    # Rate metrics: weighted average
                    for c in rate_cols:
                        if w_sum > 0:
                            row[c] = (grp[c].fillna(0) * w).sum() / w_sum
                        else:
                            row[c] = grp[c].mean()
                    result_frames.append(row)
                return pd.DataFrame(result_frames)
            else:
                return peer_subset.groupby("REPDTE")[numeric_cols].mean().reset_index()

    def _create_peer_composite(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Creates composite 'dummy' banks for each defined Peer Group (90001-90006)
        and the MSPBNA+MSBNA Combined entity.
        """
        logging.info("Generating historical composites for defined Peer Groups...")

        method = self._resolve_composite_method(df)
        composites = []
        base_dummy_cert = 90000
        numeric_cols = df.select_dtypes(include=np.number).columns.drop("CERT", errors="ignore")

        for group_key, group_info in PEER_GROUPS.items():
            group_certs = group_info["certs"]
            peer_subset = df[df["CERT"].isin(group_certs)]
            n_found = peer_subset["CERT"].nunique() if not peer_subset.empty else 0

            if peer_subset.empty:
                logging.warning(
                    f"No data for {group_info['name']}; skipping composite."
                )
                continue

            if n_found < MIN_PEER_MEMBERS:
                logging.warning(
                    f"{group_info['name']}: only {n_found} member(s) found "
                    f"(need {MIN_PEER_MEMBERS}). Missing CERTs: "
                    f"{set(group_certs) - set(peer_subset['CERT'].unique())}. "
                    f"Creating composite anyway."
                )

            group_avg = self._compute_group_avg(peer_subset, numeric_cols, method)

            dummy_cert = base_dummy_cert + group_info["display_order"]
            group_avg["CERT"] = dummy_cert
            group_avg["NAME"] = f"AVG: {group_info['name']}"
            group_avg["HQ_STATE"] = "AVG"

            composites.append(group_avg)
            logging.info(f"Created composite for {group_info['name']} (CERT {dummy_cert})")

        # --- MSPBNA+MSBNA Combined entity ---
        ms_pair = df[df["CERT"].isin([MSPBNA_CERT, MSBNA_CERT])]
        if not ms_pair.empty and ms_pair["CERT"].nunique() == 2:
            logging.info("Creating MSPBNA+MSBNA Combined entity...")
            combined_rows = []
            weight_col = "LNLS" if "LNLS" in ms_pair.columns else (
                "ASSET" if "ASSET" in ms_pair.columns else None
            )
            rate_cols = [c for c in numeric_cols if self._RATE_PATTERNS.search(c)]
            level_cols = [c for c in numeric_cols if c not in rate_cols]

            for repdte, grp in ms_pair.groupby("REPDTE"):
                if grp["CERT"].nunique() < 2:
                    continue
                row = {"REPDTE": repdte}
                # Level fields: sum
                for c in level_cols:
                    row[c] = grp[c].sum()
                # Rate fields: weighted average (or simple mean if no weights)
                w = grp[weight_col].fillna(0) if weight_col else pd.Series([1]*len(grp))
                w_sum = w.sum()
                for c in rate_cols:
                    if w_sum > 0:
                        row[c] = (grp[c].fillna(0) * w).sum() / w_sum
                    else:
                        row[c] = grp[c].mean()
                combined_rows.append(row)

            if combined_rows:
                combined_df = pd.DataFrame(combined_rows)
                combined_df["CERT"] = MS_COMBINED_CERT
                combined_df["NAME"] = "MSPBNA+MSBNA Combined"
                combined_df["HQ_STATE"] = "NY"
                composites.append(combined_df)
                logging.info(f"Created MSPBNA+MSBNA Combined entity (CERT {MS_COMBINED_CERT})")
        else:
            logging.warning(
                f"Cannot create combined entity: need both MSPBNA ({MSPBNA_CERT}) "
                f"and MSBNA ({MSBNA_CERT}) in data."
            )

        if composites:
            df = pd.concat([df] + composites, ignore_index=True)

        return df

    def run(self) -> Dict[str, Any]:
        logging.info("Starting dashboard generation...")
        fdic_df, _ = self.fdic_fetcher.fetch_all_banks()
        if fdic_df.empty: raise ValueError("No FDIC data retrieved.")

        all_certs = fdic_df['CERT'].unique().tolist()
        locations_df = get_bank_locations(all_certs)
        if 'NAME' in locations_df.columns:
            locations_df.drop(columns=['NAME'], inplace=True)
        fdic_df = pd.merge(fdic_df, locations_df, on='CERT', how='left')
        logging.info("Location data merged successfully.")

        fdic_analysis = self._analyze_fdic_data_availability(fdic_df)

        proc_df = self.processor.create_derived_metrics(fdic_df)
        proc_df_with_ttm = self.processor.calculate_ttm_metrics(proc_df)
        proc_df_with_peers = self._create_peer_composite(proc_df_with_ttm)

        peer_comp_df = self.analyzer.create_peer_comparison(proc_df_with_peers)
        norm_comp_df = self.analyzer.create_normalized_comparison(proc_df_with_peers)  # NEW: Normalized view
        snapshot_df = self.processor.create_latest_snapshot(proc_df_with_peers)
        avg_8q_all_metrics_df = self.processor.calculate_8q_averages(proc_df_with_peers)

        # === FRED DATA FETCHING - NO DUPLICATES ===
        logging.info("Fetching FRED macroeconomic data asynchronously...")
        start_time = time.perf_counter()

        # Define calculated series that should NOT be fetched
        CALCULATED_SERIES = ['SOFR3MTB3M']

        # Build list - EXCLUDING calculated series
        # Build list - EXCLUDING calculated series
        series_ids_to_fetch = []
        for category_dict in FRED_SERIES_TO_FETCH.values():
            for sid in category_dict.keys():
                if sid not in CALCULATED_SERIES:
                    series_ids_to_fetch.append(sid)

        logger.info(f"Fetching {len(series_ids_to_fetch)} series (excluding {CALCULATED_SERIES})")

        # Build descriptions DataFrame for ALL series (including calculated)
        # Use a completely unique variable name to avoid any collisions
        _temp_series_metadata_dict = {}
        for category, series_dict in FRED_SERIES_TO_FETCH.items():
            # Validate that series_dict is actually a dict of series
            if not isinstance(series_dict, dict):
                logger.error(f"Category '{category}' has invalid structure: {type(series_dict)}")
                continue

            for series_id, metadata in series_dict.items():
                # Skip if this looks like a metadata key, not a series ID
                if series_id in ['short', 'long', 'annualized', 'best_technical', 'technical_reason']:
                    logger.warning(f"Skipping metadata key '{series_id}' in category '{category}'")
                    continue

                if not isinstance(metadata, dict):
                    logger.error(f"Invalid metadata for {series_id} in category '{category}': {type(metadata)}")
                    continue

                _temp_series_metadata_dict[series_id] = {
                    'Series ID': series_id,
                    'Category': category,
                    **metadata
                }

        series_descriptions_df = pd.DataFrame.from_dict(_temp_series_metadata_dict, orient='index')
        logger.info(f"Created descriptions dataframe with {len(series_descriptions_df)} rows")

        # Handle event loop properly for Spyder/Jupyter
        try:
            loop = asyncio.get_running_loop()
            import nest_asyncio
            nest_asyncio.apply()
            fred_df, fred_desc_df, failed_fred_series, fred_metadata_df = asyncio.run(
                self.fred_fetcher.fetch_all_series_async(
                    series_ids=series_ids_to_fetch,
                    series_descriptions=series_descriptions_df
                )
            )
            # === PATCH S-RAW3a: pull stashed raw observations DF ===
            fred_obs_df = getattr(self.fred_fetcher, 'last_fred_obs_df', pd.DataFrame())
            # === END PATCH S-RAW3a ===

        except RuntimeError:
            fred_df, fred_desc_df, failed_fred_series, fred_metadata_df = asyncio.run(
                self.fred_fetcher.fetch_all_series_async(
                    series_ids=series_ids_to_fetch,
                    series_descriptions=series_descriptions_df
                )
            )
            # === PATCH S-RAW3a: pull stashed raw observations DF ===
            fred_obs_df = getattr(self.fred_fetcher, 'last_fred_obs_df', pd.DataFrame())
            # === END PATCH S-RAW3a ===

        end_time = time.perf_counter()
        logging.info(f"Asynchronous FRED data fetching completed in {end_time - start_time:.2f} seconds.")
        # Merge frequency information into descriptions
        # Merge frequency information into descriptions
        # === PATCH: enrich FRED metadata with DateBasis & QuarterOffset and apply overrides ===
        # Enrich FRED metadata with DateBasis / QuarterOffset and merge into descriptions
        # Defaults if metadata is missing
        if fred_metadata_df is None or fred_metadata_df.empty:
            fred_metadata_df = pd.DataFrame(columns=['Series ID','frequency','frequency_short','units','DateBasis','QuarterOffset'])
        else:
            fred_metadata_df = fred_metadata_df.copy()
            # Ensure the new columns exist with safe defaults
            if 'DateBasis' not in fred_metadata_df.columns:
                fred_metadata_df['DateBasis'] = 'period_start'
            else:
                fred_metadata_df['DateBasis'] = fred_metadata_df['DateBasis'].fillna('period_start')

            if 'QuarterOffset' not in fred_metadata_df.columns:
                fred_metadata_df['QuarterOffset'] = 0
            fred_metadata_df['QuarterOffset'] = pd.to_numeric(fred_metadata_df['QuarterOffset'], errors='coerce').fillna(0).astype(int)

        # Known publication-dated quarterlies that represent the PREVIOUS quarter
        _fred_overrides = {
            # Tightening CRE Lending Standards – publication month maps to prior quarter
            'DRCRELEXFACBS': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
            'DRBLACBS': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
            'DRTSCILM': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
            'DRTSCLCC': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
        }

        # Apply overrides
        for _sid, _vals in _fred_overrides.items():
            _m = fred_metadata_df['Series ID'].astype(str).eq(_sid)
            for _k, _v in _vals.items():
                fred_metadata_df.loc[_m, _k] = _v

        # Merge the enriched metadata into descriptions
        fred_desc_df = pd.merge(
            fred_desc_df,
            fred_metadata_df[['Series ID','frequency','frequency_short','units','DateBasis','QuarterOffset']],
            on='Series ID',
            how='left'
        )


        # Set index and create calculated series
        if 'date' in fred_df.columns:
            fred_df.set_index('date', inplace=True)

        if 'SOFR' in fred_df.columns and 'TB3MS' in fred_df.columns:
            fred_df['SOFR'] = fred_df['SOFR'].ffill()
            fred_df['TB3MS'] = fred_df['TB3MS'].ffill()
            fred_df['SOFR3MTB3M'] = fred_df['SOFR'] - fred_df['TB3MS']
            if fred_desc_df is not None and fred_desc_df[fred_desc_df['Series ID'] == 'SOFR3MTB3M'].empty:
                new_row_data = {'Series ID': 'SOFR3MTB3M', 'Category': 'Middle Market, Healthcare, & Funding Indicators', 'short': 'SOFR vs T-Bill Spread', 'long': 'Calculated Spread: SOFR minus 3-Month T-Bill Rate'}
                new_row = pd.DataFrame([new_row_data])
                fred_desc_df = pd.concat([fred_desc_df, new_row], ignore_index=True)

        # Process technical indicators
        enhanced_analyzer = MacroTrendAnalyzer(self.config)
        processed_data = {}
        validation_reports = {}

        for category, series_in_category in FRED_SERIES_TO_FETCH.items():
            for series_id in series_in_category.keys():
                if series_id in CALCULATED_SERIES:
                    continue
                if series_id in fred_df.columns:
                    validation_reports[series_id] = enhanced_analyzer.validate_series_data(
                        series_id, fred_df[series_id]
                    )
                    if validation_reports[series_id]['status'] != 'Error':
                        result = enhanced_analyzer.calculate_technical_indicators(
                            fred_df[series_id], series_id
                        )
                        if not result.empty:
                            processed_data[series_id] = result

        # Process calculated series
        for series_id in CALCULATED_SERIES:
            if series_id in fred_df.columns:
                validation_reports[series_id] = enhanced_analyzer.validate_series_data(
                    series_id, fred_df[series_id]
                )
                if validation_reports[series_id]['status'] != 'Error':
                    result = enhanced_analyzer.calculate_technical_indicators(
                        fred_df[series_id], series_id
                    )
                    if not result.empty:
                        processed_data[series_id] = result

        validation_df = pd.DataFrame.from_dict(validation_reports, orient='index')
        powerbi_macro_df = enhanced_analyzer.generate_powerbi_output(processed_data)


        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f"{self.config.output_dir}/Bank_Performance_Dashboard_{ts}.xlsx"
        peer_comp_df = self._optimize_df_dtypes(peer_comp_df)
        norm_comp_df = self._optimize_df_dtypes(norm_comp_df)  # NEW: Optimize normalized view
        snapshot_df = self._optimize_df_dtypes(snapshot_df)
        proc_df_with_peers = self._optimize_df_dtypes(proc_df_with_peers)
        avg_8q_all_metrics_df = self._optimize_df_dtypes(avg_8q_all_metrics_df)
        powerbi_macro_df = self._optimize_df_dtypes(powerbi_macro_df)
        # === FDIC-META: units & scaling for email tables ===
        # === FDIC-META (with Basis): tells consumer how to scale/format ===
        fdic_meta_df = pd.DataFrame([
            # Dollars in $000 → display $M (scale 1e-3)
            # --- Normalized ACL Metrics ---
            {"MetricCode":"Norm_ACL_Balance",                 "Display":"Normalized ACL Balance",       "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":2, "Basis":"level"},
            {"MetricCode":"Norm_Risk_Adj_Allowance_Coverage", "Display":"Norm: Risk-Adj Coverage (Ex-SBL)", "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent",    "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_RESI_ACL_Coverage",           "Display":"Resi Reserve % of Norm ACL",   "DisplayUnit":"%",  "Scale":1.0,  "Fmt":"percent",    "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_CRE_ACL_Coverage",            "Display":"CRE Reserve % of Norm ACL",    "DisplayUnit":"%",  "Scale":1.0,  "Fmt":"percent",    "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Comm_ACL_Coverage",           "Display":"C&I Reserve % of Norm Loans",  "DisplayUnit":"%",  "Scale":1.0,  "Fmt":"percent",    "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_CRE_Investment_Composition", "Display":"Norm: CRE Invest. % (Ex-ADC)", "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_CRE_OO_Composition",         "Display":"Norm: CRE Owner-Occ %",        "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"ASSET", "Display":"Assets",                      "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"LNLS",  "Display":"Gross Loans",                 "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            # --- NEW: Risk-Adjusted Coverage (Ratio, e.g. 1.5x) ---
            {"MetricCode":"RIC_CRE_Risk_Adj_Coverage", "Display":"CRE Risk-Adj Coverage", "DisplayUnit":"x", "Scale":1.0, "Fmt":"ratio", "Decimals":2, "Basis":"ratio"},
            {"MetricCode":"RIC_Resi_Risk_Adj_Coverage", "Display":"Resi Risk-Adj Coverage", "DisplayUnit":"x", "Scale":1.0, "Fmt":"ratio", "Decimals":2, "Basis":"ratio"},
            {"MetricCode":"RIC_Comm_Risk_Adj_Coverage", "Display":"C&I Risk-Adj Coverage", "DisplayUnit":"x", "Scale":1.0, "Fmt":"ratio", "Decimals":2, "Basis":"ratio"},

            # --- NEW: Burn Rate (Years) ---
            {"MetricCode":"RIC_CRE_Years_of_Reserves", "Display":"CRE Years of Reserves", "DisplayUnit":"Yrs", "Scale":1.0, "Fmt":"number", "Decimals":1, "Basis":"ratio"},

            # --- NEW: Growth Rates (Percent) ---
            {"MetricCode":"CRE_Growth_TTM", "Display":"CRE Growth (YoY)", "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"SBL_Growth_TTM", "Display":"SBL Growth (YoY)", "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},

            # Ratios: mark as 'percent' (already %) OR 'fraction' (0–1) depending on your upstream
            # Given your HTML shows 8.86% (too big), your upstream is ALREADY percent → use Basis='percent' and Scale=1.0
            {"MetricCode":"CI_to_Capital_Risk",             "Display":"C&I to Capital Risk (%)",             "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"CRE_Concentration_Capital_Risk", "Display":"CRE Concentration Risk (%)",          "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"IDB_CRE_Growth_TTM",             "Display":"CRE Growth TTM (%)",                  "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"IDB_CRE_Growth_36M",             "Display":"CRE Growth 36M (%)",                  "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"TTM_NCO_Rate",                  "Display":"TTM NCO Rate (%)",                    "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"NPL_to_Gross_Loans_Rate",       "Display":"NPL to Gross Loans (%)",              "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Allowance_to_Gross_Loans_Rate", "Display":"Allowance to Gross Loans (%)",        "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"TTM_Past_Due_Rate",             "Display":"TTM Past Due Rate (%)",               "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"TTM_PD30_Rate",                 "Display":"TTM Past Due (30-90 Days) Rate (%)",  "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"TTM_PD90_Rate",                 "Display":"TTM Past Due (90+ Days) Rate (%)",    "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},

            # --- NORMALIZED METRICS (Ex-Commercial/Ex-Consumer) ---
            {"MetricCode":"Norm_NCO_Rate",                 "Display":"Normalized NCO Rate (%)",             "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Nonaccrual_Rate",          "Display":"Normalized Nonaccrual Rate (%)",      "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Delinquency_Rate",         "Display":"Normalized Delinquency Rate (%)",     "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_ACL_Coverage",             "Display":"Normalized ACL Coverage (%)",         "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Gross_Loans",              "Display":"Normalized Gross Loans",              "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Norm_Total_NCO",                "Display":"Normalized Total NCOs (TTM)",         "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":2, "Basis":"level"},
            {"MetricCode":"Norm_Total_Nonaccrual",         "Display":"Normalized Total Nonaccruals",        "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Excluded_Balance",              "Display":"Total Excluded Balance",              "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Norm_Exclusion_Pct",            "Display":"Exclusion % of Gross Loans",          "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_SBL_Composition",          "Display":"SBL % of Norm Loans",                 "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Fund_Finance_Composition", "Display":"Fund Finance % of Norm Loans",        "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Wealth_Resi_Composition",  "Display":"Wealth Resi % of Norm Loans",         "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_CRE_OO_Composition",         "Display":"CRE Owner-Occ % of Norm Loans",       "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_CRE_Investment_Composition", "Display":"CRE Invest. % of Norm Loans",         "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Loan_Yield",               "Display":"Normalized Loan Yield (%)",           "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Provision_Rate",           "Display":"Normalized Provision Rate (%)",       "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Loss_Adj_Yield",           "Display":"Normalized Loss-Adj Yield (%)",       "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Norm_Risk_Adj_Return",          "Display":"Normalized Risk-Adj Return (%)",      "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},

            # --- RESTORED LIQUIDITY METRICS ---
            {"MetricCode":"Liquid_Assets",                 "Display":"Liquid Assets",                       "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Liquidity_Ratio",               "Display":"Liquidity Ratio (%)",                 "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"HQLA",                          "Display":"High Quality Liquid Assets",          "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"HQLA_Ratio",                    "Display":"HQLA Ratio (%)",                      "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Cash_to_Assets",                "Display":"Cash to Assets (%)",                  "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Securities_to_Assets",          "Display":"Securities to Assets (%)",            "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Loans_to_Deposits",             "Display":"Loans to Deposits (%)",               "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Cash_and_Balances",             "Display":"Cash & Due From Banks",               "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Securities_HTM",                "Display":"HTM Securities",                      "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Securities_AFS",                "Display":"AFS Securities",                      "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Unused_Commitments",            "Display":"Unused Commitments",                  "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Unused_Commitment_Ratio",       "Display":"Unused Commitment Ratio (%)",         "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},

            # --- RESTORED CAPITAL METRICS ---
            {"MetricCode":"Total_Equity_Raw",              "Display":"Total Equity Capital",                "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"Equity_to_Assets",              "Display":"Equity to Assets (%)",                "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Leverage_Ratio",                "Display":"Leverage Ratio (Assets/Equity)",      "DisplayUnit":"x", "Scale":1.0, "Fmt":"number", "Decimals":1, "Basis":"level"},

            # --- RESTORED RAW PROFITABILITY METRICS ---
            {"MetricCode":"Net_Income_Raw",                "Display":"Net Income (Raw)",                    "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"ROA_Raw",                       "Display":"ROA (Raw) (%)",                       "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"ROE_Raw",                       "Display":"ROE (Raw) (%)",                       "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"NIM_Raw",                       "Display":"Net Interest Margin (Raw) (%)",       "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Efficiency_Ratio_Raw",          "Display":"Efficiency Ratio (Raw) (%)",          "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Yield_on_Loans_Raw",            "Display":"Yield on Loans (Raw) (%)",            "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Cost_of_Deposits_Raw",          "Display":"Cost of Deposits (Raw) (%)",          "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},

        ])
        # === END FDIC-META ===


        self.output_gen.write_excel_output(
            file_path=fname,
            Summary_Dashboard=peer_comp_df,
            Normalized_Comparison=norm_comp_df,  # NEW: Ex-Commercial/Ex-Consumer view
            Latest_Peer_Snapshot=snapshot_df,
            Averages_8Q_All_Metrics=avg_8q_all_metrics_df,
            FDIC_Metadata=fdic_meta_df,
            Macro_Analysis=powerbi_macro_df,
            FDIC_Data=proc_df_with_peers,
            FRED_Data=fred_df.reset_index(),
            FRED_Metadata=fred_metadata_df,
            FRED_Descriptions=fred_desc_df,
            Data_Validation_Report=validation_df,
        )


        # === ENHANCED DIAGNOSTIC: Inspect Critical FDIC Data for Subject Bank ===
        print("\n" + "="*80)
        print("CRITICAL FDIC DATA INSPECTION FOR YOUR BANK")
        print("="*80)

        try:
            # Get the latest data for your bank
            subject_bank_data = proc_df_with_peers[
                proc_df_with_peers['CERT'] == self.config.subject_bank_cert
            ].sort_values('REPDTE')

            if not subject_bank_data.empty:
                latest_record = subject_bank_data.iloc[-1]
                bank_name = latest_record.get('NAME', 'Unknown Bank')
                report_date = latest_record.get('REPDTE', 'Unknown Date')

                print(f"Bank: {bank_name} (CERT: {self.config.subject_bank_cert})")
                print(f"Latest Report Date: {report_date}")
                print("-" * 80)

                # Critical columns to inspect
                critical_columns = {
                    'ASSET': 'Total Assets',
                    'EQ': 'Total Equity Capital',
                    'LNLS': 'Gross Loans & Leases',
                    'LNCI': 'C&I Loans',
                    'LNRENROW': 'Owner-Occ CRE',
                    'LNRECONS': 'Construction & Development',
                    'LNRERES': '1-4 Family Residential',
                    'LNRENROT': 'Nonfarm CRE '
                }

                print("CRITICAL BALANCE SHEET ITEMS:")
                for col_code, description in critical_columns.items():
                    value = latest_record.get(col_code)
                    if pd.isna(value) or value is None:
                        print(f"  {col_code:12} ({description:30}): *** MISSING DATA ***")
                    elif value == 0:
                        print(f"  {col_code:12} ({description:30}): *** ZERO VALUE ***")
                    else:
                        # Format as currency in thousands
                        formatted_value = f"${value:,.0f}"
                        print(f"  {col_code:12} ({description:30}): {formatted_value}")

                # === NEW: Investigate C&I Loans specifically ===
                print("\n" + "="*60)
                print("C&I LOANS INVESTIGATION")
                print("="*60)

                # Check if LNCI exists in the raw data at all
                if 'LNCI' not in subject_bank_data.columns:
                    print("❌ LNCI column does not exist in the fetched data")
                    print("   This suggests FDIC API did not return this field")
                else:
                    print("✓ LNCI column exists in the data")

                    # Check LNCI values across all quarters
                    lnci_summary = subject_bank_data['LNCI'].describe()
                    print(f"   LNCI Data Summary across all quarters:")
                    print(f"     Count (non-null): {subject_bank_data['LNCI'].notna().sum()}")
                    print(f"     Count (null):     {subject_bank_data['LNCI'].isna().sum()}")
                    print(f"     Count (zero):     {(subject_bank_data['LNCI'] == 0).sum()}")
                    if subject_bank_data['LNCI'].notna().sum() > 0:
                        print(f"     Min value:        ${subject_bank_data['LNCI'].min():,.0f}")
                        print(f"     Max value:        ${subject_bank_data['LNCI'].max():,.0f}")
                        print(f"     Latest non-null:  ${subject_bank_data['LNCI'].dropna().iloc[-1] if len(subject_bank_data['LNCI'].dropna()) > 0 else 0:,.0f}")

                # Check alternative C&I loan fields that might be available
                print("\n   Checking for alternative C&I loan series:")
                ci_alternatives = {
                    'LNCI': 'Commercial and Industrial Loans',
                    'LNCIDOM': 'C&I Loans - Domestic',
                    'LNCINUM': 'C&I Loans - Number of Loans',
                    'LNCON': 'Consumer Loans (sometimes includes C&I)',
                    'BUSLOANS': 'Total Business Loans (broader category)'
                }

                for field, description in ci_alternatives.items():
                    if field in subject_bank_data.columns:
                        latest_val = latest_record.get(field)
                        if pd.notna(latest_val) and latest_val != 0:
                            print(f"   ✓ {field:10}: {description:35} = ${latest_val:,.0f}")
                        else:
                            print(f"   ⚠ {field:10}: {description:35} = Missing/Zero")
                    else:
                        print(f"   ❌ {field:10}: {description:35} = Not available")

                # Check what fields ARE available for this bank
                print(f"\n   Available loan fields for your bank:")
                loan_fields = [col for col in subject_bank_data.columns if col.startswith('LN') and latest_record.get(col, 0) > 0]
                for field in sorted(loan_fields):
                    value = latest_record.get(field, 0)
                    print(f"     {field}: ${value:,.0f}")

                # Check if this is a reporting threshold issue
                total_assets = latest_record.get('ASSET', 0)
                print(f"\n   Bank Size Analysis:")
                print(f"     Total Assets: ${total_assets:,.0f}")
                if total_assets < 300_000:  # Less than $300M
                    print("     ⚠ Small bank - may have different reporting requirements")
                elif total_assets < 1_000_000:  # Less than $1B
                    print("     ℹ Community bank - standard reporting applies")
                else:
                    print("     ℹ Large bank - full reporting requirements")

                # Calculate what C&I might be if it's combined with other categories
                total_loans = latest_record.get('LNLS', 0)
                known_loans = sum([
                    latest_record.get('LNRENROW', 0),  # CRE Non-Owner
                    latest_record.get('LNRECONS', 0),  # Construction
                    latest_record.get('LNRERES', 0),   # Residential
                    latest_record.get('LNRENROT', 0),  # CRE Owner-Occ
                    latest_record.get('LNRELOC', 0),   # HELOC
                    latest_record.get('LNCRCD', 0),    # Credit Cards
                    latest_record.get('LNAUTO', 0),    # Auto
                    latest_record.get('LNOTHER', 0),   # Other
                    latest_record.get('LS', 0),        # Leases
                ])

                implied_ci = total_loans - known_loans
                print(f"\n   Implied C&I Calculation:")
                print(f"     Total Loans:        ${total_loans:,.0f}")
                print(f"     Known Categories:   ${known_loans:,.0f}")
                print(f"     Implied C&I:        ${implied_ci:,.0f}")
                print(f"     Percentage of Total: {(implied_ci/total_loans)*100:.1f}%" if total_loans > 0 else "     Percentage: N/A")

                # Rest of your existing diagnostic code...
                print("\n" + "="*60)
                print("OTHER CALCULATED FIELDS")
                print("="*60)
                calc_fields = {
                    'Total_Capital': 'Total Capital (T1 + T2)',
                    'AOBS': 'Off-Balance Sheet Allowance',
                    'Total_ACL': 'Total Allowance for Credit Losses'
                }

                for col_code, description in calc_fields.items():
                    value = latest_record.get(col_code)
                    if pd.isna(value) or value is None:
                        print(f"  {col_code:15} ({description:30}): *** CALCULATION FAILED ***")
                    else:
                        formatted_value = f"${value:,.0f}"
                        print(f"  {col_code:15} ({description:30}): {formatted_value}")

            else:
                print(f"❌ NO DATA FOUND for your bank (CERT: {self.config.subject_bank_cert})")

        except Exception as e:
            print(f"❌ ERROR during diagnostic inspection: {e}")

        print("="*80)
        return {
            "output_file": fname,
            "powerbi_macro_df": powerbi_macro_df,
            "fred_desc_df": fred_desc_df,
            "failed_fred_series": failed_fred_series,
            "fdic_analysis": fdic_analysis,
        }



def load_config() -> DashboardConfig:
    """
    Loads configuration from environment variables with fallback to defaults.

    Create a .env file with:
        FRED_API_KEY=your_key_here
        SUBJECT_BANK_CERT=34221
        PEER_CERTS=26876,9396,18221,...

    Returns:
        DashboardConfig: Configuration object

    Raises:
        ValueError: If FRED_API_KEY is missing
    """
    import os
    from pathlib import Path

    # Try to load .env file (install python-dotenv: pip install python-dotenv)
    try:
        from dotenv import load_dotenv
        env_path = Path(__file__).parent / '.env'
        if env_path.exists():
            load_dotenv(env_path)
            logger.info(f"Loaded configuration from {env_path}")
        else:
            logger.warning(f".env file not found at {env_path}, using environment variables")
    except ImportError:
        logger.warning("python-dotenv not installed, reading from environment variables only")

    # Required
    fred_api_key = os.getenv('FRED_API_KEY')
    if not fred_api_key:
        raise ValueError(
            "FRED_API_KEY not found. Set it in .env file or environment:\n"
            "  export FRED_API_KEY='your_key_here'"
        )


    subject_cert = int(os.getenv('SUBJECT_BANK_CERT', str(MSPBNA_CERT)))

    # UPDATED: Derive unique peer certs from the PEER_GROUPS dictionary
    # This ignores the .env PEER_CERTS list to ensure consistency with your logic
    peer_certs = get_all_peer_certs()

    # Remove subject bank from peer list if present to avoid fetching twice
    if subject_cert in peer_certs:
        peer_certs.remove(subject_cert)

    logger.info(f"Configuration loaded: Subject CERT={subject_cert}, {len(peer_certs)} unique peers from {len(PEER_GROUPS)} groups.")

    return DashboardConfig(
        fred_api_key=fred_api_key,
        subject_bank_cert=subject_cert,
        peer_bank_certs=peer_certs
    )


def main():
    run_results = {}
    try:
        config = load_config()
        dash = BankPerformanceDashboard(config)
        run_results = dash.run()
        print("\n" + "="*80)
        print("DASHBOARD GENERATION COMPLETE")
        print(f"Output file located at: {run_results.get('output_file', 'N/A')}")
        print("="*80)

        print("\n--- FDIC DATA AVAILABILITY REPORT ---")
        fdic_analysis = run_results.get('fdic_analysis', {})

        empty_series = fdic_analysis.get('empty_series', [])
        sparse_series = fdic_analysis.get('sparse_series', [])
        available_series = fdic_analysis.get('available_series', [])

        if empty_series:
            print(f"\n❌ EMPTY FDIC SERIES ({len(empty_series)} series with no data):")
            for series in empty_series:
                description = _get_metric_short_name(series)
                print(f"  - {series}: {description}")
        else:
            print("\n✅ All FDIC series returned data.")

        if sparse_series:
            print(f"\n⚠️  SPARSE FDIC SERIES ({len(sparse_series)} series with limited data):")
            for item in sparse_series:
                print(f"  - {item['field']}: {item['description']} ({item['coverage_pct']}% coverage)")

        print(f"\n✅ AVAILABLE FDIC SERIES: {len(available_series)} series with good data coverage")

        total_requested = len(FDIC_FIELDS_TO_FETCH)
        total_available = len(available_series)
        total_sparse = len(sparse_series)
        total_empty = len(empty_series)

        print(f"\nSUMMARY:")
        print(f"  Total FDIC series requested: {total_requested}")
        print(f"  Available (good coverage):   {total_available} ({(total_available/total_requested)*100:.1f}%)")
        print(f"  Sparse (limited coverage):   {total_sparse} ({(total_sparse/total_requested)*100:.1f}%)")
        print(f"  Empty (no data):             {total_empty} ({(total_empty/total_requested)*100:.1f}%)")

        print("\n--- POWER BI INTEGRATION ---")
        try:
            macro_df = run_results.get('powerbi_macro_df')
            desc_df = run_results.get('fred_desc_df')
            if macro_df is not None and not macro_df.empty and desc_df is not None:
                powerbi_generator = PowerBIGenerator()
                powerbi_generator.generate_powerbi_setup_file(
                    output_dir=config.output_dir,
                    powerbi_macro_df=macro_df,
                    fred_desc_df=desc_df
                )
                powerbi_setup_file = Path(config.output_dir) / "PowerBI_Setup.txt"
                print(f"[+] Power BI setup file generated successfully.")
                print(f"    -> Location: {powerbi_setup_file.resolve()}")
            else:
                print("[-] Power BI setup file was not generated because macro data was missing.")

        except Exception as e:
            logger.error(f"Failed to generate Power BI setup file: {e}", exc_info=True)
            print("[-] Power BI setup file was not generated due to an unexpected error.")

    except Exception as e:
        logger.error(f"Application failed to complete: {e}", exc_info=True)
        sys.exit(1)

    finally:
        print("\n--- DATA SERIES VERIFICATION ---")
        failed_series = run_results.get("failed_fred_series", [])
        if failed_series:
            print("\nWARNING: The following FRED series failed to fetch and are NOT in the output:")
            for series in failed_series:
                print(f"  - {series}")
        else:
            print("\nAll FRED series fetched successfully.")

        logging.shutdown()

if __name__ == '__main__':
    main()