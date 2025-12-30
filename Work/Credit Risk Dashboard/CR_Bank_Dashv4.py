import os
import sys
import logging
import time
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass
from pathlib import Path
import warnings

import requests
import pandas as pd
import numpy as np
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)
# ==================================================================================
#  1. SCRIPT CONFIGURATION & SETUP
# ==================================================================================

def setup_logging(log_dir: str = "logs") -> logging.Logger:
    """Setup comprehensive logging configuration."""
    Path(log_dir).mkdir(exist_ok=True)
    log_filename = f"{log_dir}/bank_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    # Remove existing handlers to avoid duplicate logs in interactive environments
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[logging.FileHandler(log_filename), logging.StreamHandler(sys.stdout)]
    )
    logger = logging.getLogger(__name__)
    logger.info(f"Logging initialized. Log file: {log_filename}")
    return logger

logger = setup_logging()

@dataclass
class DashboardConfig:
    """Clean configuration class to hold runtime parameters."""
    fred_api_key: str
    subject_bank_cert: int
    peer_bank_certs: List[int]
    quarters_back: int = 30
    fred_years_back: int = 15
    output_dir: str = "output"
    fdic_api_base: str = "https://banks.data.fdic.gov/api"
    fred_api_base: str = "https://api.stlouisfed.org/fred"

# ==================================================================================
#  2. GLOBAL DATA DICTIONARIES & CONSTANTS
# ==================================================================================
FDIC_FIELDS_TO_FETCH = [
    "CERT", "NAME", "REPDTE", "ASSET", "DEP", "LIAB", "EQ", "LNLSNET", "ROA", "ROE", "NIMY",
    "EEFFR", "NONIIAY", "ELNATRY", "RBCT1CER", "RBC1RWAJ", "RBCRWAJ", "RBC1AAJ", "LNATRES",
    "NCLNLS", "NTLNLS", "P3LNLS", "P9LNLS", "LNLS", "EINTEXP", "FREPP", "OTHBOR",
    "RBCT2", # Tier 2 Capital (includes allowances eligible for inclusion)
    "EQ", # Total Bank Equity Capital
    "RBC1RWAJ", # Tier 1 Capital Ratio
    "EQCS", # Common Stock
    "EQSUR", # EQSUR (exclude all EQSUR related to preferred stock)
    "EQUP", # Retained Earnings
    "EQCCOMPI", # Accumulated Other Comprehensive Income
    "EQO", # Other Equity Capital Components
    "EQPP", # Perpetual Preferred Stock and Related EQSUR
    "RBCT1J", # Tier One (Core) Capital
    "RBCRWAJ", # Total Risk-Based Capital Ratio,
    "RB2LNRES",#Loan Loss Allowance Included in Tier 2 Capital,
    "RWA",#Total Risk Weighted Assets,
    "MUTUAL",

    # Loan Categories
    "LNCI", "LNRENROW", "LNRECONS", "LNREMULT", "LNRENROT", "LNREAG", "LNRERES", "LNRELOC",
    "LNCON", "LNCRCD", "LNAUTO", "LNOTHER", "LS", "LNAG",
    # Add series for Non-Owner-Occupied (Income-Producing) Nonfarm Nonresidential - Note: May be sparse.
    "LNREOTH", "NTRENRES", "P9RENRES", "NARENRES", "P3LRENRS",
    # NCOs
    "NTCI", "NTRECONS", "NTREMULT", "NTRENROT", "NTREAG", "NTRERES", "NTRELOC",
    "NTCON", "NTCRCD", "NTAUTO", "NTLS", "NTOTHER", "NTAG",
    # Past Due 30-89
    "P3CI", "P3RENROW", "P3RECONS", "P3LREMUL", "P3RENROT", "P3REAG", "P3RERES",
    "P3RELOC", "P3CON", "P3CRCD", "P3AUTO", "P3OTHLN", "P3LS", "P3AG",
    # Past Due 90+
    "P9CI", "P9RENROW", "P9RECONS", "P9REMULT", "P9RENROT", "P9REAG", "P9RERES",
    "P9RELOC", "P9CON", "P9CRCD", "P9AUTO", "P9OTHLN", "P9LS", "P9AG",
    # Nonaccrual
    "NACI", "NARENROW", "NARECONS", "NAREMULT", "NARENROT", "NAREAG", "NARERES",
    "NARELOC", "NACON", "NACRCD", "NAAUTO", "NAOTHLN", "NALS", "NAAG","NALRERES"
]

FDIC_FIELD_DESCRIPTIONS = {
    # Key Balance Sheet & Income Statement
    "ASSET":    {"short": "Total Assets", "long": "Total assets held by the institution."},
    "DEP":      {"short": "Total Deposits", "long": "Total deposits held by the institution."},
    "LIAB":     {"short": "Total Liabilities", "long": "Total liabilities of the institution."},
    "EQ":       {"short": "Total Bank Equity Capital", "long": "Total equity capital of the institution."},
    "MUTUAL":       {"short": "Ownership Type", "long": "A non-stock institution, or mutual institution, is owned and controlled solely by its depositors. A mutual does not issue capital stock. (1=Yes, a mutual) An institution which sells stock to raise capital is called a stock institution. It is owned by the shareholders who benefit from profits earned by the institution. (0=No, not a mutual.)"},
    "RWA":       {"short": "Total Risk Weighted Assets", "long": "Total risk-weighted assets calculated under the standardized approach. (Schedule RC-R, Part II, item 31)"},
    "LNLS":     {"short": "Gross Loans & Leases", "long": "Total loans and leases, gross, before deducting the allowance for loan and lease losses."},
    "LNATRES":  {"short": "Loan Loss Allowance (no OBS)", "long": "Allowance for Loan and Lease Losses without OBS."},
    "AOBS": {"short": "Allowance for Off-Balance Sheet Exposures", "long": "Allowance for Credit Losses on Off-Balance Sheet Credit Exposures."},
    "NONIIAY":  {"short": "Noninterest Income (YTD)", "long": "Total noninterest income, year-to-date."},
    "ELNATRY":  {"short": "Noninterest Expense (YTD)", "long": "Total noninterest expense, year-to-date."},
    "EINTEXP":  {"short": "Interest Expense (YTD)", "long": "Total interest expense, year-to-date."},
    "Total_ACL":{"short": "Total Allowance for Credit Losses", "long": "Sum of the allowance for on-balance-sheet loans and off-balance-sheet credit exposures."},
    "LNREOTH":  {"short": "Income-Prod. CRE Bal", "long": "Loans Secured by Other Nonfarm Nonresidential Properties (Income-Producing)."},


    # Key Ratios & Capital
    "ROA":      {"short": "ROA", "long": "Return on Assets, annualized."},
    "ROE":      {"short": "ROE", "long": "Return on Equity, annualized."},
    "NIMY":     {"short": "Net Interest Margin", "long": "Net Interest Margin, annualized, year-to-date."},
    "EEFFR":    {"short": "Efficiency Ratio", "long": "Efficiency Ratio."},
    "RBCT1CER": {"short": "Common Equity Tier 1 Capital Ratio", "long": "Common Equity Tier 1 Capital Ratio."},
    # Capital
    "RBCT2": {"short": "Tier 2 Capital", "long": "Tier 2 risk-based capital including qualifying allowances for credit losses."},
    "RB2LNRES": {"short": "Total Loan Loss Allowance", "long": "( YTD, $ ) Loan Loss Allowance Included in Tier 2 Capital"},
    "RBCRWAJ": {"short": "Total Risk-Based Capital Ratio", "long": "Total risk-based capital (Tier 1 + Tier 2)."},
    "RBC1RWAJ": {"short": "Tier 1 Capital Ratio", "long": "Tier 1 risk-based capital."},
    "RBCT1J": {"short": "Tier 1 Capital", "long": "Tier 1 risk-based capital."},
    "EQCS": {"short": "Common Stock", "long": "Common stock par value and paid-in capital."},
    "EQSUR": {"short": "EQSUR", "long": "EQSUR (exclude all EQSUR related to preferred stock)."},
    "EQUP": {"short": "Retained Earnings", "long": "Retained earnings."},
    "LEVRATIO": {"short": "Tier 1 Leverage Ratio", "long": "Tier 1 capital as a percentage of average total assets. (Schedule RC-R, Part I, item 31)"},
    "CET1R": {"short": "Common Equity Tier 1 Capital Ratio", "long": "CET1 capital as a percentage of risk-weighted assets. (Schedule RC-R, Part I, item 28)"},
    "EQCCOMPI": {"short": "Accumulated Other Comprehensive Income", "long": "Accumulated other comprehensive income (can be negative)."},
    "EQO": {"short": "Other Equity Capital Components", "long": "Other equity capital components."},
    "EQPP": {"short": "Perpetual Preferred Stock", "long": "Perpetual preferred stock and related EQSUR."},


    "NTLNLS":   {"short": "Total Net Charge-Offs (YTD)", "long": "Total net charge-offs on loans and leases, year-to-date."},
    "P3LNLS":   {"short": "Total Loans PD 30-89 Days", "long": "Total loans and leases 30-89 days past due."},
    "P9LNLS":   {"short": "Total Loans PD 90+ Days", "long": "Total loans and leases 90 or more days past due."},
    # Additional derived metrics
    "Total_Capital": {"short": "Total Capital", "long": "Total risk-based capital available for regulatory purposes."},
    "Common_Stock_Pct": {"short": "Common Stock %", "long": "Common stock as a percentage of total equity capital."},
    "EQSUR_Pct": {"short": "EQSUR %", "long": "EQSUR as a percentage of total equity capital."},
    "Retained_Earnings_Pct": {"short": "Retained Earnings %", "long": "Retained earnings (net of EQCCOMPI) as a percentage of total equity capital."},
    "Preferred_Stock_Pct": {"short": "Preferred Stock %", "long": "Perpetual preferred stock as a percentage of total equity capital."},
    "CI_to_Capital_Risk": {"short": "C&I / Total Capital", "long": "Commercial & Industrial loans as a percentage of total capital."},

    # TTM Capital Composition
    "TTM_Common_Stock_Pct": {"short": "TTM Common Stock %", "long": "Trailing 12-month average common stock percentage of total equity."},
    "TTM_EQSUR_Pct": {"short": "TTM EQSUR %", "long": "Trailing 12-month average EQSUR percentage of total equity."},
    "TTM_Retained_Earnings_Pct": {"short": "TTM Retained Earnings %", "long": "Trailing 12-month average retained earnings percentage of total equity."},
    "TTM_Preferred_Stock_Pct": {"short": "TTM Preferred Stock %", "long": "Trailing 12-month average preferred stock percentage of total equity."},
    # Asset Quality - Top of House
    "NCLNLS":   {"short": "Total Noncurrent Loans", "long": "Total loans and leases past due 90 days or more plus nonaccrual loans."},

    # ===== DERIVED METRICS (MODIFIED & NEW) =====
    "Cost_of_Funds": {"short": "Cost of Funds", "long": "Annualized cost of interest-bearing liabilities (Quarterly Interest Expense * 4 / Average Interest-Bearing Liabilities)."},
    "Allowance_to_Gross_Loans_Rate": {"short": "Total Allowance / Gross Loans", "long": "Total Allowance for Credit Losses (on- and off-balance sheet) as a percentage of gross loans."},
    "TTM_NCO_Rate": {"short": "TTM NCO / Avg Loans", "long": "Trailing 12-Month sum of net charge-offs as a percentage of TTM average gross loans."},
    "TTM_PD30_Rate": {"short": "TTM PD 30-89 / Avg Loans", "long": "Trailing 12-Month average of loans 30-89 days past due as a percentage of TTM average gross loans."},
    "TTM_PD90_Rate": {"short": "TTM PD 90+ / Avg Loans", "long": "Trailing 12-Month average of loans 90+ days past due as a percentage of TTM average gross loans."},
    "TTM_Past_Due_Rate": {"short": "Total TTM PD Rate", "long": "Trailing 12-Month average of total loans 30+ days past due as a percentage of TTM average gross loans."},
    "Nonaccrual_to_Gross_Loans_Rate": {"short": "Nonaccrual / Gross Loans", "long": "Total nonaccrual loans as a percentage of total gross loans."},

    # Granular TTM PD Rates
    "IDB_CI_TTM_PD30_Rate": {"short": "TTM C&I PD 30-89 Rate", "long": "TTM avg C&I loans 30-89 days past due as a % of TTM avg C&I loans."},
    "IDB_CI_TTM_PD90_Rate": {"short": "TTM C&I PD 90+ Rate", "long": "TTM avg C&I loans 90+ days past due as a % of TTM avg C&I loans."},
    "IDB_CRE_TTM_PD30_Rate": {"short": "TTM CRE PD 30-89 Rate", "long": "TTM avg CRE loans 30-89 days past due as a % of TTM avg CRE loans."},
    "IDB_CRE_TTM_PD90_Rate": {"short": "TTM CRE PD 90+ Rate", "long": "TTM avg CRE loans 90+ days past due as a % of TTM avg CRE loans."},
    "IDB_Consumer_TTM_PD30_Rate": {"short": "TTM Consumer PD 30-89 Rate", "long": "TTM avg Consumer loans 30-89 days past due as a % of TTM avg Consumer loans."},
    "IDB_Consumer_TTM_PD90_Rate": {"short": "TTM Consumer PD 90+ Rate", "long": "TTM avg Consumer loans 90+ days past due as a % of TTM avg Consumer loans."},
    "IDB_Resi_TTM_PD30_Rate": {"short": "TTM Resi. PD 30-89 Rate", "long": "TTM avg Residential loans 30-89 days past due as a % of TTM avg Residential loans."},
    "IDB_Resi_TTM_PD90_Rate": {"short": "TTM Resi. PD 90+ Rate", "long": "TTM avg Residential loans 90+ days past due as a % of TTM avg Residential loans."},
    "IDB_Other_TTM_PD30_Rate": {"short": "TTM Other PD 30-89 Rate", "long": "TTM avg Other loans 30-89 days past due as a % of TTM avg Other loans."},
    "IDB_Other_TTM_PD90_Rate": {"short": "TTM Other PD 90+ Rate", "long": "TTM avg Other loans 90+ days past due as a % of TTM avg Other loans."},

    # [MODIFIED & NEW] Loan Composition and Growth Metrics
    "IDB_CI_Composition": {"short": "C&I Comp.", "long": "Commercial & Industrial (including Owner-Occupied CRE) loans as a percentage of total gross loans."},
    "IDB_CRE_Composition": {"short": "CRE Comp.", "long": "Commercial Real Estate (Construction, Multifamily, Farmland, Income-Producing) loans as a percentage of total gross loans."},
    "IDB_Consumer_Composition": {"short": "Consumer Comp.", "long": "Consumer loans as a percentage of total gross loans."},
    "IDB_Resi_Composition": {"short": "Residential Comp.", "long": "Residential Real Estate (including 1-4 Family Investor) loans as a percentage of total gross loans."},
    "IDB_Other_Composition": {"short": "Other Comp.", "long": "Other loans and leases as a percentage of total gross loans."},
    "IDB_CRE_Growth_TTM": {"short": "TTM CRE Growth", "long": "Trailing 12-Month (Year-over-Year) growth rate of the Commercial Real Estate loan portfolio."},
    "IDB_CRE_Growth_36M": {"short": "36M CRE Growth", "long": "36-Month (3-Year) growth rate of the Commercial Real Estate loan portfolio."},
    "CRE_Concentration_Capital_Risk": {"short": "CRE / (T1C + Total ACL)", "long": "Total Commercial Real Estate Loans as a percentage of Tier 1 Capital plus the Total Allowance for Credit Losses."},

    # ===== LOAN BALANCES =====
    "LNCI":     {"short": "C&I Loan Balances", "long": "Commercial and Industrial Loans."},

    "LNRECONS": {"short": "Construction & Dev Loan Bal", "long": "Construction and Land Development Loans."},
    "LNREMULT": {"short": "Multifamily CRE Loan Bal", "long": "Loans Secured by Multifamily (5 or more) Residential Properties."},
    "LNRENROW": {"short": "Nonfarm CRE Owner-Occ Bal", "long": "Loans Secured by Nonfarm Nonresidential Properties (Owner-Occupied)."},
    "LNREAG":   {"short": "Farmland Loan Bal", "long": "Loans Secured by Farmland."},
    "LNRERES":  {"short": "1-4 Family Resi Loan Bal", "long": "Loans Secured by 1-4 Family Residential Properties."},
    "LNRELOC":  {"short": "HELOC Balances", "long": "Home Equity Lines of Credit."},
    "LNCON":    {"short": "Consumer Loan Balances", "long": "Loans to Individuals for Household, Family, and Other Personal Expenditures."},
    "LNCRCD":   {"short": "Credit Card Loan Balances", "long": "Credit Card Loans."},
    "LNAUTO":   {"short": "Auto Loan Balances", "long": "Automobile Loans."},
    "LNOTHER":  {"short": "Other Loan Balances", "long": "Other Loans."},
    "LS":       {"short": "Lease Balances", "long": "Lease Financing Receivables."},
    "LNAG":     {"short": "Agricultural Loan Balances", "long": "Agricultural Loans."},

    # ===== NET CHARGE-OFFS (YTD) BY CATEGORY =====
    "NTCI":     {"short": "C&I NCOs", "long": "Net Charge-Offs on Commercial and Industrial Loans."},
    "NTRECONS": {"short": "Construction & Dev NCOs", "long": "Net Charge-Offs on Construction and Land Development Loans."},
    "NTREMULT": {"short": "Multifamily CRE NCOs", "long": "Net Charge-Offs on Loans Secured by Multifamily Residential Properties."},
    "NTLSNFOO": {"short": "Nonfarm CRE Owner-Occ NCOs", "long": "Net Charge-Offs on Loans Secured by Nonfarm Nonresidential Properties (Owner-Occupied)."},
    "NTREAG":   {"short": "Farmland NCOs", "long": "Net Charge-Offs on Loans Secured by Farmland."},
    "NTRERES":  {"short": "1-4 Family Resi NCOs", "long": "Net Charge-Offs on Loans Secured by 1-4 Family Residential Properties."},
    "NTRELOC":  {"short": "HELOC NCOs", "long": "Net Charge-Offs on Home Equity Lines of Credit."},
    "NTCON":    {"short": "Consumer NCOs", "long": "Net Charge-Offs on Loans to Individuals."},
    "NTCRCD":   {"short": "Credit Card NCOs", "long": "Net Charge-Offs on Credit Card Loans."},
    "NTAUTO":   {"short": "Auto Loan NCOs", "long": "Net Charge-Offs on Automobile Loans."},
    "NTLS":     {"short": "Lease NCOs", "long": "Net Charge-Offs on Leases."},
    "NTOTHER":  {"short": "Other Loan NCOs", "long": "Net Charge-Offs on Other Loans."},
    "NTAG":     {"short": "Agricultural NCOs", "long": "Net Charge-Offs on Agricultural Loans."},

    # ===== PAST DUE 30-89 DAYS BY CATEGORY =====
    "P3CI":     {"short": "C&I PD 30-89", "long": "Commercial and Industrial Loans 30-89 Days Past Due."},
    "P3RENROW": {"short": "1-4 Fam CRE Non-Owner PD 30-89", "long": "Loans Secured by 1-4 Family Residential (Non-owner Occupied) 30-89 Days Past Due."},
    "P3RECONS": {"short": "Construction PD 30-89", "long": "Construction and Land Development Loans 30-89 Days Past Due."},
    "P3LREMUL": {"short": "Multifamily PD 30-89", "long": "Loans Secured by Multifamily Residential Properties 30-89 Days Past Due."},
    "P3LRENRS": {"short": "Nonfarm Nonres Income CRE 30-89", "long": "Loans Secured by Multifamily Residential Properties 30-89 Days Past Due."},
    "P3RENROT": {"short": "Nonfarm CRE Owner-Occ PD 30-89", "long": "Loans Secured by Nonfarm Nonresidential (Owner-Occupied) 30-89 Days Past Due."},
    "P3REAG":   {"short": "Farmland PD 30-89", "long": "Loans Secured by Farmland 30-89 Days Past Due."},
    "P3RERES":  {"short": "1-4 Family Resi PD 30-89", "long": "Loans Secured by 1-4 Family Residential Properties 30-89 Days Past Due."},
    "P3RELOC":  {"short": "HELOC PD 30-89", "long": "Home Equity Lines of Credit 30-89 Days Past Due."},
    "P3CON":    {"short": "Consumer PD 30-89", "long": "Loans to Individuals 30-89 Days Past Due."},
    "P3CRCD":   {"short": "Credit Card PD 30-89", "long": "Credit Card Loans 30-89 Days Past Due."},
    "P3AUTO":   {"short": "Auto PD 30-89", "long": "Automobile Loans 30-89 Days Past Due."},
    "P3OTHLN":  {"short": "Other Loans PD 30-89", "long": "Other Loans 30-89 Days Past Due."},
    "P3LS":     {"short": "Leases PD 30-89", "long": "Leases 30-89 Days Past Due."},
    "P3AG":     {"short": "Agricultural PD 30-89", "long": "Agricultural Loans 30-89 Days Past Due."},

    # ===== PAST DUE 90+ DAYS BY CATEGORY =====
    "P9CI":     {"short": "C&I PD 90+", "long": "Commercial and Industrial Loans 90+ Days Past Due."},
    "P9RENROW": {"short": "1-4 Fam CRE Non-Owner PD 90+", "long": "Loans Secured by 1-4 Family Residential (Non-owner Occupied) 90+ Days Past Due."},
    "P9RECONS": {"short": "Construction PD 90+", "long": "Construction and Land Development Loans 90+ Days Past Due."},
    "P9REMULT": {"short": "Multifamily PD 90+", "long": "Loans Secured by Multifamily Residential Properties 90+ Days Past Due."},
    "P9RENROT": {"short": "Nonfarm CRE Owner-Occ PD 90+", "long": "Loans Secured by Nonfarm Nonresidential (Owner-Occupied) 90+ Days Past Due."},
    "P9RENRES": {"short": "Nonfarm CRE Owner-Occ PD 90+", "long": "Loans Secured by Nonfarm Nonresidential (Income Producing) 90+ Days Past Due."},
    "P9REAG":   {"short": "Farmland PD 90+", "long": "Loans Secured by Farmland 90+ Days Past Due."},
    "P9RERES":  {"short": "1-4 Family Resi PD 90+", "long": "Loans Secured by 1-4 Family Residential Properties 90+ Days Past Due."},
    "P9RELOC":  {"short": "HELOC PD 90+", "long": "Home Equity Lines of Credit 90+ Days Past Due."},
    "P9CON":    {"short": "Consumer PD 90+", "long": "Loans to Individuals 90+ Days Past Due."},
    "P9CRCD":   {"short": "Credit Card PD 90+", "long": "Credit Card Loans 90+ Days Past Due."},
    "P9AUTO":   {"short": "Auto PD 90+", "long": "Automobile Loans 90+ Days Past Due."},
    "P9OTHLN":  {"short": "Other Loans PD 90+", "long": "Other Loans 90+ Days Past Due."},
    "P9LS":     {"short": "Leases PD 90+", "long": "Leases 90+ Days Past Due."},
    "P9AG":     {"short": "Agricultural PD 90+", "long": "Agricultural Loans 90+ Days Past Due."},

    # ===== NONACCRUAL BY CATEGORY =====
    "NACI":     {"short": "C&I Nonaccrual", "long": "Commercial and Industrial Loans on Nonaccrual Status."},
    "NALRERES": {"short": "1-4 REsi Nonaccrual", "long": "Loans Secured by 1-4 Family Residential (Non-owner Occupied) on Nonaccrual Status."},
    "NARECONS": {"short": "Construction Nonaccrual", "long": "Construction and Land Development Loans on Nonaccrual Status."},
    "NAREMULT": {"short": "Multifamily Nonaccrual", "long": "Loans Secured by Multifamily Residential Properties on Nonaccrual Status."},
    "NARENROW": {"short": "Owner-Occupied CRE Owner-Occ Nonaccrual", "long": "Loans Secured by Nonfarm Nonresidential (Owner-Occupied) on Nonaccrual Status."},
    "NARENRES": {"short": "Nonfarm Nonres Income CRE Nonaccrual", "long": "Loans Secured by Nonfarm Nonresidential (Income Producing) on Nonaccrual Status."},
    "NAREAG":   {"short": "Farmland Nonaccrual", "long": "Loans Secured by Farmland on Nonaccrual Status."},
    "NARERES":  {"short": "1-4 Family Resi Nonaccrual", "long": "Loans Secured by 1-4 Family Residential Properties on Nonaccrual Status."},
    "NARELOC":  {"short": "HELOC Nonaccrual", "long": "Home Equity Lines of Credit on Nonaccrual Status."},
    "NACON":    {"short": "Consumer Nonaccrual", "long": "Loans to Individuals on Nonaccrual Status."},
    "NACRCD":   {"short": "Credit Card Nonaccrual", "long": "Credit Card Loans on Nonaccrual Status."},
    "NAAUTO":   {"short": "Auto Nonaccrual", "long": "Automobile Loans on Nonaccrual Status."},
    "NAOTHLN":  {"short": "Other Loans Nonaccrual", "long": "Other Loans on Nonaccrual Status."},
    "NALS":     {"short": "Leases Nonaccrual", "long": "Leases on Nonaccrual Status."},
    "NAAG":     {"short": "Agricultural Nonaccrual", "long": "Agricultural Loans on Nonaccrual Status."}
}

LOAN_CATEGORIES = {
    "CI":       {"balance": ["LNCI", "LNRENROW","LNOTHER"], "nco": ["NTCI", "NTLSNFOO","NTOTHER"], "pd30": ["P3CI", "P3RENROW","P3OTHLN"], "pd90": ["P9CI","P9RENROW","P9OTHLN"], "na": ["NACI", "NARENROW","NAOTHLN"]},
    "CRE":      {"balance": ["LNRECONS", "LNREMULT", "LNREAG", "LNRENROT"], "nco": ["NTRECONS", "NTREMULT", "NTREAG", "NTRENRES"], "pd30": ["P3RECONS", "P3LREMUL", "P3REAG", "P3RENROT"], "pd90": ["P9RECONS", "P9REMULT", "P9REAG", "P9RENROT"], "na": ["NARECONS", "NAREMULT", "NAREAG", "NARENROT"]},
    "Consumer": {"balance": ["LNCON" ], "nco": ["NTCON"], "pd30": ["P3CON"], "pd90": ["P9CON"], "na": ["NACON"]},
    "Resi":     {"balance": ["LNRERES", "LNRELOC",], "nco": ["NTRERES", "NTRELOC",], "pd30": ["P3RERES", "P3RELOC"], "pd90": ["P9RERES", "P9RELOC"], "na": ["NALRERES", "NARELOC"]},
    "Other":    {"balance": ["LS", "LNAG"], "nco": [ "NTLS", "NTAG"], "pd30": [ "P3LS", "P3AG"], "pd90": [ "P9LS", "P9AG"], "na": [ "NALS", "NAAG"]}
}
FRED_SERIES_TO_FETCH = {
    "Key Economic Indicators": {
        "GDPC1":    {"short": "Real GDP", "long": "Real Gross Domestic Product"},
        "A191RL1Q225SBEA": {"short": "Real GDP Growth", "long": "Real Gross Domestic Product, Percent Change from Preceding Period"},
        "UNRATE":   {"short": "Unemployment Rate", "long": "Civilian Unemployment Rate"},
        "CPIAUCSL": {"short": "CPI Inflation", "long": "Consumer Price Index for All Urban Consumers: All Items"},
        "UMCSENT":  {"short": "Consumer Sentiment", "long": "University of Michigan: Consumer Sentiment"},
        "ICSA":     {"short": "Initial Jobless Claims", "long": "Initial Claims, Insured Unemployment"},
        "USSLIND":  {"short": "Leading Economic Index", "long": "Conference Board Leading Economic Index for the United States"},
        "RSXFS":    {"short": "Retail Sales", "long": "Advance Retail Sales: Retail Trade"}
    },
    "Interest Rates & Yield Curve": {
        "FEDFUNDS": {"short": "Effective Fed Funds", "long": "Federal Funds Effective Rate"},
        "DFF":      {"short": "Fed Funds Rate", "long": "Federal Funds Effective Rate (Daily)"},
        "DPRIME":   {"short": "Bank Prime Loan Rate", "long": "Bank Prime Loan Rate"},
        "MORTGAGE30US": {"short": "30Y Mortgage Rate", "long": "30-Year Fixed Rate Mortgage Average in the United States"},
        "DGS30":    {"short": "30-Yr Treasury", "long": "Market Yield on U.S. Treasury at 30-Year Constant Maturity"},
        "DGS20":    {"short": "20-Yr Treasury", "long": "Market Yield on U.S. Treasury at 20-Year Constant Maturity"},
        "DGS10":    {"short": "10-Yr Treasury", "long": "Market Yield on U.S. Treasury at 10-Year Constant Maturity"},
        "DGS7":     {"short": "7-Yr Treasury", "long": "Market Yield on U.S. Treasury at 7-Year Constant Maturity"},
        "DGS5":     {"short": "5-Yr Treasury", "long": "Market Yield on U.S. Treasury at 5-Year Constant Maturity"},
        "DGS3":     {"short": "3-Yr Treasury", "long": "Market Yield on U.S. Treasury at 3-Year Constant Maturity"},
        "DGS2":     {"short": "2-Yr Treasury", "long": "Market Yield on U.S. Treasury at 2-Year Constant Maturity"},
        "DGS1":     {"short": "1-Yr Treasury", "long": "Market Yield on U.S. Treasury at 1-Year Constant Maturity"},
        "DGS6MO":   {"short": "6-Mo Treasury", "long": "Market Yield on U.S. Treasury at 6-Month Constant Maturity"},
        "DGS3MO":   {"short": "3-Mo Treasury", "long": "Market Yield on U.S. Treasury at 3-Month Constant Maturity"},
        "DGS1MO":   {"short": "1-Mo Treasury", "long": "Market Yield on U.S. Treasury at 1-Month Constant Maturity"},
        "T10Y2Y":   {"short": "10Y-2Y Spread", "long": "10-Year Minus 2-Year Treasury Spread"},
        "T10Y3M":   {"short": "10Y-3M Spread", "long": "10-Year Minus 3-Month Treasury Spread"}
    },
    "Credit Spreads & Lending Standards": {
        "DBAA":       {"short": "Baa Corp. Yield", "long": "Moody's Seasoned Baa Corporate Bond Yield"},
        "BAMLH0A0HYM2": {"short": "High-Yield Spread", "long": "ICE BofA US High Yield Index Option-Adjusted Spread"},
        "BAMLH0A0HYM2EY": {"short": "HY Effective Yield", "long": "ICE BofA US High Yield Index Effective Yield"},
        "BAMLH0A1HYBB": {"short": "BB-Rated Spread", "long": "ICE BofA BB US High Yield Index Option-Adjusted Spread"},
        "BAMLH0A2HYB":  {"short": "B-Rated Spread", "long": "ICE BofA B US High Yield Index Option-Adjusted Spread"},
        "BAMLH0A3HYC":  {"short": "CCC-Rated Spread", "long": "ICE BofA CCC & Lower US High Yield Index Option-Adjusted Spread"},
        "BAMLC0A0CM":   {"short": "IG Corp. Spread", "long": "ICE BofA US Corporate Index Option-Adjusted Spread"},
        "BAMLC0A1CAAAEY": {"short": "AAA-Rated Eff. Yield", "long": "ICE BofA AAA US Corporate Index Effective Yield"},
        "BAMLC0A4CBBB": {"short": "BBB-Rated Spread", "long": "ICE BofA BBB US Corporate Index Option-Adjusted Spread"},
        "DRBLACBS":   {"short": "Biz Loan Delinquency", "long": "Delinquency Rate on Business Loans, All Commercial Banks"},
        "DRTSCILM":   {"short": "C&I Lending Standards", "long": "Net Pct of Banks Tightening Standards for C&I Loans"},
        "DRTSCLCC":   {"short": "CRE Lending Standards", "long": "Net Pct of Banks Tightening Standards for CRE Loans"}
    },
    "Financial Stress & Risk": {
        "STLFSI4":      {"short": "Financial Stress Index", "long": "St. Louis Fed Financial Stress Index"},
        "VIXCLS":       {"short": "VIX", "long": "CBOE Volatility Index"},
        "NFCI":         {"short": "Nat'l Financial Conditions", "long": "Chicago Fed National Financial Conditions Index"}
    },
    "Global Benchmarks": {
        "FPCPITOTLZGWLD": {"short": "World Inflation", "long": "Inflation, CPI for World"},
        "FPCPITOTLZGHIC": {"short": "High-Income Inflation", "long": "Inflation, CPI for High-Income Countries"},
        "FPCPITOTLZGLMY": {"short": "Low/Mid-Income Inflation", "long": "Inflation, CPI for Low & Middle Income Countries"},
        "NYGDPMKTPCDWLD": {"short": "World GDP", "long": "GDP at Market Prices for World"}
    },
    "Real Estate & Housing": {
        "HOUST":         {"short": "Housing Starts", "long": "New Privately-Owned Housing Units Started"},
        "PERMIT":        {"short": "Building Permits", "long": "New Private Housing Units Authorized by Building Permits"},
        "NYXRSA": {"short": "NY Metro Home Price Index", "long": "S&P/Case-Shiller NY Home Price Index"},
        "LXXRSA": {"short": "LA Metro Home Price Index", "long": "S&P/Case-Shiller LA Home Price Index"},
        "MIXRNSA": {"short": "Miami Metro Home Price Index", "long": "S&P/Case-Shiller Miami Home Price Index"},
        "CASTHPI": {"short": "California HPI", "long": "All-Transactions House Price Index for California"},
        "FLSTHPI": {"short": "Florida HPI", "long": "All-Transactions House Price Index for Florida"}
    },
    "Banking Sector Aggregates": {
        "BUSLOANS":   {"short": "Business Loans (All Banks)", "long": "Business Loans, All Commercial Banks"},
        "REALLN":     {"short": "Real Estate Loans (All Banks)", "long": "Real Estate Loans, All Commercial Banks"},
        "CONSUMER":   {"short": "Consumer Loans (All Banks)", "long": "Consumer Loans, All Commercial Banks"},
        "DPSACBW027SBOG": {"short": "Total Deposits (All Banks)", "long": "Total Deposits, All Commercial Banks"},
        "DRCRELEXFACBS": {"short": "CRE Delinquency Rate", "long": "Delinquency Rate on CRE Loans (Excluding Farmland)"},
        "CORBLACBS": {"short": "Business Loan Charge-offs", "long": "Charge-off Rate on Business Loans, Annualized"},
        "DRALACBS": {"short": "All Loans Delinquency", "long": "Delinquency Rate on All Loans"},
        "CORALACBN": {"short": "All Loans Charge-offs", "long": "Charge-off Rate on All Loans, Annualized"}
    },
    "Leading Indicators": {
        "USSLIND": {"short": "US Leading Index", "long": "The Conference Board Leading Economic Index® (LEI) for the U.S."},
        "USALOLITONOSTSAM": {"short": "OECD CLI US", "long": "OECD Composite Leading Indicator for US"},
        "PAYEMS": {"short": "Nonfarm Employment", "long": "All Employees: Total Nonfarm"}
    },
    "Middle Market, Healthcare, & Funding Indicators": {
    "DRTSCIS": {"short": "Small Firm C&I Standards", "long": "Net Pct Banks Tightening Standards - Small Firms"},
    "SOFR":    {"short": "SOFR", "long": "Secured Overnight Financing Rate"},
    "TB3MS":   {"short": "3-Month T-Bill", "long": "3-Month Treasury Bill Secondary Market Rate"},
    "SOFR3MTB3M": {"short": "SOFR vs T-Bill Spread", "long": "Calculated Spread: SOFR minus 3-Month T-Bill Rate"},
    "MPCT04XXS": {"short": "Healthcare Construction", "long": "Total Construction Spending: Health Care"},
}
}

# ==================================================================================
#  3. HELPER CLASSES
# ==================================================================================
def get_bank_locations(cert_numbers: list) -> pd.DataFrame:
    """
    Fetches the primary (HQ) and all operating states for a list of banks.

    Args:
        cert_numbers: A list of bank CERT numbers.

    Returns:
        A pandas DataFrame with CERT, NAME, HQ_STATE, and ALL_OPERATING_STATES.
    """
    institution_data = []
    session = requests.Session()

    for cert in cert_numbers:
        logger.info(f"Fetching location data for CERT: {cert}...")
        hq_state = "N/A"
        all_states = set()
        bank_name = f"Bank with CERT {cert}"

        # 1. Get Institution details (for HQ State)
        try:
            inst_url = f"https://banks.data.fdic.gov/api/institutions?filters=CERT%3A%20{cert}&fields=NAME,STALP"
            response = session.get(inst_url, timeout=20)
            response.raise_for_status()
            data = response.json().get('data', [])
            if data:
                bank_name = data[0]['data']['NAME']
                hq_state = data[0]['data']['STALP']
                all_states.add(hq_state)

        except requests.exceptions.RequestException as e:
            logger.error(f"  - Could not fetch institution data for CERT {cert}: {e}")


        # 2. Get all branch locations (for all operating states)
        try:
            # Increase limit to get all branches, assuming no bank has > 10000 branches
            loc_url = f"https://banks.data.fdic.gov/api/locations?filters=CERT%3A%20{cert}&fields=STALP&limit=10000"
            response = session.get(loc_url, timeout=20)
            response.raise_for_status()
            locations = response.json().get('data', [])
            for loc in locations:
                all_states.add(loc['data']['STALP'])

        except requests.exceptions.RequestException as e:
            logger.error(f"  - Could not fetch branch location data for CERT {cert}: {e}")

        institution_data.append({
            "CERT": cert,
            "NAME": bank_name,
            "HQ_STATE": hq_state,
            "ALL_OPERATING_STATES": ", ".join(sorted(list(all_states)))
        })
        time.sleep(0.1) # Small delay to be polite to the API

    return pd.DataFrame(institution_data)
class FDICDataFetcher:
    def __init__(self, config: DashboardConfig):
        self.config = config
        self.session = requests.Session()

    def fetch_lnci_separately(self, certs_to_fetch: List[int]) -> pd.DataFrame:
        """Fetch LNCI field separately since it doesn't work in bulk requests."""
        lnci_data = []

        for cert in certs_to_fetch:
            try:
                # Request LNCI specifically with minimal other fields
                params = {
                    "filters": f"CERT:{cert}",
                    "fields": "CERT,REPDTE,LNCI",  # Only essential fields
                    "sort_by": "REPDTE",
                    "sort_order": "DESC",
                    "limit": self.config.quarters_back + 4,
                    "format": "json"
                }

                response = self.session.get(f"{self.config.fdic_api_base}/financials", params=params, timeout=30)
                response.raise_for_status()

                data = [item.get('data', {}) for item in response.json().get('data', []) if item.get('data')]
                if data:
                    df = pd.DataFrame(data)
                    df['CERT'] = cert

                    # DEBUG: Log LNCI values for this bank
                    lnci_values = df['LNCI'].dropna()
                    if not lnci_values.empty:
                        logger.info(f"🔍 CERT {cert} LNCI values: {lnci_values.iloc[:3].tolist()}")
                    else:
                        logger.warning(f"⚠️ CERT {cert} has LNCI column but all values are NaN")

                    lnci_data.append(df[['CERT', 'REPDTE', 'LNCI']])
                    logger.info(f"✅ LNCI fetched successfully for CERT {cert}")
                else:
                    logger.warning(f"⚠️ No LNCI data for CERT {cert}")

            except Exception as e:
                logger.error(f"❌ Error fetching LNCI for CERT {cert}: {e}")

            time.sleep(0.2)

        if lnci_data:
            combined_lnci = pd.concat(lnci_data, ignore_index=True)

            # DEBUG: Log the combined LNCI data
            logger.info(f"🔍 Combined LNCI data shape: {combined_lnci.shape}")
            logger.info(f"🔍 LNCI non-null count: {combined_lnci['LNCI'].notna().sum()}")
            logger.info(f"🔍 Sample LNCI values: {combined_lnci['LNCI'].dropna().head().tolist()}")

            # CRITICAL FIX: Convert REPDTE to datetime consistently
            combined_lnci['REPDTE'] = pd.to_datetime(combined_lnci['REPDTE'])
            combined_lnci['LNCI'] = pd.to_numeric(combined_lnci['LNCI'], errors='coerce')

            # DEBUG: Log after conversion
            logger.info(f"🔍 After conversion - LNCI non-null count: {combined_lnci['LNCI'].notna().sum()}")

            return combined_lnci

        return pd.DataFrame()

    def fetch_all_banks(self) -> Tuple[pd.DataFrame, List[int]]:
        certs_to_fetch = [self.config.subject_bank_cert] + self.config.peer_bank_certs
        all_bank_data, failed_certs = [], []

        # Step 1: Fetch all fields EXCEPT LNCI (remove it from the main request)
        main_fields = [f for f in FDIC_FIELDS_TO_FETCH if f != 'LNCI']

        for cert in certs_to_fetch:
            try:
                params = {
                    "filters": f"CERT:{cert}",
                    "fields": ",".join(main_fields),  # <-- REMOVED LNCI
                    "sort_by": "REPDTE", "sort_order": "DESC",
                    "limit": self.config.quarters_back + 4, "format": "json"
                }
                response = self.session.get(f"{self.config.fdic_api_base}/financials", params=params, timeout=30)
                response.raise_for_status()
                data = [item.get('data', {}) for item in response.json().get('data', []) if item.get('data')]
                if data:
                    df = pd.DataFrame(data)
                    df['CERT'] = cert
                    all_bank_data.append(df)
            except Exception as e:
                logger.error(f"Error fetching FDIC data for CERT {cert}: {e}")
                failed_certs.append(cert)
            time.sleep(0.2)

        if not all_bank_data:
            return pd.DataFrame(), failed_certs

        # Step 2: Combine main data
        combined_df = pd.concat(all_bank_data, ignore_index=True)

        # CRITICAL FIX: Convert REPDTE to datetime FIRST, before fetching LNCI
        combined_df['REPDTE'] = pd.to_datetime(combined_df['REPDTE'])

        # Step 3: Fetch LNCI separately and merge
        logger.info("Fetching LNCI data separately...")
        lnci_df = self.fetch_lnci_separately(certs_to_fetch)

        if not lnci_df.empty:
            # DEBUG: Log the data types and sample values before merge
            logger.info(f"🔍 Main DF REPDTE dtype: {combined_df['REPDTE'].dtype}")
            logger.info(f"🔍 LNCI DF REPDTE dtype: {lnci_df['REPDTE'].dtype}")
            logger.info(f"🔍 Main DF sample dates: {combined_df['REPDTE'].head(3).tolist()}")
            logger.info(f"🔍 LNCI DF sample dates: {lnci_df['REPDTE'].head(3).tolist()}")

            # Check for overlapping keys before merge
            main_keys = set(zip(combined_df['CERT'], combined_df['REPDTE']))
            lnci_keys = set(zip(lnci_df['CERT'], lnci_df['REPDTE']))
            overlapping_keys = main_keys.intersection(lnci_keys)
            logger.info(f"🔍 Overlapping merge keys: {len(overlapping_keys)} out of {len(main_keys)} main records")

            if len(overlapping_keys) == 0:
                logger.error("❌ NO OVERLAPPING KEYS - merge will fail!")
                # Log a few examples from each dataset
                logger.info(f"🔍 Main DF key examples: {list(main_keys)[:5]}")
                logger.info(f"🔍 LNCI DF key examples: {list(lnci_keys)[:5]}")

            # Merge LNCI data back into main DataFrame
            combined_df = pd.merge(
                combined_df,
                lnci_df,
                on=['CERT', 'REPDTE'],
                how='left'
            )

            # DEBUG: Check LNCI after merge
            lnci_after_merge = combined_df['LNCI'].notna().sum()
            logger.info(f"🔍 After merge - LNCI non-null count: {lnci_after_merge}")

            # Check specific bank
            bank_19977_lnci = combined_df[combined_df['CERT'] == 19977]['LNCI'].dropna()
            if not bank_19977_lnci.empty:
                logger.info(f"🔍 Bank 19977 LNCI values after merge: {bank_19977_lnci.head().tolist()}")
            else:
                logger.warning(f"⚠️ Bank 19977 has no LNCI values after merge!")

            logger.info("✅ LNCI data merged successfully")
        else:
            # Add empty LNCI column if fetch failed
            combined_df['LNCI'] = np.nan
            logger.warning("⚠️ LNCI fetch failed, using NaN values")

        # Step 4: Continue with existing processing
        for field in FDIC_FIELDS_TO_FETCH:
            if field not in combined_df.columns:
                combined_df[field] = np.nan
        # Convert numeric columns (REPDTE is already datetime)
        num_cols = [c for c in combined_df.columns if c not in ['CERT', 'NAME', 'REPDTE']]
        combined_df[num_cols] = combined_df[num_cols].apply(pd.to_numeric, errors='coerce')

        return combined_df.sort_values(['CERT', 'REPDTE']), failed_certs


class FREDDataFetcher:
    def __init__(self, config: DashboardConfig):
        self.config = config
        self.session = requests.Session()

    # MODIFIED: The return signature now includes a list of failed series strings
    def fetch_all_series(self) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
        logger.info("Starting FRED data fetch...")
        endpoint_url = f"{self.config.fred_api_base}/series/observations"
        all_series_data, descriptions_list = {}, []
        failed_series = []
        start_date = (datetime.now() - timedelta(days=self.config.fred_years_back * 365)).strftime('%Y-%m-%d')
        for category, series_info in FRED_SERIES_TO_FETCH.items():
            for series_id, details in series_info.items():
                descriptions_list.append({'Series ID': series_id, 'Category': category, **details})
                try:
                    params = {'series_id': series_id, 'api_key': self.config.fred_api_key, 'file_type': 'json', 'observation_start': start_date}
                    response = self.session.get(endpoint_url, params=params, timeout=15)
                    response.raise_for_status()
                    data = response.json().get('observations', [])
                    if data:
                        df = pd.DataFrame(data)[['date', 'value']]
                        df['date'] = pd.to_datetime(df['date'])
                        df['value'] = pd.to_numeric(df['value'], errors='coerce')
                        all_series_data[series_id] = df.set_index('date')['value']
                except Exception as e:
                    logger.error(f"Failed to fetch FRED series {series_id}: {e}")
                    failed_series.append(series_id)
                time.sleep(0.5)
        if not all_series_data:
            return pd.DataFrame(), pd.DataFrame(), failed_series

        final_df = pd.DataFrame(all_series_data)

        return final_df.reset_index(), pd.DataFrame(descriptions_list), failed_series

class BankMetricsProcessor:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config

    def _get_series(self, df: pd.DataFrame, series_list: List[str]) -> pd.Series:
        total = pd.Series(0, index=df.index)
        missing_cols = []
        for col in series_list:
            if col in df:
                total += df[col].fillna(0)
            else:
                missing_cols.append(col)

        if missing_cols:
            logger.warning(f"Missing columns in _get_series: {missing_cols}")
        return total

    def _safe_divide(self, numerator, denominator):
        if np.isscalar(numerator) and np.isscalar(denominator):
            return numerator / denominator if denominator != 0 else 0.0
        if isinstance(denominator, pd.Series):
            denominator = denominator.replace(0, np.nan)
        return (numerator / denominator).fillna(0)

    def create_derived_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        logging.info("Creating quarterly derived metrics...")
        df_processed = df.copy()

        # Define field lists for aggregations
        all_na_fields = [
            'NACI', 'NALRERES', 'NARECONS', 'NAREMULT', 'NARENROW', 'NARENRES',
            'NAREAG', 'NARERES', 'NARELOC', 'NACON', 'NACRCD', 'NAAUTO',
            'NAOTHLN', 'NALS', 'NAAG'
        ]

        all_acl_fields = [
            'LNATRES',  # Loan Loss Allowance (on-balance sheet)
            'AOBS'      # Off-balance sheet allowance (will be calculated)
        ]

        tier1_fields = ['RBCT1J']  # Tier 1 Capital
        tier2_fields = ['RBCT2']   # Tier 2 Capital
        total_cap_fields = ['RBCT1J', 'RBCT2']  # Total Capital = Tier 1 + Tier 2

        # [1] Convert all Year-to-Date (YTD) series to discrete quarterly values
        ytd_fields = [col for col in df_processed.columns if col.startswith('NT') or 'EINTEXP' in col]
        for cert, group in df_processed.groupby('CERT'):
            group = group.sort_values('REPDTE')
            for col in ytd_fields:
                quarterly_vals = group[col].diff()
                q1_mask = group['REPDTE'].dt.quarter == 1
                quarterly_vals.loc[q1_mask] = group.loc[q1_mask, col]
                df_processed.loc[group.index, f"{col}_Q"] = quarterly_vals

        # [2] Calculate off-balance sheet allowance (AOBS) before using it
        allowance_loans = df_processed.get('LNATRES', 0).fillna(0)
        rb2_ln_allowance = df_processed.get('RB2LNRES', 0).fillna(0)
        df_processed['AOBS'] = rb2_ln_allowance - allowance_loans

        # [3] Calculate aggregated series once, before the main loop
        df_processed['Total_Nonaccrual'] = self._get_series(df_processed, all_na_fields)
        df_processed['Total_ACL'] = self._get_series(df_processed, all_acl_fields)
        df_processed['Total_Capital'] = self._get_series(df_processed, total_cap_fields)
        df_processed['Tier1_Capital'] = self._get_series(df_processed, tier1_fields)
        df_processed['Tier2_Capital'] = self._get_series(df_processed, tier2_fields)

        # [4] Calculate bank-specific ratios within each bank's data
        for cert, group in df_processed.groupby('CERT'):
            idx = group.index

            # Get bank-specific values for this group
            loans = group['LNLS'].fillna(0)
            pd30 = group['P3LNLS'].fillna(0)
            pd90 = group['P9LNLS'].fillna(0)
            past_due = pd30 + pd90
            total_acl = group['Total_ACL'].fillna(0)
            nonaccrual = group['Total_Nonaccrual'].fillna(0)
            nplloans = group['NCLNLS'].fillna(0)
            total_equity = group['EQ'].fillna(0)

            # Calculate ratios for this specific bank
            df_processed.loc[idx, 'Total_Past_Due_to_Book'] = self._safe_divide(past_due, loans) * 100
            df_processed.loc[idx, 'Allowance_to_Gross_Loans_Rate'] = self._safe_divide(total_acl, loans) * 100
            df_processed.loc[idx, 'Nonaccrual_to_Gross_Loans_Rate'] = self._safe_divide(nonaccrual, loans) * 100
            df_processed.loc[idx, 'NPL_to_Gross_Loans_Rate'] = self._safe_divide(nplloans, loans) * 100
            df_processed.loc[idx, 'NPL_to_Past_Due'] = self._safe_divide(nplloans, past_due) * 100
            df_processed.loc[idx, 'NPL_to_ACL'] = self._safe_divide(nplloans, total_acl) * 100

            # Capital composition percentages for this bank
            net_retained_earnings = group['EQUP'].fillna(0) - group['EQCCOMPI'].fillna(0)
            df_processed.loc[idx, 'Net_Retained_Earnings'] = net_retained_earnings
            df_processed.loc[idx, 'Common_Stock_Pct'] = self._safe_divide(group.get('EQCS', 0), total_equity) * 100
            df_processed.loc[idx, 'EQSUR_Pct'] = self._safe_divide(group.get('EQSUR', 0), total_equity) * 100
            df_processed.loc[idx, 'Preferred_Stock_Pct'] = self._safe_divide(group.get('EQPP', 0), total_equity) * 100
            df_processed.loc[idx, 'Retained_Earnings_Pct'] = self._safe_divide(net_retained_earnings, total_equity) * 100

        # [5] Calculate custom loan category balances and rates
        category_balances = {}

        # First pass: Calculate all balances
        for cat_name, cat_details in LOAN_CATEGORIES.items():
            balance = self._get_series(df_processed, cat_details.get('balance', []))
            df_processed[f'IDB_{cat_name}_Balance'] = balance
            category_balances[cat_name] = balance

        # Calculate total categorized loans
        total_categorized_loans = pd.Series(0, index=df_processed.index)
        for balance in category_balances.values():
            total_categorized_loans += balance.fillna(0)

        # Second pass: Calculate compositions and NCO rates by bank
        for cert, group in df_processed.groupby('CERT'):
            idx = group.index

            for cat_name, cat_details in LOAN_CATEGORIES.items():
                balance = group[f'IDB_{cat_name}_Balance']
                total_cat_loans_for_bank = group[f'IDB_{cat_name}_Balance'].fillna(0)

                # Get total categorized loans for this bank
                bank_total_categorized = total_categorized_loans.loc[idx]

                # Composition relative to total categorized loans for this bank
                df_processed.loc[idx, f'IDB_{cat_name}_Composition'] = self._safe_divide(balance, bank_total_categorized) * 100

                # NCO Rate calculation for this bank
                nco_q_val = self._get_series(group, [f"{s}_Q" for s in cat_details.get('nco', []) if f"{s}_Q" in group])
                df_processed.loc[idx, f'IDB_{cat_name}_NCO_Rate'] = self._safe_divide(nco_q_val, balance) * 400

        # [6] Cost of Funds calculation by bank
        for cert, group in df_processed.groupby('CERT'):
            idx = group.index

            # Calculate interest bearing liabilities
            interest_bearing_liab = self._get_series(group, ['DEP', 'FREPP', 'OTHBOR'])

            # Calculate average interest bearing liabilities (current + previous quarter)
            prev_interest_bearing_liab = interest_bearing_liab.shift(1)
            avg_interest_bearing_liab = (interest_bearing_liab + prev_interest_bearing_liab) / 2

            # Calculate cost of funds for this bank
            eintexp_q = group.get('EINTEXP_Q', 0)
            df_processed.loc[idx, 'Cost_of_Funds'] = self._safe_divide(eintexp_q * 4, avg_interest_bearing_liab) * 100

        # [7] Capital risk ratios by bank
        for cert, group in df_processed.groupby('CERT'):
            idx = group.index

            cre_balance = group.get('IDB_CRE_Balance', 0)
            ci_balance = group.get('IDB_CI_Balance', 0)
            tier1_capital = group['Tier1_Capital'].fillna(0)
            total_capital = group['Total_Capital'].fillna(0)
            total_acl = group['Total_ACL'].fillna(0)

            df_processed.loc[idx, 'CRE_Concentration_Capital_Risk'] = self._safe_divide(cre_balance, (tier1_capital + total_acl)) * 100
            df_processed.loc[idx, 'CI_to_Capital_Risk'] = self._safe_divide(ci_balance, total_capital) * 100

        return df_processed

    def calculate_ttm_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        logging.info("Calculating granular TTM metrics...")

        all_banks_data = []
        for cert, group in df.groupby('CERT'):
            bank_df = group.sort_values("REPDTE").copy()

            if len(bank_df) >= 4:
                # Portfolio-wide TTM calcs
                avg_loans_ttm = bank_df['LNLS'].rolling(window=4, min_periods=4).mean()
                bank_df['TTM_NCO_Rate'] = self._safe_divide(bank_df['NTLNLS_Q'].rolling(window=4).sum(), avg_loans_ttm) * 100
                bank_df['TTM_PD30_Rate'] = self._safe_divide(bank_df['P3LNLS'].rolling(window=4).mean(), avg_loans_ttm) * 100
                bank_df['TTM_PD90_Rate'] = self._safe_divide(bank_df['P9LNLS'].rolling(window=4).mean(), avg_loans_ttm) * 100
                bank_df['TTM_Past_Due_Rate'] = bank_df['TTM_PD30_Rate'] + bank_df['TTM_PD90_Rate']

                # Granular IDB category TTM calcs
                for cat_name, cat_details in LOAN_CATEGORIES.items():
                    avg_cat_balance_ttm = bank_df[f'IDB_{cat_name}_Balance'].rolling(window=4).mean()

                    pd30_balance = self._get_series(bank_df, cat_details.get('pd30', []))
                    avg_pd30_ttm = pd30_balance.rolling(window=4).mean()
                    bank_df[f'IDB_{cat_name}_TTM_PD30_Rate'] = self._safe_divide(avg_pd30_ttm, avg_cat_balance_ttm) * 100

                    pd90_balance = self._get_series(bank_df, cat_details.get('pd90', []))
                    avg_pd90_ttm = pd90_balance.rolling(window=4).mean()
                    bank_df[f'IDB_{cat_name}_TTM_PD90_Rate'] = self._safe_divide(avg_pd90_ttm, avg_cat_balance_ttm) * 100

                    nco_balance = self._get_series(bank_df, cat_details.get('nco', []))
                    avg_nco_ttm = nco_balance.rolling(window=4).mean()
                    bank_df[f'IDB_{cat_name}_TTM_NCO_Rate'] = self._safe_divide(avg_nco_ttm, avg_cat_balance_ttm) * 100

                bank_df['IDB_CRE_Growth_TTM'] = bank_df['IDB_CRE_Balance'].pct_change(periods=4) * 100
                bank_df['TTM_Common_Stock_Pct'] = bank_df['Common_Stock_Pct'].rolling(window=4).mean()
                bank_df['TTM_EQSUR_Pct'] = bank_df['EQSUR_Pct'].rolling(window=4).mean()
                bank_df['TTM_Retained_Earnings_Pct'] = bank_df['Retained_Earnings_Pct'].rolling(window=4).mean()
                bank_df['TTM_Preferred_Stock_Pct'] = bank_df['Preferred_Stock_Pct'].rolling(window=4).mean()

            if len(bank_df) >= 12:
                bank_df['IDB_CRE_Growth_36M'] = bank_df['IDB_CRE_Balance'].pct_change(periods=12) * 100

            all_banks_data.append(bank_df)

        return pd.concat(all_banks_data, ignore_index=True) if all_banks_data else df

    def calculate_8q_averages(self, proc_df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculates 8-quarter rolling averages for ALL numeric metrics.
        """
        logging.info("Calculating 8-quarter averages for all numeric metrics...")
        if proc_df.empty:
            return pd.DataFrame()

        metrics_to_average = proc_df.select_dtypes(include=np.number).columns.tolist()
        if 'CERT' in metrics_to_average:
            metrics_to_average.remove('CERT')

        avg_results = []
        for cert, group in proc_df.groupby('CERT'):
            if len(group) < 8:
                continue

            group = group.sort_values('REPDTE')
            averages = group[metrics_to_average].rolling(window=8, min_periods=8).mean()
            latest_averages = averages.iloc[-1]

            record = {"CERT": cert, "NAME": group['NAME'].iloc[-1]}
            record.update(latest_averages.to_dict())
            avg_results.append(record)

        if not avg_results:
            return pd.DataFrame()

        avg_df = pd.DataFrame(avg_results).set_index('CERT')

        # Add back location data from the processed dataframe
        location_cols = ['HQ_STATE', 'ALL_OPERATING_STATES']
        if all(c in proc_df.columns for c in location_cols):
            location_info = proc_df[['CERT'] + location_cols].dropna(
                subset=['HQ_STATE']
            ).drop_duplicates(subset='CERT', keep='last').set_index('CERT')
            avg_df = avg_df.merge(location_info, left_index=True, right_index=True, how='left')

        all_cols = avg_df.columns.tolist()

        # Move NAME and location columns to the front for better readability
        front_cols = ['NAME']
        if 'HQ_STATE' in all_cols:
            front_cols.append('HQ_STATE')
            all_cols.remove('HQ_STATE')
        if 'ALL_OPERATING_STATES' in all_cols:
            front_cols.append('ALL_OPERATING_STATES')
            all_cols.remove('ALL_OPERATING_STATES')

        if 'NAME' in all_cols:
            all_cols.remove('NAME')

        new_col_order = front_cols + sorted(all_cols)
        avg_df = avg_df[new_col_order]
        return avg_df

    def create_latest_snapshot(self, proc_df: pd.DataFrame) -> pd.DataFrame:
        if proc_df.empty:
            return pd.DataFrame()

        latest_date = proc_df['REPDTE'].max()
        latest_data = proc_df[proc_df['REPDTE'] == latest_date].copy()

        # Include location columns if they exist
        snapshot_base_cols = ['CERT', 'NAME', 'REPDTE']
        if 'HQ_STATE' in latest_data.columns:
            snapshot_base_cols.append('HQ_STATE')
        if 'ALL_OPERATING_STATES' in latest_data.columns:
            snapshot_base_cols.append('ALL_OPERATING_STATES')
        snapshot = latest_data[snapshot_base_cols].copy()

        # Add all relevant metrics to the snapshot
        metrics_to_include = [
            "LNLS", "Total_ACL", "Tier1_Capital", "Tier2_Capital", "Total_Capital",
            "NPL_to_Gross_Loans_Rate", "Nonaccrual_to_Gross_Loans_Rate", "Allowance_to_Gross_Loans_Rate", "TTM_NCO_Rate",
            "TTM_PD30_Rate", "TTM_PD90_Rate", "TTM_Past_Due_Rate", "IDB_CRE_Growth_TTM",
            "IDB_CRE_Growth_36M", "CRE_Concentration_Capital_Risk", "CI_to_Capital_Risk", "Cost_of_Funds",
            "Common_Stock_Pct", "EQSUR_Pct", "Retained_Earnings_Pct", "Preferred_Stock_Pct",
            "TTM_Common_Stock_Pct", "TTM_EQSUR_Pct", "TTM_Retained_Earnings_Pct", "TTM_Preferred_Stock_Pct"
        ]

        # Add compositions and granular TTM rates
        for cat in LOAN_CATEGORIES.keys():
            snapshot[f'{cat}_Composition'] = latest_data.get(f'IDB_{cat}_Composition', np.nan)
            snapshot[f'{cat}_TTM_PD30_Rate'] = latest_data.get(f'IDB_{cat}_TTM_PD30_Rate', np.nan)
            snapshot[f'{cat}_TTM_PD90_Rate'] = latest_data.get(f'IDB_{cat}_TTM_PD90_Rate', np.nan)

        # Add single value metrics
        for metric in metrics_to_include:
            snapshot[metric] = latest_data.get(metric, np.nan)

        snapshot.rename(columns={'REPDTE': 'Reporting_Date', 'LNLS': 'Gross_Loans'}, inplace=True)
        return snapshot.set_index('CERT')

class PeerAnalyzer:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.metric_descriptions = FDIC_FIELD_DESCRIPTIONS.copy()

    def create_peer_comparison(self, processed_df: pd.DataFrame) -> pd.DataFrame:
        logging.info("Creating peer comparison analysis...")
        if processed_df.empty: return pd.DataFrame()

        latest_date = processed_df["REPDTE"].max()
        latest_data = processed_df[processed_df["REPDTE"] == latest_date].copy()

        subject_data = latest_data[latest_data["CERT"] == self.config.subject_bank_cert]
        peer_data = latest_data[latest_data["CERT"].isin(self.config.peer_bank_certs)]

        if subject_data.empty or peer_data.empty: return pd.DataFrame()

        metrics_to_compare = list(self.metric_descriptions.keys())

        comparison_list = []
        for metric in metrics_to_compare:
            if metric not in latest_data.columns: continue

            subject_value = subject_data[metric].iloc[0] if not subject_data.empty else np.nan
            peer_values = peer_data[metric].dropna()
            if peer_values.empty: continue

            record = {
                "Metric Code": metric,
                "Metric Name": self.metric_descriptions.get(metric, {}).get("short", metric),
                "Your_Bank": subject_value,
                "Peer_Mean": peer_values.mean(),
                "Peer_Median": peer_values.median(),
                "Peer_25th": peer_values.quantile(0.25),
                "Peer_75th": peer_values.quantile(0.75),
            }

            if pd.notna(subject_value):
                record["Your_Percentile"] = stats.percentileofscore(peer_values, subject_value, kind='rank')
            else:
                record["Your_Percentile"] = np.nan

            comparison_list.append(record)

        if not comparison_list: return pd.DataFrame()

        comparison_df = pd.DataFrame(comparison_list)
        comparison_df["Performance_Flag"] = comparison_df.apply(self._get_performance_flag, axis=1)
        return comparison_df

    def _get_performance_flag(self, row: pd.Series) -> str:
        if pd.isna(row.get("Your_Percentile")): return "N/A"
        percentile = row["Your_Percentile"]
        metric = row["Metric Code"].lower()
        # For these metrics, lower is better
        if any(term in metric for term in ["nco", "npl", "past_due", "eeffr", "cost_of_funds", "pd30", "pd90"]):
            if percentile <= 25: return "Top Quartile"
            if percentile <= 50: return "Better than Median"
            return "Bottom Quartile"
        # For all other metrics, higher is better
        else:
            if percentile >= 75: return "Top Quartile"
            if percentile >= 50: return "Better than Median"
            return "Bottom Quartile"

class MacroTrendAnalyzer:
    """
    Analyzes macroeconomic time series data, calculates technical indicators,
    and generates structured output for Power BI dashboards.
    """
    def __init__(self, config: 'DashboardConfig'):
        """
        Initializes the analyzer with configuration, metadata, and technical parameters.
        """
        self.config = config
        self.indicator_metadata = self._initialize_metadata()

        self.default_technical_params = self._get_default_tech_params()
        self.technical_params = self._set_technical_parameters()

    def _get_default_tech_params(self) -> Dict:
        """Centralizes the definition of default technical parameters."""
        return {
            'sma_period': 12, 'ema_period': 12,
            'bb_period': 12, 'bb_std': 2.0,
            'rsi_period': 14, 'z_score_period': 12
        }
    def _initialize_metadata(self) -> Dict:
        """
        Defines a comprehensive metadata repository for each macroeconomic indicator.
        This dictionary is the "brain" of the automated analysis, providing context
        for interpretation, classification, and commentary generation.

        Returns:
            Dict]: A nested dictionary where keys are FRED series IDs
                                       and values are dictionaries of metadata attributes.
        """
        return {
            # Leading Indicators
            "USSLIND": {
                "type": "leading", "horizon": "medium", "sectors": ["economy_wide"],
                "benchmark": {"recession": 98.0, "expansion": 102.0}, "annualized": False,
                "best_technical": "EMA", "technical_reason": "Highly responsive to economic turning points."
            },
            "T10Y2Y": { # Assuming this was an existing series
                "type": "leading", "horizon": "long", "sectors": ["banking", "real_estate", "economy_wide"],
                "benchmark": {"inversion": 0.0, "normal": 1.5}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Stable trend indicator for yield curve shape."
            },
            "USALOLITONOSTSAM": {
                "type": "leading", "horizon": "medium", "sectors": ["economy_wide", "global"],
                "benchmark": {"contraction": 99.5, "expansion": 100.5}, "annualized": False,
                "best_technical": "EMA", "technical_reason": "Captures momentum shifts in OECD economic outlook."
            },
            "DRTSCIS": {
                "type": "leading", "horizon": "short", "sectors": ["banking", "credit", "middle_market"],
                "benchmark": {"tightening": 20.0, "easing": -10.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Breakouts signal significant shifts in lending sentiment."
            },

            # Banking & Credit Quality
            "DRCRELEXFACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "real_estate", "credit"],
                "benchmark": {"stress": 3.0, "normal": 1.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Volatility in delinquencies indicates credit cycle stress."
            },
            "CORBLACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "credit", "corporate"],
                "benchmark": {"stress": 1.5, "normal": 0.5}, "annualized": True,
                "best_technical": "SMA", "technical_reason": "Identifies the underlying trend in realized credit losses."
            },
            "DRALACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "credit", "consumer"],
                "benchmark": {"stress": 3.5, "normal": 2.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Smooths quarterly data to reveal the broad credit quality trend."
            },
            "CORALACBN": {
                "type": "lagging", "horizon": "long", "sectors": ["banking", "credit", "economy_wide"],
                "benchmark": {"stress": 2.0, "normal": 0.75}, "annualized": True,
                "best_technical": "SMA", "technical_reason": "Tracks the long-term cycle of aggregate bank loan losses."
            },
            "SOFR3MTB3M": {
                "type": "coincident", "horizon": "short", "sectors": ["banking", "funding", "credit"],
                "benchmark": {"stress": 0.5, "extreme_stress": 1.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Band expansion signals rising funding stress and counterparty risk."
            },

            # Employment
            "PAYEMS": {
                "type": "coincident", "horizon": "short", "sectors": ["economy_wide", "employment"],
                "benchmark": {"contraction_yoy_pct": 0.0}, "annualized": False,
                "best_technical": "RSI", "technical_reason": "Identifies momentum exhaustion in labor market growth."
            },
            "CEU6562000101": {
                "type": "coincident", "horizon": "medium", "sectors": ["healthcare", "employment"],
                "benchmark": {"contraction_yoy_pct": 0.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Healthcare employment is a stable trend, best captured by SMA."
            },

            # Real Estate & Construction
            "NYXRSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "LXXRSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "MIXRNSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "CASTHPI": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Ideal for smoothing quarterly state-level HPI data."},
            "FLSTHPI": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Ideal for smoothing quarterly state-level HPI data."},
            "MPCT04XXS": {
                "type": "coincident", "horizon": "medium", "sectors": ["healthcare", "construction"],
                "benchmark": {"contraction": 0.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Construction spending data is volatile; SMA reveals the underlying investment trend."
            },

            # Global Benchmarks
            "FPCPITOTLZGWLD": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 5.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Annual data requires a long-term moving average to identify global inflation regimes."},
            "FPCPITOTLZGHIC": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 4.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Tracks inflation trends in developed economies."},
            "FPCPITOTLZGLMY": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 6.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Tracks inflation trends in emerging economies."},
            "NYGDPMKTPCDWLD": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Annual data requires a long-term moving average to identify global growth regimes."},
        }

    def _set_technical_parameters(self) -> Dict:
        """
        Sets default and series-specific parameters for technical indicator calculations.
        These can be tuned to optimize for different data frequencies and volatilities.

        Returns:
            Dict]: A dictionary of parameters for each indicator.
        """
        # Default parameters are chosen based on common practice for economic data
        default_params = self.default_technical_params

        specific_params = {
            "T10Y2Y": {'sma_period': 50, 'ema_period': 20, 'bb_period': 50, 'rsi_period': 14, 'z_score_period': 50},
            "USSLIND": {'sma_period': 6, 'ema_period': 6, 'bb_period': 6, 'rsi_period': 14, 'z_score_period': 6},
            "PAYEMS": {'sma_period': 12, 'ema_period': 12, 'bb_period': 20, 'rsi_period': 14, 'z_score_period': 12},
        }

        params = {}
        for series_id in self.indicator_metadata.keys():
            params[series_id] = default_params.copy()
            if series_id in specific_params:
                params[series_id].update(specific_params[series_id])
        return params
    def _calculate_rsi(self, series: pd.Series, period: int) -> pd.Series:
        """
        Helper function to calculate the Relative Strength Index (RSI).
        """
        delta = series.diff(1)
        gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()

        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return rsi
    def _safe_divide(self, numerator, denominator):
        if isinstance(denominator, pd.Series):
            denominator = denominator.replace(0, np.nan)
        return (numerator / denominator).fillna(0)
    def _compute_trend_slope(self, data: pd.Series, window: int) -> float:
        """
        Computes the linear trend slope over a specified window.
        Returns the slope as an annualized percentage change.
        """
        if len(data) < window or data.isna().all():
            return 0.0
        y = data.iloc[-window:].values
        x = np.arange(len(y))
        # Remove NaN values
        mask = ~np.isnan(y)
        if mask.sum() < 2:  # Need at least 2 points
            return 0.0
        x_clean = x[mask]
        y_clean = y[mask]
        try:
            slope, _ = np.polyfit(x_clean, y_clean, 1)
            # Convert to annualized percentage
            mean_val = np.mean(y_clean)
            if mean_val != 0:
                return (slope * 12 / mean_val) * 100
            else:
                return 0.0
        except:
            return 0.0

    def calculate_technical_indicators(self, series: pd.Series, series_id: str) -> pd.DataFrame:
        """
        Calculates a suite of technical indicators for a given time series.
        Handles mixed-frequency data by resampling to a consistent monthly frequency
        before calculation.

        Args:
            series (pd.Series): The input time series data.
            series_id (str): The FRED series ID for parameter lookup.

        Returns:
            pd.DataFrame: A DataFrame with the original value and all calculated indicators.
        """
        # --- Pre-processing: Handle mixed frequencies ---
        # Resample to a consistent monthly frequency to make indicators comparable
        # Use linear interpolation for upsampling quarterly/annual data to avoid step-changes
        series_monthly = series.resample('MS').ffill()
        params = self.technical_params.get(series_id, self.default_technical_params)
        df = pd.DataFrame({series_id: series_monthly})
        df.index.name = 'date'

        df[f'{series_id}_SMA'] = series_monthly.rolling(window=params['sma_period'], min_periods=1).mean()
        df[f'{series_id}_EMA'] = series_monthly.ewm(span=params['ema_period'], adjust=False, min_periods=1).mean()

        rolling_mean = series_monthly.rolling(window=params['bb_period'], min_periods=1).mean()
        rolling_std = series_monthly.rolling(window=params['bb_period'], min_periods=1).std()
        df[f'{series_id}_BB_Upper'] = rolling_mean + (params['bb_std'] * rolling_std)
        df[f'{series_id}_BB_Lower'] = rolling_mean - (params['bb_std'] * rolling_std)
        df[f'{series_id}_BB_Width'] = self._safe_divide((df[f'{series_id}_BB_Upper'] - df[f'{series_id}_BB_Lower']), rolling_mean)

        df[f'{series_id}_RSI'] = self._calculate_rsi(series_monthly, params['rsi_period'])

        z_score_mean = series_monthly.rolling(window=params['z_score_period'], min_periods=1).mean()
        z_score_std = series_monthly.rolling(window=params['z_score_period'], min_periods=1).std()
        df[f'{series_id}_ZScore'] = self._safe_divide((series_monthly - z_score_mean), z_score_std)

        return df


    def _get_optimal_technical_indicator(self, series_id: str) -> Dict[str, Any]:
        """
        Selects the optimal technical indicator and provides reasoning based on the
        economic characteristics of the series.
        Args:
            series_id (str): The FRED series ID.
        Returns:
            Dict[str, Any]: A dictionary containing the name of the best indicator
                            and the reasoning for its selection.
        """
        # 1. This dictionary now contains the reasons for choosing each indicator.
        # This fixes the original SyntaxError.
        selection_rules = {
            "SMA": {
                "reason": "Best for identifying long-term trends in stable, less volatile series."
            },
            "EMA": {
                "reason": "Best for tracking recent price changes and identifying short-term trends."
            },
            "Bollinger Bands": {
                "reason": "Best for assessing volatility and identifying overbought/oversold trend conditions."
            },
            "RSI": {
                "reason": "Best for measuring momentum and identifying potential trend reversals."
            }
        }

        # 2. Determine which indicator is best for the given series_id.
        # We still get this from the metadata, with "SMA" as a safe default.
        metadata = self.indicator_metadata.get(series_id, {})
        best_indicator = metadata.get("best_technical", "SMA")

        # 3. Look up the specific reason from the rules dictionary using the best_indicator.
        # We use .get() to provide a default reason in case the indicator isn't in our rules.
        default_rule = {"reason": "Default trend-following indicator used as no specific rule was found."}
        reason = selection_rules.get(best_indicator, default_rule)["reason"]

        # 4. Return the final output, which is now ready to be used in your required column.
        return {
            "indicator": best_indicator,
            "reason": reason
        }
    def calculate_oecd_aggregates(self, fred_data: pd.DataFrame) -> pd.DataFrame:
        """
        Calculates or extracts aggregates for OECD vs. non-OECD country groups
        using World Bank proxies available on FRED.

        Args:
            fred_data (pd.DataFrame): The DataFrame containing all fetched FRED series.

        Returns:
            pd.DataFrame: A DataFrame with the aggregated OECD and non-OECD series.
        """
        # 1. Define constants for the specific series IDs being used as proxies.
        # These are the actual FRED series for inflation in different income groups.
        HI_INCOME_SERIES = 'FPCPITOTLZGHI'   # Proxy for OECD countries
        LOW_MID_INCOME_SERIES = 'FPCPITOTLZGLM' # Proxy for Non-OECD countries
        WORLD_SERIES = 'FPCPITOTLZGOW'   # World aggregate

        # 2. Assign the list of required columns to the variable. This fixes the SyntaxError.
        required_cols = [HI_INCOME_SERIES, LOW_MID_INCOME_SERIES, WORLD_SERIES]

        # Check if all required columns are present in the input DataFrame.
        if not all(col in fred_data.columns for col in required_cols):
            print("Warning: Not all global inflation series found for OECD aggregation.")
            return pd.DataFrame()

        # 3. Create the output by selecting the CORRECT column for each key.
        # This fixes the logical error/ValueError.
        output_df = pd.DataFrame({
            'OECD_Proxy_Inflation': fred_data[HI_INCOME_SERIES],
            'NonOECD_Proxy_Inflation': fred_data[LOW_MID_INCOME_SERIES],
            'World_Inflation': fred_data[WORLD_SERIES]
        })

        # The data is annual, so forward-fill to create a step-function series
        # that can be merged with higher-frequency data.
        output_df.ffill(inplace=True)

        return output_df
    def _detect_frequency(self, series: pd.Series) -> str:
        """
        Infers the frequency of a time series index.
        """
        series = series.dropna()
        if len(series.index) < 3:
            return "Undetermined"
        return pd.infer_freq(series.index) or "Irregular"

    def validate_series_data(self, series_id: str, data: pd.Series) -> Dict[str, Any]:
        """
        Performs data quality checks on a time series and returns a validation summary.
        """
        total = len(data)
        missing = data.isna().sum()
        pct_missing = missing / total * 100 if total else 100
        validation_results = {
            "status": "OK",
            "missing_data_pct": pct_missing,
            "last_update": data.last_valid_index(),
            "data_frequency": self._detect_frequency(data),
        }
        z_scores = np.abs(stats.zscore(data.fillna(0)))
        validation_results["outliers_detected"] = int((z_scores > 3).sum())
        return validation_results
    def _get_technical_signal(self, series_id: str, data: pd.DataFrame, horizon: str = 'medium') -> str:
        """
        Generates a sophisticated signal by analyzing the state, duration, and
        recent events of the optimal technical indicator.

        Args:
            series_id (str): The FRED series ID (e.g., 'DGS10').
            data (pd.DataFrame): DataFrame containing the series and its technical indicators.
            horizon (str): The lookback horizon ('short', 'medium', 'long') to define
                           the period for signal confirmation.

        Returns:
            str: A descriptive signal string.
        """
        # 1. Map the horizon string to a lookback period (number of data points)
        horizon_map = {'short': 2, 'medium': 5, 'long': 10}
        lookback = horizon_map.get(horizon, 5)

        # Ensure we have enough data to look back
        if len(data) < lookback + 1:
            return "Awaiting Data (Insufficient History)"

        # --- Determine the optimal indicator to use ---
        metadata = self.indicator_metadata.get(series_id, {})
        best_tech = metadata.get("best_technical", "SMA")

        # --- SMA / EMA Logic ---
        if best_tech in ["SMA", "EMA"]:
            col_name = f"{series_id}_{best_tech}"
            if col_name not in data.columns: return "Signal N/A"

            # Get the last two points for crossover detection
            price_now = data[series_id].iloc[-1]
            price_prev = data[series_id].iloc[-2]
            tech_now = data[col_name].iloc[-1]
            tech_prev = data[col_name].iloc[-2]

            if pd.isna(price_now) or pd.isna(tech_now): return "Awaiting Data"

            # Crossover Detection
            if price_prev <= tech_prev and price_now > tech_now:
                return f"Bullish Crossover ({best_tech})"
            if price_prev >= tech_prev and price_now < tech_now:
                return f"Bearish Crossover ({best_tech})"

            # Confirmed Trend Detection (checks the entire lookback period)
            if all(data[series_id].iloc[-lookback:] > data[col_name].iloc[-lookback:]):
                return f"Confirmed Bullish Trend (Above {best_tech} for {lookback} periods)"
            if all(data[series_id].iloc[-lookback:] < data[col_name].iloc[-lookback:]):
                return f"Confirmed Bearish Trend (Below {best_tech} for {lookback} periods)"

            return "Neutral (Consolidating)"

        # --- Bollinger Bands Logic ---
        elif best_tech == "Bollinger Bands":
            upper_col, lower_col = f"{series_id}_BB_Upper", f"{series_id}_BB_Lower"
            if upper_col not in data.columns or lower_col not in data.columns: return "Signal N/A"

            price_now = data[series_id].iloc[-1]
            upper_now, lower_now = data[upper_col].iloc[-1], data[lower_col].iloc[-1]

            if any(pd.isna([price_now, upper_now, lower_now])): return "Awaiting Data"

            # Check for breakouts or reversions
            if price_now > upper_now:
                # Check if it's a new breakout or a sustained one
                if data[series_id].iloc[-2] <= data[upper_col].iloc[-2]:
                    return "Volatility Breakout (High Stress)"
                else:
                    return "Sustained High Volatility"
            elif price_now < lower_now:
                if data[series_id].iloc[-2] >= data[lower_col].iloc[-2]:
                     return "Volatility Breakout (Low/Recession)"
                else:
                    return "Sustained Low Pressure"

            return "Neutral (Range-bound)"

        # --- RSI Logic ---
        elif best_tech == "RSI":
            rsi_col = f"{series_id}_RSI"
            if rsi_col not in data.columns: return "Signal N/A"

            rsi_val = data[rsi_col].iloc[-1]
            if pd.isna(rsi_val): return "Awaiting Data"

            # Add magnitude to the signal
            if rsi_val > 80: return "Extreme Overbought (High Momentum)"
            if rsi_val > 70: return "Overbought"
            if rsi_val < 20: return "Extreme Oversold (Low Momentum)"
            if rsi_val < 30: return "Oversold"

            # Check for momentum change
            avg_rsi_lookback = data[rsi_col].iloc[-lookback:].mean()
            if rsi_val > avg_rsi_lookback and rsi_val > 50: return "Gaining Bullish Momentum"
            if rsi_val < avg_rsi_lookback and rsi_val < 50: return "Gaining Bearish Momentum"

            return "Neutral Momentum"

        return "Not Implemented"

    def _generate_short_term_analysis(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates markdown-formatted short-term (1-3 months) analysis for a series.
        This is applied row-wise to generate commentary for each point in time.
        """
        def generate_row_analysis(row):
            current_value = row[series_id]
            # Get previous row's value for change calculation
            prev_row_index = data.index.get_loc(row.name) - 1
            if prev_row_index < 0:
                change = 0.0
            else:
                prior_value = data[series_id].iloc[prev_row_index]
                change = current_value - prior_value

            if pd.isna(current_value):
                return "No data available for this period."

            tech_signal = self._get_technical_signal(series_id, data.loc[:row.name], 'short')
            annualized_text = ' (annualized)' if metadata.get('annualized', False) else ''

            # Simple outlook logic
            outlook = "Stable"
            if "Bullish" in tech_signal or "High" in tech_signal:
                outlook = "Improving"
            elif "Bearish" in tech_signal or "Low" in tech_signal:
                outlook = "Worsening"

            return (
                f"**Current Level**: {current_value:,.2f}{annualized_text}\n\n"
                f"**Change (MoM)**: {change:+.2f}\n\n"
                f"**Technical Signal**: {tech_signal}\n\n"
                f"**Outlook**: Short-term trend appears to be **{outlook}**."
            )

        # Apply the function to each row of the DataFrame
        return data.apply(generate_row_analysis, axis=1)


    def _generate_long_term_analysis(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates markdown-formatted long-term (1-3 years) analysis for a series.
        Includes trends, momentum, historical context, and regime identification.
        """
        def generate_row_analysis(row):
            current_value = row[series_id]
            current_date = row.name
            if pd.isna(current_value):
                return "No data available for this period."
            # Get historical data up to current point
            historical_data = data.loc[:current_date, series_id].dropna()
            if len(historical_data) < 12:  # Need at least 1 year of data
                return "Insufficient historical data for long-term analysis."
            # 1. Calculate long-term trends (1-3 year windows)
            analysis_parts = []
            # Year-over-year change
            if len(historical_data) >= 12:
                yoy_value = historical_data.iloc[-13] if len(historical_data) >= 13 else historical_data.iloc[0]
                yoy_change = ((current_value - yoy_value) / yoy_value) * 100 if yoy_value != 0 else 0
                yoy_change_abs = current_value - yoy_value
            else:
                yoy_change = 0
                yoy_change_abs = 0
            # Calculate long-term moving averages
            ma_12m = historical_data.rolling(window=12, min_periods=12).mean().iloc[-1] if len(historical_data) >= 12 else current_value
            ma_24m = historical_data.rolling(window=24, min_periods=24).mean().iloc[-1] if len(historical_data) >= 24 else ma_12m
            # Calculate trend slope (annualized percentage change over 12-24 months)
            if len(historical_data) >= 24:
                window_data = historical_data.iloc[-24:]
                x = np.arange(len(window_data))
                y = window_data.values
                if len(y) > 0 and not np.all(np.isnan(y)):
                    slope, _ = np.polyfit(x, y, 1)
                    # Annualize the slope as percentage of mean
                    trend_slope_pct = (slope * 12 / np.mean(y)) * 100 if np.mean(y) != 0 else 0
                else:
                    trend_slope_pct = 0
            else:
                trend_slope_pct = 0
            # Calculate z-score over 3-year window
            if len(historical_data) >= 36:
                rolling_mean = historical_data.rolling(window=36).mean().iloc[-1]
                rolling_std = historical_data.rolling(window=36).std().iloc[-1]
                z_score_3y = (current_value - rolling_mean) / rolling_std if rolling_std > 0 else 0
            else:
                z_score_3y = 0
            # Calculate historical percentile (5-year window)
            if len(historical_data) >= 60:
                percentile_rank = stats.percentileofscore(historical_data.iloc[-60:], current_value)
            elif len(historical_data) >= 12:
                percentile_rank = stats.percentileofscore(historical_data, current_value)
            else:
                percentile_rank = 50
            # Determine regime and context based on metadata
            annualized_text = ' (annualized)' if metadata.get('annualized', False) else ''
            benchmark = metadata.get('benchmark', {})
            # Build analysis text
            analysis_parts.append(f"**Current Level vs Historical Trend:** As of {current_date.strftime('%Y-%m')}, "
                                f"{metadata.get('short', series_id)} stands at {current_value:,.2f}{annualized_text} "
                                f"(YoY: {yoy_change:+.1f}%), which is in the {percentile_rank:.0f}th percentile "
                                f"of its 5-year historical range.")
            # Trend analysis
            trend_direction = "upward" if trend_slope_pct > 0.5 else "downward" if trend_slope_pct < -0.5 else "sideways"
            ma_signal = "above" if current_value > ma_12m else "below"
            analysis_parts.append(f"**Long-Term Trend:** The 24-month trend shows {trend_direction} momentum "
                                f"with an annualized slope of {trend_slope_pct:+.1f}%. "
                                f"Current value is {ma_signal} its 12-month moving average ({ma_12m:,.2f}).")
            # Volatility and regime analysis
            if abs(z_score_3y) > 2:
                volatility_status = "extreme high" if z_score_3y > 2 else "extreme low"
                analysis_parts.append(f"**Volatility/Alerts:** Recent readings have achieved a 3-year z-score of {z_score_3y:.2f}, "
                                    f"indicating {volatility_status} levels relative to historical norms.")
            elif abs(z_score_3y) > 1:
                volatility_status = "elevated" if z_score_3y > 1 else "depressed"
                analysis_parts.append(f"**Volatility/Alerts:** The indicator shows {volatility_status} levels "
                                    f"with a z-score of {z_score_3y:.2f}.")
            # Economic context based on benchmarks and series type
            context_parts = []
            # Check against specific benchmarks
            if benchmark:
                for regime, threshold in benchmark.items():
                    if 'stress' in regime and current_value > threshold:
                        context_parts.append(f"Value exceeds stress threshold of {threshold:.2f}")
                    elif 'recession' in regime and current_value < threshold:
                        context_parts.append(f"Below recession threshold of {threshold:.2f}")
                    elif 'expansion' in regime and current_value > threshold:
                        context_parts.append(f"In expansion territory (above {threshold:.2f})")
                    elif 'inversion' in regime and current_value < threshold:
                        context_parts.append(f"Yield curve inverted (below {threshold:.2f})")
            # Add series-specific context
            if 'inflation' in series_id.lower() or series_id == 'CPIAUCSL':
                if current_value > 2.5:
                    context_parts.append("Consumer inflation is above the Fed's 2% target")
                else:
                    context_parts.append("Inflation remains near or below target")
            if 'unemployment' in metadata.get('long', '').lower() or series_id == 'UNRATE':
                if percentile_rank < 25:
                    context_parts.append("Unemployment near historical lows suggests tight labor market")
                elif percentile_rank > 75:
                    context_parts.append("Elevated unemployment indicates labor market stress")
            if context_parts:
                analysis_parts.append(f"**Economic Context:** {'. '.join(context_parts)}.")
            return '\n\n'.join(analysis_parts)
        # Apply the function to each row
        return data.apply(generate_row_analysis, axis=1)

    def _generate_risk_assessment(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates risk assessment with red/yellow/green flags and numeric scores (0-100).
        Evaluates deviation from benchmarks, volatility, and momentum.
        """
        def generate_row_assessment(row):
            current_value = row[series_id]
            current_date = row.name
            if pd.isna(current_value):
                return "**Risk Score:** N/A - No data available."
            # Get historical data
            historical_data = data.loc[:current_date, series_id].dropna()
            if len(historical_data) < 12:
                return "**Risk Score:** N/A - Insufficient data for risk assessment."
            # Initialize risk components
            risk_score = 0
            risk_factors = []
            risk_color = "Green"
            # 1. Deviation from historical norms (0-40 points)
            if len(historical_data) >= 36:
                rolling_mean = historical_data.rolling(window=36).mean().iloc[-1]
                rolling_std = historical_data.rolling(window=36).std().iloc[-1]
                z_score = abs((current_value - rolling_mean) / rolling_std) if rolling_std > 0 else 0
            else:
                rolling_mean = historical_data.mean()
                rolling_std = historical_data.std()
                z_score = abs((current_value - rolling_mean) / rolling_std) if rolling_std > 0 else 0
            # Map z-score to risk points (0-40)
            deviation_score = min(40, (z_score / 3) * 40)
            risk_score += deviation_score
            if z_score > 2:
                risk_factors.append(f"Extreme deviation ({z_score:.1f}σ)")
            elif z_score > 1:
                risk_factors.append(f"Moderate deviation ({z_score:.1f}σ)")
            # 2. Benchmark breach assessment (0-30 points)
            benchmark = metadata.get('benchmark', {})
            benchmark_score = 0
            for regime, threshold in benchmark.items():
                if 'stress' in regime.lower() or 'extreme' in regime.lower():
                    if current_value > threshold:
                        benchmark_score = 30
                        risk_factors.append(f"Exceeds {regime} level ({threshold:.2f})")
                    elif current_value > threshold * 0.8:  # Within 20% of stress
                        benchmark_score = max(benchmark_score, 20)
                        risk_factors.append(f"Approaching {regime} level")
                elif 'normal' in regime.lower():
                    if abs(current_value - threshold) / threshold > 0.5:
                        benchmark_score = max(benchmark_score, 15)
            risk_score += benchmark_score
            # 3. Volatility assessment (0-20 points)
            if len(historical_data) >= 12:
                recent_volatility = historical_data.iloc[-12:].std()
                long_term_volatility = historical_data.std()
                vol_ratio = recent_volatility / long_term_volatility if long_term_volatility > 0 else 1
                volatility_score = min(20, max(0, (vol_ratio - 1) * 20))
                risk_score += volatility_score
                if vol_ratio > 1.5:
                    risk_factors.append("High recent volatility")
            # 4. Momentum/trend assessment (0-10 points)
            if len(historical_data) >= 6:
                recent_trend = historical_data.iloc[-6:].values
                x = np.arange(len(recent_trend))
                if len(recent_trend) > 1:
                    slope, _ = np.polyfit(x, recent_trend, 1)
                    # For certain indicators, negative trends increase risk
                    indicator_type = metadata.get('type', '')
                    if indicator_type == 'leading' and slope < 0:
                        momentum_score = 10
                        risk_factors.append("Negative momentum in leading indicator")
                    elif 'delinquency' in metadata.get('long', '').lower() and slope > 0:
                        momentum_score = 10
                        risk_factors.append("Rising delinquency trend")
                    else:
                        momentum_score = 0
                    risk_score += momentum_score
            # 5. Special conditions for specific series
            if series_id == 'T10Y2Y' and current_value < 0:
                risk_score = max(risk_score, 80)  # Yield curve inversion is high risk
                risk_factors.append("Yield curve inverted")
            if 'spread' in series_id.lower() and series_id.startswith('BAM'):
                percentile = stats.percentileofscore(historical_data, current_value)
                if percentile > 90:
                    risk_score = max(risk_score, 70)
                    risk_factors.append("Credit spreads at extreme highs")
            # Determine color flag based on final score
            if risk_score >= 70:
                risk_color = "Red"
            elif risk_score >= 40:
                risk_color = "Yellow"
            else:
                risk_color = "Green"
            # Format the assessment
            assessment_parts = [
                f"**Risk Score:** {risk_score:.0f}/100 ({risk_color})"
            ]
            if risk_factors:
                assessment_parts.append(f"**Risk Drivers:** {', '.join(risk_factors)}")
            # Add specific risk interpretation
            if risk_color == "Red":
                assessment_parts.append("**Assessment:** High risk - Indicator signals significant stress or deviation from normal conditions.")
            elif risk_color == "Yellow":
                assessment_parts.append("**Assessment:** Moderate risk - Indicator shows concerning trends requiring close monitoring.")
            else:
                assessment_parts.append("**Assessment:** Low risk - Indicator within normal operating parameters.")
            return ' '.join(assessment_parts)
        # Apply the function to each row
        return data.apply(generate_row_assessment, axis=1)
    def generate_powerbi_output(self, processed_data: Dict) -> pd.DataFrame:
        """
        Generates a Power BI-optimized "long" format DataFrame from the processed
        technical indicator data. Also attaches metadata and generated commentary.
        """
        all_series_dfs = []
        for series_id, df in processed_data.items():
            if df.empty:
                continue
            # Use a copy to avoid side effects
            df_copy = df.copy()
            # --- Generate Commentary ---
            metadata = self.indicator_metadata.get(series_id, {})
            df_copy['Short_Term_Analysis'] = self._generate_short_term_analysis(series_id, df_copy, metadata)
            df_copy['Long_Term_Analysis'] = self._generate_long_term_analysis(series_id, df_copy, metadata)
            df_copy['Risk_Assessment'] = self._generate_risk_assessment(series_id, df_copy, metadata)

            # --- Add Metadata Columns ---
            commentary_cols = ['Short_Term_Analysis', 'Long_Term_Analysis', 'Risk_Assessment']
            df_copy['Type'] = metadata.get('type', 'N/A')
            df_copy['Horizon'] = metadata.get('horizon', 'N/A')
            df_copy['Sectors'] = '|'.join(metadata.get('sectors', []))
            metadata_cols = ['Type', 'Horizon', 'Sectors']
            date_col_name = df_copy.index.name
            df_copy.reset_index(inplace=True)
            id_vars = [date_col_name] + commentary_cols + metadata_cols

            value_vars = [col for col in df_copy.columns if col not in id_vars and col != series_id]

            base_df = df_copy[id_vars + [series_id]].copy()
            base_df['Indicator_Name'] = 'Actual Value'
            base_df.rename(columns={series_id: 'Indicator_Value', date_col_name: 'date'}, inplace=True)

            melted_df = pd.DataFrame()
            if value_vars:
                melted_df = df_copy.melt(
                    id_vars=id_vars,
                    value_vars=value_vars,
                    var_name='Indicator_Name',
                    value_name='Indicator_Value'
                )
                melted_df.rename(columns={date_col_name: 'date'}, inplace=True)
            series_final_df = pd.concat([base_df, melted_df], ignore_index=True)
            series_final_df['Series_ID'] = series_id
            all_series_dfs.append(series_final_df)
        if not all_series_dfs:
            return pd.DataFrame()
        # Concatenate all the processed series DataFrames into one big one
        final_output_df = pd.concat(all_series_dfs, ignore_index=True)
        return final_output_df


class ExcelOutputGenerator:
    """Generates the final Excel dashboard output file."""

    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.fdic_desc_df = pd.DataFrame.from_dict(FDIC_FIELD_DESCRIPTIONS, orient='index').reset_index()
        self.fdic_desc_df.rename(columns={'index': 'Metric Code', 'short': 'Metric Name', 'long': 'Description'}, inplace=True)

    def write_excel_output(self, file_path: str, **kwargs):
        """Writes all DataFrames to a single styled Excel file with multiple sheets."""
        logging.info(f"Writing final dashboard to: {file_path}")
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in kwargs.items():
                if isinstance(df, pd.DataFrame) and not df.empty:
                    n_rows, n_cols = df.shape
                    logging.info(f"Sheet '{sheet_name}' contains {n_rows} rows and {n_cols} cols.")
                    if n_rows > 1048576:
                        logging.error(f"Sheet '{sheet_name}' too large ({n_rows} rows).")
                    write_index = sheet_name in ["Latest_Peer_Snapshot", "Averages_8Q_All_Metrics", "Data_Validation_Report"]
                    df.to_excel(writer, sheet_name=sheet_name, index=write_index)
            logging.info("All data written, starting styling...")
            self._style_metric_descriptions_sheet(writer)
            self._apply_summary_styles(writer, kwargs.get("Summary_Dashboard"))
            self._apply_snapshot_styles(writer, kwargs.get("Latest_Peer_Snapshot"))
            self._apply_macro_analysis_styles(writer, kwargs.get("Macro_Analysis"))
        logging.info("Excel file written and styled successfully.")

    def _style_metric_descriptions_sheet(self, writer):
        """Styles the FDIC metric descriptions sheet."""
        sheet_name = 'FDIC_Metric_Descriptions'
        if sheet_name not in writer.sheets: return

        logging.info(f"Styling {sheet_name} sheet...")
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 25 # Metric Code
        worksheet.column_dimensions['B'].width = 35 # Metric Name
        worksheet.column_dimensions['C'].width = 80 # Description

    def _apply_summary_styles(self, writer, df: pd.DataFrame):
        """Applies conditional formatting to the Summary_Dashboard sheet using openpyxl."""
        if df is None or df.empty: return
        sheet_name = 'Summary_Dashboard'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")

        # Define styles using openpyxl objects
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        green_font = Font(color='006100')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        yellow_font = Font(color='9C6500')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006')

        # Set column widths
        worksheet.column_dimensions['A'].width = 15 # Metric Code
        worksheet.column_dimensions['B'].width = 35 # Metric Name
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
            worksheet.column_dimensions[col_letter].width = 18

        try:
            flag_col_index = df.columns.get_loc('Performance_Flag') + 1 # +1 for 0-based to 1-based
            flag_col_letter = openpyxl.utils.get_column_letter(flag_col_index)

            # Define conditional formatting rules
            green_rule = FormulaRule(formula=[f'SEARCH("Top Quartile",${flag_col_letter}2)'], fill=green_fill, font=green_font)
            yellow_rule = FormulaRule(formula=[f'SEARCH("Better than Median",${flag_col_letter}2)'], fill=yellow_fill, font=yellow_font)
            red_rule = FormulaRule(formula=[f'SEARCH("Bottom Quartile",${flag_col_letter}2)'], fill=red_fill, font=red_font)

            data_range = f"A2:{openpyxl.utils.get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            worksheet.conditional_formatting.add(data_range, green_rule)
            worksheet.conditional_formatting.add(data_range, yellow_rule)
            worksheet.conditional_formatting.add(data_range, red_rule)

        except KeyError:
            logger.warning("Styling skipped: 'Performance_Flag' column not found in Summary_Dashboard.")

    def _apply_snapshot_styles(self, writer, df: pd.DataFrame):
        """Applies number formatting to the Latest_Peer_Snapshot sheet using openpyxl."""
        if df is None or df.empty: return
        sheet_name = 'Latest_Peer_Snapshot'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")

        worksheet.column_dimensions['A'].width = 15  # CERT index
        worksheet.column_dimensions['B'].width = 35  # NAME

        percent_format = '0.00%'

        # Start from column 2 because to_excel writes the index
        for col_num, col_name in enumerate(df.columns, 2):
            col_letter = openpyxl.utils.get_column_letter(col_num)

            if col_letter not in ['A', 'B']:
                 worksheet.column_dimensions[col_letter].width = 20

            if any(term in col_name for term in ["Rate", "Pct", "Ratio", "Comp", "Growth", "Risk", "Funds"]):
                for cell in worksheet[col_letter]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = percent_format

    def _apply_macro_analysis_styles(self, writer, df: pd.DataFrame):
        if df is None or df.empty: return
        sheet_name = 'Macro_Analysis'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")
        worksheet.column_dimensions['A'].width = 25 # date
        worksheet.column_dimensions['H'].width = 30 # Series_ID
        worksheet.column_dimensions['I'].width = 25 # Indicator_Name
        worksheet.column_dimensions['B'].width = 40 # Short_Term_Analysis
        worksheet.column_dimensions['C'].width = 40 # Long_Term_Analysis
        worksheet.column_dimensions['D'].width = 40 # Risk_Assessment
class PowerBIGenerator:
    """Generates M and DAX scripts for a Power BI Star Schema."""

    def generate_powerbi_setup_file(self, output_dir: str, powerbi_macro_df: pd.DataFrame, fred_desc_df: pd.DataFrame):
        logger.info("Generating Power BI setup file...")
        Path(output_dir).mkdir(exist_ok=True)
        output_path = Path(output_dir) / "PowerBI_Setup.txt"

        m_code_part = self._get_m_language_code()
        dax_code_part = self._get_dax_measures_code()
        bonus_code_part = self._get_bonus_m_code()
        instructions_part = self._get_instructions()

        full_content = (
            f"{instructions_part}\n"
            f"//--- M LANGUAGE QUERIES ---\n"
            f"//--- Paste each query into a new Blank Query in Power BI's Power Query Editor ---\n\n"
            f"{m_code_part}\n"
            f"//--- DATA MODEL RELATIONSHIPS ---\n"
            f"//--- Power BI should auto-detect these. If not, create them manually in the 'Model' view. ---\n\n"
            f"// Relationship 1: Dates\n"
            f"//   Table 1: fMacro_Data (Many side)\n"
            f"//   Table 2: dDate (One side)\n"
            f"//   Column:  DateKey\n"
            f"//   Cardinality: Many to One (*:1)\n"
            f"//   Cross-filter direction: Single\n\n"
            f"// Relationship 2: Series\n"
            f"//   Table 1: fMacro_Data (Many side)\n"
            f"//   Table 2: dSeries (One side)\n"
            f"//   Column:  SeriesKey\n"
            f"//   Cardinality: Many to One (*:1)\n"
            f"//   Cross-filter direction: Single\n\n"
            f"{dax_code_part}\n"
            f"{bonus_code_part}\n"
        )

        with open(output_path, "w") as f:
            f.write(full_content)
        logger.info(f"Successfully wrote Power BI setup file to {output_path}")

    def _get_instructions(self) -> str:
        return f"""
//==============================================================================
// Power BI Setup Script
// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
//==============================================================================

// INSTRUCTIONS:
// 1. In Power BI, click "Get data" -> "Excel workbook" and select the generated Excel file.
// 2. In the Navigator, select the 'Macro_Analysis' and 'FRED_Descriptions' sheets. Click "Transform Data".
//    This loads the data into the Power Query Editor.
// 3. Once in the Power Query Editor, create a new "Blank Query" for each of the M Language Queries below.
//    (Right-click in the Queries pane -> New Query -> Blank Query).
// 4. Copy and paste the code for each query into its corresponding blank query's Advanced Editor.
// 5. Rename the queries exactly as specified in the titles (e.g., "dDate", "dSeries", "fMacro_Data").
// 6. Click "Close & Apply" to load the data model.
// 7. Once loaded, verify the relationships as described below in the 'Model' view.
// 8. Finally, go to the "Data" view, select the 'fMacro_Data' table, and create each DAX measure.
"""

    def _get_m_language_code(self) -> str:
        dDate_m = """
// --- Dimension Table: dDate ---
// This query generates a complete and dynamic date table based on your data.
let
    // **FIXED**: Point to the correct 'Macro_Analysis' sheet
    SourceTable = Macro_Analysis,
    MinDate = List.Min(Table.Column(SourceTable, "Date")),
    MaxDate = List.Max(Table.Column(SourceTable, "Date")),

    DayCount = Duration.Days(MaxDate - MinDate) + 1,
    DateList = List.Dates(MinDate, DayCount, #duration(1, 0, 0, 0)),
    DateTable = Table.FromList(DateList, Splitter.SplitByNothing(), {"Date"}),

    #"Changed Type to Date" = Table.TransformColumnTypes(DateTable,{{"Date", type date}}),
    #"Inserted Year" = Table.AddColumn(#"Changed Type to Date", "Year", each Date.Year([Date]), Int64.Type),
    #"Inserted Quarter Num" = Table.AddColumn(#"Inserted Year", "QuarterNum", each Date.QuarterOfYear([Date]), Int64.Type),
    #"Added Quarter" = Table.AddColumn(#"Inserted Quarter Num", "Quarter", each "Q" & Text.From([QuarterNum]), type text),
    #"Inserted Month" = Table.AddColumn(#"Added Quarter", "Month", each Date.Month([Date]), Int64.Type),
    #"Inserted Month Name" = Table.AddColumn(#"Inserted Month", "MonthName", each Date.ToText([Date], "MMMM"), type text),
    #"Inserted Day" = Table.AddColumn(#"Inserted Month Name", "Day", each Date.Day([Date]), Int64.Type),
    #"Inserted Day Name" = Table.AddColumn(#"Inserted Day", "DayName", each Date.ToText([Date], "dddd"), type text),
    #"Inserted Week of Year" = Table.AddColumn(#"Inserted Day Name", "WeekOfYear", each Date.WeekOfYear([Date]), Int64.Type),
    #"Added DateKey" = Table.AddColumn(#"Inserted Week of Year", "DateKey", each [Year] * 10000 + [Month] * 100 + [Day], Int64.Type),

    #"Reordered Columns" = Table.ReorderColumns(#"Added DateKey",{"DateKey", "Date", "Year", "Quarter", "Month", "MonthName", "Day", "DayName", "WeekOfYear"}),
    #"Removed QuarterNum" = Table.RemoveColumns(#"Reordered Columns",{"QuarterNum"})
in
    #"Removed QuarterNum"
"""
        dSeries_m = """
// --- Dimension Table: dSeries ---
// It assumes you have loaded the 'FRED_Descriptions' and 'Macro_Analysis' sheets from Excel.
let
    SourceDescriptions = FRED_Descriptions,
    SourceData = Macro_Analysis,

    // Get the distinct list of Series IDs from the main data table to ensure relevance
    MacroDataSeries = Table.Distinct(Table.SelectColumns(SourceData, {"Series_ID"})),

    // Join descriptions with the actual series present in the data
    #"Merged Queries" = Table.NestedJoin(SourceDescriptions, {"Series ID"}, MacroDataSeries, {"Series_ID"}, "Join", JoinKind.Inner),
    #"Removed Columns" = Table.RemoveColumns(#"Merged Queries",{"Join"}),

    // Add a unique integer key for the dimension, which will be the primary key
    #"Added Index" = Table.AddIndexColumn(#"Removed Columns", "SeriesKey", 1, 1, Int64.Type),

    // Select and reorder for clarity
    #"Selected Columns" = Table.SelectColumns(#"Added Index",{"SeriesKey", "Series ID", "long", "short", "Category", "Sectors"}),
    #"Renamed Columns" = Table.RenameColumns(#"Selected Columns",{{"long", "Description"}, {"short", "Series Name"}})
in
    #"Renamed Columns"
"""
        fMacro_Data_m = """
// --- Fact Table: fMacro_Data ---
// It assumes you have the queries 'Macro_Analysis', and 'dSeries' already created.
let
    Source = Macro_Analysis,

    #"Added DateKey" = Table.AddColumn(Source, "DateKey", each Date.Year([Date]) * 10000 + Date.Month([Date]) * 100 + Date.Day([Date]), Int64.Type),

    #"Merged with dSeries" = Table.NestedJoin(#"Added DateKey", {"Series_ID"}, dSeries, {"Series ID"}, "dSeries", JoinKind.LeftOuter),
    #"Expanded SeriesKey" = Table.ExpandTableColumn(#"Merged with dSeries", "dSeries", {"SeriesKey"}, {"SeriesKey"}),

    #"Selected Columns" = Table.SelectColumns(#"Expanded SeriesKey",{"DateKey", "SeriesKey", "Indicator_Value"}),
    #"Changed Type" = Table.TransformColumnTypes(#"Selected Columns",{{"Indicator_Value", type number}, {"DateKey", Int64.Type}, {"SeriesKey", Int64.Type}})
in
    #"Changed Type"
"""
        return f"{dDate_m}\n{dSeries_m}\n{fMacro_Data_m}"

    def _get_dax_measures_code(self) -> str:
        return """
//--- DAX MEASURES & COLUMNS ---
//--- Select the 'fMacro_Data' table, then click 'New Measure' for each item. ---

[Selected Value] = SUM('fMacro_Data'[Indicator_Value])

[Latest Value] =
CALCULATE(
    [Selected Value],
    LASTDATE('dDate'[Date])
)

[Previous Month Value] =
CALCULATE(
    [Selected Value],
    DATEADD('dDate'[Date], -1, MONTH)
)

[Month-over-Month Change] =
VAR CurrentValue = [Selected Value]
VAR PreviousValue = [Previous Month Value]
RETURN
    IF(NOT ISBLANK(CurrentValue) && NOT ISBLANK(PreviousValue), CurrentValue - PreviousValue, BLANK())

[Year-over-Year Value] =
CALCULATE(
    [Selected Value],
    SAMEPERIODLASTYEAR('dDate'[Date])
)

[Year-over-Year Change] =
VAR CurrentValue = [Selected Value]
VAR PreviousValue = [Year-over-Year Value]
RETURN
    IF(NOT ISBLANK(CurrentValue) && NOT ISBLANK(PreviousValue), CurrentValue - PreviousValue, BLANK())
"""

    def _get_bonus_m_code(self) -> str:
        return """
//--- BONUS: M-LANGUAGE FOR BRIDGE TABLE ---
// --- Bridge Table: dSeries_Sectors ---
let
    Source = dSeries,
    #"Removed Other Columns" = Table.SelectColumns(Source,{"SeriesKey", "Sectors"}),
    #"Split Column by Delimiter" = Table.ExpandListColumn(Table.TransformColumns(#"Removed Other Columns", {{"Sectors", Splitter.SplitTextByDelimiter("|", QuoteStyle.None), let itemType = (type nullable text) meta [Serialized.Text = true] in type {itemType}}}), "Sectors"),
    #"Renamed Columns" = Table.RenameColumns(#"Split Column by Delimiter",{{"Sectors", "Sector"}}),
    #"Trimmed Text" = Table.TransformColumns(#"Renamed Columns",{{"Sector", Text.Trim, type text}}),
    #"Filtered Rows" = Table.SelectRows(#"Trimmed Text", each ([Sector] <> null and [Sector] <> ""))
in
    #"Filtered Rows"
"""



# ==================================================================================
#  4. MAIN DASHBOARD CLASS & EXECUTION
# ==================================================================================

class BankPerformanceDashboard:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.fdic_fetcher = FDICDataFetcher(config)
        self.fred_fetcher = FREDDataFetcher(config)
        self.processor = BankMetricsProcessor(config)
        self.analyzer = PeerAnalyzer(config)
        self.macro_analyzer = MacroTrendAnalyzer(config)
        self.output_gen = ExcelOutputGenerator(config)
        Path(config.output_dir).mkdir(exist_ok=True)
    def _analyze_fdic_data_availability(self, fdic_df: pd.DataFrame) -> Dict[str, Any]:
        """
        Analyzes FDIC data availability and returns a report of missing/empty series.
        """
        # Exclude calculated series from analysis
        CALCULATED_FDIC_SERIES = ['Total_Capital', 'AOBS']
        fields_to_analyze = [f for f in FDIC_FIELDS_TO_FETCH if f not in CALCULATED_FDIC_SERIES]

        if fdic_df.empty:
            return {"empty_series": fields_to_analyze, "sparse_series": [], "available_series": []}

        analysis_report = {
            "empty_series": [],      # Completely missing or all NaN
            "sparse_series": [],     # Present but >75% missing data
            "available_series": []   # Present with reasonable data coverage
        }

        total_records = len(fdic_df)

        for field in fields_to_analyze:
            if field not in fdic_df.columns:
                analysis_report["empty_series"].append(field)
            else:
                non_null_count = fdic_df[field].notna().sum()
                coverage_pct = (non_null_count / total_records) * 100

                if coverage_pct == 0:
                    analysis_report["empty_series"].append(field)
                elif coverage_pct < 25:
                    analysis_report["sparse_series"].append({
                        "field": field,
                        "coverage_pct": round(coverage_pct, 1),
                        "description": FDIC_FIELD_DESCRIPTIONS.get(field, {}).get("short", "Unknown")
                    })
                else:
                    analysis_report["available_series"].append({
                        "field": field,
                        "coverage_pct": round(coverage_pct, 1)
                    })

        return analysis_report
    def _optimize_df_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Reduces DataFrame memory usage and file size by downcasting numeric types
        and converting object columns with low cardinality to 'category'.
        """
        for col in df.columns:
            if df[col].dtype == 'float64':
                df[col] = pd.to_numeric(df[col], downcast='float')
            elif df[col].dtype == 'int64':
                df[col] = pd.to_numeric(df[col], downcast='integer')
            elif df[col].dtype == 'object':
                if len(df[col].unique()) / len(df[col]) < 0.5:
                    df[col] = df[col].astype('category')
        return df
    def _create_peer_composite(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Creates TWO peer composites:
        1. Original peer composite (CERT 99999) - includes all peers
        2. Selective peer composite (CERT 99998) - excludes Flagstar (32541) and Valley (9396)
        """
        # Original peer composite (all peers)
        peer_df_all = df[df['CERT'].isin(self.config.peer_bank_certs)]
        if not peer_df_all.empty:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", FutureWarning)
                numeric_cols = df.select_dtypes(include=np.number).columns.drop('CERT', errors='ignore')
                peer_avg_all = peer_df_all.groupby('REPDTE')[numeric_cols].mean().reset_index()
                peer_avg_all['CERT'] = 99999
                peer_avg_all['NAME'] = "Peers' Average"
                df = pd.concat([df, peer_avg_all], ignore_index=True)

        # Selective peer composite (excluding Flagstar and Valley)
        peers_excl_flagstar_valley = [cert for cert in self.config.peer_bank_certs if cert not in [32541, 9396]]
        peer_df_selective = df[df['CERT'].isin(peers_excl_flagstar_valley)]
        if not peer_df_selective.empty:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", FutureWarning)
                numeric_cols = df.select_dtypes(include=np.number).columns.drop('CERT', errors='ignore')
                peer_avg_selective = peer_df_selective.groupby('REPDTE')[numeric_cols].mean().reset_index()
                peer_avg_selective['CERT'] = 99998
                peer_avg_selective['NAME'] = "Peers' Average (Ex. Flagstar & Valley)"
                df = pd.concat([df, peer_avg_selective], ignore_index=True)

        return df

    def run(self) -> Dict[str, Any]:
        logging.info("Starting dashboard generation...")
        fdic_df, _ = self.fdic_fetcher.fetch_all_banks()
        if fdic_df.empty: raise ValueError("No FDIC data retrieved.")

        all_certs = fdic_df['CERT'].unique().tolist()
        locations_df = get_bank_locations(all_certs)
        if 'NAME' in locations_df.columns:
            locations_df.drop(columns=['NAME'], inplace=True)
        fdic_df = pd.merge(fdic_df, locations_df, on='CERT', how='left')
        logging.info("Location data merged successfully.")

        fdic_analysis = self._analyze_fdic_data_availability(fdic_df)

        proc_df = self.processor.create_derived_metrics(fdic_df)
        proc_df_with_ttm = self.processor.calculate_ttm_metrics(proc_df)
        proc_df_with_peers = self._create_peer_composite(proc_df_with_ttm)

        peer_comp_df = self.analyzer.create_peer_comparison(proc_df_with_peers)
        snapshot_df = self.processor.create_latest_snapshot(proc_df_with_peers)
        avg_8q_all_metrics_df = self.processor.calculate_8q_averages(proc_df_with_peers)

        fred_df, fred_desc_df, failed_fred_series = self.fred_fetcher.fetch_all_series()
        if 'date' in fred_df.columns:
            fred_df.set_index('date', inplace=True)

        if 'SOFR' in fred_df.columns and 'TB3MS' in fred_df.columns:
            # --- Use non-inplace assignment to avoid FutureWarning ---
            fred_df['SOFR'] = fred_df['SOFR'].ffill()
            fred_df['TB3MS'] = fred_df['TB3MS'].ffill()
            fred_df['SOFR3MTB3M'] = fred_df['SOFR'] - fred_df['TB3MS']
            if fred_desc_df is not None and fred_desc_df[fred_desc_df['Series ID'] == 'SOFR3MTB3M'].empty:
                 new_row_data = {'Series ID': 'SOFR3MTB3M', 'Category': 'Middle Market, Healthcare, & Funding Indicators', 'short': 'SOFR vs T-Bill Spread', 'long': 'Calculated Spread: SOFR minus 3-Month T-Bill Rate'}
                 new_row = pd.DataFrame([new_row_data])
                 fred_desc_df = pd.concat([fred_desc_df, new_row], ignore_index=True)

        enhanced_analyzer = MacroTrendAnalyzer(self.config)
        processed_data = {}
        validation_reports = {}

        # Define calculated series that should not be fetched from FRED
        CALCULATED_SERIES = ['SOFR3MTB3M']
        CALCULATED_FDIC_SERIES = ['Total_Capital','AOBS']

        for category, series_in_category in FRED_SERIES_TO_FETCH.items():
            for series_id in series_in_category.keys():
                # Skip validation and processing for calculated series until after they're created
                if series_id in CALCULATED_SERIES:
                    continue

                if series_id in fred_df.columns:
                    validation_reports[series_id] = enhanced_analyzer.validate_series_data(
                        series_id, fred_df[series_id]
                    )
                    if validation_reports[series_id]['status'] != 'Error':
                        processed_data[series_id] = enhanced_analyzer.calculate_technical_indicators(
                            fred_df[series_id], series_id
                        )

        # Now process calculated series after they've been created
        for series_id in CALCULATED_SERIES:
            if series_id in fred_df.columns:
                validation_reports[series_id] = enhanced_analyzer.validate_series_data(
                    series_id, fred_df[series_id]
                )
                if validation_reports[series_id]['status'] != 'Error':
                    processed_data[series_id] = enhanced_analyzer.calculate_technical_indicators(
                        fred_df[series_id], series_id
                    )

        validation_df = pd.DataFrame.from_dict(validation_reports, orient='index')

        powerbi_macro_df = enhanced_analyzer.generate_powerbi_output(processed_data)

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f"{self.config.output_dir}/Bank_Performance_Dashboard_{ts}.xlsx"
        peer_comp_df = self._optimize_df_dtypes(peer_comp_df)
        snapshot_df = self._optimize_df_dtypes(snapshot_df)
        proc_df_with_peers = self._optimize_df_dtypes(proc_df_with_peers)
        avg_8q_all_metrics_df = self._optimize_df_dtypes(avg_8q_all_metrics_df)
        powerbi_macro_df = self._optimize_df_dtypes(powerbi_macro_df)
        self.output_gen.write_excel_output(
            file_path=fname,
            Summary_Dashboard=peer_comp_df,
            Latest_Peer_Snapshot=snapshot_df,
            Averages_8Q_All_Metrics=avg_8q_all_metrics_df,
            FDIC_Metric_Descriptions=self.output_gen.fdic_desc_df,
            Macro_Analysis=powerbi_macro_df,
            FDIC_Data=proc_df_with_peers,
            FRED_Data=fred_df.reset_index(), # Pass dataframe with date as column
            FRED_Descriptions=fred_desc_df,
            Data_Validation_Report=validation_df,
        )
        # === ENHANCED DIAGNOSTIC: Inspect Critical FDIC Data for Subject Bank ===
        print("\n" + "="*80)
        print("CRITICAL FDIC DATA INSPECTION FOR YOUR BANK")
        print("="*80)

        try:
            # Get the latest data for your bank
            subject_bank_data = proc_df_with_peers[
                proc_df_with_peers['CERT'] == self.config.subject_bank_cert
            ].sort_values('REPDTE')

            if not subject_bank_data.empty:
                latest_record = subject_bank_data.iloc[-1]
                bank_name = latest_record.get('NAME', 'Unknown Bank')
                report_date = latest_record.get('REPDTE', 'Unknown Date')

                print(f"Bank: {bank_name} (CERT: {self.config.subject_bank_cert})")
                print(f"Latest Report Date: {report_date}")
                print("-" * 80)

                # Critical columns to inspect
                critical_columns = {
                    'ASSET': 'Total Assets',
                    'EQ': 'Total Equity Capital',
                    'LNLS': 'Gross Loans & Leases',
                    'LNCI': 'C&I Loans',
                    'LNRENROW': 'Owner-Occ CRE',
                    'LNRECONS': 'Construction & Development',
                    'LNRERES': '1-4 Family Residential',
                    'LNRENROT': 'Nonfarm CRE '
                }

                print("CRITICAL BALANCE SHEET ITEMS:")
                for col_code, description in critical_columns.items():
                    value = latest_record.get(col_code)
                    if pd.isna(value) or value is None:
                        print(f"  {col_code:12} ({description:30}): *** MISSING DATA ***")
                    elif value == 0:
                        print(f"  {col_code:12} ({description:30}): *** ZERO VALUE ***")
                    else:
                        # Format as currency in thousands
                        formatted_value = f"${value:,.0f}"
                        print(f"  {col_code:12} ({description:30}): {formatted_value}")

                # === NEW: Investigate C&I Loans specifically ===
                print("\n" + "="*60)
                print("C&I LOANS INVESTIGATION")
                print("="*60)

                # Check if LNCI exists in the raw data at all
                if 'LNCI' not in subject_bank_data.columns:
                    print("❌ LNCI column does not exist in the fetched data")
                    print("   This suggests FDIC API did not return this field")
                else:
                    print("✓ LNCI column exists in the data")

                    # Check LNCI values across all quarters
                    lnci_summary = subject_bank_data['LNCI'].describe()
                    print(f"   LNCI Data Summary across all quarters:")
                    print(f"     Count (non-null): {subject_bank_data['LNCI'].notna().sum()}")
                    print(f"     Count (null):     {subject_bank_data['LNCI'].isna().sum()}")
                    print(f"     Count (zero):     {(subject_bank_data['LNCI'] == 0).sum()}")
                    if subject_bank_data['LNCI'].notna().sum() > 0:
                        print(f"     Min value:        ${subject_bank_data['LNCI'].min():,.0f}")
                        print(f"     Max value:        ${subject_bank_data['LNCI'].max():,.0f}")
                        print(f"     Latest non-null:  ${subject_bank_data['LNCI'].dropna().iloc[-1] if len(subject_bank_data['LNCI'].dropna()) > 0 else 0:,.0f}")

                # Check alternative C&I loan fields that might be available
                print("\n   Checking for alternative C&I loan series:")
                ci_alternatives = {
                    'LNCI': 'Commercial and Industrial Loans',
                    'LNCIDOM': 'C&I Loans - Domestic',
                    'LNCINUM': 'C&I Loans - Number of Loans',
                    'LNCON': 'Consumer Loans (sometimes includes C&I)',
                    'BUSLOANS': 'Total Business Loans (broader category)'
                }

                for field, description in ci_alternatives.items():
                    if field in subject_bank_data.columns:
                        latest_val = latest_record.get(field)
                        if pd.notna(latest_val) and latest_val != 0:
                            print(f"   ✓ {field:10}: {description:35} = ${latest_val:,.0f}")
                        else:
                            print(f"   ⚠ {field:10}: {description:35} = Missing/Zero")
                    else:
                        print(f"   ❌ {field:10}: {description:35} = Not available")

                # Check what fields ARE available for this bank
                print(f"\n   Available loan fields for your bank:")
                loan_fields = [col for col in subject_bank_data.columns if col.startswith('LN') and latest_record.get(col, 0) > 0]
                for field in sorted(loan_fields):
                    value = latest_record.get(field, 0)
                    print(f"     {field}: ${value:,.0f}")

                # Check if this is a reporting threshold issue
                total_assets = latest_record.get('ASSET', 0)
                print(f"\n   Bank Size Analysis:")
                print(f"     Total Assets: ${total_assets:,.0f}")
                if total_assets < 300_000:  # Less than $300M
                    print("     ⚠ Small bank - may have different reporting requirements")
                elif total_assets < 1_000_000:  # Less than $1B
                    print("     ℹ Community bank - standard reporting applies")
                else:
                    print("     ℹ Large bank - full reporting requirements")

                # Calculate what C&I might be if it's combined with other categories
                total_loans = latest_record.get('LNLS', 0)
                known_loans = sum([
                    latest_record.get('LNRENROW', 0),  # CRE Non-Owner
                    latest_record.get('LNRECONS', 0),  # Construction
                    latest_record.get('LNRERES', 0),   # Residential
                    latest_record.get('LNRENROT', 0),  # CRE Owner-Occ
                    latest_record.get('LNRELOC', 0),   # HELOC
                    latest_record.get('LNCRCD', 0),    # Credit Cards
                    latest_record.get('LNAUTO', 0),    # Auto
                    latest_record.get('LNOTHER', 0),   # Other
                    latest_record.get('LS', 0),        # Leases
                ])

                implied_ci = total_loans - known_loans
                print(f"\n   Implied C&I Calculation:")
                print(f"     Total Loans:        ${total_loans:,.0f}")
                print(f"     Known Categories:   ${known_loans:,.0f}")
                print(f"     Implied C&I:        ${implied_ci:,.0f}")
                print(f"     Percentage of Total: {(implied_ci/total_loans)*100:.1f}%" if total_loans > 0 else "     Percentage: N/A")

                # Rest of your existing diagnostic code...
                print("\n" + "="*60)
                print("OTHER CALCULATED FIELDS")
                print("="*60)
                calc_fields = {
                    'Total_Capital': 'Total Capital (T1 + T2)',
                    'AOBS': 'Off-Balance Sheet Allowance',
                    'Total_ACL': 'Total Allowance for Credit Losses'
                }

                for col_code, description in calc_fields.items():
                    value = latest_record.get(col_code)
                    if pd.isna(value) or value is None:
                        print(f"  {col_code:15} ({description:30}): *** CALCULATION FAILED ***")
                    else:
                        formatted_value = f"${value:,.0f}"
                        print(f"  {col_code:15} ({description:30}): {formatted_value}")

            else:
                print(f"❌ NO DATA FOUND for your bank (CERT: {self.config.subject_bank_cert})")

        except Exception as e:
            print(f"❌ ERROR during diagnostic inspection: {e}")

        print("="*80)
        return {
            "output_file": fname,
            "powerbi_macro_df": powerbi_macro_df,
            "fred_desc_df": fred_desc_df,
            "failed_fred_series": failed_fred_series,
            "fdic_analysis": fdic_analysis,
        }



def load_config() -> DashboardConfig:
    """Loads configuration from a hardcoded source."""
    return DashboardConfig(
        fred_api_key="de4ecbcab88d5d5c5c330d772c08bcfe",
        subject_bank_cert=19977,
        peer_bank_certs=[26876,9396,18221,16068,22953,57919,20234,58647,26610,32541, 32172,24045]
    )


def main():
    run_results = {}
    try:
        config = load_config()
        dash = BankPerformanceDashboard(config)
        run_results = dash.run()
        print("\n" + "="*80)
        print("DASHBOARD GENERATION COMPLETE")
        print(f"Output file located at: {run_results.get('output_file', 'N/A')}")
        print("="*80)

        print("\n--- FDIC DATA AVAILABILITY REPORT ---")
        fdic_analysis = run_results.get('fdic_analysis', {})

        empty_series = fdic_analysis.get('empty_series', [])
        sparse_series = fdic_analysis.get('sparse_series', [])
        available_series = fdic_analysis.get('available_series', [])

        if empty_series:
            print(f"\n❌ EMPTY FDIC SERIES ({len(empty_series)} series with no data):")
            for series in empty_series:
                description = FDIC_FIELD_DESCRIPTIONS.get(series, {}).get("short", "Unknown")
                print(f"  - {series}: {description}")
        else:
            print("\n✅ All FDIC series returned data.")

        if sparse_series:
            print(f"\n⚠️  SPARSE FDIC SERIES ({len(sparse_series)} series with limited data):")
            for item in sparse_series:
                print(f"  - {item['field']}: {item['description']} ({item['coverage_pct']}% coverage)")

        print(f"\n✅ AVAILABLE FDIC SERIES: {len(available_series)} series with good data coverage")

        total_requested = len(FDIC_FIELDS_TO_FETCH)
        total_available = len(available_series)
        total_sparse = len(sparse_series)
        total_empty = len(empty_series)

        print(f"\nSUMMARY:")
        print(f"  Total FDIC series requested: {total_requested}")
        print(f"  Available (good coverage):   {total_available} ({(total_available/total_requested)*100:.1f}%)")
        print(f"  Sparse (limited coverage):   {total_sparse} ({(total_sparse/total_requested)*100:.1f}%)")
        print(f"  Empty (no data):             {total_empty} ({(total_empty/total_requested)*100:.1f}%)")

        print("\n--- POWER BI INTEGRATION ---")
        try:
            macro_df = run_results.get('powerbi_macro_df')
            desc_df = run_results.get('fred_desc_df')
            if macro_df is not None and not macro_df.empty and desc_df is not None:
                powerbi_generator = PowerBIGenerator()
                powerbi_generator.generate_powerbi_setup_file(
                    output_dir=config.output_dir,
                    powerbi_macro_df=macro_df,
                    fred_desc_df=desc_df
                )
                powerbi_setup_file = Path(config.output_dir) / "PowerBI_Setup.txt"
                print(f"[+] Power BI setup file generated successfully.")
                print(f"    -> Location: {powerbi_setup_file.resolve()}")
            else:
                print("[-] Power BI setup file was not generated because macro data was missing.")

        except Exception as e:
            logger.error(f"Failed to generate Power BI setup file: {e}", exc_info=True)
            print("[-] Power BI setup file was not generated due to an unexpected error.")

    except Exception as e:
        logger.error(f"Application failed to complete: {e}", exc_info=True)
        sys.exit(1)

    finally:
        print("\n--- DATA SERIES VERIFICATION ---")
        failed_series = run_results.get("failed_fred_series", [])
        if failed_series:
            print("\nWARNING: The following FRED series failed to fetch and are NOT in the output:")
            for series in failed_series:
                print(f"  - {series}")
        else:
            print("\nAll FRED series fetched successfully.")

        logging.shutdown()

if __name__ == '__main__':
    main()
import matplotlib.pyplot as plt

import pandas as pd

import seaborn as sns

import pandas as pd


# Set style for better-looking charts
plt.style.use('seaborn-v0_8-whitegrid')
sns.set_palette("husl")

# ==================================================================================
# UPDATED CHART FUNCTION WITH NEW LABELING AND COLORS
# ==================================================================================

def create_credit_deterioration_chart_v3(proc_df_with_peers, subject_bank_cert=19977, show_both_peer_groups=True):
    """
    Creates a chart comparing TTM NCO Rate and NPL-to-Gross Loans Rate
    Updated with reversed colors and every 4 periods labeling from latest backward
    """

    # Define entities to plot based on options
    entities_to_plot = [subject_bank_cert]  # Always include subject bank

    if show_both_peer_groups:
        entities_to_plot.extend([99999, 99998])  # Both peer groups
        # REVERSED COLORS
        colors = {
            subject_bank_cert: '#FFA500',  # Orange for IDB (was blue)
            99999: '#5B9BD5',              # Blue for all peers (was orange)
            99998: '#70AD47'               # Green for selective peers (same)
        }
        entity_names = {
            subject_bank_cert: "IDB",
            99999: "All Peers",
            99998: "Peers (Ex. F&V)"
        }
    else:
        entities_to_plot.append(99998)  # Only selective peer group
        # REVERSED COLORS
        colors = {
            subject_bank_cert: '#FFA500',  # Orange for IDB (was blue)
            99998: '#5B9BD5'               # Blue for selective peers (was orange)
        }
        entity_names = {
            subject_bank_cert: "IDB",
            99998: "Peers (Ex. F&V)"
        }

    chart_df = proc_df_with_peers[proc_df_with_peers['CERT'].isin(entities_to_plot)].copy()

    if chart_df.empty:
        print("No data found for charting")
        return None, None

    # Sort by date
    chart_df = chart_df.sort_values('REPDTE')

    # Filter for dates after Q3 2019 (2019-09-30)
    chart_df = chart_df[chart_df['REPDTE'] >= '2019-10-01']

    if chart_df.empty:
        print("No data found after Q3 2019")
        return None, None

    # Create period labels - show Q#-YY format
    chart_df['Quarter'] = chart_df['REPDTE'].dt.quarter
    chart_df['Year'] = chart_df['REPDTE'].dt.year
    chart_df['Period_Label'] = 'Q' + chart_df['Quarter'].astype(str) + '-' + (chart_df['Year'] % 100).astype(str).str.zfill(2)

    # NEW LABELING LOGIC: Show every 4 periods starting from latest going backward
    subject_data = chart_df[chart_df['CERT'] == subject_bank_cert].reset_index(drop=True)
    num_periods = len(subject_data)

    # Create mask for showing labels - every 4 periods from the end
    show_label_mask = [False] * num_periods
    for i in range(num_periods - 1, -1, -4):  # Start from latest, go backward by 4
        show_label_mask[i] = True

    # Also show Q4 and Q2 periods regardless
    for i, row in subject_data.iterrows():
        quarter = row['Quarter']
        if quarter in [2, 4]:  # Q2 and Q4
            show_label_mask[i] = True

    # Apply the mask to all entities
    chart_df['Show_Label'] = False  # Initialize all as False
    for cert in entities_to_plot:
        cert_data = chart_df[chart_df['CERT'] == cert].reset_index(drop=True)
        if len(cert_data) == len(show_label_mask):
            # Get the original indices for this cert
            cert_indices = chart_df[chart_df['CERT'] == cert].index
            chart_df.loc[cert_indices, 'Show_Label'] = show_label_mask

    # Create the chart
    fig, ax = plt.subplots(figsize=(16, 8))

    # Get the subject bank data for x-axis positioning
    x_positions = np.arange(len(subject_data))

    # Calculate bar width and positions based on number of entities
    num_entities = len(entities_to_plot)
    bar_width = 0.8 / num_entities
    bar_positions = {}

    for i, cert in enumerate(entities_to_plot):
        offset = (i - (num_entities - 1) / 2) * bar_width
        bar_positions[cert] = x_positions + offset

    # Plot TTM NCO Rate (bars)
    for cert in entities_to_plot:
        entity_data = chart_df[chart_df['CERT'] == cert].reset_index(drop=True)
        if not entity_data.empty:
            label = entity_names.get(cert, f"CERT {cert}")

            # Plot TTM NCO Rate as bars
            nco_rate = entity_data['TTM_NCO_Rate'].fillna(0) / 100  # Convert to decimal

            bars = ax.bar(bar_positions[cert], nco_rate, alpha=0.7, color=colors[cert],
                         label=f'{label} TTM NCO Rate', width=bar_width)

    # Create second y-axis for NPL rates (lines)
    ax2 = ax.twinx()

    # Plot NPL-to-Gross Loans Rate (lines)
    line_styles = ['-', '--', '-.']  # Different line styles for each entity
    for i, cert in enumerate(entities_to_plot):
        entity_data = chart_df[chart_df['CERT'] == cert].reset_index(drop=True)
        if not entity_data.empty:
            label = entity_names.get(cert, f"CERT {cert}")

            # Plot NPL rate as line
            npl_rate = entity_data['NPL_to_Gross_Loans_Rate'].fillna(0) / 100  # Convert to decimal

            line_style = line_styles[i % len(line_styles)]
            ax2.plot(x_positions, npl_rate, color=colors[cert],
                    linestyle=line_style, marker='o', linewidth=2.5,
                    label=f'{label} NPL-to-Book', markersize=5)

    # Format the chart
    ax.set_xlabel('Reporting Period', fontsize=12, fontweight='bold')
    ax.set_ylabel('TTM NCO Rate', fontsize=12, fontweight='bold', color='black')
    ax2.set_ylabel('NPL-to-Gross Loans Rate', fontsize=12, fontweight='bold', color='black')

    # Format y-axes as percentages
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: '{:.2%}'.format(y)))
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: '{:.2%}'.format(y)))

    # Set x-axis labels - only show labels where Show_Label is True
    ax.set_xticks(x_positions)

    # Create labels array - empty string for periods we don't want to show
    labels = []
    for i, (_, row) in enumerate(subject_data.iterrows()):
        if show_label_mask[i]:
            labels.append(row['Period_Label'])
        else:
            labels.append('')  # Empty string for hidden labels

    ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=10)

    # Add value labels on bars and points (only for labeled periods)
    for cert in entities_to_plot:
        entity_data = chart_df[chart_df['CERT'] == cert].reset_index(drop=True)
        if not entity_data.empty:
            # NCO Rate labels (bars) - only for shown periods
            nco_rate = entity_data['TTM_NCO_Rate'].fillna(0) / 100
            entity_show_labels = entity_data['Show_Label'].values if 'Show_Label' in entity_data.columns else [False] * len(entity_data)

            for i, rate in enumerate(nco_rate):
                if i < len(entity_show_labels) and entity_show_labels[i] and not np.isnan(rate) and rate > 0:
                    ax.text(bar_positions[cert][i], rate + 0.0005, f'{rate:.2%}',
                           ha='center', va='bottom', fontsize=8, fontweight='bold')

            # NPL Rate labels (lines) - only for shown periods and only for IDB to avoid clutter
            if cert == subject_bank_cert:
                npl_rate = entity_data['NPL_to_Gross_Loans_Rate'].fillna(0) / 100
                for i, rate in enumerate(npl_rate):
                    if i < len(entity_show_labels) and entity_show_labels[i] and not np.isnan(rate) and rate > 0:
                        ax2.text(i, rate + 0.0005, f'{rate:.2%}',
                                ha='center', va='bottom', fontsize=8, fontweight='bold',
                                bbox=dict(boxstyle='round,pad=0.2', facecolor='white', alpha=0.7))

    # Combine legends
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, loc='upper left', frameon=True, fancybox=True)

    # Set title
    title = 'A Comparative Look at Credit Deterioration\nTTM NCO Rate (bars) vs NPL-to-Book (lines)'
    if show_both_peer_groups:
        title += '\n(F&V = Flagstar & Valley excluded from green series)'
    plt.title(title, fontsize=14, fontweight='bold', pad=20)

    # Adjust layout
    plt.tight_layout()

    # Set grid
    ax.grid(True, alpha=0.3)
    ax2.grid(False)

    # Improve spacing
    plt.subplots_adjust(bottom=0.15)

    return fig, ax

# ==================================================================================
# EMAIL TABLE GENERATION FUNCTION
# ==================================================================================

def generate_credit_metrics_email_table(proc_df_with_peers, subject_bank_cert=19977):
    """
    Generates an HTML table for email with 10 most important credit metrics
    comparing IDB vs All Peers vs Selective Peers (latest quarter)
    """

    # Define the 10 most important credit metrics
    important_metrics = {
        'TTM_NCO_Rate': 'TTM NCO Rate (%)',
        'NPL_to_Gross_Loans_Rate': 'NPL to Gross Loans (%)',
        'TTM_Past_Due_Rate': 'TTM Past Due Rate (%)',
        'CRE_Concentration_Capital_Risk': 'CRE Concentration Risk (%)',
        'IDB_CRE_Growth_TTM': 'CRE Growth TTM (%)',
        'Allowance_to_Gross_Loans_Rate': 'Allowance to Gross Loans (%)',
    }

    # Get latest quarter data
    latest_date = proc_df_with_peers['REPDTE'].max()
    latest_data = proc_df_with_peers[proc_df_with_peers['REPDTE'] == latest_date]

    # Get data for each entity
    idb_data = latest_data[latest_data['CERT'] == subject_bank_cert]
    all_peers_data = latest_data[latest_data['CERT'] == 99999]
    selective_peers_data = latest_data[latest_data['CERT'] == 99998]

    if idb_data.empty or all_peers_data.empty or selective_peers_data.empty:
        print("Missing data for one or more entities")
        return None

    # Create the comparison table
    table_data = []

    for metric_code, metric_name in important_metrics.items():
        if metric_code in idb_data.columns:
            idb_value = idb_data[metric_code].iloc[0]
            all_peers_value = all_peers_data[metric_code].iloc[0] if metric_code in all_peers_data.columns else np.nan
            selective_peers_value = selective_peers_data[metric_code].iloc[0] if metric_code in selective_peers_data.columns else np.nan

            # Calculate differences
            diff_vs_all = idb_value - all_peers_value if not np.isnan(all_peers_value) else np.nan
            diff_vs_selective = idb_value - selective_peers_value if not np.isnan(selective_peers_value) else np.nan

            table_data.append({
                'Metric': metric_name,
                'IDB': f"{idb_value:.2f}%" if not np.isnan(idb_value) else "N/A",
                'All Peers': f"{all_peers_value:.2f}%" if not np.isnan(all_peers_value) else "N/A",
                'Selective Peers': f"{selective_peers_value:.2f}%" if not np.isnan(selective_peers_value) else "N/A",
                'Diff vs All': f"{diff_vs_all:+.2f}%" if not np.isnan(diff_vs_all) else "N/A",
                'Diff vs Selective': f"{diff_vs_selective:+.2f}%" if not np.isnan(diff_vs_selective) else "N/A"
            })

    # Convert to DataFrame for easier manipulation
    df = pd.DataFrame(table_data)

    # Generate HTML table
    html_table = generate_html_email_table(df, latest_date)

    return html_table, df

def generate_html_email_table(df, report_date):
    """
    Generates a properly formatted HTML table for email
    """

    # Format the report date
    formatted_date = report_date.strftime('%B %d, %Y') if hasattr(report_date, 'strftime') else str(report_date)

    html = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 20px;
                background-color: #f5f5f5;
            }}
            .email-container {{
                background-color: white;
                padding: 30px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                max-width: 900px;
                margin: 0 auto;
            }}
            .header {{
                text-align: center;
                margin-bottom: 30px;
                color: #2c3e50;
            }}
            .header h2 {{
                margin: 0;
                font-size: 24px;
                font-weight: 600;
            }}
            .header p {{
                margin: 5px 0 0 0;
                color: #7f8c8d;
                font-size: 14px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
                font-size: 12px;
            }}
            th {{
                background-color: #34495e;
                color: white;
                padding: 12px 8px;
                text-align: center;
                font-weight: 600;
                border: 1px solid #2c3e50;
            }}
            td {{
                padding: 10px 8px;
                text-align: center;
                border: 1px solid #bdc3c7;
            }}
            tr:nth-child(even) {{
                background-color: #f8f9fa;
            }}
            tr:hover {{
                background-color: #e3f2fd;
            }}
            .metric-name {{
                text-align: left !important;
                font-weight: 500;
                color: #2c3e50;
            }}
            .idb-value {{
                background-color: #fff3cd;
                font-weight: 600;
                color: #856404;
            }}
            .positive {{
                color: #d32f2f;
                font-weight: 600;
            }}
            .negative {{
                color: #388e3c;
                font-weight: 600;
            }}
            .neutral {{
                color: #5d4037;
            }}
            .footer {{
                margin-top: 30px;
                padding-top: 20px;
                border-top: 2px solid #ecf0f1;
                font-size: 12px;
                color: #7f8c8d;
                text-align: center;
            }}
            .legend {{
                margin: 20px 0;
                padding: 15px;
                background-color: #f8f9fa;
                border-radius: 5px;
                font-size: 11px;
                color: #495057;
            }}
            .legend strong {{
                color: #2c3e50;
            }}
        </style>
    </head>
    <body>
        <div class="email-container">
            <div class="header">
                <h2>Credit Risk Metrics Comparison</h2>
                <p>IDB vs Peer Groups Analysis | Report Date: {formatted_date}</p>
            </div>

            <table>
                <thead>
                    <tr>
                        <th style="width: 30%;">Credit Metric</th>
                        <th style="width: 12%;">IDB</th>
                        <th style="width: 12%;">All Peers</th>
                        <th style="width: 15%;">Selective Peers<br><small>(Ex. Flagstar & Valley)</small></th>
                        <th style="width: 15%;">Diff vs<br>All Peers</th>
                        <th style="width: 16%;">Diff vs<br>Selective Peers</th>
                    </tr>
                </thead>
                <tbody>
    """

    # Add table rows
    for _, row in df.iterrows():
        # Determine styling for difference columns
        diff_all_class = get_diff_class(row['Diff vs All'])
        diff_selective_class = get_diff_class(row['Diff vs Selective'])

        html += f"""
                    <tr>
                        <td class="metric-name">{row['Metric']}</td>
                        <td class="idb-value">{row['IDB']}</td>
                        <td>{row['All Peers']}</td>
                        <td>{row['Selective Peers']}</td>
                        <td class="{diff_all_class}">{row['Diff vs All']}</td>
                        <td class="{diff_selective_class}">{row['Diff vs Selective']}</td>
                    </tr>
        """

    html += """
                </tbody>
            </table>

            <div class="legend">
                <strong>Legend:</strong><br>
                • <span style="color: #d32f2f; font-weight: 600;">Red values</span>: IDB performs worse than peer group (higher is worse for most credit metrics)<br>
                • <span style="color: #388e3c; font-weight: 600;">Green values</span>: IDB performs better than peer group (lower is better for most credit metrics)<br>
                • <strong>Selective Peers</strong>: Excludes Flagstar Bank and Valley National Bank for cleaner comparison<br>
                • All values are as of the latest available quarter
            </div>

            <div class="footer">
                <p>This report is generated automatically from FDIC call report data.<br>
                For questions or additional analysis, please contact the Credit Risk team.</p>
            </div>
        </div>
    </body>
    </html>
    """

    return html

def get_diff_class(diff_str):
    """
    Determines CSS class based on difference value
    """
    if diff_str == "N/A":
        return "neutral"

    try:
        # Remove % sign and convert to float
        diff_value = float(diff_str.replace('%', '').replace('+', ''))

        if diff_value > 0.05:  # More than 5 basis points worse
            return "positive"  # Red (worse performance)
        elif diff_value < -0.05:  # More than 5 basis points better
            return "negative"  # Green (better performance)
        else:
            return "neutral"   # Brown (neutral)
    except:
        return "neutral"

# ==================================================================================
# COMPLETE IMPLEMENTATION FUNCTION
# ==================================================================================

def create_complete_credit_analysis(show_both_peer_groups=True, save_table_file=True):
    """
    Complete function that creates both chart and email table
    """
    # Load configuration
    config = load_config()

    # Create dashboard instance
    dashboard = BankPerformanceDashboard(config)

    # Run dashboard and get results
    results = dashboard.run()

    # Load the data from the Excel file that was created
    excel_file = results["output_file"]

    try:
        # Read the FDIC_Data sheet
        proc_df_with_peers = pd.read_excel(excel_file, sheet_name='FDIC_Data')

        # Convert REPDTE back to datetime if it's not already
        if 'REPDTE' in proc_df_with_peers.columns:
            proc_df_with_peers['REPDTE'] = pd.to_datetime(proc_df_with_peers['REPDTE'])

        # Create the chart
        print("Creating credit deterioration chart...")
        fig, ax = create_credit_deterioration_chart_v3(
            proc_df_with_peers,
            config.subject_bank_cert,
            show_both_peer_groups=show_both_peer_groups
        )

        # Generate email table
        print("Generating email table...")
        html_table, table_df = generate_credit_metrics_email_table(
            proc_df_with_peers,
            config.subject_bank_cert
        )

        # Display the chart
        plt.show()

        # Save files
        base_filename = excel_file.replace('.xlsx', '')

        # Save chart
        chart_filename = f"{base_filename}_credit_chart_v3.png"
        plt.savefig(chart_filename, dpi=300, bbox_inches='tight')
        print(f"Chart saved as: {chart_filename}")

        # Save HTML table
        if save_table_file:
            table_filename = f"{base_filename}_credit_table.html"
            with open(table_filename, 'w', encoding='utf-8') as f:
                f.write(html_table)
            print(f"Email table saved as: {table_filename}")

        # Print table for copying
        print("\n" + "="*80)
        print("EMAIL TABLE (Copy the HTML below for email):")
        print("="*80)
        print(html_table)
        print("="*80)

        # Also print summary
        print(f"\nSUMMARY TABLE (DataFrame):")
        print(table_df.to_string(index=False))

        return fig, ax, html_table, table_df, proc_df_with_peers

    except Exception as e:
        print(f"Error creating analysis: {e}")
        return None, None, None, None, None

# ==================================================================================
# USAGE INSTRUCTIONS
# ==================================================================================

print("""
UPDATED USAGE:

1. For complete analysis with both chart and email table:
   fig, ax, html_table, table_df, data = create_complete_credit_analysis(show_both_peer_groups=True)

2. The function will:
   - Create the updated chart with reversed colors
   - Generate an HTML table ready for email
   - Save both chart and table files
   - Print the HTML for copying

3. Chart updates:
   - Colors reversed: IDB is now orange, peers are blue/green
   - Labels show every 4 periods from latest backward
   - Also shows Q2 and Q4 regardless of the 4-period rule

4. Email table features:
   - Professional HTML formatting
   - 10 most important credit metrics
   - Comparison against both peer groups
   - Color-coded differences (red=worse, green=better)
   - Ready to copy into email
""")
fig, ax, html_table, table_df, data = create_complete_credit_analysis(show_both_peer_groups=True)
# After running the complete analysis

#%%


# Load your data
config = load_config()
dashboard = BankPerformanceDashboard(config)
results = dashboard.run()
excel_file = results["output_file"]
proc_df_with_peers = pd.read_excel(excel_file, sheet_name='FDIC_Data')
proc_df_with_peers['REPDTE'] = pd.to_datetime(proc_df_with_peers['REPDTE'])

# Create chart directly
fig, ax = create_credit_deterioration_chart_v3(
    proc_df_with_peers,
    config.subject_bank_cert,
    show_both_peer_groups=True
)

# Check if chart has data
if fig is not None:
    plt.figure(fig.number)
    plt.show()
    plt.savefig('credit_chart_direct.png', dpi=300, bbox_inches='tight')
    print("Chart created successfully")
else:
    print("Chart creation failed - check your data")

# Then save with explicit parameters
plt.savefig('credit_chart_fixed.png',
            dpi=300,
            bbox_inches='tight',
            facecolor='white',
            edgecolor='black',
            format='png')

print("Chart should now be visible and saved properly")

#%%
import pandas as pd
import numpy as np
from typing import Dict, List, Any

# This helper function determines the color for the difference cell.
def _get_cell_class(value: float, metric_code: str, metric_descriptions: Dict[str, Dict[str, str]]) -> str:
    """
    Determines CSS class for a cell based on its value and metric type.

    Args:
        value (float): The value to evaluate.
        metric_code (str): The code of the metric (e.g., 'TTM_NCO_Rate').
        metric_descriptions (Dict): A dictionary of metric metadata.

    Returns:
        str: A CSS class ('positive' for green, 'negative' for red, 'neutral' for no color).
    """
    if pd.isna(value):
        return "" # Don't apply color to missing values

    # Define metrics where lower is better (e.g., risk metrics)
    lower_is_better_terms = [
        'NCO', 'NPL', 'Past Due', 'Nonaccrual', 'Concentration',
        'Risk', 'Cost', 'Ratio', 'Allowance'
    ]

    metric_name = metric_descriptions.get(metric_code, {}).get("short", "")
    is_lower_better = any(term in metric_name for term in lower_is_better_terms)

    # Use a small threshold to classify as neutral
    if abs(value) < 0.05:
        return ""

    if (is_lower_better and value > 0) or (not is_lower_better and value < 0):
        return "negative" # Red (worse)
    else:
        return "positive" # Green (better)

def generate_flexible_html_table(
    proc_df_with_peers: pd.DataFrame,
    avg_df: pd.DataFrame,
    subject_bank_cert: int,
    metrics_to_display: Dict[str, str],
    title: str,
    col_names: Dict[str, str] = None,
    fdic_field_descriptions: Dict[str, Dict[str, str]] = None
) -> str:
    """
    Generates a professional, mobile-responsive HTML table for email or web display,
    combining latest quarter and 8-quarter rolling average data.

    Args:
        proc_df_with_peers (pd.DataFrame): The main processed DataFrame containing
                                           all metrics for all banks and peers.
        avg_df (pd.DataFrame): The DataFrame containing the 8-quarter rolling averages.
        subject_bank_cert (int): The CERT number for the subject bank.
        metrics_to_display (Dict[str, str]): A dictionary where keys are the
                                             metric codes (e.g., 'TTM_NCO_Rate')
                                             and values are the display names for the columns.
        title (str): The main title for the HTML table.
        col_names (Dict[str, str], optional): A dictionary for custom column
                                               headers. Defaults to a standard set.
        fdic_field_descriptions (Dict[str, Dict[str, str]]): A dictionary with
                                                             full metric descriptions.

    Returns:
        str: A complete HTML string of the formatted table.
    """
    if proc_df_with_peers.empty or not metrics_to_display:
        return "<p>No data or metrics provided for table generation.</p>"

    # Get latest quarter data
    latest_date = proc_df_with_peers['REPDTE'].max()
    latest_data = proc_df_with_peers[proc_df_with_peers['REPDTE'] == latest_date]

    # Get all banks and peer groups to be included in the table
    certs_to_include = [subject_bank_cert] + [99999, 99998]

    # Define default column names
    if col_names is None:
        col_names = {
            'Latest': 'Latest Value',
            '8Q_Avg': '8-Qtr Average'
        }

    # Generate HTML string with updated styling
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 0; background-color: #f3f4f6; }}
            .email-container {{ background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); max-width: 1200px; margin: 20px auto; }}
            .header {{ text-align: center; margin-bottom: 20px; }}
            .header h2 {{ margin: 0; font-size: 20px; color: #2d3748; }}
            .header p {{ margin: 5px 0 0 0; color: #718096; font-size: 12px; }}
            table {{ width: 100%; border-collapse: separate; border-spacing: 0; overflow: hidden; }}
            .table-header {{ background-color: #f7a81b; color: #ffffff; font-weight: bold; }}
            .sub-header {{ background-color: #f7a81b; color: #ffffff; font-weight: bold; text-align: center; }}
            th, td {{ padding: 12px 15px; text-align: left; border: 1px solid #e2e8f0; }}
            th {{ font-size: 14px; text-transform: uppercase; letter-spacing: 0.05em; }}
            .metric-name {{ font-weight: normal; color: #2d3748; }}
            .value-cell {{ text-align: center; }}
            tr:last-child {{ background-color: #f7a81b; font-weight: bold; color: white; }}
            .positive {{ background-color: #b9e7c9; }} /* Green */
            .negative {{ background-color: #f8b8b8; }} /* Red */
        </style>
    </head>
    <body>
        <div class="email-container">
            <div class="header">
                <h2>{title}</h2>
                <p>Latest Report Date: {latest_date.strftime('%B %d, %Y')}</p>
            </div>
            <table>
                <thead>
                    <tr class="table-header">
                        <th rowspan="2">NAME</th>
                        <th rowspan="2">HQ</th>
                        <th rowspan="2">Operating States</th>
    """

    # Dynamically generate the top row of headers
    for display_name in metrics_to_display.values():
        html += f"""
                        <th colspan="2" class="value-cell">{display_name}</th>
        """

    html += """
                    </tr>
                    <tr class="sub-header">
    """

    # Dynamically generate the sub-header row
    for _ in metrics_to_display:
        html += f"""
                        <th class="value-cell">{col_names.get('Latest', 'Latest')}</th>
                        <th class="value-cell">{col_names.get('8Q_Avg', '8Q Avg')}</th>
        """

    html += """
                    </tr>
                </thead>
                <tbody>
    """
    for cert in certs_to_include:
        latest_row = latest_data[latest_data['CERT'] == cert].iloc[0] if not latest_data[latest_data['CERT'] == cert].empty else pd.Series()
        avg_row = avg_df[avg_df['CERT'] == cert].iloc[0] if not avg_df[avg_df['CERT'] == cert].empty else pd.Series()

        name = latest_row.get('NAME', 'N/A')
        hq = latest_row.get('HQ_STATE', 'N/A')
        states = latest_row.get('ALL_OPERATING_STATES', 'N/A')

        html += f"""
                    <tr>
                        <td class="metric-name">{name}</td>
                        <td>{hq}</td>
                        <td>{states}</td>
        """

        for metric_code in metrics_to_display.keys():
            latest_val = latest_row.get(metric_code, np.nan)
            avg_val = avg_row.get(metric_code, np.nan)

            latest_class = _get_cell_class(latest_val, metric_code, fdic_field_descriptions)
            avg_class = _get_cell_class(avg_val, metric_code, fdic_field_descriptions)

            html += f"""
                        <td class="value-cell {latest_class}">{latest_val if not pd.isna(latest_val) else 'N/A':.2f}%</td>
                        <td class="value-cell {avg_class}">{avg_val if not pd.isna(avg_val) else 'N/A':.2f}%</td>
            """

        html += """
                    </tr>
        """

    html += """
                </tbody>
            </table>
        </div>
    </body>
    </html>
    """
    return html

def create_complete_html_report(
    proc_df_with_peers: pd.DataFrame,
    avg_df: pd.DataFrame,
    subject_bank_cert: int,
    fdic_field_descriptions: Dict[str, Dict[str, str]]
) -> str:
    """
    A sample function to demonstrate how to use the flexible generator.
    This generates a comprehensive credit risk report.
    """
    print(f"The total number of rows from the FDIC data table is: {len(proc_df_with_peers)}")

    # Define a custom list of metrics that match the image
    image_metrics = {
        'ASSET': 'Assets',
        'LNLS': 'Gross Loans',
        'CRE_Concentration_Capital_Risk': 'CRE-to-Capital',
        'NPL_to_Gross_Loans_Rate': 'NPL to Gross Loans',
        'TTM_NCO_Rate': 'TTM NCO Rate',
        'TTM_Past_Due_Rate': 'TTM Past Due (30+90) Rate',
        'IDB_CRE_Growth_TTM': 'CRE 36M Growth %'
    }

    # Define custom column display names
    custom_col_names = {
        'Latest': 'Latest Value',
        '8Q_Avg': '8-Qtr Average'
    }

    # Generate the HTML table using the new function
    html_output = generate_flexible_html_table(
        proc_df_with_peers,
        avg_df,
        subject_bank_cert,
        image_metrics,
        "Credit Risk Metrics Dashboard",
        custom_col_names,
        fdic_field_descriptions
    )

    return html_output

def main():
    """Main execution block to run the dashboard generation."""
    try:
        # Load the data from the Excel file created by your main script.
        # This assumes your main script has already been run and the file exists.
        excel_file = "output/Bank_Performance_Dashboard_...xlsx" # <-- Update this path

        # Read the FDIC_Data sheet which contains all the metrics for all banks
        proc_df_with_peers = pd.read_excel(excel_file, sheet_name='FDIC_Data')
        proc_df_with_peers['REPDTE'] = pd.to_datetime(proc_df_with_peers['REPDTE'])

        # Read the 8-quarter averages sheet
        avg_df = pd.read_excel(excel_file, sheet_name='Averages_8Q_All_Metrics')

        # Define a sample subject bank CERT number
        subject_bank_cert = 19977

        # Your FDIC field descriptions from CR_Bank_Dashv4.py
        FDIC_FIELD_DESCRIPTIONS = {
            # Key Balance Sheet & Income Statement
            "ASSET":    {"short": "Total Assets", "long": "Total assets held by the institution."},
            "LNLS":     {"short": "Gross Loans & Leases", "long": "Total loans and leases, gross, before deducting the allowance for loan and lease losses."},
            "NPL_to_Gross_Loans_Rate": {"short": "NPL to Gross Loans", "long": "Total loans and leases past due 90 days or more plus nonaccrual loans."},
            "TTM_NCO_Rate": {"short": "TTM NCO / Avg Loans", "long": "Trailing 12-Month sum of net charge-offs as a percentage of TTM average gross loans."},
            "TTM_Past_Due_Rate": {"short": "Total TTM PD Rate", "long": "Trailing 12-Month average of total loans 30+ days past due as a percentage of TTM average gross loans."},
            "IDB_CRE_Growth_TTM": {"short": "TTM CRE Growth", "long": "Trailing 12-Month (Year-over-Year) growth rate of the Commercial Real Estate loan portfolio."},
            "CRE_Concentration_Capital_Risk": {"short": "CRE / (T1C + Total ACL)", "long": "Total Commercial Real Estate Loans as a percentage of Tier 1 Capital plus the Total Allowance for Credit Losses."},
        }

        # --- Example 1: Generate the image-style report ---
        print("Generating a table that mimics the image provided...")
        image_metrics = {
            'ASSET': 'Assets',
            'LNLS': 'Gross Loans',
            'CRE_Concentration_Capital_Risk': 'CRE-to-Capital',
            'NPL_to_Gross_Loans_Rate': 'NPL to Gross Loans',
            'TTM_NCO_Rate': 'TTM NCO Rate',
            'TTM_Past_Due_Rate': 'TTM Past Due (30+90) Rate',
            'IDB_CRE_Growth_TTM': 'CRE 36M Growth %'
        }

        html_report_1 = generate_flexible_html_table(
            proc_df_with_peers,
            avg_df,
            subject_bank_cert,
            image_metrics,
            "Credit Risk Metrics Dashboard",
            fdic_field_descriptions=FDIC_FIELD_DESCRIPTIONS
        )
        with open("output/image_style_report.html", "w") as f:
            f.write(html_report_1)
        print("Report 1 saved to output/image_style_report.html")

        # --- Example 2: Generate a custom report with different metrics ---
        print("\nGenerating a custom report with different metrics...")
        custom_metrics = {
            'RBCT1CER': 'Tier 1 Ratio',
            'NPL_to_Gross_Loans_Rate': 'NPL Rate',
            'Cost_of_Funds': 'Cost of Funds',
            'Retained_Earnings_Pct': 'Retained Earnings %',
        }

        html_report_2 = generate_flexible_html_table(
            proc_df_with_peers,
            avg_df,
            subject_bank_cert,
            custom_metrics,
            "Custom Capital & Risk Report"
        )
        with open("output/custom_report.html", "w") as f:
            f.write(html_report_2)
        print("Report 2 saved to output/custom_report.html")

    except FileNotFoundError:
        print("Error: Excel file not found. Please ensure the main script (CR_Bank_Dashv4.py) has been run first.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()



