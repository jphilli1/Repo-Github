# Standard library imports
import asyncio
import csv
import io
import logging
import os
import re
import sys
import time
import warnings
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Third-party imports
import aiohttp
import numpy as np
import openpyxl
import pandas as pd
import requests
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats
from tqdm import tqdm
from tqdm.asyncio import tqdm as tqdm_asyncio



script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)
# ==================================================================================
#  1. SCRIPT CONFIGURATION & SETUP
# ==================================================================================

def setup_logging(log_dir: str = "logs") -> logging.Logger:
    """Setup comprehensive logging configuration."""
    Path(log_dir).mkdir(exist_ok=True)
    log_filename = f"{log_dir}/bank_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    # Remove existing handlers to avoid duplicate logs in interactive environments
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[logging.FileHandler(log_filename), logging.StreamHandler(sys.stdout)]
    )
    logger = logging.getLogger(__name__)
    logger.info(f"Logging initialized. Log file: {log_filename}")
    return logger

logger = setup_logging()

@dataclass
class DashboardConfig:
    """Clean configuration class to hold runtime parameters."""
    fred_api_key: str
    subject_bank_cert: int
    peer_bank_certs: List[int]
    quarters_back: int = 30
    fred_years_back: int = 15
    output_dir: str = "output"
    fdic_api_base: str = "https://banks.data.fdic.gov/api"
    fred_api_base: str = "https://api.stlouisfed.org/fred"

# ==================================================================================
#  2. GLOBAL DATA DICTIONARIES & CONSTANTS
# ==================================================================================

# [NEW] FIELD RESOLUTION LAYER
# Maps conceptual/synthetic fields to authoritative Call Report series.
# Keys = The legacy/conceptual name you want to use.
# Values = The list of authoritative series to check in order (Consolidated -> Domestic -> Proxy).
FDIC_FALLBACK_MAP = {
    # --- SBL & Fund Finance (FFIEC Bulk / Consolidated) ---
    "LNOTHPCS":    ["RCFD1545", "RCON1545", "LNOTHER"], # Legacy SBL -> Consolidated SBL
    "LNOTHNONDEP": ["RCFDJ454", "RCONJ454"],            # Legacy Fund Fin -> Consolidated Fund Fin

    # --- RI-C Disaggregated Allowance (FFIEC Bulk Only) ---
    "RCFDJ466": ["FFIEC"], # Construction
    "RCFDJ467": ["FFIEC"], # Commercial RE
    "RCFDJ468": ["FFIEC"], # Residential
    "RCFDJ469": ["FFIEC"], # C&I
    "RCFDJ470": ["FFIEC"], # Credit Cards
    "RCFDJ471": ["FFIEC"], # Other Consumer
    "RCFDJ472": ["FFIEC"], # Unallocated
    "RCFDJ474": ["FFIEC"], # Other Loans (SBL Proxy)

    # --- Consumer Granular (Derived) ---
    "LNCONAUTO": ["LNAUTO"],
    "LNCONCC":   ["LNCRCD"],
    "LNCONOTHX": ["DERIVED"], # Calculated as LNCON - LNAUTO - LNCRCD

    # --- Legacy Real Estate Aliases (Mapping to standard codes) ---
    "LNREOTH":  ["LNRENROT"], # "Income Producing" -> Non-Owner Occupied Nonfarm
    "NALRERES": ["NARERES"],  # Legacy Resi NA -> Total Resi NA
    "P9RENRES": ["P9RENROT"], # Legacy Nonfarm NA -> Non-Owner Nonfarm NA
}

# [UPDATED] MASTER FETCH LIST
# Includes ALL valid API fields + Legacy placeholders (resolved later)
FDIC_FIELDS_TO_FETCH = [
    # --- Identifiers & Dates ---
    "CERT", "NAME", "REPDTE",

    # --- Balance Sheet Totals ---
    "ASSET", "LIAB", "DEP", "EQ",

    # --- Profitability ---
    "ROA", "ROE", "NIMY", "EEFFR", "NONIIAY", "ELNATRY", "EINTEXP",

    # --- Capital & RWA ---
    "RBCT1CER", "RBCT1J", "RBCT2", "RWAJ", "RBCRWAJ", "RB2LNRES",
    "EQCS", "EQSUR", "EQUP", "EQCCOMPI", "EQPP",

    # --- Reserves & Funding ---
    "LNATRES", "OTHBOR", "MUTUAL", "FREPP",

    # --- LOAN BALANCES (Authoritative) ---
    "LNLS", "LNLSNET",
    "LNCI",       # C&I
    "LNRECONS",   # Construction
    "LNREMULT",   # Multifamily
    "LNRENROW",   # Owner-Occ CRE
    "LNRENROT",   # Non-Owner CRE (Income Producing)
    "LNREAG",     # Farmland
    "LNRERES",    # 1-4 Family Resi Total
    "LNRELOC",    # HELOC
    "LNCON",      # Consumer Total
    "LNCRCD",     # Credit Cards
    "LNAUTO",     # Auto
    "LNOTHER",    # All Other Loans
    "LS",         # Leases
    "LNAG",       # Ag Loans

    # --- LEGACY BALANCES (Mapped via Fallback) ---
    "LNREOTH",    # Will map to LNRENROT
    "LNOTHPCS",   # Will map to RCFD1545
    "LNOTHNONDEP",# Will map to RCFDJ454
    "LNCONAUTO", "LNCONCC", "LNCONOTHX", # Will map/derive

    # --- NET CHARGE-OFFS ---
    "NTLNLS", "NCLNLS",
    "NTCI",
    "NTRECONS", "NTREMULT", "NTRENROT", "NTREAG", "NTRENRES",
    "NTRERES", "NTRELOC",
    "NTCON", "NTCRCD", "NTAUTO", "NTLS", "NTOTHER", "NTAG",
    "NTCONOTH", # Other Consumer NCO

    # --- PAST DUE 30-89 ---
    "P3LNLS", "P3CI",
    "P3RECONS", "P3LREMUL", "P3RENROT", "P3RENROW", "P3REAG", "P3RENRES",
    "P3RERES", "P3RELOC",
    "P3CON", "P3CRCD", "P3AUTO", "P3LS", "P3AG", "P3OTHLN",
    "P3CONOTH",

    # --- PAST DUE 90+ ---
    "P9LNLS", "P9CI",
    "P9RECONS", "P9REMULT", "P9RENROT", "P9RENROW", "P9REAG", "P9RENRES",
    "P9RERES", "P9RELOC",
    "P9CON", "P9CRCD", "P9AUTO", "P9LS", "P9AG", "P9OTHLN",
    "P9CONOTH",

    # --- NONACCRUAL ---
    "NACI",
    "NARECONS", "NAREMULT", "NARENROT", "NARENROW", "NAREAG", "NARENRES",
    "NARERES", "NARELOC",
    "NACON", "NACRCD", "NAAUTO", "NALS", "NAAG", "NAOTHLN",
    "NACONOTH",
    "NALRERES", # Legacy
]


FDIC_FIELD_DESCRIPTIONS = {
    # Key Balance Sheet & Income Statement
    "ASSET":    {"short": "Total Assets", "long": "Total assets held by the institution."},
    "DEP":      {"short": "Total Deposits", "long": "Total deposits held by the institution."},
    "LIAB":     {"short": "Total Liabilities", "long": "Total liabilities of the institution."},
    "EQ":       {"short": "Total Bank Equity Capital", "long": "Total equity capital of the institution."},
    "MUTUAL":       {"short": "Ownership Type", "long": "A non-stock institution, or mutual institution, is owned and controlled solely by its depositors. A mutual does not issue capital stock. (1=Yes, a mutual) An institution which sells stock to raise capital is called a stock institution. It is owned by the shareholders who benefit from profits earned by the institution. (0=No, not a mutual.)"},
    "RWA":       {"short": "Total Risk Weighted Assets", "long": "Total risk-weighted assets calculated under the standardized approach. (Schedule RC-R, Part II, item 31)"},
    "LNLS":     {"short": "Gross Loans & Leases", "long": "Total loans and leases, gross, before deducting the allowance for loan and lease losses."},
    "LNATRES":  {"short": "Loan Loss Allowance (no OBS)", "long": "Allowance for Loan and Lease Losses without OBS."},
    "AOBS": {"short": "Allowance for Off-Balance Sheet Exposures", "long": "Allowance for Credit Losses on Off-Balance Sheet Credit Exposures."},
    "NONIIAY":  {"short": "Noninterest Income (YTD)", "long": "Total noninterest income, year-to-date."},
    "ELNATRY":  {"short": "Noninterest Expense (YTD)", "long": "Total noninterest expense, year-to-date."},
    "EINTEXP":  {"short": "Interest Expense (YTD)", "long": "Total interest expense, year-to-date."},
    "Total_ACL":{"short": "Total Allowance for Credit Losses", "long": "Sum of the allowance for on-balance-sheet loans and off-balance-sheet credit exposures."},
    "LNREOTH":  {"short": "Income-Prod. CRE Bal", "long": "Loans Secured by Other Nonfarm Nonresidential Properties (Income-Producing)."},


    # Key Ratios & Capital
    "ROA":      {"short": "ROA", "long": "Return on Assets, annualized."},
    "ROE":      {"short": "ROE", "long": "Return on Equity, annualized."},
    "NIMY":     {"short": "Net Interest Margin", "long": "Net Interest Margin, annualized, year-to-date."},
    "EEFFR":    {"short": "Efficiency Ratio", "long": "Efficiency Ratio."},
    "RBCT1CER": {"short": "Common Equity Tier 1 Capital Ratio", "long": "Common Equity Tier 1 Capital Ratio."},
    # Capital
    "RBCT2": {"short": "Tier 2 Capital", "long": "Tier 2 risk-based capital including qualifying allowances for credit losses."},
    "RB2LNRES": {"short": "Total Loan Loss Allowance", "long": "( YTD, $ ) Loan Loss Allowance Included in Tier 2 Capital"},
    "RBCRWAJ": {"short": "Total Risk-Based Capital Ratio", "long": "Total risk-based capital (Tier 1 + Tier 2)."},
    "RWAJ": {"short": "RWA", "long": "Risk Weighted Assets"},
    "RBC1RWAJ": {"short": "Tier 1 Capital Ratio", "long": "Tier 1 risk-based capital."},
    "RBCT1J": {"short": "Tier 1 Capital", "long": "Tier 1 risk-based capital."},
    "EQCS": {"short": "Common Stock", "long": "Common stock par value and paid-in capital."},
    "EQSUR": {"short": "EQSUR", "long": "EQSUR (exclude all EQSUR related to preferred stock)."},
    "EQUP": {"short": "Retained Earnings", "long": "Retained earnings."},
    "LEVRATIO": {"short": "Tier 1 Leverage Ratio", "long": "Tier 1 capital as a percentage of average total assets. (Schedule RC-R, Part I, item 31)"},
    "CET1R": {"short": "Common Equity Tier 1 Capital Ratio", "long": "CET1 capital as a percentage of risk-weighted assets. (Schedule RC-R, Part I, item 28)"},
    "EQCCOMPI": {"short": "Accumulated Other Comprehensive Income", "long": "Accumulated other comprehensive income (can be negative)."},
    "EQO": {"short": "Other Equity Capital Components", "long": "Other equity capital components."},
    "EQPP": {"short": "Perpetual Preferred Stock", "long": "Perpetual preferred stock and related EQSUR."},


    "NTLNLS":   {"short": "Total Net Charge-Offs (YTD)", "long": "Total net charge-offs on loans and leases, year-to-date."},
    "P3LNLS":   {"short": "Total Loans PD 30-89 Days", "long": "Total loans and leases 30-89 days past due."},
    "P9LNLS":   {"short": "Total Loans PD 90+ Days", "long": "Total loans and leases 90 or more days past due."},
    # Additional derived metrics
    "Total_Capital": {"short": "Total Capital", "long": "Total risk-based capital available for regulatory purposes."},
    "Common_Stock_Pct": {"short": "Common Stock %", "long": "Common stock as a percentage of total equity capital."},
    "EQSUR_Pct": {"short": "EQSUR %", "long": "EQSUR as a percentage of total equity capital."},
    "Retained_Earnings_Pct": {"short": "Retained Earnings %", "long": "Retained earnings (net of EQCCOMPI) as a percentage of total equity capital."},
    "Preferred_Stock_Pct": {"short": "Preferred Stock %", "long": "Perpetual preferred stock as a percentage of total equity capital."},
    "CI_to_Capital_Risk": {"short": "C&I / Total Capital", "long": "Commercial & Industrial loans as a percentage of total capital."},

    # TTM Capital Composition
    "TTM_Common_Stock_Pct": {"short": "TTM Common Stock %", "long": "Trailing 12-month average common stock percentage of total equity."},
    "TTM_EQSUR_Pct": {"short": "TTM EQSUR %", "long": "Trailing 12-month average EQSUR percentage of total equity."},
    "TTM_Retained_Earnings_Pct": {"short": "TTM Retained Earnings %", "long": "Trailing 12-month average retained earnings percentage of total equity."},
    "TTM_Preferred_Stock_Pct": {"short": "TTM Preferred Stock %", "long": "Trailing 12-month average preferred stock percentage of total equity."},
    # Asset Quality - Top of House
    "NCLNLS":   {"short": "Total Noncurrent Loans", "long": "Total loans and leases past due 90 days or more plus nonaccrual loans."},

    # ===== DERIVED METRICS (MSPBNA V6 PRIVATE BANK VIEW) =====

    # --- 1. BANK LEVEL PROFITABILITY & COVERAGE ---
    "Cost_of_Funds": {"short": "Cost of Funds", "long": "Annualized cost of interest-bearing liabilities (Quarterly Interest Expense * 4 / Average Interest-Bearing Liabilities)."},
    "Allowance_to_Gross_Loans_Rate": {"short": "ACL / Total Loans", "long": "Total Allowance for Credit Losses as a percentage of Total Gross Loans."},
    "Risk_Adj_Allowance_Coverage": {"short": "ACL / (Loans - SBL)", "long": "Risk-Adjusted Coverage: Total Allowance divided by (Total Loans minus Securities-Based Lending). Excludes SBL as it is fully collateralized by marketable securities and requires minimal reserves."},
    "Nonaccrual_to_Gross_Loans_Rate": {"short": "Nonaccrual / Loans", "long": "Total nonaccrual loans as a percentage of total gross loans."},
    "TTM_NCO_Rate": {"short": "Total NCO Rate", "long": "Trailing 12-Month sum of Net Charge-Offs as a percentage of TTM Average Gross Loans."},

    # --- 2. LOAN COMPOSITION (PORTFOLIO MIX) ---
    "SBL_Composition": {"short": "SBL %", "long": "Securities-Based Lending (Loans for purchasing or carrying securities) as a % of Total Loans."},
    "Fund_Finance_Composition": {"short": "Fund Finance %", "long": "Loans to Nondepository Financial Institutions (PE/VC Capital Call Lines) as a % of Total Loans."},
    "Wealth_Resi_Composition": {"short": "Wealth Resi %", "long": "Wealth Residential (Jumbo 1-4 Family First Liens + HELOCs) as a % of Total Loans."},
    "Corp_CI_Composition": {"short": "Corp C&I %", "long": "Traditional Commercial & Industrial loans to operating companies as a % of Total Loans."},
    "CRE_OO_Composition": {"short": "CRE Owner-Occ %", "long": "Owner-Occupied Nonfarm Nonresidential CRE (Business Cash Flow dependent) as a % of Total Loans."},
    "CRE_Investment_Composition": {"short": "CRE Invest. %", "long": "Investment CRE (Construction, Multifamily, Non-OO Nonfarm) as a % of Total Loans."},
    "Consumer_Auto_Composition": {"short": "Auto %", "long": "Automobile loans as a % of Total Loans."},
    "Consumer_Other_Composition": {"short": "Cons. Other %", "long": "Other Consumer loans (Credit Cards, Unsecured, Other Revolving) as a % of Total Loans."},

    # --- 3. SEGMENT RISK: SECURITIES BASED LENDING (SBL) ---
    "SBL_TTM_NCO_Rate": {"short": "SBL NCO Rate", "long": "Net Charge-offs for SBL (Estimated/Allocated from All Other Loans) as a % of SBL Loans."},
    "SBL_TTM_PD30_Rate": {"short": "SBL PD 30-89", "long": "Past Due 30-89 Days for SBL (Proxy: All Other Loans) as a % of SBL Loans."},
    "SBL_NA_Rate": {"short": "SBL Nonaccrual", "long": "Nonaccrual Rate for SBL (Proxy: All Other Loans)."},

    # --- 4. SEGMENT RISK: FUND FINANCE ---
    "Fund_Finance_TTM_PD30_Rate": {"short": "Fund Fin. PD 30-89", "long": "Past Due 30-89 Days for Nondepository Fin. Institutions (Memo Item)."},
    "Fund_Finance_NA_Rate": {"short": "Fund Fin. Nonaccrual", "long": "Nonaccrual Rate for Nondepository Fin. Institutions (Memo Item)."},

    # --- 5. SEGMENT RISK: WEALTH RESIDENTIAL ---
    "Wealth_Resi_TTM_NCO_Rate": {"short": "Wealth Resi NCOs", "long": "Net Charge-offs for 1-4 Family First Liens & HELOCs as a % of Avg Segment Loans."},
    "Wealth_Resi_TTM_PD30_Rate": {"short": "Wealth Resi PD 30-89", "long": "Past Due 30-89 Days for 1-4 Family First Liens & HELOCs."},
    "Wealth_Resi_NA_Rate": {"short": "Wealth Resi Nonaccrual", "long": "Nonaccrual Rate for 1-4 Family First Liens & HELOCs."},

    # --- 6. SEGMENT RISK: COMMERCIAL & INDUSTRIAL ---
    "Corp_CI_TTM_NCO_Rate": {"short": "C&I NCO Rate", "long": "Net Charge-offs for C&I Loans as a % of Avg C&I Loans."},
    "Corp_CI_TTM_PD30_Rate": {"short": "C&I PD 30-89", "long": "Past Due 30-89 Days for C&I Loans."},
    "Corp_CI_NA_Rate": {"short": "C&I Nonaccrual", "long": "Nonaccrual Rate for C&I Loans."},

    # --- 7. SEGMENT RISK: CRE (OWNER OCCUPIED) ---
    "CRE_OO_TTM_NCO_Rate": {"short": "CRE OO NCO Rate", "long": "Net Charge-offs for Owner-Occupied CRE as a % of Avg OO CRE Loans."},
    "CRE_OO_TTM_PD30_Rate": {"short": "CRE OO PD 30-89", "long": "Past Due 30-89 Days for Owner-Occupied CRE."},
    "CRE_OO_NA_Rate": {"short": "CRE OO Nonaccrual", "long": "Nonaccrual Rate for Owner-Occupied CRE."},

    # --- 8. SEGMENT RISK: CRE (INVESTMENT) ---
    "CRE_Investment_TTM_NCO_Rate": {"short": "CRE Inv. NCO Rate", "long": "Net Charge-offs for Investment CRE (Constr/Multi/Non-OO) as a % of Avg Inv. CRE Loans."},
    "CRE_Investment_TTM_PD30_Rate": {"short": "CRE Inv. PD 30-89", "long": "Past Due 30-89 Days for Investment CRE."},
    "CRE_Investment_NA_Rate": {"short": "CRE Inv. Nonaccrual", "long": "Nonaccrual Rate for Investment CRE."},
    "CRE_Concentration_Capital_Risk": {"short": "Inv. CRE / Capital", "long": "Total Investment CRE Loans (Construction + Multifamily + Non-OO) as a percentage of Tier 1 Capital + ACL. (Proxy for Reg Concentration)."},

    # --- 9. SEGMENT RISK: CONSUMER (AUTO & OTHER) ---
    "Consumer_Auto_TTM_NCO_Rate": {"short": "Auto NCO Rate", "long": "Net Charge-offs for Auto Loans as a % of Avg Auto Loans."},
    "Consumer_Other_TTM_NCO_Rate": {"short": "Cons. Other NCOs", "long": "Net Charge-offs for Credit Cards & Other Consumer Loans as a % of Avg Segment Loans."},
    # --- 10. SEGMENT GROWTH METRICS (TTM / YEAR-OVER-YEAR) ---
    "Total_Loan_Growth_TTM": {"short": "Total Loan Growth", "long": "Trailing 12-Month (Year-over-Year) growth rate of Total Gross Loans."},
    "SBL_Growth_TTM": {"short": "SBL Growth", "long": "TTM growth rate of the Securities-Based Lending portfolio."},
    "Fund_Finance_Growth_TTM": {"short": "Fund Finance Growth", "long": "TTM growth rate of Loans to Nondepository Financial Institutions."},
    "Wealth_Resi_Growth_TTM": {"short": "Wealth Resi Growth", "long": "TTM growth rate of the Wealth Residential portfolio (1-4 Family First Liens + HELOCs)."},
    "Corp_CI_Growth_TTM": {"short": "C&I Growth", "long": "TTM growth rate of the Commercial & Industrial portfolio."},
    "CRE_OO_Growth_TTM": {"short": "CRE OO Growth", "long": "TTM growth rate of Owner-Occupied Commercial Real Estate."},
    "CRE_Investment_Growth_TTM": {"short": "CRE Inv. Growth", "long": "TTM growth rate of Investment CRE (Construction + Multifamily + Non-OO)."},
    "Consumer_Auto_Growth_TTM": {"short": "Auto Growth", "long": "TTM growth rate of the Automobile Loan portfolio."},
    "Consumer_Other_Growth_TTM": {"short": "Cons. Other Growth", "long": "TTM growth rate of Other Consumer Loans (Credit Cards + Unsecured)."},

    # ===== LOAN BALANCES =====
    "LNCI":     {"short": "C&I Loan Balances", "long": "Commercial and Industrial Loans."},

    "LNRECONS": {"short": "Construction & Dev Loan Bal", "long": "Construction and Land Development Loans."},
    "LNREMULT": {"short": "Multifamily CRE Loan Bal", "long": "Loans Secured by Multifamily (5 or more) Residential Properties."},
    "LNRENROW": {"short": "Nonfarm CRE Owner-Occ Bal", "long": "Loans Secured by Nonfarm Nonresidential Properties (Owner-Occupied)."},
    "LNREAG":   {"short": "Farmland Loan Bal", "long": "Loans Secured by Farmland."},
    "LNRERES":  {"short": "1-4 Family Resi Loan Bal", "long": "Loans Secured by 1-4 Family Residential Properties."},
    "LNRELOC":  {"short": "HELOC Balances", "long": "Home Equity Lines of Credit."},
    "LNCON":    {"short": "Consumer Loan Balances", "long": "Loans to Individuals for Household, Family, and Other Personal Expenditures."},
    "LNCRCD":   {"short": "Credit Card Loan Balances", "long": "Credit Card Loans."},
    "LNAUTO":   {"short": "Auto Loan Balances", "long": "Automobile Loans."},
    "LNOTHER":  {"short": "Other Loan Balances", "long": "Other Loans."},
    "LS":       {"short": "Lease Balances", "long": "Lease Financing Receivables."},
    "LNAG":     {"short": "Agricultural Loan Balances", "long": "Agricultural Loans."},

    # ===== NET CHARGE-OFFS (YTD) BY CATEGORY =====
    "NTCI":     {"short": "C&I NCOs", "long": "Net Charge-Offs on Commercial and Industrial Loans."},
    "NTRECONS": {"short": "Construction & Dev NCOs", "long": "Net Charge-Offs on Construction and Land Development Loans."},
    "NTREMULT": {"short": "Multifamily CRE NCOs", "long": "Net Charge-Offs on Loans Secured by Multifamily Residential Properties."},
    "NTLSNFOO": {"short": "Nonfarm CRE Owner-Occ NCOs", "long": "Net Charge-Offs on Loans Secured by Nonfarm Nonresidential Properties (Owner-Occupied)."},
    "NTREAG":   {"short": "Farmland NCOs", "long": "Net Charge-Offs on Loans Secured by Farmland."},
    "NTRERES":  {"short": "1-4 Family Resi NCOs", "long": "Net Charge-Offs on Loans Secured by 1-4 Family Residential Properties."},
    "NTRELOC":  {"short": "HELOC NCOs", "long": "Net Charge-Offs on Home Equity Lines of Credit."},
    "NTCON":    {"short": "Consumer NCOs", "long": "Net Charge-Offs on Loans to Individuals."},
    "NTCRCD":   {"short": "Credit Card NCOs", "long": "Net Charge-Offs on Credit Card Loans."},
    "NTAUTO":   {"short": "Auto Loan NCOs", "long": "Net Charge-Offs on Automobile Loans."},
    "NTLS":     {"short": "Lease NCOs", "long": "Net Charge-Offs on Leases."},
    "NTOTHER":  {"short": "Other Loan NCOs", "long": "Net Charge-Offs on Other Loans."},
    "NTAG":     {"short": "Agricultural NCOs", "long": "Net Charge-Offs on Agricultural Loans."},

    # ===== PAST DUE 30-89 DAYS BY CATEGORY =====
    "P3CI":     {"short": "C&I PD 30-89", "long": "Commercial and Industrial Loans 30-89 Days Past Due."},
    "P3RENROW": {"short": "1-4 Fam CRE Non-Owner PD 30-89", "long": "Loans Secured by 1-4 Family Residential (Non-owner Occupied) 30-89 Days Past Due."},
    "P3RECONS": {"short": "Construction PD 30-89", "long": "Construction and Land Development Loans 30-89 Days Past Due."},
    "P3LREMUL": {"short": "Multifamily PD 30-89", "long": "Loans Secured by Multifamily Residential Properties 30-89 Days Past Due."},
    "P3LRENRS": {"short": "Nonfarm Nonres Income CRE 30-89", "long": "Loans Secured by Multifamily Residential Properties 30-89 Days Past Due."},
    "P3RENROT": {"short": "Nonfarm CRE Owner-Occ PD 30-89", "long": "Loans Secured by Nonfarm Nonresidential (Owner-Occupied) 30-89 Days Past Due."},
    "P3REAG":   {"short": "Farmland PD 30-89", "long": "Loans Secured by Farmland 30-89 Days Past Due."},
    "P3RERES":  {"short": "1-4 Family Resi PD 30-89", "long": "Loans Secured by 1-4 Family Residential Properties 30-89 Days Past Due."},
    "P3RELOC":  {"short": "HELOC PD 30-89", "long": "Home Equity Lines of Credit 30-89 Days Past Due."},
    "P3CON":    {"short": "Consumer PD 30-89", "long": "Loans to Individuals 30-89 Days Past Due."},
    "P3CRCD":   {"short": "Credit Card PD 30-89", "long": "Credit Card Loans 30-89 Days Past Due."},
    "P3AUTO":   {"short": "Auto PD 30-89", "long": "Automobile Loans 30-89 Days Past Due."},
    "P3OTHLN":  {"short": "Other Loans PD 30-89", "long": "Other Loans 30-89 Days Past Due."},
    "P3LS":     {"short": "Leases PD 30-89", "long": "Leases 30-89 Days Past Due."},
    "P3AG":     {"short": "Agricultural PD 30-89", "long": "Agricultural Loans 30-89 Days Past Due."},

    # ===== PAST DUE 90+ DAYS BY CATEGORY =====
    "P9CI":     {"short": "C&I PD 90+", "long": "Commercial and Industrial Loans 90+ Days Past Due."},
    "P9RENROW": {"short": "1-4 Fam CRE Non-Owner PD 90+", "long": "Loans Secured by 1-4 Family Residential (Non-owner Occupied) 90+ Days Past Due."},
    "P9RECONS": {"short": "Construction PD 90+", "long": "Construction and Land Development Loans 90+ Days Past Due."},
    "P9REMULT": {"short": "Multifamily PD 90+", "long": "Loans Secured by Multifamily Residential Properties 90+ Days Past Due."},
    "P9RENROT": {"short": "Nonfarm CRE Owner-Occ PD 90+", "long": "Loans Secured by Nonfarm Nonresidential (Owner-Occupied) 90+ Days Past Due."},
    "P9RENRES": {"short": "Nonfarm CRE Owner-Occ PD 90+", "long": "Loans Secured by Nonfarm Nonresidential (Income Producing) 90+ Days Past Due."},
    "P9REAG":   {"short": "Farmland PD 90+", "long": "Loans Secured by Farmland 90+ Days Past Due."},
    "P9RERES":  {"short": "1-4 Family Resi PD 90+", "long": "Loans Secured by 1-4 Family Residential Properties 90+ Days Past Due."},
    "P9RELOC":  {"short": "HELOC PD 90+", "long": "Home Equity Lines of Credit 90+ Days Past Due."},
    "P9CON":    {"short": "Consumer PD 90+", "long": "Loans to Individuals 90+ Days Past Due."},
    "P9CRCD":   {"short": "Credit Card PD 90+", "long": "Credit Card Loans 90+ Days Past Due."},
    "P9AUTO":   {"short": "Auto PD 90+", "long": "Automobile Loans 90+ Days Past Due."},
    "P9OTHLN":  {"short": "Other Loans PD 90+", "long": "Other Loans 90+ Days Past Due."},
    "P9LS":     {"short": "Leases PD 90+", "long": "Leases 90+ Days Past Due."},
    "P9AG":     {"short": "Agricultural PD 90+", "long": "Agricultural Loans 90+ Days Past Due."},

    # ===== NONACCRUAL BY CATEGORY =====
    "NACI":     {"short": "C&I Nonaccrual", "long": "Commercial and Industrial Loans on Nonaccrual Status."},
    "NALRERES": {"short": "1-4 REsi Nonaccrual", "long": "Loans Secured by 1-4 Family Residential (Non-owner Occupied) on Nonaccrual Status."},
    "NARECONS": {"short": "Construction Nonaccrual", "long": "Construction and Land Development Loans on Nonaccrual Status."},
    "NAREMULT": {"short": "Multifamily Nonaccrual", "long": "Loans Secured by Multifamily Residential Properties on Nonaccrual Status."},
    "NARENROW": {"short": "Owner-Occupied CRE Owner-Occ Nonaccrual", "long": "Loans Secured by Nonfarm Nonresidential (Owner-Occupied) on Nonaccrual Status."},
    "NARENRES": {"short": "Nonfarm Nonres Income CRE Nonaccrual", "long": "Loans Secured by Nonfarm Nonresidential (Income Producing) on Nonaccrual Status."},
    "NAREAG":   {"short": "Farmland Nonaccrual", "long": "Loans Secured by Farmland on Nonaccrual Status."},
    "NARERES":  {"short": "1-4 Family Resi Nonaccrual", "long": "Loans Secured by 1-4 Family Residential Properties on Nonaccrual Status."},
    "NARELOC":  {"short": "HELOC Nonaccrual", "long": "Home Equity Lines of Credit on Nonaccrual Status."},
    "NACON":    {"short": "Consumer Nonaccrual", "long": "Loans to Individuals on Nonaccrual Status."},
    "NACRCD":   {"short": "Credit Card Nonaccrual", "long": "Credit Card Loans on Nonaccrual Status."},
    "NAAUTO":   {"short": "Auto Nonaccrual", "long": "Automobile Loans on Nonaccrual Status."},
    "NAOTHLN":  {"short": "Other Loans Nonaccrual", "long": "Other Loans on Nonaccrual Status."},
    "NALS":     {"short": "Leases Nonaccrual", "long": "Leases on Nonaccrual Status."},
    "NAAG":     {"short": "Agricultural Nonaccrual", "long": "Agricultural Loans on Nonaccrual Status."}
}
from enum import Enum

# =============================================================================
#  PEER GROUP DEFINITIONS
# =============================================================================

class PeerGroupType(str, Enum):
    CORE_PRIVATE_BANK = "Core_Private_Bank"
    MS_FAMILY_PLUS = "MS_Family_Plus"
    ALL_PEERS = "All_Peers"

PEER_GROUPS = {
    PeerGroupType.CORE_PRIVATE_BANK: {
        "name": "Core Private Bank Peers",
        "short_name": "Core PB",
        "description": "True private banking comparables - SBL, wealth management, UHNW focus",
        "certs": [34221, 33124, 57565],  # MSPBNA, GS, UBS
        "use_case": "Best for SBL/wealth product comparisons, NCO benchmarking",
        "display_order": 1
    },
    PeerGroupType.MS_FAMILY_PLUS: {
        "name": "Morgan Stanley + Extended Wealth",
        "short_name": "MS+Wealth",
        "description": "MS sister bank plus wealth management peers",
        "certs": [34221, 32992, 33124, 57565, 57450, 17281],  # MS banks + Schwab + City National
        "use_case": "Internal MS comparison plus broader wealth industry view",
        "display_order": 2
    },
    PeerGroupType.ALL_PEERS: {
        "name": "Full Peer Universe",
        "short_name": "Full Peer Set",
        "description": "Complete peer set including G-SIBs for size/scale context",
        "certs": [34221, 32992, 33124, 57565, 57450, 17281, 628, 3511, 7213, 3510],
        "use_case": "Regulatory comparison, market share analysis, full industry context",
        "display_order": 3
    }
}

# Helper to ensure we fetch data for ALL distinct certs mentioned in any group
def get_all_peer_certs():
    all_certs = set()
    for group in PEER_GROUPS.values():
        all_certs.update(group['certs'])
    return list(all_certs)
# MSPBNA V6 SEGMENTATION (Adapted for V5 Structure)
# NOTE: SBL and Fund Finance NCOs are not separately reported in Call Reports.
# They are lumped into 'NTOTH' (All Other). They are left empty [] here to
# prevent double-counting. You must calculate them via allocation if needed.
LOAN_CATEGORIES = {
    # 1. SBL & LIQUIDITY
    "SBL": {
        "balance": ["SBL_Balance"], # Resolved in Processor
        "nco": [], "pd30": [], "pd90": [], "na": []
    },
    # 2. FUND FINANCE
    "Fund_Finance": {
        "balance": ["Fund_Finance_Balance"], # Resolved in Processor
        "nco": [], "pd30": ["P3NDFI"], "pd90": ["P9NDFI"], "na": ["NANDFI"]
    },
    # 3. WEALTH RESIDENTIAL
    "Wealth_Resi": {
        "balance": ["Wealth_Resi_Balance"], # Collapsed in Processor
        "nco": ["NTRERES", "NTRELOC"],
        "pd30": ["P3RERES", "P3RELOC"],
        "pd90": ["P9RERES", "P9RELOC"],
        "na": ["NARERES", "NARELOC"]
    },
    # 4. TRADITIONAL C&I
    "Corp_CI": {
        "balance": ["Corp_CI_Balance"],
        "nco": ["NTCI"], "pd30": ["P3CI"], "pd90": ["P9CI"], "na": ["NACI"]
    },
    # 5. CRE: OWNER-OCCUPIED
    "CRE_OO": {
        "balance": ["CRE_OO_Balance"],
        "nco": ["NTRENROT"], "pd30": ["P3RENROW"], "pd90": ["P9RENROW"], "na": ["NARENROW"]
    },
    # 6. CRE: INVESTMENT
    "CRE_Investment": {
        "balance": ["CRE_Investment_Balance"],
        "nco": ["NTRECONS", "NTREMULT"],
        "pd30": ["P3RECONS", "P3LREMUL"],
        "pd90": ["P9RECONS", "P9REMULT"],
        "na": ["NARECONS", "NAREMULT"]
    },
    # 7. CONSUMER: AUTO
    "Consumer_Auto": {
        "balance": ["Consumer_Auto_Balance"],
        "nco": ["NTAUTO"], "pd30": ["P3AUTO"], "pd90": ["P9AUTO"], "na": ["NAAUTO"]
    },
    # 8. CONSUMER: OTHER
    "Consumer_Other": {
        "balance": ["Consumer_Other_Balance"], # Derived Residual in Processor
        "nco": ["NTCON", "NTCRCD"],
        "pd30": ["P3CON", "P3CRCD"],
        "pd90": ["P9CON", "P9CRCD"],
        "na": ["NACON", "NACRCD"]
    }
}
FRED_SERIES_TO_FETCH = {
    "Key Economic Indicators": {
        "GDPC1":    {"short": "Real GDP", "long": "Real Gross Domestic Product"},
        "A191RL1Q225SBEA": {"short": "Real GDP Growth", "long": "Real Gross Domestic Product, Percent Change from Preceding Period"},
        "UNRATE":   {"short": "Unemployment Rate", "long": "Civilian Unemployment Rate"},
        "CPIAUCSL": {"short": "CPI Inflation", "long": "Consumer Price Index for All Urban Consumers: All Items"},
        "UMCSENT":  {"short": "Consumer Sentiment", "long": "University of Michigan: Consumer Sentiment"},
        "ICSA":     {"short": "Initial Jobless Claims", "long": "Initial Claims, Insured Unemployment"},
        "USSLIND":  {"short": "Leading Economic Index", "long": "Conference Board Leading Economic Index for the United States"},
        "RSXFS":    {"short": "Retail Sales", "long": "Advance Retail Sales: Retail Trade"}
    },
    "Interest Rates & Yield Curve": {
        "FEDFUNDS": {"short": "Effective Fed Funds", "long": "Federal Funds Effective Rate"},
        "DFF":      {"short": "Fed Funds Rate", "long": "Federal Funds Effective Rate (Daily)"},
        "DPRIME":   {"short": "Bank Prime Loan Rate", "long": "Bank Prime Loan Rate"},
        "MORTGAGE30US": {"short": "30Y Mortgage Rate", "long": "30-Year Fixed Rate Mortgage Average in the United States"},
        "DGS30":    {"short": "30-Yr Treasury", "long": "Market Yield on U.S. Treasury at 30-Year Constant Maturity"},
        "DGS20":    {"short": "20-Yr Treasury", "long": "Market Yield on U.S. Treasury at 20-Year Constant Maturity"},
        "DGS10":    {"short": "10-Yr Treasury", "long": "Market Yield on U.S. Treasury at 10-Year Constant Maturity"},
        "DGS7":     {"short": "7-Yr Treasury", "long": "Market Yield on U.S. Treasury at 7-Year Constant Maturity"},
        "DGS5":     {"short": "5-Yr Treasury", "long": "Market Yield on U.S. Treasury at 5-Year Constant Maturity"},
        "DGS3":     {"short": "3-Yr Treasury", "long": "Market Yield on U.S. Treasury at 3-Year Constant Maturity"},
        "DGS2":     {"short": "2-Yr Treasury", "long": "Market Yield on U.S. Treasury at 2-Year Constant Maturity"},
        "DGS1":     {"short": "1-Yr Treasury", "long": "Market Yield on U.S. Treasury at 1-Year Constant Maturity"},
        "DGS6MO":   {"short": "6-Mo Treasury", "long": "Market Yield on U.S. Treasury at 6-Month Constant Maturity"},
        "DGS3MO":   {"short": "3-Mo Treasury", "long": "Market Yield on U.S. Treasury at 3-Month Constant Maturity"},
        "DGS1MO":   {"short": "1-Mo Treasury", "long": "Market Yield on U.S. Treasury at 1-Month Constant Maturity"},
        "T10Y2Y":   {"short": "10Y-2Y Spread", "long": "10-Year Minus 2-Year Treasury Spread"},
        "T10Y3M":   {"short": "10Y-3M Spread", "long": "10-Year Minus 3-Month Treasury Spread"}
    },
    "Credit Spreads & Lending Standards": {
        "DBAA":       {"short": "Baa Corp. Yield", "long": "Moody's Seasoned Baa Corporate Bond Yield"},
        "BAMLH0A0HYM2": {"short": "High-Yield Spread", "long": "ICE BofA US High Yield Index Option-Adjusted Spread"},
        "BAMLH0A0HYM2EY": {"short": "HY Effective Yield", "long": "ICE BofA US High Yield Index Effective Yield"},
        "BAMLH0A1HYBB": {"short": "BB-Rated Spread", "long": "ICE BofA BB US High Yield Index Option-Adjusted Spread"},
        "BAMLH0A2HYB":  {"short": "B-Rated Spread", "long": "ICE BofA B US High Yield Index Option-Adjusted Spread"},
        "BAMLH0A3HYC":  {"short": "CCC-Rated Spread", "long": "ICE BofA CCC & Lower US High Yield Index Option-Adjusted Spread"},
        "BAMLC0A0CM":   {"short": "IG Corp. Spread", "long": "ICE BofA US Corporate Index Option-Adjusted Spread"},
        "BAMLC0A1CAAAEY": {"short": "AAA-Rated Eff. Yield", "long": "ICE BofA AAA US Corporate Index Effective Yield"},
        "BAMLC0A4CBBB": {"short": "BBB-Rated Spread", "long": "ICE BofA BBB US Corporate Index Option-Adjusted Spread"},
        "DRBLACBS":   {"short": "Biz Loan Delinquency", "long": "Delinquency Rate on Business Loans, All Commercial Banks"},
        "DRTSCILM":   {"short": "C&I Lending Standards", "long": "Net Pct of Banks Tightening Standards for C&I Loans"},
        "DRTSCLCC":   {"short": "CRE Lending Standards", "long": "Net Pct of Banks Tightening Standards for CRE Loans"}
    },
    "Financial Stress & Risk": {
        "STLFSI4":      {"short": "Financial Stress Index", "long": "St. Louis Fed Financial Stress Index"},
        "VIXCLS":       {"short": "VIX", "long": "CBOE Volatility Index"},
        "NFCI":         {"short": "Nat'l Financial Conditions", "long": "Chicago Fed National Financial Conditions Index"}
    },
    "Global Benchmarks": {
        "FPCPITOTLZGWLD": {"short": "World Inflation", "long": "Inflation, CPI for World"},
        "FPCPITOTLZGHIC": {"short": "High-Income Inflation", "long": "Inflation, CPI for High-Income Countries"},
        "FPCPITOTLZGLMY": {"short": "Low/Mid-Income Inflation", "long": "Inflation, CPI for Low & Middle Income Countries"},
        "NYGDPMKTPCDWLD": {"short": "World GDP", "long": "GDP at Market Prices for World"}
    },
    "Real Estate & Housing": {
        "HOUST":         {"short": "Housing Starts", "long": "New Privately-Owned Housing Units Started"},
        "NYXRSA": {"short": "NY Metro Home Price Index", "long": "S&P/Case-Shiller NY Home Price Index"},
        "LXXRSA": {"short": "LA Metro Home Price Index", "long": "S&P/Case-Shiller LA Home Price Index"},
        "MIXRNSA": {"short": "Miami Metro Home Price Index", "long": "S&P/Case-Shiller Miami Home Price Index"},
        "CASTHPI": {"short": "California HPI", "long": "All-Transactions House Price Index for California"},
        "TLCOMCONS":         {"short": "Total US Construction Spending: Commercial", "long": "Total Construction Spending: Commercial in the United States, Monthly, Seasonally Adjusted Annual Rate"},
        "TLHLTHCONS":        {"short": "Total US Construction Spending: Health Care", "long": "Total Construction Spending: Health Care in the United States, Monthly, Seasonally Adjusted Annual Rate"},
        "TLNRESCONS": {"short": "Total US Construction Spending: Nonresidential", "long": "Total Construction Spending: Nonresidential in the United States, Monthly, Seasonally Adjusted Annual Rate"},
        "TLOFCONS": {"short": "Total US Construction Spending: Office", "long": "Total Construction Spending: Office in the United States, Monthly, Seasonally Adjusted Annual Rate"},
        "TLRESCONS": {"short": "Total US Construction Spending: Residential", "long": "Total Construction Spending: Residential in the United States, Monthly, Seasonally Adjusted Annual Rate"},
        "TTLCONS": {"short": "Total US Construction Spending", "long": "Total Construction Spending: Total Construction in the United States, Monthly, Seasonally Adjusted Annual Rate"},
        "FLSTHPI": {"short": "Florida HPI", "long": "All-Transactions House Price Index for Florida"}
    },
    "Banking Sector Aggregates": {
        "BUSLOANS":   {"short": "Business Loans (All Banks)", "long": "Business Loans, All Commercial Banks"},
        "REALLN":     {"short": "Real Estate Loans (All Banks)", "long": "Real Estate Loans, All Commercial Banks"},
        "CONSUMER":   {"short": "Consumer Loans (All Banks)", "long": "Consumer Loans, All Commercial Banks"},
        "DPSACBW027SBOG": {"short": "Total Deposits (All Banks)", "long": "Total Deposits, All Commercial Banks"},
        "DRCRELEXFACBS": {"short": "CRE Delinquency Rate", "long": "Delinquency Rate on CRE Loans (Excluding Farmland)"},
        "CORBLACBS": {"short": "Business Loan Charge-offs", "long": "Charge-off Rate on Business Loans, Annualized"},
        "DRALACBS": {"short": "All Loans Delinquency", "long": "Delinquency Rate on All Loans"},
        "DRCCLACBS": {"short": "Credit Card Delinquency Rate", "long": "Delinquency Rate on Credit Card Loans, All Commercial Banks "},
        "CORBLACBS": {"short": "Business Loan Charge-offs", "long": "Charge-off Rate on Business Loans, Annualized"},
        "DRSFRMACBS": {"short": "1-4 Resi Loans Delinquency", "long": "Delinquency Rate on Single-Family Residential Mortgages, Booked in Domestic Offices, All Commercial Banks"},
        "CORALACBN": {"short": "All Loans Charge-offs", "long": "Charge-off Rate on All Loans, Annualized"}
    },
    "Leading Indicators": {
        "USSLIND": {"short": "US Leading Index", "long": "The Conference Board Leading Economic Index® (LEI) for the U.S."},
        "USALOLITONOSTSAM": {"short": "OECD CLI US", "long": "OECD Composite Leading Indicator for US"},
        "PAYEMS": {"short": "Nonfarm Employment", "long": "All Employees: Total Nonfarm"},
        "PERMIT":        {"short": "Building Permits", "long": "New Privately-Owned Housing Units Authorized in Permit-Issuing Places: Total Units"},
    },

    "Middle Market, Healthcare, & Funding Indicators": {
    "DRTSCIS": {"short": "Small Firm C&I Standards", "long": "Net Pct Banks Tightening Standards - Small Firms"},
    "TCU":    {"short": "Capacity Utilization: Total Industry", "long": "Capacity Utilization: Total Index"},
    "INDPRO":    {"short": "Industrial Production Index", "long": "Industrial Production: Total Index"},
    "NEWORDER": {"short": "Manufacturers' New Orders: Nondefense Capital Goods", "long": "Manufacturers' New Orders: Nondefense Capital Goods Excluding Aircraft"},
    "SOFR":    {"short": "SOFR", "long": "Secured Overnight Financing Rate"},
    "TB3MS":   {"short": "3-Month T-Bill", "long": "3-Month Treasury Bill Secondary Market Rate"},
    "SOFR3MTB3M": {"short": "SOFR vs T-Bill Spread", "long": "Calculated Spread: SOFR minus 3-Month T-Bill Rate"},
    "MPCT04XXS": {"short": "Healthcare Construction", "long": "Total Construction Spending: Health Care"},
}
}
FDIC_FIELDS_TO_FETCH =list(dict.fromkeys(FDIC_FIELDS_TO_FETCH))
class FFIECBulkLoader:
    """
    Dual-Track Data Engine: Fetches granular 'Private Bank' fields from
    FFIEC CDR Bulk Data when FDIC API fails.
    """

    def __init__(self, output_dir="data/ffiec_cache"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.debug_dir = self.output_dir / "debug"
        self.debug_dir.mkdir(parents=True, exist_ok=True)

        self.base_url = "https://cdr.ffiec.gov/public/pws/downloadbulkdata.aspx"

        # Critical Private Banking Fields
        self.target_fields = [
            'RCFD1545', 'RCON1545', 'RCFDJ454', 'RCONJ454', # SBL / Fund Finance
            'RCFDJ466', 'RCONJ466', 'RCFDJ467', 'RCONJ467', # RI-C Real Estate
            'RCFDJ468', 'RCONJ468', 'RCFDJ469', 'RCONJ469',
            'RCFDJ470', 'RCONJ470', 'RCFDJ471', 'RCONJ471',
            'RCFDJ472', 'RCONJ472', 'RCFDJ474', 'RCONJ474'
        ]

    def _extract_all_hidden_fields(self, html):
        """Parses all <input type='hidden'> fields to ensure valid ViewState."""
        fields = {}
        matches = re.findall(r'<input[^>]*?type=["\']hidden["\'][^>]*?name=["\']([^"\']+)["\'][^>]*?value=["\']([^"\']*)["\']', html, re.IGNORECASE)
        for name, value in matches:
            fields[name] = value
        return fields

    def _save_debug_html(self, content, tag):
        try:
            ts = datetime.now().strftime("%H%M%S")
            fname = self.debug_dir / f"ffiec_{tag}_{ts}.html"
            with open(fname, "wb") as f:
                f.write(content)
            return str(fname)
        except Exception:
            return "Save failed"

    def _download_with_webforms(self, date_str, report_type="Reports of Condition and Income"):
        """
        STRICT 3-STEP FLOW:
        1. GET Page -> Get Session + State A
        2. POST Date Selection -> Get State B (CRITICAL MISSING STEP)
        3. POST Download Button -> Get ZIP
        """
        # Create ONE session per attempt to maintain cookies/ViewState continuity
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://cdr.ffiec.gov/public/pws/downloadbulkdata.aspx'
        })

        try:
            # --- STEP 1: INITIAL GET ---
            r1 = session.get(self.base_url, timeout=30)
            if r1.status_code != 200:
                logging.error(f"      [FFIEC] Step 1 GET failed: {r1.status_code}")
                return None

            fields_step1 = self._extract_all_hidden_fields(r1.text)
            if '__VIEWSTATE' not in fields_step1:
                logging.error("      [FFIEC] Failed to parse ViewState from Step 1.")
                return None

            # --- STEP 2: SELECT DATE (Prime the Server) ---
            # We must tell the server "We selected a new date".
            # This returns a NEW ViewState valid for that date.
            payload_step2 = fields_step1.copy()
            payload_step2.update({
                '__EVENTTARGET': 'DatesDropDownList', # The element that triggered the postback
                '__EVENTARGUMENT': '',
                'ListBox1': report_type,
                'DatesDropDownList': date_str,
                'ExportFormatDropDownList': 'TXT'
            })

            # This is NOT the download yet, just the UI update
            r2 = session.post(self.base_url, data=payload_step2, timeout=30)
            if r2.status_code != 200:
                logging.error(f"      [FFIEC] Step 2 POST (Date Select) failed: {r2.status_code}")
                return None

            fields_step2 = self._extract_all_hidden_fields(r2.text)
            if '__VIEWSTATE' not in fields_step2:
                logging.error("      [FFIEC] Failed to parse updated ViewState from Step 2.")
                return None

            # --- STEP 3: CLICK DOWNLOAD ---
            payload_step3 = fields_step2.copy()
            payload_step3.update({
                'ListBox1': report_type,
                'DatesDropDownList': date_str,
                'ExportFormatDropDownList': 'TXT',
                'Download_0': 'Download' # The button name
            })
            # Remove event target for button clicks
            if '__EVENTTARGET' in payload_step3: del payload_step3['__EVENTTARGET']

            r3 = session.post(self.base_url, data=payload_step3, stream=True, timeout=120)

            if r3.status_code == 200:
                # Robust ZIP detection
                is_zip = False
                if r3.content.startswith(b'PK'): is_zip = True
                elif 'zip' in r3.headers.get('Content-Type', '').lower(): is_zip = True
                elif 'octet-stream' in r3.headers.get('Content-Type', '').lower(): is_zip = True

                if is_zip:
                    return r3
                else:
                    debug_file = self._save_debug_html(r3.content, f"fail_{date_str.replace('/','')}")
                    logging.warning(f"      [FFIEC] Step 3 returned HTML, not ZIP. Saved to {debug_file}")
            else:
                logging.error(f"      [FFIEC] Step 3 Download POST failed: {r3.status_code}")

        except Exception as e:
            logging.error(f"      [FFIEC] Exception: {e}")

        return None

    def _quarter_needs_patching(self, df_fdic, date_obj):
        if df_fdic.empty: return True
        q_data = df_fdic[df_fdic['REPDTE'] == date_obj]
        if q_data.empty: return False

        # Heuristic: If SBL (RCFD1545) is missing/zero, we need FFIEC
        check_cols = [c for c in ['RCFD1545', 'LNOTHPCS', 'RCFDJ466'] if c in q_data.columns]
        if not check_cols: return True

        missing_or_zero = 0
        total_cells = 0
        for col in check_cols:
            vals = q_data[col]
            missing_or_zero += ((vals.isna()) | (vals == 0)).sum()
            total_cells += len(vals)

        return (missing_or_zero / total_cells) > 0.5 if total_cells > 0 else True

    def fetch_quarter_data(self, date_obj, peer_certs):
        date_fmt = date_obj.strftime("%m/%d/%Y")
        file_date_sig = date_obj.strftime("%Y%m%d")
        cache_file = self.output_dir / f"FFIEC_Bulk_Call_{file_date_sig}.csv"

        if cache_file.exists():
            try:
                df_cache = pd.read_csv(cache_file)
                if 'CERT' in df_cache.columns:
                    # Filter cache by CERT (not IDRSSD)
                    df_cache = df_cache[df_cache['CERT'].isin(peer_certs)]
                    print(f"      [Cache Hit] {date_fmt} ({len(df_cache)} records)")
                    return df_cache
            except Exception:
                print(f"      [Cache Error] Corrupt file for {date_fmt}, re-downloading.")

        print(f"      [Downloading] FFIEC Bulk Data for {date_fmt}...")

        # Retry Logic wrapping the strict flow
        r = None
        for attempt in range(3):
            r = self._download_with_webforms(date_fmt)
            if r: break
            time.sleep(2)

        if not r:
            return pd.DataFrame()

        try:
            with zipfile.ZipFile(io.BytesIO(r.content)) as z:
                target_files = [n for n in z.namelist() if ('Schedule RC-C I' in n or 'Schedule RI-C' in n) and n.endswith('.txt')]

                cert_data_map = {}

                for t_file in target_files:
                    with z.open(t_file) as f:
                        content = io.TextIOWrapper(f, encoding='latin-1')
                        _ = content.readline() # Metadata
                        reader = csv.DictReader(content, delimiter='\t')

                        for row in reader:
                            cert_str = row.get('FDIC Certificate Number', '0')
                            if not cert_str.isdigit(): continue
                            cert = int(cert_str)

                            if cert in peer_certs:
                                if cert not in cert_data_map:
                                    cert_data_map[cert] = {'CERT': cert, 'REPDTE': date_obj}

                                for field in self.target_fields:
                                    # Strict existence check
                                    if field in row and row[field] is not None and row[field].strip() != '':
                                        try:
                                            cert_data_map[cert][field] = float(row[field].replace(',', ''))
                                        except: pass

                if cert_data_map:
                    df_q = pd.DataFrame(list(cert_data_map.values()))
                    df_q.to_csv(cache_file, index=False)
                    return df_q

        except Exception as e:
            logging.error(f"      [Error] Zip processing failed: {e}")

        return pd.DataFrame()

    def heal_dataset(self, df_fdic, peer_certs):
        print("\n" + "="*60)
        print("DUAL-TRACK DATA RECOVERY: FFIEC BULK API")
        print("="*60)

        if df_fdic.empty: return df_fdic

        dates = sorted(df_fdic['REPDTE'].unique(), reverse=True)[:8]
        ffiec_frames = []

        for dt in dates:
            dt_obj = pd.to_datetime(dt)
            if not self._quarter_needs_patching(df_fdic, dt_obj):
                print(f"      [Skip] Data sufficient for {dt_obj.strftime('%Y-%m-%d')}")
                continue

            df_ffiec_q = self.fetch_quarter_data(dt_obj, peer_certs)
            if not df_ffiec_q.empty:
                ffiec_frames.append(df_ffiec_q)
            else:
                logging.warning(f"      [Failed] No FFIEC data for {dt_obj.date()}")

        if not ffiec_frames:
            print("      No FFIEC data merged.")
            df_fdic['FFIEC_RIC_STATUS'] = 'FAILED'
            return df_fdic

        print("      Merging FFIEC Granular Data...")
        df_patch = pd.concat(ffiec_frames, ignore_index=True)

        df_fdic['REPDTE'] = pd.to_datetime(df_fdic['REPDTE'])
        df_patch['REPDTE'] = pd.to_datetime(df_patch['REPDTE'])

        df_main = df_fdic.set_index(['CERT', 'REPDTE'])
        df_p = df_patch.set_index(['CERT', 'REPDTE'])

        # Merge: Overwrite NaN, or 0 if FFIEC has data
        for col in self.target_fields:
            if col in df_p.columns:
                if col not in df_main.columns:
                    df_main[col] = df_p[col]
                else:
                    df_main[col] = df_main[col].fillna(df_p[col])
                    mask_fix = (df_main[col] == 0) & (df_p[col].notna()) & (df_p[col] != 0)
                    if mask_fix.any():
                        df_main.loc[mask_fix, col] = df_p.loc[mask_fix, col]

        res = df_main.reset_index()
        res['FFIEC_RIC_STATUS'] = 'SUCCESS'
        return res
# ==================================================================================
#  3. HELPER CLASSES
# ==================================================================================
def get_bank_locations(cert_numbers: list) -> pd.DataFrame:
    """
    Fetches the primary (HQ) and all operating states for a list of banks.

    Args:
        cert_numbers: A list of bank CERT numbers.

    Returns:
        A pandas DataFrame with CERT, NAME, HQ_STATE, and ALL_OPERATING_STATES.
    """
    institution_data = []
    session = requests.Session()

    for cert in cert_numbers:
        logger.info(f"Fetching location data for CERT: {cert}...")
        hq_state = "N/A"
        all_states = set()
        bank_name = f"Bank with CERT {cert}"

        # 1. Get Institution details (for HQ State)
        try:
            inst_url = f"https://banks.data.fdic.gov/api/institutions?filters=CERT%3A%20{cert}&fields=NAME,STALP"
            response = session.get(inst_url, timeout=20)
            response.raise_for_status()
            data = response.json().get('data', [])
            if data:
                bank_name = data[0]['data']['NAME']
                hq_state = data[0]['data']['STALP']
                all_states.add(hq_state)

        except requests.exceptions.RequestException as e:
            logger.error(f"  - Could not fetch institution data for CERT {cert}: {e}")


        # 2. Get all branch locations (for all operating states)
        try:
            # Increase limit to get all branches, assuming no bank has > 10000 branches
            loc_url = f"https://banks.data.fdic.gov/api/locations?filters=CERT%3A%20{cert}&fields=STALP&limit=10000"
            response = session.get(loc_url, timeout=20)
            response.raise_for_status()
            locations = response.json().get('data', [])
            for loc in locations:
                all_states.add(loc['data']['STALP'])

        except requests.exceptions.RequestException as e:
            logger.error(f"  - Could not fetch branch location data for CERT {cert}: {e}")

        institution_data.append({
            "CERT": cert,
            "NAME": bank_name,
            "HQ_STATE": hq_state,
            "ALL_OPERATING_STATES": ", ".join(sorted(list(all_states)))
        })
        time.sleep(0.1) # Small delay to be polite to the API

    return pd.DataFrame(institution_data)
class FDICDataFetcher:
    def __init__(self, config: DashboardConfig):
        self.config = config
        self.session = requests.Session()

    def fetch_lnci_separately(self, certs_to_fetch: List[int]) -> pd.DataFrame:
        """Fetch LNCI field separately since it doesn't work in bulk requests."""
        lnci_data = []

        for cert in tqdm(certs_to_fetch, desc="Fetching LNCI data", unit="bank"):
            try:
                params = {
                    "filters": f"CERT:{cert}",
                    "fields": "CERT,REPDTE,LNCI",
                    "sort_by": "REPDTE",
                    "sort_order": "DESC",
                    "limit": self.config.quarters_back + 4,
                    "format": "json"
                }

                response = self.session.get(f"{self.config.fdic_api_base}/financials", params=params, timeout=30)
                response.raise_for_status()

                data = [item.get('data', {}) for item in response.json().get('data', []) if item.get('data')]
                if data:
                    df = pd.DataFrame(data)
                    df['CERT'] = cert

                    # CRITICAL FIX: Use same datetime conversion as main fetch
                    df['REPDTE'] = pd.to_datetime(df['REPDTE'], format='ISO8601', utc=True).dt.tz_localize(None)
                    df['LNCI'] = pd.to_numeric(df['LNCI'], errors='coerce')

                    # Diagnostic
                    lnci_values = df['LNCI'].dropna()
                    if not lnci_values.empty:
                        logger.info(f"📊 CERT {cert} LNCI: {len(lnci_values)} non-null values, range ${lnci_values.min():,.0f} - ${lnci_values.max():,.0f}")
                    else:
                        logger.warning(f"⚠️ CERT {cert}: LNCI column exists but all values are NaN")

                    lnci_data.append(df[['CERT', 'REPDTE', 'LNCI']])
                    logger.info(f"✅ LNCI fetched for CERT {cert}")
                else:
                    logger.warning(f"⚠️ No LNCI data returned for CERT {cert}")

            except Exception as e:
                logger.error(f"❌ Error fetching LNCI for CERT {cert}: {e}")

            time.sleep(0.2)

        if lnci_data:
            combined = pd.concat(lnci_data, ignore_index=True)
            logger.info(f"📊 Combined LNCI: {len(combined)} rows, {combined['LNCI'].notna().sum()} non-null")
            return combined

        return pd.DataFrame()

    def fetch_all_banks(self) -> Tuple[pd.DataFrame, List[int]]:
        certs_to_fetch = [self.config.subject_bank_cert] + self.config.peer_bank_certs
        all_bank_data, failed_certs = [], []

        # Step 1: Fetch all fields EXCEPT LNCI
        main_fields = [f for f in FDIC_FIELDS_TO_FETCH if f != 'LNCI']

        for cert in tqdm(certs_to_fetch, desc="Fetching FDIC data", unit="bank"):
            try:
                params = {
                    "filters": f"CERT:{cert}",
                    "fields": ",".join(main_fields),
                    "sort_by": "REPDTE", "sort_order": "DESC",
                    "limit": self.config.quarters_back + 4, "format": "json"
                }
                response = self.session.get(f"{self.config.fdic_api_base}/financials", params=params, timeout=30)
                response.raise_for_status()
                data = [item.get('data', {}) for item in response.json().get('data', []) if item.get('data')]
                if data:
                    df = pd.DataFrame(data)
                    df['CERT'] = cert
                    all_bank_data.append(df)
            except Exception as e:
                logger.error(f"Error fetching FDIC data for CERT {cert}: {e}")
                failed_certs.append(cert)
            time.sleep(0.2)

        if not all_bank_data:
            return pd.DataFrame(), failed_certs

        # Step 2: Combine main data
        combined_df = pd.concat(all_bank_data, ignore_index=True)

        # CRITICAL FIX: Convert REPDTE to datetime WITH EXPLICIT FORMAT
        # This ensures consistent datetime representation for merging
        combined_df['REPDTE'] = pd.to_datetime(combined_df['REPDTE'], format='ISO8601', utc=True).dt.tz_localize(None)

        # Step 3: Fetch LNCI separately and merge
        logger.info("Fetching LNCI data separately...")
        lnci_df = self.fetch_lnci_separately(certs_to_fetch)

        if not lnci_df.empty:
            # DIAGNOSTIC: Check merge compatibility
            logger.info(f"📊 Pre-merge diagnostics:")
            logger.info(f"   Main DF REPDTE dtype: {combined_df['REPDTE'].dtype}")
            logger.info(f"   LNCI DF REPDTE dtype: {lnci_df['REPDTE'].dtype}")
            logger.info(f"   Main DF sample: {combined_df[['CERT', 'REPDTE']].head(2).to_dict('records')}")
            logger.info(f"   LNCI DF sample: {lnci_df[['CERT', 'REPDTE']].head(2).to_dict('records')}")

            # Check for overlapping keys
            main_keys = set(zip(combined_df['CERT'], combined_df['REPDTE'].astype(str)))
            lnci_keys = set(zip(lnci_df['CERT'], lnci_df['REPDTE'].astype(str)))
            overlap = main_keys.intersection(lnci_keys)
            logger.info(f"   Overlapping keys: {len(overlap)} out of {len(main_keys)}")

            if len(overlap) == 0:
                logger.error("❌ NO OVERLAPPING KEYS - merge will fail!")
                logger.info(f"   Main key examples: {list(main_keys)[:3]}")
                logger.info(f"   LNCI key examples: {list(lnci_keys)[:3]}")

            # Merge LNCI data
            combined_df = pd.merge(
                combined_df,
                lnci_df,
                on=['CERT', 'REPDTE'],
                how='left'
            )

            # POST-MERGE DIAGNOSTIC
            lnci_count = combined_df['LNCI'].notna().sum()
            logger.info(f"✅ After merge - LNCI non-null count: {lnci_count}/{len(combined_df)}")

            if lnci_count == 0:
                logger.warning("⚠️ LNCI merge produced all NaN - investigating...")
                # Try alternative merge strategy
                logger.info("   Attempting alternative merge with string dates...")
                combined_df['REPDTE_STR'] = combined_df['REPDTE'].astype(str)
                lnci_df['REPDTE_STR'] = lnci_df['REPDTE'].astype(str)

                combined_df = combined_df.drop(columns=['LNCI'], errors='ignore')
                combined_df = pd.merge(
                    combined_df,
                    lnci_df[['CERT', 'REPDTE_STR', 'LNCI']],
                    on=['CERT', 'REPDTE_STR'],
                    how='left'
                )
                combined_df = combined_df.drop(columns=['REPDTE_STR'])

                lnci_count_retry = combined_df['LNCI'].notna().sum()
                logger.info(f"   Retry result: {lnci_count_retry} non-null LNCI values")

        else:
            combined_df['LNCI'] = np.nan
            logger.warning("⚠️ LNCI fetch failed, using NaN values")

        # =========================================================
        # [STEP 2] DUAL-TRACK DATA RECOVERY (FFIEC BULK HEALER)
        # =========================================================
        try:
            print("\n[Dual-Track] Initiating FFIEC Bulk Data Recovery...")
            ffiec_loader = FFIECBulkLoader(output_dir="data/ffiec_cache")
            all_certs_set = set(certs_to_fetch)

            # Run Healer
            combined_df = ffiec_loader.heal_dataset(combined_df, all_certs_set)
            # [D] VALIDATOR FOR CERT 34221
            try:
                target_cert = 34221
                # Check latest available date in the combined frame
                if not combined_df.empty:
                    latest_dt = combined_df['REPDTE'].max()
                    test_row = combined_df[(combined_df["CERT"] == target_cert) & (combined_df["REPDTE"] == latest_dt)]

                    if not test_row.empty:
                        logging.info(f"\n[VALIDATOR] Checking RI-C fields for CERT {target_cert} at {latest_dt.date()}:")
                        check_cols = ["RCFDJ466", "RCONJ466", "RCFDJ474", "RCONJ474"]
                        for col in check_cols:
                            val = test_row[col].iloc[0] if col in test_row.columns else "MISSING"
                            logging.info(f"  - {col}: {val}")

                        if 'FFIEC_RIC_STATUS' in test_row.columns and test_row['FFIEC_RIC_STATUS'].iloc[0] == 'FAILED':
                            logging.warning("  -> FFIEC Download Failed for this date. Check debug/ folder.")
                    else:
                        logging.warning(f"[VALIDATOR] CERT {target_cert} not found in latest data.")
            except Exception as e:
                logging.error(f"[VALIDATOR] Exception: {e}")

            # ===== POST-FFIEC HEAL DIAGNOSTIC (RI-C) =====
            # Check if columns actually exist after the merge
            ric_cols = [
                "RCFDJ466","RCONJ466","RCFDJ467","RCONJ467","RCFDJ468","RCONJ468","RCFDJ469","RCONJ469",
                "RCFDJ470","RCONJ470","RCFDJ471","RCONJ471","RCFDJ472","RCONJ472","RCFDJ474","RCONJ474"
            ]
            missing = [c for c in ric_cols if c not in combined_df.columns]
            logger.info(f"[RI-C DIAG] Columns present: {len(ric_cols)-len(missing)}/{len(ric_cols)}")
            if missing:
                logger.warning(f"[RI-C DIAG] Missing RI-C source columns: {missing}")

            # Non-null counts
            counts = {c: int(combined_df[c].notna().sum()) for c in ric_cols if c in combined_df.columns}
            logger.info(f"[RI-C DIAG] Non-null counts: {counts}")

            # Validation Row (MSPBNA)
            try:
                test_cert = 34221
                # Find the latest available date in the df to test against
                latest_dt = combined_df['REPDTE'].max()
                test = combined_df[(combined_df["CERT"] == test_cert) & (combined_df["REPDTE"] == latest_dt)]

                if test.empty:
                    logger.warning(f"[RI-C DIAG] Validation row missing for CERT={test_cert}.")
                else:
                    show_cols = ["CERT","REPDTE"] + [c for c in ric_cols if c in combined_df.columns]
                    # Print first 200 chars of dict to avoid massive log spam
                    row_data = test[show_cols].iloc[0].to_dict()
                    logger.info(f"[RI-C DIAG] CERT={test_cert} values: {row_data}")
            except Exception as e:
                logger.error(f"[RI-C DIAG] Exception during validation row check: {e}")
            # ===========================================

        except Exception as e:
            logger.error(f"FFIEC Healer failed: {e}")
            print(f"WARNING: FFIEC Healer skipped due to error: {e}")
        # =========================================================


        # Step 4: Continue with existing processing
        for field in FDIC_FIELDS_TO_FETCH:
            if field not in combined_df.columns:
                combined_df[field] = np.nan

        # Convert numeric columns
        num_cols = [c for c in combined_df.columns if c not in ['CERT', 'NAME', 'REPDTE']]
        combined_df[num_cols] = combined_df[num_cols].apply(pd.to_numeric, errors='coerce')

        return combined_df.sort_values(['CERT', 'REPDTE']), failed_certs



class FREDDataFetcher:
    """
    Asynchronous data fetcher for the FRED API.
    Handles concurrent requests with rate limiting and retries.
    """
    BASE_URL = "https://api.stlouisfed.org/fred/series/observations"

    def __init__(self, config: 'DashboardConfig', max_concurrent: int = 3, rate_limit_delay: float = 1.0):
        self.config = config
        self.api_key = self.config.fred_api_key
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.rate_limit_delay = rate_limit_delay
        self.logger = logging.getLogger(__name__)
        self.last_fred_obs_df = pd.DataFrame() # Initialize storage for raw obs

    async def _fetch_series_metadata(
        self, session: aiohttp.ClientSession, series_id: str
    ) -> Tuple[str, Optional[Dict]]:
        """Fetches metadata for a FRED series including frequency."""
        params = {
            "series_id": series_id,
            "api_key": self.api_key,
            "file_type": "json"
        }

        async with self.semaphore:
            try:
                async with session.get(
                    "https://api.stlouisfed.org/fred/series",
                    params=params
                ) as response:
                    response.raise_for_status()
                    data = await response.json()

                    if "seriess" in data and data["seriess"]:
                        metadata = data["seriess"][0]
                        return series_id, {
                            'frequency': metadata.get('frequency', 'Unknown'),
                            'frequency_short': metadata.get('frequency_short', 'Unknown'),
                            'units': metadata.get('units', 'Unknown'),
                            'seasonal_adjustment': metadata.get('seasonal_adjustment', 'Unknown'),
                            'last_updated': metadata.get('last_updated', 'Unknown')
                        }
                    return series_id, None
            except Exception as e:
                self.logger.error(f"Error fetching metadata for {series_id}: {e}")
                return series_id, None
            finally:
                await asyncio.sleep(self.rate_limit_delay)

    async def _fetch_single_series(
        self, session: aiohttp.ClientSession, series_id: str, start_date: str
    ) -> Tuple[str, Optional[pd.DataFrame]]:
        """
        Fetches a single time series from the FRED API asynchronously.
        """
        params = {
            "series_id": series_id,
            "api_key": self.api_key,
            "file_type": "json",
            "observation_start": start_date,
            "sort_order": "asc",
        }

        async with self.semaphore:
            try:
                async with session.get(self.BASE_URL, params=params) as response:
                    response.raise_for_status()
                    data = await response.json()

                    if "observations" in data and data["observations"]:
                        df = pd.DataFrame(data["observations"])
                        df = df[['date', 'value']]
                        df['date'] = pd.to_datetime(df['date'])
                        df['value'] = pd.to_numeric(df['value'], errors='coerce')
                        df = df.set_index('date').rename(columns={'value': series_id})
                        return series_id, df
                    else:
                        self.logger.warning(f"No observations returned for series {series_id}.")
                        return series_id, None
            except aiohttp.ClientError as e:
                self.logger.error(f"HTTP error fetching {series_id}: {e}")
                return series_id, None
            except Exception as e:
                self.logger.error(f"An unexpected error occurred for {series_id}: {e}")
                return series_id, None
            finally:
                await asyncio.sleep(self.rate_limit_delay)

    async def fetch_all_series_async(
        self,
        series_ids: List[str],
        series_descriptions: Optional[pd.DataFrame] = None
    ) -> Tuple[pd.DataFrame, pd.DataFrame, List[str], pd.DataFrame]:
        """
        Async fetch for all FRED series + metadata.
        Consolidated method that handles optional description inputs.
        """
        start_date = (datetime.now() - pd.DateOffset(years=self.config.fred_years_back)).strftime('%Y-%m-%d')

        if not series_ids:
             return pd.DataFrame(), pd.DataFrame(), [], pd.DataFrame()

        async with aiohttp.ClientSession() as session:
            # Fetch both data and metadata concurrently
            data_tasks = [
                self._fetch_single_series(session, series_id, start_date)
                for series_id in series_ids
            ]
            metadata_tasks = [
                self._fetch_series_metadata(session, series_id)
                for series_id in series_ids
            ]

            all_tasks = data_tasks + metadata_tasks
            # Use tqdm_asyncio for async progress bar (fixes AttributeError)
            results = await tqdm_asyncio.gather(*all_tasks, desc="Fetching FRED Series & Metadata")

        # Split results
        num_series = len(series_ids)
        data_results = results[:num_series]
        metadata_results = results[num_series:]

        successful_dfs = []
        failed_series = []
        metadata_dict = {}
        raw_df_map = {}  # keep original (pre-resample) series by id

        # Process data results
        for series_id, df in data_results:
            if df is not None and not df.empty:
                successful_dfs.append(df)          # this is your resample-ready frame (DATE, sid as column)
                raw_df_map[series_id] = df.copy()  # keep the original obs-dated frame
            else:
                failed_series.append(series_id)

        # Process metadata results
        for series_id, metadata in metadata_results:
            if metadata:
                metadata_dict[series_id] = metadata

        if not successful_dfs:
            self.logger.error("No FRED series were successfully fetched.")
            # Return empty structure matching signature, using passed descriptions if available
            desc_return = series_descriptions if series_descriptions is not None else pd.DataFrame()
            return pd.DataFrame(), desc_return, failed_series, pd.DataFrame()

        merged_df = pd.concat(successful_dfs, axis=1)
        merged_df = merged_df.resample('D').asfreq().ffill()

        # Create metadata DataFrame
        metadata_df = pd.DataFrame.from_dict(metadata_dict, orient='index')
        metadata_df.index.name = 'Series ID'
        metadata_df.reset_index(inplace=True)

        # === PATCH S-RAW2: build long raw observations (DATE, SeriesID, VALUE) and stash ===
        raw_long = []
        for sid, rdf in raw_df_map.items():
            r = rdf.copy().reset_index()                             # index is 'date'
            date_col = 'date' if 'date' in r.columns else r.columns[0]
            r = r.rename(columns={date_col: 'DATE', sid: 'VALUE'})
            r = r[['DATE', 'VALUE']]
            r['SeriesID'] = str(sid)
            raw_long.append(r)
        FRED_Obs_df = (pd.concat(raw_long, ignore_index=True)
                       if raw_long else
                       pd.DataFrame(columns=['DATE','VALUE','SeriesID']))
        FRED_Obs_df['DATE'] = pd.to_datetime(FRED_Obs_df['DATE'], errors='coerce')
        FRED_Obs_df['VALUE'] = pd.to_numeric(FRED_Obs_df['VALUE'], errors='coerce')
        FRED_Obs_df['SeriesID'] = FRED_Obs_df['SeriesID'].astype(str)
        FRED_Obs_df = FRED_Obs_df.dropna(subset=['DATE']).sort_values(['SeriesID','DATE'])
        # expose for downstream writer
        self.last_fred_obs_df = FRED_Obs_df
        # === END PATCH S-RAW2 ===

        # === PATCH FREQ-2: infer frequency for series with missing/unknown metadata ===
        def _infer_freq_from_index(idx: pd.DatetimeIndex) -> tuple[str, str]:
            """Return ('daily'|'monthly'|'quarterly', 'D'|'M'|'Q') using raw index cadence."""
            idx = pd.DatetimeIndex(idx).sort_values().unique()
            if len(idx) < 3:
                return ("monthly", "M")  # conservative default
            guess = pd.infer_freq(idx)
            if guess:
                g = guess.upper()
                if g.startswith(("Q", "QS")): return ("quarterly", "Q")
                if g.startswith(("M", "MS")): return ("monthly", "M")
                if g.startswith(("D", "B", "C", "W")): return ("daily", "D")
            # Heuristic fallback: med observations per year
            vc = pd.Series(idx.year).value_counts()
            med_per_year = float(vc.median()) if not vc.empty else 0.0
            if med_per_year >= 200: return ("daily", "D")
            if 6 <= med_per_year <= 15: return ("monthly", "M")
            if 3 <= med_per_year <= 5: return ("quarterly", "Q")
            # last resort: bucket by month uniqueness
            months_per_year = pd.Series(list(zip(idx.year, idx.month))).drop_duplicates().groupby(lambda x: x[0]).size().median()
            if months_per_year and months_per_year >= 6: return ("monthly", "M")
            return ("quarterly", "Q")

        # Build a quick lookup of existing metadata
        if metadata_df is None or metadata_df.empty:
            metadata_df = pd.DataFrame(columns=["Series ID","frequency","frequency_short","units"])

        present = set(metadata_df["Series ID"].astype(str)) if not metadata_df.empty else set()

        # Candidates: (a) not in metadata_df, or (b) frequency missing/Unknown
        def _needs_infer(row) -> bool:
            f = str(row.get("frequency", "")).strip().lower()
            fs = str(row.get("frequency_short", "")).strip().lower()
            return (f in ("", "nan", "none", "unknown")) and (fs in ("", "nan", "none", "unknown"))

        need_rows = []
        for sid in merged_df.columns:
            ss = str(sid)
            if ss not in raw_df_map:  # no raw series to infer from
                continue
            if ss not in present:
                need_rows.append(ss)
            else:
                # check if this metadata row needs inference
                r = metadata_df.loc[metadata_df["Series ID"].astype(str)==ss]
                if r.empty or _needs_infer(r.iloc[0].to_dict()):
                    need_rows.append(ss)

        inferred_records = []
        for sid in need_rows:
            raw = raw_df_map.get(sid)
            if raw is None or raw.empty:
                continue
            idx = raw.index
            freq_long, freq_short = _infer_freq_from_index(idx)
            inferred_records.append({
                "Series ID": sid,
                "frequency": freq_long.title() if freq_long else "Unknown",
                "frequency_short": freq_short.upper() if freq_short else "",
                "units": np.nan  # leave units; your downstream formatter/heuristic handles it
            })

        if inferred_records:
            inferred_df = pd.DataFrame(inferred_records)
            # Update/append intelligently
            if metadata_df.empty:
                metadata_df = inferred_df
            else:
                md = metadata_df.set_index("Series ID")
                idf = inferred_df.set_index("Series ID")
                md.update(idf)  # update rows that exist
                to_add = idf.loc[~idf.index.isin(md.index)]
                metadata_df = pd.concat([md, to_add]).reset_index()

        self.logger.info(f"[FREQ] Metadata rows: {len(metadata_df)} (after inference)")
        # === END PATCH FREQ-2 ===

        # 3. Compile Descriptions DataFrame
        # Logic: If caller provided descriptions, use them. Otherwise return empty or fetched ones (if implemented).
        if series_descriptions is not None and not series_descriptions.empty:
            desc_df = series_descriptions
        else:
            desc_df = pd.DataFrame()

        self.logger.info(f"Successfully fetched {len(successful_dfs)} series with metadata.")

        return merged_df.reset_index(), desc_df, failed_series, metadata_df
class BankMetricsProcessor:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config

    def _get_series(self, df: pd.DataFrame, series_list: List[str]) -> pd.Series:
        total = pd.Series(0, index=df.index)
        for col in series_list:
            if col in df:
                total += df[col].fillna(0)
        return total

    def _safe_divide(self, numerator, denominator, fill_value=0.0):
        if np.isscalar(numerator) and np.isscalar(denominator):
            return numerator / denominator if denominator != 0 else fill_value

        if isinstance(denominator, pd.Series):
            denom_safe = denominator.replace(0, np.nan)
            return (numerator / denom_safe).fillna(fill_value)

        try:
            return (numerator / denominator).fillna(fill_value)
        except:
            return fill_value

    def create_derived_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty: return df
        df_processed = df.copy()

        # [GUARD LOG] Confirm FFIEC data made it this far
        ric_source_cols = ["RCFDJ469", "RCONJ469", "RCFDJ474", "RCONJ474"]
        present = [c for c in ric_source_cols if c in df_processed.columns]
        logging.info(f"[RIC DERIVE] Source cols present at derive-time: {present}")

        # [E] RI-C DISAGGREGATED ALLOWANCE (Robust Logic)
        # -----------------------------------------------------------------------
        required_ric = [
            "RCFDJ466", "RCONJ466", "RCFDJ467", "RCONJ467", "RCFDJ468", "RCONJ468",
            "RCFDJ469", "RCONJ469", "RCFDJ470", "RCONJ470", "RCFDJ471", "RCONJ471",
            "RCFDJ472", "RCONJ472", "RCFDJ474", "RCONJ474"
        ]

        # Check if we actually have data to work with
        has_ric_data = any(c in df_processed.columns for c in required_ric)

        if not has_ric_data:
            logging.warning("[RIC DERIVE] No RI-C columns present. Setting all to NaN.")
            df_processed['RIC_SOURCE_AVAILABLE'] = 0
            for suffix in ['Constr', 'CommRE', 'Resi', 'Comm', 'Card', 'OthCons', 'Unalloc', 'Other']:
                df_processed[f'RIC_{suffix}_Best'] = np.nan
        else:
            df_processed['RIC_SOURCE_AVAILABLE'] = 1

            def best_of(df, cons_col, dom_col):
                s_cons = df[cons_col] if cons_col in df.columns else pd.Series(np.nan, index=df.index)
                s_dom  = df[dom_col]  if dom_col  in df.columns else pd.Series(np.nan, index=df.index)
                return s_cons.fillna(s_dom)

            df_processed['RIC_Constr_Best']  = best_of(df_processed, 'RCFDJ466', 'RCONJ466')
            df_processed['RIC_CommRE_Best']  = best_of(df_processed, 'RCFDJ467', 'RCONJ467')
            df_processed['RIC_Resi_Best']    = best_of(df_processed, 'RCFDJ468', 'RCONJ468')
            df_processed['RIC_Comm_Best']    = best_of(df_processed, 'RCFDJ469', 'RCONJ469')
            df_processed['RIC_Card_Best']    = best_of(df_processed, 'RCFDJ470', 'RCONJ470')
            df_processed['RIC_OthCons_Best'] = best_of(df_processed, 'RCFDJ471', 'RCONJ471')
            df_processed['RIC_Unalloc_Best'] = best_of(df_processed, 'RCFDJ472', 'RCONJ472')
            df_processed['RIC_Other_Best']   = best_of(df_processed, 'RCFDJ474', 'RCONJ474')

        # [A] Field Resolution Layer
        logging.info("Applying Field Resolution Layer...")
        for field, sources in FDIC_FALLBACK_MAP.items():
            if "FFIEC" in sources: continue

            def resolve(row, src=sources, fld=field):
                if "DERIVED" in src:
                    if fld == "LNCONOTHX": return max(0, row.get('LNCON',0)-row.get('LNAUTO',0)-row.get('LNCRCD',0))
                    return 0
                for s in src:
                    val = row.get(s)
                    if pd.notna(val) and val != 0: return val
                return 0

            df_processed[field] = df_processed.apply(resolve, axis=1)

        # [B] YTD -> Quarterly
        ytd_fields = [col for col in df_processed.columns if col.startswith('NT') or 'EINTEXP' in col]
        for cert, group in df_processed.groupby('CERT'):
            group = group.sort_values('REPDTE')
            for col in ytd_fields:
                quarterly_vals = group[col].diff()
                q1_mask = group['REPDTE'].dt.quarter == 1
                quarterly_vals.loc[q1_mask] = group.loc[q1_mask, col]
                df_processed.loc[group.index, f"{col}_Q"] = quarterly_vals

        # [C] Top Level
        df_processed['Total_Nonaccrual'] = self._get_series(df_processed, ['NACI','NARENROT','NARECONS','NARERES','NACON'])
        df_processed['Total_ACL'] = df_processed.get('LNATRES', 0).fillna(0) + \
                                    (df_processed.get('RB2LNRES', 0).fillna(0) - df_processed.get('LNATRES', 0).fillna(0))
        df_processed['Total_Capital'] = df_processed.get('RBCT1J', 0) + df_processed.get('RBCT2', 0)

        # [D] Categories
        category_balances = {}
        category_balances['SBL'] = df_processed.get('LNOTHPCS', 0)
        df_processed['SBL_Balance'] = category_balances['SBL']

        category_balances['Fund_Finance'] = df_processed.get('LNOTHNONDEP', 0)
        df_processed['Fund_Finance_Balance'] = category_balances['Fund_Finance']

        df_processed['Wealth_Resi_Balance'] = self._get_series(df_processed, ['LNRERES', 'LNRELOC'])
        category_balances['Wealth_Resi'] = df_processed['Wealth_Resi_Balance']

        df_processed['Consumer_Auto_Balance'] = df_processed.get('LNAUTO', 0)
        category_balances['Consumer_Auto'] = df_processed['Consumer_Auto_Balance']

        df_processed['Consumer_Other_Balance'] = df_processed.get('LNCONOTHX', 0) + df_processed.get('LNCRCD', 0)
        category_balances['Consumer_Other'] = df_processed['Consumer_Other_Balance']

        df_processed['Corp_CI_Balance'] = df_processed.get('LNCI', 0)
        category_balances['Corp_CI'] = df_processed['Corp_CI_Balance']

        df_processed['CRE_OO_Balance'] = df_processed.get('LNRENROW', 0)
        category_balances['CRE_OO'] = df_processed['CRE_OO_Balance']

        df_processed['CRE_Investment_Balance'] = self._get_series(df_processed, ['LNRECONS', 'LNREMULT', 'LNRENROT'])
        category_balances['CRE_Investment'] = df_processed['CRE_Investment_Balance']

        total_categorized = sum(category_balances.values())

        # [F] Ratios
        ric_targets = ['RIC_Constr_Best', 'RIC_CommRE_Best', 'RIC_Resi_Best', 'RIC_Comm_Best',
                       'RIC_Card_Best', 'RIC_OthCons_Best', 'RIC_Unalloc_Best', 'RIC_Other_Best']

        ric_total = pd.Series(0.0, index=df_processed.index)
        for col in ric_targets:
            ric_total += df_processed[col].fillna(0)

        for cert, group in df_processed.groupby('CERT'):
            idx = group.index
            loans = group['LNLS'].fillna(0)
            acl = group['Total_ACL'].fillna(0)

            df_processed.loc[idx, 'Allowance_to_Gross_Loans_Rate'] = self._safe_divide(acl, loans)
            df_processed.loc[idx, 'Nonaccrual_to_Gross_Loans_Rate'] = self._safe_divide(group['Total_Nonaccrual'], loans)

            sbl_bal = group['SBL_Balance']
            df_processed.loc[idx, 'Risk_Adj_Allowance_Coverage'] = self._safe_divide(acl, (loans - sbl_bal))

            denom_acl = ric_total.loc[idx].replace(0, np.nan)

            df_processed.loc[idx, 'RIC_Constr_ACL_Pct'] = self._safe_divide(group['RIC_Constr_Best'], denom_acl)
            df_processed.loc[idx, 'RIC_CommRE_ACL_Pct'] = self._safe_divide(group['RIC_CommRE_Best'], denom_acl)
            df_processed.loc[idx, 'RIC_Resi_ACL_Pct'] = self._safe_divide(group['RIC_Resi_Best'], denom_acl)
            df_processed.loc[idx, 'RIC_Comm_ACL_Pct'] = self._safe_divide(group['RIC_Comm_Best'], denom_acl)
            df_processed.loc[idx, 'RIC_CreditCard_ACL_Pct'] = self._safe_divide(group['RIC_Card_Best'], denom_acl)
            df_processed.loc[idx, 'RIC_OtherCons_ACL_Pct'] = self._safe_divide(group['RIC_OthCons_Best'], denom_acl)
            df_processed.loc[idx, 'RIC_Other_ACL_Pct'] = self._safe_divide(group['RIC_Other_Best'], denom_acl)
            df_processed.loc[idx, 'RIC_Unallocated_ACL_Pct'] = self._safe_divide(group['RIC_Unalloc_Best'], denom_acl)

            # Scatter Groups
            total_loans_bank = total_categorized.loc[idx]

            comm_loan = group['Corp_CI_Balance'] + group['Fund_Finance_Balance']
            comm_acl = group['RIC_Comm_Best']
            df_processed.loc[idx, 'Group_Commercial_Loan_Share'] = self._safe_divide(comm_loan, total_loans_bank)
            df_processed.loc[idx, 'Group_Commercial_ACL_Share'] = self._safe_divide(comm_acl, denom_acl)

            resi_loan = group['Wealth_Resi_Balance']
            resi_acl = group['RIC_Resi_Best']
            df_processed.loc[idx, 'Group_Residential_Loan_Share'] = self._safe_divide(resi_loan, total_loans_bank)
            df_processed.loc[idx, 'Group_Residential_ACL_Share'] = self._safe_divide(resi_acl, denom_acl)

            cre_loan = group['CRE_OO_Balance'] + group['CRE_Investment_Balance']
            cre_acl = group['RIC_Constr_Best'] + group['RIC_CommRE_Best']
            df_processed.loc[idx, 'Group_CRE_Loan_Share'] = self._safe_divide(cre_loan, total_loans_bank)
            df_processed.loc[idx, 'Group_CRE_ACL_Share'] = self._safe_divide(cre_acl, denom_acl)

            other_loan = group['SBL_Balance'] + group['Consumer_Auto_Balance'] + group['Consumer_Other_Balance']
            other_acl = group['RIC_Other_Best'] + group['RIC_Card_Best'] + group['RIC_OthCons_Best']
            df_processed.loc[idx, 'Group_OtherSBL_Loan_Share'] = self._safe_divide(other_loan, total_loans_bank)
            df_processed.loc[idx, 'Group_OtherSBL_ACL_Share'] = self._safe_divide(other_acl, denom_acl)

        return df_processed

    def calculate_ttm_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty: return df

        # [FIX] Safe Growth that handles 0 -> 0 case
        def calc_safe_growth(curr, prev):
            if pd.isna(curr) or pd.isna(prev): return np.nan
            if prev == 0:
                return 0.0 if curr == 0 else np.nan # 0->0 is 0%
            return (curr - prev) / abs(prev) * 100

        all_banks_data = []
        for cert, group in df.groupby('CERT'):
            bank_df = group.sort_values("REPDTE").copy()

            if len(bank_df) >= 4:
                avg_loans = bank_df['LNLS'].rolling(4).mean()

                # TTM Growth
                for cat in LOAN_CATEGORIES.keys():
                    col_bal = f'{cat}_Balance'
                    if col_bal in bank_df.columns:
                        prev_bal = bank_df[col_bal].shift(4)
                        bank_df[f'{cat}_Growth_TTM'] = bank_df.apply(
                            lambda row: calc_safe_growth(row[col_bal], prev_bal.loc[row.name])
                            if row.name in prev_bal.index else np.nan, axis=1
                        )

                # Granular TTM Nonaccrual Rates
                for cat_name, cat_details in LOAN_CATEGORIES.items():
                    col_bal = f'{cat_name}_Balance'
                    col_na = f'{cat_name}_NA_Balance' # Ensure these exist or use mapped fields
                    # (Mapping handled in derivation if needed, usually NA cols are direct)

                    # For V6, we often map NA cols directly. If specific _NA_Balance cols
                    # aren't created, skip this or adapt.
                    # Assuming standard NA cols exist from previous steps:
                    pass

                # Standard TTM Metrics
                bank_df['TTM_NCO_Rate'] = self._safe_divide(bank_df['NTLNLS_Q'].rolling(4).sum(), avg_loans)
                bank_df['TTM_Past_Due_Rate'] = self._safe_divide(
                    (bank_df.get('P3LNLS', 0) + bank_df.get('P9LNLS', 0)).rolling(4).mean(),
                    avg_loans
                )

            all_banks_data.append(bank_df)

        return pd.concat(all_banks_data, ignore_index=True) if all_banks_data else df


    def calculate_8q_averages(self, proc_df: pd.DataFrame) -> pd.DataFrame:
        if proc_df.empty: return pd.DataFrame()
        metrics = proc_df.select_dtypes(include=np.number).columns.tolist()
        if 'CERT' in metrics: metrics.remove('CERT')
        results = []
        for cert, group in proc_df.groupby('CERT'):
            if len(group) < 8: continue
            group = group.sort_values('REPDTE')
            avgs = group[metrics].rolling(8).mean().iloc[-1]
            rec = {"CERT": cert, "NAME": group['NAME'].iloc[-1]}
            rec.update(avgs.to_dict())
            results.append(rec)
        return pd.DataFrame(results).set_index('CERT') if results else pd.DataFrame()

    def create_latest_snapshot(self, proc_df: pd.DataFrame) -> pd.DataFrame:
        if proc_df.empty: return pd.DataFrame()
        latest = proc_df[proc_df['REPDTE'] == proc_df['REPDTE'].max()].copy()
        snapshot = latest[['CERT', 'NAME', 'REPDTE']].copy()

        # 1. Base Metrics
        for m in ['LNLS', 'Total_ACL', 'Total_Capital', 'Allowance_to_Gross_Loans_Rate', 'TTM_NCO_Rate']:
             snapshot[m] = latest.get(m, np.nan)

        # 2. V6 Segment Metrics
        for cat in LOAN_CATEGORIES.keys():
            snapshot[f'{cat}_TTM_NA_Rate'] = latest.get(f'{cat}_TTM_NA_Rate', np.nan)
            snapshot[f'{cat}_Composition'] = latest.get(f'{cat}_Composition', np.nan)

        # 3. Scatter Plot Metrics (Group ACL Share vs Loan Share)
        groups = ['Commercial', 'Residential', 'CRE', 'OtherSBL']
        for g in groups:
            snapshot[f'Group_{g}_ACL_Share'] = latest.get(f'Group_{g}_ACL_Share', np.nan) * 100 # Convert to %
            snapshot[f'Group_{g}_Loan_Share'] = latest.get(f'Group_{g}_Loan_Share', np.nan) * 100 # Convert to %

        # 4. RRI Metrics
        for rri in ['RRI_Commercial', 'RRI_Residential', 'RRI_CRE', 'RRI_Other_SBL']:
            snapshot[rri] = latest.get(rri, np.nan)

        # 5. Raw RI-C Pcts (Optional, for reference)
        for col in latest.columns:
            if col.startswith('RIC_'):
                snapshot[col] = latest[col] * 100

        return snapshot.set_index('CERT')

class PeerAnalyzer:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.metric_descriptions = FDIC_FIELD_DESCRIPTIONS.copy()

    def create_peer_comparison(self, processed_df: pd.DataFrame) -> pd.DataFrame:
        logging.info("Creating multi-group peer comparison analysis...")
        if processed_df.empty: return pd.DataFrame()

        # Get latest date data
        latest_date = processed_df["REPDTE"].max()
        latest_data = processed_df[processed_df["REPDTE"] == latest_date].copy()

        # Isolate Subject Bank
        subject_data = latest_data[latest_data["CERT"] == self.config.subject_bank_cert]
        if subject_data.empty: return pd.DataFrame()

        metrics_to_compare = list(self.metric_descriptions.keys())
        comparison_list = []

        for metric in metrics_to_compare:
            if metric not in latest_data.columns: continue

            # 1. Base Record (Subject Bank Data)
            subject_value = subject_data[metric].iloc[0]
            record = {
                "Metric Code": metric,
                "Metric Name": self.metric_descriptions.get(metric, {}).get("short", metric),
                "Your_Bank": subject_value
            }

            # 2. Iterate through EACH Peer Group
            primary_percentile = None # To store percentile for the "Primary" (Core) group for flagging

            for group_key, group_info in PEER_GROUPS.items():
                group_name = group_info['short_name']
                group_certs = group_info['certs']

                # Filter data for this specific group
                group_data = latest_data[latest_data["CERT"].isin(group_certs)][metric].dropna()

                if group_data.empty: continue

                # Calculate Stats
                record[f"{group_name} Median"] = group_data.median()
                record[f"{group_name} Mean"] = group_data.mean()

                # Calculate Percentile (Rank)
                if pd.notna(subject_value):
                    pct = stats.percentileofscore(group_data, subject_value, kind='rank')
                    record[f"{group_name} Pct"] = pct

                    # Use CORE group as the driver for the "Performance Flag"
                    if group_key == PeerGroupType.CORE_PRIVATE_BANK:
                        primary_percentile = pct

            # 3. Generate Performance Flag (Based on Core Private Bank Peers)
            if primary_percentile is not None:
                record["Performance_Flag"] = self._get_performance_flag(metric, primary_percentile)
            else:
                record["Performance_Flag"] = "N/A"

            comparison_list.append(record)

        if not comparison_list: return pd.DataFrame()

        comparison_df = pd.DataFrame(comparison_list)
        return comparison_df

    def _get_performance_flag(self, metric_code: str, percentile: float) -> str:
        """Determines if high/low percentile is good/bad based on metric type."""
        metric = metric_code.lower()

        # Metrics where LOWER is better
        lower_is_better_terms = ["nco", "npl", "past_due", "nonaccrual", "cost_of_funds", "pd30", "pd90", "risk"]
        is_risk_metric = any(term in metric for term in lower_is_better_terms)

        if is_risk_metric:
            if percentile <= 25: return "Top Quartile (Low Risk)"
            if percentile <= 50: return "Better than Median"
            return "Bottom Quartile (High Risk)"

        # Metrics where HIGHER is better (Profitability, Capital, Growth)
        else:
            if percentile >= 75: return "Top Quartile (Strong)"
            if percentile >= 50: return "Better than Median"
            return "Bottom Quartile (Weak)"

class MacroTrendAnalyzer:
    """
    Analyzes macroeconomic time series data, calculates technical indicators,
    and generates structured output for Power BI dashboards.
    """
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.indicator_metadata = self._initialize_metadata()
        self.default_technical_params = self._get_default_tech_params()
        self.technical_params = self._set_technical_parameters()



    def _get_default_tech_params(self) -> Dict:
        """Centralizes the definition of default technical parameters."""
        return {
            'sma_period': 12, 'ema_period': 12,
            'bb_period': 12, 'bb_std': 2.0,
            'rsi_period': 14, 'z_score_period': 12
        }
    def _initialize_metadata(self) -> Dict:
        """
        Defines a comprehensive metadata repository for each macroeconomic indicator.
        This dictionary is the "brain" of the automated analysis, providing context
        for interpretation, classification, and commentary generation.

        Returns:
            Dict]: A nested dictionary where keys are FRED series IDs
                                       and values are dictionaries of metadata attributes.
        """
        return {
            # Leading Indicators
            "USSLIND": {
                "type": "leading", "horizon": "medium", "sectors": ["economy_wide"],
                "benchmark": {"recession": 98.0, "expansion": 102.0}, "annualized": False,
                "best_technical": "EMA", "technical_reason": "Highly responsive to economic turning points."
            },
            "T10Y2Y": { # Assuming this was an existing series
                "type": "leading", "horizon": "long", "sectors": ["banking", "real_estate", "economy_wide"],
                "benchmark": {"inversion": 0.0, "normal": 1.5}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Stable trend indicator for yield curve shape."
            },
            "USALOLITONOSTSAM": {
                "type": "leading", "horizon": "medium", "sectors": ["economy_wide", "global"],
                "benchmark": {"contraction": 99.5, "expansion": 100.5}, "annualized": False,
                "best_technical": "EMA", "technical_reason": "Captures momentum shifts in OECD economic outlook."
            },
            "DRTSCIS": {
                "type": "leading", "horizon": "short", "sectors": ["banking", "credit", "middle_market"],
                "benchmark": {"tightening": 20.0, "easing": -10.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Breakouts signal significant shifts in lending sentiment."
            },

            # Banking & Credit Quality
            "DRCRELEXFACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "real_estate", "credit"],
                "benchmark": {"stress": 3.0, "normal": 1.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Volatility in delinquencies indicates credit cycle stress."
            },
            "CORBLACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "credit", "corporate"],
                "benchmark": {"stress": 1.5, "normal": 0.5}, "annualized": True,
                "best_technical": "SMA", "technical_reason": "Identifies the underlying trend in realized credit losses."
            },
            "DRALACBS": {
                "type": "lagging", "horizon": "medium", "sectors": ["banking", "credit", "consumer"],
                "benchmark": {"stress": 3.5, "normal": 2.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Smooths quarterly data to reveal the broad credit quality trend."
            },
            "CORALACBN": {
                "type": "lagging", "horizon": "long", "sectors": ["banking", "credit", "economy_wide"],
                "benchmark": {"stress": 2.0, "normal": 0.75}, "annualized": True,
                "best_technical": "SMA", "technical_reason": "Tracks the long-term cycle of aggregate bank loan losses."
            },
            "SOFR3MTB3M": {
                "type": "coincident", "horizon": "short", "sectors": ["banking", "funding", "credit"],
                "benchmark": {"stress": 0.5, "extreme_stress": 1.0}, "annualized": False,
                "best_technical": "Bollinger Bands", "technical_reason": "Band expansion signals rising funding stress and counterparty risk."
            },

            # Employment
            "PAYEMS": {
                "type": "coincident", "horizon": "short", "sectors": ["economy_wide", "employment"],
                "benchmark": {"contraction_yoy_pct": 0.0}, "annualized": False,
                "best_technical": "RSI", "technical_reason": "Identifies momentum exhaustion in labor market growth."
            },
            "CEU6562000101": {
                "type": "coincident", "horizon": "medium", "sectors": ["healthcare", "employment"],
                "benchmark": {"contraction_yoy_pct": 0.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Healthcare employment is a stable trend, best captured by SMA."
            },

            # Real Estate & Construction
            "NYXRSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "LXXRSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "MIXRNSA": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Smooths monthly price volatility to show the primary trend."},
            "CASTHPI": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Ideal for smoothing quarterly state-level HPI data."},
            "FLSTHPI": {"type": "coincident", "horizon": "medium", "sectors": ["real_estate"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Ideal for smoothing quarterly state-level HPI data."},
            "MPCT04XXS": {
                "type": "coincident", "horizon": "medium", "sectors": ["healthcare", "construction"],
                "benchmark": {"contraction": 0.0}, "annualized": False,
                "best_technical": "SMA", "technical_reason": "Construction spending data is volatile; SMA reveals the underlying investment trend."
            },

            # Global Benchmarks
            "FPCPITOTLZGWLD": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 5.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Annual data requires a long-term moving average to identify global inflation regimes."},
            "FPCPITOTLZGHIC": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 4.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Tracks inflation trends in developed economies."},
            "FPCPITOTLZGLMY": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {"high_inflation": 6.0}, "annualized": True, "best_technical": "SMA", "technical_reason": "Tracks inflation trends in emerging economies."},
            "NYGDPMKTPCDWLD": {"type": "lagging", "horizon": "long", "sectors": ["global", "economy_wide"], "benchmark": {}, "annualized": False, "best_technical": "SMA", "technical_reason": "Annual data requires a long-term moving average to identify global growth regimes."},
        }

    def _set_technical_parameters(self) -> Dict:
        """
        Sets default and series-specific parameters for technical indicator calculations.
        These can be tuned to optimize for different data frequencies and volatilities.

        Returns:
            Dict]: A dictionary of parameters for each indicator.
        """
        # Default parameters are chosen based on common practice for economic data
        default_params = self.default_technical_params

        specific_params = {
            "T10Y2Y": {'sma_period': 50, 'ema_period': 20, 'bb_period': 50, 'rsi_period': 14, 'z_score_period': 50},
            "USSLIND": {'sma_period': 6, 'ema_period': 6, 'bb_period': 6, 'rsi_period': 14, 'z_score_period': 6},
            "PAYEMS": {'sma_period': 12, 'ema_period': 12, 'bb_period': 20, 'rsi_period': 14, 'z_score_period': 12},
        }

        params = {}
        for series_id in self.indicator_metadata.keys():
            params[series_id] = default_params.copy()
            if series_id in specific_params:
                params[series_id].update(specific_params[series_id])
        return params
    def _calculate_rsi(self, series: pd.Series, period: int) -> pd.Series:
        """
        Helper function to calculate the Relative Strength Index (RSI).
        """
        delta = series.diff(1)
        gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()

        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return rsi
    def _safe_divide(self, numerator, denominator):
        if isinstance(denominator, pd.Series):
            denominator = denominator.replace(0, np.nan)
        return (numerator / denominator).fillna(0)
    def _compute_trend_slope(self, data: pd.Series, window: int) -> float:
        """
        Computes the linear trend slope over a specified window.
        Returns the slope as an annualized percentage change.
        """
        if len(data) < window or data.isna().all():
            return 0.0
        y = data.iloc[-window:].values
        x = np.arange(len(y))
        # Remove NaN values
        mask = ~np.isnan(y)
        if mask.sum() < 2:  # Need at least 2 points
            return 0.0
        x_clean = x[mask]
        y_clean = y[mask]
        try:
            slope, _ = np.polyfit(x_clean, y_clean, 1)
            # Convert to annualized percentage
            mean_val = np.mean(y_clean)
            if mean_val != 0:
                return (slope * 12 / mean_val) * 100
            else:
                return 0.0
        except:
            return 0.0
    def calculate_technical_indicators(self, series: pd.Series, series_id: str) -> pd.DataFrame:
        """
        Calculates a suite of technical indicators for a given time series.
        """
        # Validate input
        if series is None or series.empty or len(series) == 0:
            logger.warning(f"Empty or invalid series data for {series_id}")
            return pd.DataFrame()

        # Ensure series has a datetime index
        if not isinstance(series.index, pd.DatetimeIndex):
            logger.warning(f"Series {series_id} does not have DatetimeIndex, skipping")
            return pd.DataFrame()

        # Drop NaN values before resampling
        series = series.dropna()

        if len(series) < 2:  # Need at least 2 data points
            logger.warning(f"Series {series_id} has fewer than 2 data points after dropping NaN")
            return pd.DataFrame()

        # Resample to monthly frequency
        try:
            series_monthly = series.resample('MS').ffill()
        except Exception as e:
            logger.warning(f"Resampling failed for {series_id}: {e}")
            return pd.DataFrame()

        # Verify resampling produced valid output
        if not isinstance(series_monthly, pd.Series):
            logger.warning(f"Resampling produced non-Series output for {series_id}")
            return pd.DataFrame()

        if series_monthly.empty or len(series_monthly) < 2:
            logger.warning(f"Resampling resulted in insufficient data for {series_id}")
            return pd.DataFrame()

        # Additional check: ensure it's not a scalar
        if series_monthly.ndim != 1:
            logger.warning(f"Series {series_id} has invalid dimensions after resampling")
            return pd.DataFrame()

        params = self.technical_params.get(series_id, self.default_technical_params)

        # Create DataFrame - this should now be safe
        try:
            df = pd.DataFrame({series_id: series_monthly})
            df.index.name = 'date'
        except Exception as e:
            logger.error(f"Failed to create DataFrame for {series_id}: {e}")
            return pd.DataFrame()

        # Calculate indicators (rest of your existing code)
        df[f'{series_id}_SMA'] = series_monthly.rolling(window=params['sma_period'], min_periods=1).mean()
        df[f'{series_id}_EMA'] = series_monthly.ewm(span=params['ema_period'], adjust=False, min_periods=1).mean()

        rolling_mean = series_monthly.rolling(window=params['bb_period'], min_periods=1).mean()
        rolling_std = series_monthly.rolling(window=params['bb_period'], min_periods=1).std()
        df[f'{series_id}_BB_Upper'] = rolling_mean + (params['bb_std'] * rolling_std)
        df[f'{series_id}_BB_Lower'] = rolling_mean - (params['bb_std'] * rolling_std)
        df[f'{series_id}_BB_Width'] = self._safe_divide((df[f'{series_id}_BB_Upper'] - df[f'{series_id}_BB_Lower']), rolling_mean)

        df[f'{series_id}_RSI'] = self._calculate_rsi(series_monthly, params['rsi_period'])

        z_score_mean = series_monthly.rolling(window=params['z_score_period'], min_periods=1).mean()
        z_score_std = series_monthly.rolling(window=params['z_score_period'], min_periods=1).std()
        df[f'{series_id}_ZScore'] = self._safe_divide((series_monthly - z_score_mean), z_score_std)

        return df

    def _get_optimal_technical_indicator(self, series_id: str) -> Dict[str, Any]:
        """
        Selects the optimal technical indicator and provides reasoning based on the
        economic characteristics of the series.
        Args:
            series_id (str): The FRED series ID.
        Returns:
            Dict[str, Any]: A dictionary containing the name of the best indicator
                            and the reasoning for its selection.
        """
        # 1. This dictionary now contains the reasons for choosing each indicator.
        # This fixes the original SyntaxError.
        selection_rules = {
            "SMA": {
                "reason": "Best for identifying long-term trends in stable, less volatile series."
            },
            "EMA": {
                "reason": "Best for tracking recent price changes and identifying short-term trends."
            },
            "Bollinger Bands": {
                "reason": "Best for assessing volatility and identifying overbought/oversold trend conditions."
            },
            "RSI": {
                "reason": "Best for measuring momentum and identifying potential trend reversals."
            }
        }

        # 2. Determine which indicator is best for the given series_id.
        # We still get this from the metadata, with "SMA" as a safe default.
        metadata = self.indicator_metadata.get(series_id, {})
        best_indicator = metadata.get("best_technical", "SMA")

        # 3. Look up the specific reason from the rules dictionary using the best_indicator.
        # We use .get() to provide a default reason in case the indicator isn't in our rules.
        default_rule = {"reason": "Default trend-following indicator used as no specific rule was found."}
        reason = selection_rules.get(best_indicator, default_rule)["reason"]

        # 4. Return the final output, which is now ready to be used in your required column.
        return {
            "indicator": best_indicator,
            "reason": reason
        }
    def calculate_oecd_aggregates(self, fred_data: pd.DataFrame) -> pd.DataFrame:
        """
        Calculates or extracts aggregates for OECD vs. non-OECD country groups
        using World Bank proxies available on FRED.

        Args:
            fred_data (pd.DataFrame): The DataFrame containing all fetched FRED series.

        Returns:
            pd.DataFrame: A DataFrame with the aggregated OECD and non-OECD series.
        """
        # 1. Define constants for the specific series IDs being used as proxies.
        # These are the actual FRED series for inflation in different income groups.
        HI_INCOME_SERIES = 'FPCPITOTLZGHI'   # Proxy for OECD countries
        LOW_MID_INCOME_SERIES = 'FPCPITOTLZGLM' # Proxy for Non-OECD countries
        WORLD_SERIES = 'FPCPITOTLZGOW'   # World aggregate

        # 2. Assign the list of required columns to the variable. This fixes the SyntaxError.
        required_cols = [HI_INCOME_SERIES, LOW_MID_INCOME_SERIES, WORLD_SERIES]

        # Check if all required columns are present in the input DataFrame.
        if not all(col in fred_data.columns for col in required_cols):
            print("Warning: Not all global inflation series found for OECD aggregation.")
            return pd.DataFrame()

        # 3. Create the output by selecting the CORRECT column for each key.
        # This fixes the logical error/ValueError.
        output_df = pd.DataFrame({
            'OECD_Proxy_Inflation': fred_data[HI_INCOME_SERIES],
            'NonOECD_Proxy_Inflation': fred_data[LOW_MID_INCOME_SERIES],
            'World_Inflation': fred_data[WORLD_SERIES]
        })

        # The data is annual, so forward-fill to create a step-function series
        # that can be merged with higher-frequency data.
        output_df.ffill(inplace=True)

        return output_df
    def _detect_frequency(self, series: pd.Series) -> str:
        """
        Infers the frequency of a time series index.
        """
        series = series.dropna()
        if len(series.index) < 3:
            return "Undetermined"
        return pd.infer_freq(series.index) or "Irregular"

    def validate_series_data(self, series_id: str, data: pd.Series) -> Dict[str, Any]:
        """
        Performs data quality checks on a time series and returns a validation summary.
        """
        total = len(data)
        missing = data.isna().sum()
        pct_missing = missing / total * 100 if total else 100
        validation_results = {
            "status": "OK",
            "missing_data_pct": pct_missing,
            "last_update": data.last_valid_index(),
            "data_frequency": self._detect_frequency(data),
        }
        # NEW - safer handling
        try:
            clean_data = data.dropna()
            if len(clean_data) > 3:  # Need at least a few points for z-score
                z_scores = np.abs(stats.zscore(clean_data))
                outlier_count = (z_scores > 3).sum()
                # Handle both scalar and Series results
                validation_results["outliers_detected"] = int(outlier_count) if np.isscalar(outlier_count) else int(outlier_count.item())
            else:
                validation_results["outliers_detected"] = 0
        except Exception as e:
            logger.warning(f"Could not calculate outliers for {series_id}: {e}")
            validation_results["outliers_detected"] = 0
        return validation_results
    def _get_technical_signal(self, series_id: str, data: pd.DataFrame, horizon: str = 'medium') -> str:
        """
        Generates a sophisticated signal by analyzing the state, duration, and
        recent events of the optimal technical indicator.

        Args:
            series_id (str): The FRED series ID (e.g., 'DGS10').
            data (pd.DataFrame): DataFrame containing the series and its technical indicators.
            horizon (str): The lookback horizon ('short', 'medium', 'long') to define
                           the period for signal confirmation.

        Returns:
            str: A descriptive signal string.
        """
        # 1. Map the horizon string to a lookback period (number of data points)
        horizon_map = {'short': 2, 'medium': 5, 'long': 10}
        lookback = horizon_map.get(horizon, 5)

        # Ensure we have enough data to look back
        if len(data) < lookback + 1:
            return "Awaiting Data (Insufficient History)"

        # --- Determine the optimal indicator to use ---
        metadata = self.indicator_metadata.get(series_id, {})
        best_tech = metadata.get("best_technical", "SMA")

        # --- SMA / EMA Logic ---
        if best_tech in ["SMA", "EMA"]:
            col_name = f"{series_id}_{best_tech}"
            if col_name not in data.columns: return "Signal N/A"

            # Get the last two points for crossover detection
            price_now = data[series_id].iloc[-1]
            price_prev = data[series_id].iloc[-2]
            tech_now = data[col_name].iloc[-1]
            tech_prev = data[col_name].iloc[-2]

            if pd.isna(price_now) or pd.isna(tech_now): return "Awaiting Data"

            # Crossover Detection
            if price_prev <= tech_prev and price_now > tech_now:
                return f"Bullish Crossover ({best_tech})"
            if price_prev >= tech_prev and price_now < tech_now:
                return f"Bearish Crossover ({best_tech})"

            # Confirmed Trend Detection (checks the entire lookback period)
            if all(data[series_id].iloc[-lookback:] > data[col_name].iloc[-lookback:]):
                return f"Confirmed Bullish Trend (Above {best_tech} for {lookback} periods)"
            if all(data[series_id].iloc[-lookback:] < data[col_name].iloc[-lookback:]):
                return f"Confirmed Bearish Trend (Below {best_tech} for {lookback} periods)"

            return "Neutral (Consolidating)"

        # --- Bollinger Bands Logic ---
        elif best_tech == "Bollinger Bands":
            upper_col, lower_col = f"{series_id}_BB_Upper", f"{series_id}_BB_Lower"
            if upper_col not in data.columns or lower_col not in data.columns: return "Signal N/A"

            price_now = data[series_id].iloc[-1]
            upper_now, lower_now = data[upper_col].iloc[-1], data[lower_col].iloc[-1]

            if any(pd.isna([price_now, upper_now, lower_now])): return "Awaiting Data"

            # Check for breakouts or reversions
            if price_now > upper_now:
                # Check if it's a new breakout or a sustained one
                if data[series_id].iloc[-2] <= data[upper_col].iloc[-2]:
                    return "Volatility Breakout (High Stress)"
                else:
                    return "Sustained High Volatility"
            elif price_now < lower_now:
                if data[series_id].iloc[-2] >= data[lower_col].iloc[-2]:
                     return "Volatility Breakout (Low/Recession)"
                else:
                    return "Sustained Low Pressure"

            return "Neutral (Range-bound)"

        # --- RSI Logic ---
        elif best_tech == "RSI":
            rsi_col = f"{series_id}_RSI"
            if rsi_col not in data.columns: return "Signal N/A"

            rsi_val = data[rsi_col].iloc[-1]
            if pd.isna(rsi_val): return "Awaiting Data"

            # Add magnitude to the signal
            if rsi_val > 80: return "Extreme Overbought (High Momentum)"
            if rsi_val > 70: return "Overbought"
            if rsi_val < 20: return "Extreme Oversold (Low Momentum)"
            if rsi_val < 30: return "Oversold"

            # Check for momentum change
            avg_rsi_lookback = data[rsi_col].iloc[-lookback:].mean()
            if rsi_val > avg_rsi_lookback and rsi_val > 50: return "Gaining Bullish Momentum"
            if rsi_val < avg_rsi_lookback and rsi_val < 50: return "Gaining Bearish Momentum"

            return "Neutral Momentum"

        return "Not Implemented"

    def _generate_short_term_analysis(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates markdown-formatted short-term (1-3 months) analysis for a series.
        This is applied row-wise to generate commentary for each point in time.
        """
        def generate_row_analysis(row):
            current_value = row[series_id]
            # Get previous row's value for change calculation
            prev_row_index = data.index.get_loc(row.name) - 1
            if prev_row_index < 0:
                change = 0.0
            else:
                prior_value = data[series_id].iloc[prev_row_index]
                change = current_value - prior_value

            if pd.isna(current_value):
                return "No data available for this period."

            tech_signal = self._get_technical_signal(series_id, data.loc[:row.name], 'short')
            annualized_text = ' (annualized)' if metadata.get('annualized', False) else ''

            # Simple outlook logic
            outlook = "Stable"
            if "Bullish" in tech_signal or "High" in tech_signal:
                outlook = "Improving"
            elif "Bearish" in tech_signal or "Low" in tech_signal:
                outlook = "Worsening"

            return (
                f"**Current Level**: {current_value:,.2f}{annualized_text}\n\n"
                f"**Change (MoM)**: {change:+.2f}\n\n"
                f"**Technical Signal**: {tech_signal}\n\n"
                f"**Outlook**: Short-term trend appears to be **{outlook}**."
            )

        # Apply the function to each row of the DataFrame
        return data.apply(generate_row_analysis, axis=1)


    def _generate_long_term_analysis(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates markdown-formatted long-term (1-3 years) analysis for a series.
        Includes trends, momentum, historical context, and regime identification.
        """
        def generate_row_analysis(row):
            current_value = row[series_id]
            current_date = row.name
            if pd.isna(current_value):
                return "No data available for this period."
            # Get historical data up to current point
            historical_data = data.loc[:current_date, series_id].dropna()
            if len(historical_data) < 12:  # Need at least 1 year of data
                return "Insufficient historical data for long-term analysis."
            # 1. Calculate long-term trends (1-3 year windows)
            analysis_parts = []
            # Year-over-year change
            if len(historical_data) >= 12:
                yoy_value = historical_data.iloc[-13] if len(historical_data) >= 13 else historical_data.iloc[0]
                yoy_change = ((current_value - yoy_value) / yoy_value) * 100 if yoy_value != 0 else 0
                yoy_change_abs = current_value - yoy_value
            else:
                yoy_change = 0
                yoy_change_abs = 0
            # Calculate long-term moving averages
            ma_12m = historical_data.rolling(window=12, min_periods=12).mean().iloc[-1] if len(historical_data) >= 12 else current_value
            ma_24m = historical_data.rolling(window=24, min_periods=24).mean().iloc[-1] if len(historical_data) >= 24 else ma_12m
            # Calculate trend slope (annualized percentage change over 12-24 months)
            if len(historical_data) >= 24:
                window_data = historical_data.iloc[-24:]
                x = np.arange(len(window_data))
                y = window_data.values
                if len(y) > 0 and not np.all(np.isnan(y)):
                    slope, _ = np.polyfit(x, y, 1)
                    # Annualize the slope as percentage of mean
                    trend_slope_pct = (slope * 12 / np.mean(y)) * 100 if np.mean(y) != 0 else 0
                else:
                    trend_slope_pct = 0
            else:
                trend_slope_pct = 0
            # Calculate z-score over 3-year window
            if len(historical_data) >= 36:
                rolling_mean = historical_data.rolling(window=36).mean().iloc[-1]
                rolling_std = historical_data.rolling(window=36).std().iloc[-1]
                z_score_3y = (current_value - rolling_mean) / rolling_std if rolling_std > 0 else 0
            else:
                z_score_3y = 0
            # Calculate historical percentile (5-year window)
            if len(historical_data) >= 60:
                percentile_rank = stats.percentileofscore(historical_data.iloc[-60:], current_value)
            elif len(historical_data) >= 12:
                percentile_rank = stats.percentileofscore(historical_data, current_value)
            else:
                percentile_rank = 50
            # Determine regime and context based on metadata
            annualized_text = ' (annualized)' if metadata.get('annualized', False) else ''
            benchmark = metadata.get('benchmark', {})
            # Build analysis text
            analysis_parts.append(f"**Current Level vs Historical Trend:** As of {current_date.strftime('%Y-%m')}, "
                                f"{metadata.get('short', series_id)} stands at {current_value:,.2f}{annualized_text} "
                                f"(YoY: {yoy_change:+.1f}%), which is in the {percentile_rank:.0f}th percentile "
                                f"of its 5-year historical range.")
            # Trend analysis
            trend_direction = "upward" if trend_slope_pct > 0.5 else "downward" if trend_slope_pct < -0.5 else "sideways"
            ma_signal = "above" if current_value > ma_12m else "below"
            analysis_parts.append(f"**Long-Term Trend:** The 24-month trend shows {trend_direction} momentum "
                                f"with an annualized slope of {trend_slope_pct:+.1f}%. "
                                f"Current value is {ma_signal} its 12-month moving average ({ma_12m:,.2f}).")
            # Volatility and regime analysis
            if abs(z_score_3y) > 2:
                volatility_status = "extreme high" if z_score_3y > 2 else "extreme low"
                analysis_parts.append(f"**Volatility/Alerts:** Recent readings have achieved a 3-year z-score of {z_score_3y:.2f}, "
                                    f"indicating {volatility_status} levels relative to historical norms.")
            elif abs(z_score_3y) > 1:
                volatility_status = "elevated" if z_score_3y > 1 else "depressed"
                analysis_parts.append(f"**Volatility/Alerts:** The indicator shows {volatility_status} levels "
                                    f"with a z-score of {z_score_3y:.2f}.")
            # Economic context based on benchmarks and series type
            context_parts = []
            # Check against specific benchmarks
            if benchmark:
                for regime, threshold in benchmark.items():
                    if 'stress' in regime and current_value > threshold:
                        context_parts.append(f"Value exceeds stress threshold of {threshold:.2f}")
                    elif 'recession' in regime and current_value < threshold:
                        context_parts.append(f"Below recession threshold of {threshold:.2f}")
                    elif 'expansion' in regime and current_value > threshold:
                        context_parts.append(f"In expansion territory (above {threshold:.2f})")
                    elif 'inversion' in regime and current_value < threshold:
                        context_parts.append(f"Yield curve inverted (below {threshold:.2f})")
            # Add series-specific context
            if 'inflation' in series_id.lower() or series_id == 'CPIAUCSL':
                if current_value > 2.5:
                    context_parts.append("Consumer inflation is above the Fed's 2% target")
                else:
                    context_parts.append("Inflation remains near or below target")
            if 'unemployment' in metadata.get('long', '').lower() or series_id == 'UNRATE':
                if percentile_rank < 25:
                    context_parts.append("Unemployment near historical lows suggests tight labor market")
                elif percentile_rank > 75:
                    context_parts.append("Elevated unemployment indicates labor market stress")
            if context_parts:
                analysis_parts.append(f"**Economic Context:** {'. '.join(context_parts)}.")
            return '\n\n'.join(analysis_parts)
        # Apply the function to each row
        return data.apply(generate_row_analysis, axis=1)

    def _generate_risk_assessment(self, series_id: str, data: pd.DataFrame, metadata: Dict) -> pd.Series:
        """
        Generates risk assessment with red/yellow/green flags and numeric scores (0-100).
        Evaluates deviation from benchmarks, volatility, and momentum.
        """
        def generate_row_assessment(row):
            current_value = row[series_id]
            current_date = row.name
            if pd.isna(current_value):
                return "**Risk Score:** N/A - No data available."
            # Get historical data
            historical_data = data.loc[:current_date, series_id].dropna()
            if len(historical_data) < 12:
                return "**Risk Score:** N/A - Insufficient data for risk assessment."
            # Initialize risk components
            risk_score = 0
            risk_factors = []
            risk_color = "Green"
            # 1. Deviation from historical norms (0-40 points)
            if len(historical_data) >= 36:
                rolling_mean = historical_data.rolling(window=36).mean().iloc[-1]
                rolling_std = historical_data.rolling(window=36).std().iloc[-1]
                z_score = abs((current_value - rolling_mean) / rolling_std) if rolling_std > 0 else 0
            else:
                rolling_mean = historical_data.mean()
                rolling_std = historical_data.std()
                z_score = abs((current_value - rolling_mean) / rolling_std) if rolling_std > 0 else 0
            # Map z-score to risk points (0-40)
            deviation_score = min(40, (z_score / 3) * 40)
            risk_score += deviation_score
            if z_score > 2:
                risk_factors.append(f"Extreme deviation ({z_score:.1f}σ)")
            elif z_score > 1:
                risk_factors.append(f"Moderate deviation ({z_score:.1f}σ)")
            # 2. Benchmark breach assessment (0-30 points)
            benchmark = metadata.get('benchmark', {})
            benchmark_score = 0
            for regime, threshold in benchmark.items():
                if 'stress' in regime.lower() or 'extreme' in regime.lower():
                    if current_value > threshold:
                        benchmark_score = 30
                        risk_factors.append(f"Exceeds {regime} level ({threshold:.2f})")
                    elif current_value > threshold * 0.8:  # Within 20% of stress
                        benchmark_score = max(benchmark_score, 20)
                        risk_factors.append(f"Approaching {regime} level")
                elif 'normal' in regime.lower():
                    if abs(current_value - threshold) / threshold > 0.5:
                        benchmark_score = max(benchmark_score, 15)
            risk_score += benchmark_score
            # 3. Volatility assessment (0-20 points)
            if len(historical_data) >= 12:
                recent_volatility = historical_data.iloc[-12:].std()
                long_term_volatility = historical_data.std()
                vol_ratio = recent_volatility / long_term_volatility if long_term_volatility > 0 else 1
                volatility_score = min(20, max(0, (vol_ratio - 1) * 20))
                risk_score += volatility_score
                if vol_ratio > 1.5:
                    risk_factors.append("High recent volatility")
            # 4. Momentum/trend assessment (0-10 points)
            if len(historical_data) >= 6:
                recent_trend = historical_data.iloc[-6:].values
                x = np.arange(len(recent_trend))
                if len(recent_trend) > 1:
                    slope, _ = np.polyfit(x, recent_trend, 1)
                    # For certain indicators, negative trends increase risk
                    indicator_type = metadata.get('type', '')
                    if indicator_type == 'leading' and slope < 0:
                        momentum_score = 10
                        risk_factors.append("Negative momentum in leading indicator")
                    elif 'delinquency' in metadata.get('long', '').lower() and slope > 0:
                        momentum_score = 10
                        risk_factors.append("Rising delinquency trend")
                    else:
                        momentum_score = 0
                    risk_score += momentum_score
            # 5. Special conditions for specific series
            if series_id == 'T10Y2Y' and current_value < 0:
                risk_score = max(risk_score, 80)  # Yield curve inversion is high risk
                risk_factors.append("Yield curve inverted")
            if 'spread' in series_id.lower() and series_id.startswith('BAM'):
                percentile = stats.percentileofscore(historical_data, current_value)
                if percentile > 90:
                    risk_score = max(risk_score, 70)
                    risk_factors.append("Credit spreads at extreme highs")
            # Determine color flag based on final score
            if risk_score >= 70:
                risk_color = "Red"
            elif risk_score >= 40:
                risk_color = "Yellow"
            else:
                risk_color = "Green"
            # Format the assessment
            assessment_parts = [
                f"**Risk Score:** {risk_score:.0f}/100 ({risk_color})"
            ]
            if risk_factors:
                assessment_parts.append(f"**Risk Drivers:** {', '.join(risk_factors)}")
            # Add specific risk interpretation
            if risk_color == "Red":
                assessment_parts.append("**Assessment:** High risk - Indicator signals significant stress or deviation from normal conditions.")
            elif risk_color == "Yellow":
                assessment_parts.append("**Assessment:** Moderate risk - Indicator shows concerning trends requiring close monitoring.")
            else:
                assessment_parts.append("**Assessment:** Low risk - Indicator within normal operating parameters.")
            return ' '.join(assessment_parts)
        # Apply the function to each row
        return data.apply(generate_row_assessment, axis=1)
    def generate_powerbi_output(self, processed_data: Dict) -> pd.DataFrame:
        """
        Generates a Power BI-optimized "long" format DataFrame from the processed
        technical indicator data. Also attaches metadata and generated commentary.
        """
        all_series_dfs = []
        for series_id, df in processed_data.items():
            if df.empty:
                continue
            # Use a copy to avoid side effects
            df_copy = df.copy()
            # --- Generate Commentary ---
            metadata = self.indicator_metadata.get(series_id, {})
            df_copy['Short_Term_Analysis'] = self._generate_short_term_analysis(series_id, df_copy, metadata)
            df_copy['Long_Term_Analysis'] = self._generate_long_term_analysis(series_id, df_copy, metadata)
            df_copy['Risk_Assessment'] = self._generate_risk_assessment(series_id, df_copy, metadata)

            # --- Add Metadata Columns ---
            commentary_cols = ['Short_Term_Analysis', 'Long_Term_Analysis', 'Risk_Assessment']
            df_copy['Type'] = metadata.get('type', 'N/A')
            df_copy['Horizon'] = metadata.get('horizon', 'N/A')
            df_copy['Sectors'] = '|'.join(metadata.get('sectors', []))
            metadata_cols = ['Type', 'Horizon', 'Sectors']
            date_col_name = df_copy.index.name
            df_copy.reset_index(inplace=True)
            id_vars = [date_col_name] + commentary_cols + metadata_cols

            value_vars = [col for col in df_copy.columns if col not in id_vars and col != series_id]

            base_df = df_copy[id_vars + [series_id]].copy()
            base_df['Indicator_Name'] = 'Actual Value'
            base_df.rename(columns={series_id: 'Indicator_Value', date_col_name: 'date'}, inplace=True)

            melted_df = pd.DataFrame()
            if value_vars:
                melted_df = df_copy.melt(
                    id_vars=id_vars,
                    value_vars=value_vars,
                    var_name='Indicator_Name',
                    value_name='Indicator_Value'
                )
                melted_df.rename(columns={date_col_name: 'date'}, inplace=True)
            series_final_df = pd.concat([base_df, melted_df], ignore_index=True)
            series_final_df['Series_ID'] = series_id
            all_series_dfs.append(series_final_df)
        if not all_series_dfs:
            return pd.DataFrame()
        # Concatenate all the processed series DataFrames into one big one
        final_output_df = pd.concat(all_series_dfs, ignore_index=True)
        return final_output_df


class ExcelOutputGenerator:
    """Generates the final Excel dashboard output file."""

    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.fdic_desc_df = pd.DataFrame.from_dict(FDIC_FIELD_DESCRIPTIONS, orient='index').reset_index()
        self.fdic_desc_df.rename(columns={'index': 'Metric Code', 'short': 'Metric Name', 'long': 'Description'}, inplace=True)

    def write_excel_output(self, file_path: str, **kwargs):
        """Writes all DataFrames to a single styled Excel file with multiple sheets."""
        logging.info(f"Writing final dashboard to: {file_path}")
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in kwargs.items():
                if isinstance(df, pd.DataFrame) and not df.empty:
                    n_rows, n_cols = df.shape
                    logging.info(f"Sheet '{sheet_name}' contains {n_rows} rows and {n_cols} cols.")
                    if n_rows > 1048576:
                        logging.error(f"Sheet '{sheet_name}' too large ({n_rows} rows).")
                    write_index = sheet_name in ["Latest_Peer_Snapshot", "Averages_8Q_All_Metrics", "Data_Validation_Report"]
                    df.to_excel(writer, sheet_name=sheet_name, index=write_index)
            logging.info("All data written, starting styling...")
            self._style_metric_descriptions_sheet(writer)
            self._apply_summary_styles(writer, kwargs.get("Summary_Dashboard"))
            self._apply_snapshot_styles(writer, kwargs.get("Latest_Peer_Snapshot"))
            self._apply_macro_analysis_styles(writer, kwargs.get("Macro_Analysis"))
        logging.info("Excel file written and styled successfully.")

    def _style_metric_descriptions_sheet(self, writer):
        """Styles the FDIC metric descriptions sheet."""
        sheet_name = 'FDIC_Metric_Descriptions'
        if sheet_name not in writer.sheets: return

        logging.info(f"Styling {sheet_name} sheet...")
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 25 # Metric Code
        worksheet.column_dimensions['B'].width = 35 # Metric Name
        worksheet.column_dimensions['C'].width = 80 # Description

    def _apply_summary_styles(self, writer, df: pd.DataFrame):
        """Applies conditional formatting to the Summary_Dashboard sheet using openpyxl."""
        if df is None or df.empty: return
        sheet_name = 'Summary_Dashboard'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")

        # Define styles using openpyxl objects
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        green_font = Font(color='006100')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        yellow_font = Font(color='9C6500')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006')

        # Set column widths
        worksheet.column_dimensions['A'].width = 15 # Metric Code
        worksheet.column_dimensions['B'].width = 35 # Metric Name
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
            worksheet.column_dimensions[col_letter].width = 18

        try:
            flag_col_index = df.columns.get_loc('Performance_Flag') + 1 # +1 for 0-based to 1-based
            flag_col_letter = openpyxl.utils.get_column_letter(flag_col_index)

            # Define conditional formatting rules
            green_rule = FormulaRule(formula=[f'SEARCH("Top Quartile",${flag_col_letter}2)'], fill=green_fill, font=green_font)
            yellow_rule = FormulaRule(formula=[f'SEARCH("Better than Median",${flag_col_letter}2)'], fill=yellow_fill, font=yellow_font)
            red_rule = FormulaRule(formula=[f'SEARCH("Bottom Quartile",${flag_col_letter}2)'], fill=red_fill, font=red_font)

            data_range = f"A2:{openpyxl.utils.get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            worksheet.conditional_formatting.add(data_range, green_rule)
            worksheet.conditional_formatting.add(data_range, yellow_rule)
            worksheet.conditional_formatting.add(data_range, red_rule)

        except KeyError:
            logger.warning("Styling skipped: 'Performance_Flag' column not found in Summary_Dashboard.")

    def _apply_snapshot_styles(self, writer, df: pd.DataFrame):
        """Applies number formatting to the Latest_Peer_Snapshot sheet using openpyxl."""
        if df is None or df.empty: return
        sheet_name = 'Latest_Peer_Snapshot'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")

        worksheet.column_dimensions['A'].width = 15  # CERT index
        worksheet.column_dimensions['B'].width = 35  # NAME

        percent_format = '0.00%'

        # Start from column 2 because to_excel writes the index
        for col_num, col_name in enumerate(df.columns, 2):
            col_letter = openpyxl.utils.get_column_letter(col_num)

            if col_letter not in ['A', 'B']:
                 worksheet.column_dimensions[col_letter].width = 20

            if any(term in col_name for term in ["Rate", "Pct", "Ratio", "Comp", "Growth", "Risk", "Funds"]):
                for cell in worksheet[col_letter]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = percent_format

    def _apply_macro_analysis_styles(self, writer, df: pd.DataFrame):
        if df is None or df.empty: return
        sheet_name = 'Macro_Analysis'
        if sheet_name not in writer.sheets: return

        worksheet = writer.sheets[sheet_name]
        logging.info(f"Styling {sheet_name} sheet...")
        worksheet.column_dimensions['A'].width = 25 # date
        worksheet.column_dimensions['H'].width = 30 # Series_ID
        worksheet.column_dimensions['I'].width = 25 # Indicator_Name
        worksheet.column_dimensions['B'].width = 40 # Short_Term_Analysis
        worksheet.column_dimensions['C'].width = 40 # Long_Term_Analysis
        worksheet.column_dimensions['D'].width = 40 # Risk_Assessment
class PowerBIGenerator:
    """Generates M and DAX scripts for a Power BI Star Schema."""

    def generate_powerbi_setup_file(self, output_dir: str, powerbi_macro_df: pd.DataFrame, fred_desc_df: pd.DataFrame):
        logger.info("Generating Power BI setup file...")
        Path(output_dir).mkdir(exist_ok=True)
        output_path = Path(output_dir) / "PowerBI_Setup.txt"

        m_code_part = self._get_m_language_code()
        dax_code_part = self._get_dax_measures_code()
        bonus_code_part = self._get_bonus_m_code()
        instructions_part = self._get_instructions()

        full_content = (
            f"{instructions_part}\n"
            f"//--- M LANGUAGE QUERIES ---\n"
            f"//--- Paste each query into a new Blank Query in Power BI's Power Query Editor ---\n\n"
            f"{m_code_part}\n"
            f"//--- DATA MODEL RELATIONSHIPS ---\n"
            f"//--- Power BI should auto-detect these. If not, create them manually in the 'Model' view. ---\n\n"
            f"// Relationship 1: Dates\n"
            f"//   Table 1: fMacro_Data (Many side)\n"
            f"//   Table 2: dDate (One side)\n"
            f"//   Column:  DateKey\n"
            f"//   Cardinality: Many to One (*:1)\n"
            f"//   Cross-filter direction: Single\n\n"
            f"// Relationship 2: Series\n"
            f"//   Table 1: fMacro_Data (Many side)\n"
            f"//   Table 2: dSeries (One side)\n"
            f"//   Column:  SeriesKey\n"
            f"//   Cardinality: Many to One (*:1)\n"
            f"//   Cross-filter direction: Single\n\n"
            f"{dax_code_part}\n"
            f"{bonus_code_part}\n"
        )

        with open(output_path, "w") as f:
            f.write(full_content)
        logger.info(f"Successfully wrote Power BI setup file to {output_path}")

    def _get_instructions(self) -> str:
        return f"""
//==============================================================================
// Power BI Setup Script
// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
//==============================================================================

// INSTRUCTIONS:
// 1. In Power BI, click "Get data" -> "Excel workbook" and select the generated Excel file.
// 2. In the Navigator, select the 'Macro_Analysis' and 'FRED_Descriptions' sheets. Click "Transform Data".
//    This loads the data into the Power Query Editor.
// 3. Once in the Power Query Editor, create a new "Blank Query" for each of the M Language Queries below.
//    (Right-click in the Queries pane -> New Query -> Blank Query).
// 4. Copy and paste the code for each query into its corresponding blank query's Advanced Editor.
// 5. Rename the queries exactly as specified in the titles (e.g., "dDate", "dSeries", "fMacro_Data").
// 6. Click "Close & Apply" to load the data model.
// 7. Once loaded, verify the relationships as described below in the 'Model' view.
// 8. Finally, go to the "Data" view, select the 'fMacro_Data' table, and create each DAX measure.
"""

    def _get_m_language_code(self) -> str:
        dDate_m = """
// --- Dimension Table: dDate ---
// This query generates a complete and dynamic date table based on your data.
let
    // **FIXED**: Point to the correct 'Macro_Analysis' sheet
    SourceTable = Macro_Analysis,
    MinDate = List.Min(Table.Column(SourceTable, "Date")),
    MaxDate = List.Max(Table.Column(SourceTable, "Date")),

    DayCount = Duration.Days(MaxDate - MinDate) + 1,
    DateList = List.Dates(MinDate, DayCount, #duration(1, 0, 0, 0)),
    DateTable = Table.FromList(DateList, Splitter.SplitByNothing(), {"Date"}),

    #"Changed Type to Date" = Table.TransformColumnTypes(DateTable,{{"Date", type date}}),
    #"Inserted Year" = Table.AddColumn(#"Changed Type to Date", "Year", each Date.Year([Date]), Int64.Type),
    #"Inserted Quarter Num" = Table.AddColumn(#"Inserted Year", "QuarterNum", each Date.QuarterOfYear([Date]), Int64.Type),
    #"Added Quarter" = Table.AddColumn(#"Inserted Quarter Num", "Quarter", each "Q" & Text.From([QuarterNum]), type text),
    #"Inserted Month" = Table.AddColumn(#"Added Quarter", "Month", each Date.Month([Date]), Int64.Type),
    #"Inserted Month Name" = Table.AddColumn(#"Inserted Month", "MonthName", each Date.ToText([Date], "MMMM"), type text),
    #"Inserted Day" = Table.AddColumn(#"Inserted Month Name", "Day", each Date.Day([Date]), Int64.Type),
    #"Inserted Day Name" = Table.AddColumn(#"Inserted Day", "DayName", each Date.ToText([Date], "dddd"), type text),
    #"Inserted Week of Year" = Table.AddColumn(#"Inserted Day Name", "WeekOfYear", each Date.WeekOfYear([Date]), Int64.Type),
    #"Added DateKey" = Table.AddColumn(#"Inserted Week of Year", "DateKey", each [Year] * 10000 + [Month] * 100 + [Day], Int64.Type),

    #"Reordered Columns" = Table.ReorderColumns(#"Added DateKey",{"DateKey", "Date", "Year", "Quarter", "Month", "MonthName", "Day", "DayName", "WeekOfYear"}),
    #"Removed QuarterNum" = Table.RemoveColumns(#"Reordered Columns",{"QuarterNum"})
in
    #"Removed QuarterNum"
"""
        dSeries_m = """
// --- Dimension Table: dSeries ---
// It assumes you have loaded the 'FRED_Descriptions' and 'Macro_Analysis' sheets from Excel.
let
    SourceDescriptions = FRED_Descriptions,
    SourceData = Macro_Analysis,

    // Get the distinct list of Series IDs from the main data table to ensure relevance
    MacroDataSeries = Table.Distinct(Table.SelectColumns(SourceData, {"Series_ID"})),

    // Join descriptions with the actual series present in the data
    #"Merged Queries" = Table.NestedJoin(SourceDescriptions, {"Series ID"}, MacroDataSeries, {"Series_ID"}, "Join", JoinKind.Inner),
    #"Removed Columns" = Table.RemoveColumns(#"Merged Queries",{"Join"}),

    // Add a unique integer key for the dimension, which will be the primary key
    #"Added Index" = Table.AddIndexColumn(#"Removed Columns", "SeriesKey", 1, 1, Int64.Type),

    // Select and reorder for clarity
    #"Selected Columns" = Table.SelectColumns(#"Added Index",{"SeriesKey", "Series ID", "long", "short", "Category", "Sectors"}),
    #"Renamed Columns" = Table.RenameColumns(#"Selected Columns",{{"long", "Description"}, {"short", "Series Name"}})
in
    #"Renamed Columns"
"""
        fMacro_Data_m = """
// --- Fact Table: fMacro_Data ---
// It assumes you have the queries 'Macro_Analysis', and 'dSeries' already created.
let
    Source = Macro_Analysis,

    #"Added DateKey" = Table.AddColumn(Source, "DateKey", each Date.Year([Date]) * 10000 + Date.Month([Date]) * 100 + Date.Day([Date]), Int64.Type),

    #"Merged with dSeries" = Table.NestedJoin(#"Added DateKey", {"Series_ID"}, dSeries, {"Series ID"}, "dSeries", JoinKind.LeftOuter),
    #"Expanded SeriesKey" = Table.ExpandTableColumn(#"Merged with dSeries", "dSeries", {"SeriesKey"}, {"SeriesKey"}),

    #"Selected Columns" = Table.SelectColumns(#"Expanded SeriesKey",{"DateKey", "SeriesKey", "Indicator_Value"}),
    #"Changed Type" = Table.TransformColumnTypes(#"Selected Columns",{{"Indicator_Value", type number}, {"DateKey", Int64.Type}, {"SeriesKey", Int64.Type}})
in
    #"Changed Type"
"""
        return f"{dDate_m}\n{dSeries_m}\n{fMacro_Data_m}"

    def _get_dax_measures_code(self) -> str:
        return """
//--- DAX MEASURES & COLUMNS ---
//--- Select the 'fMacro_Data' table, then click 'New Measure' for each item. ---

[Selected Value] = SUM('fMacro_Data'[Indicator_Value])

[Latest Value] =
CALCULATE(
    [Selected Value],
    LASTDATE('dDate'[Date])
)

[Previous Month Value] =
CALCULATE(
    [Selected Value],
    DATEADD('dDate'[Date], -1, MONTH)
)

[Month-over-Month Change] =
VAR CurrentValue = [Selected Value]
VAR PreviousValue = [Previous Month Value]
RETURN
    IF(NOT ISBLANK(CurrentValue) && NOT ISBLANK(PreviousValue), CurrentValue - PreviousValue, BLANK())

[Year-over-Year Value] =
CALCULATE(
    [Selected Value],
    SAMEPERIODLASTYEAR('dDate'[Date])
)

[Year-over-Year Change] =
VAR CurrentValue = [Selected Value]
VAR PreviousValue = [Year-over-Year Value]
RETURN
    IF(NOT ISBLANK(CurrentValue) && NOT ISBLANK(PreviousValue), CurrentValue - PreviousValue, BLANK())
"""

    def _get_bonus_m_code(self) -> str:
        return """
//--- BONUS: M-LANGUAGE FOR BRIDGE TABLE ---
// --- Bridge Table: dSeries_Sectors ---
let
    Source = dSeries,
    #"Removed Other Columns" = Table.SelectColumns(Source,{"SeriesKey", "Sectors"}),
    #"Split Column by Delimiter" = Table.ExpandListColumn(Table.TransformColumns(#"Removed Other Columns", {{"Sectors", Splitter.SplitTextByDelimiter("|", QuoteStyle.None), let itemType = (type nullable text) meta [Serialized.Text = true] in type {itemType}}}), "Sectors"),
    #"Renamed Columns" = Table.RenameColumns(#"Split Column by Delimiter",{{"Sectors", "Sector"}}),
    #"Trimmed Text" = Table.TransformColumns(#"Renamed Columns",{{"Sector", Text.Trim, type text}}),
    #"Filtered Rows" = Table.SelectRows(#"Trimmed Text", each ([Sector] <> null and [Sector] <> ""))
in
    #"Filtered Rows"
"""



# ==================================================================================
#  4. MAIN DASHBOARD CLASS & EXECUTION
# ==================================================================================

class BankPerformanceDashboard:
    def __init__(self, config: 'DashboardConfig'):
        self.config = config
        self.fdic_fetcher = FDICDataFetcher(config)
        self.fred_fetcher = FREDDataFetcher(config)
        self.processor = BankMetricsProcessor(config)
        self.analyzer = PeerAnalyzer(config)
        self.macro_analyzer = MacroTrendAnalyzer(config)
        self.output_gen = ExcelOutputGenerator(config)
        Path(config.output_dir).mkdir(exist_ok=True)
    def _analyze_fdic_data_availability(self, fdic_df: pd.DataFrame) -> Dict[str, Any]:
        """
        Analyzes FDIC data availability and returns a report of missing/empty series.
        """
        # Exclude calculated series from analysis
        CALCULATED_FDIC_SERIES = ['Total_Capital', 'AOBS']
        fields_to_analyze = [f for f in FDIC_FIELDS_TO_FETCH if f not in CALCULATED_FDIC_SERIES]

        if fdic_df.empty:
            return {"empty_series": fields_to_analyze, "sparse_series": [], "available_series": []}

        analysis_report = {
            "empty_series": [],      # Completely missing or all NaN
            "sparse_series": [],     # Present but >75% missing data
            "available_series": []   # Present with reasonable data coverage
        }

        total_records = len(fdic_df)

        for field in fields_to_analyze:
            if field not in fdic_df.columns:
                analysis_report["empty_series"].append(field)
            else:
                non_null_count = fdic_df[field].notna().sum()
                coverage_pct = (non_null_count / total_records) * 100

                if coverage_pct == 0:
                    analysis_report["empty_series"].append(field)
                elif coverage_pct < 25:
                    analysis_report["sparse_series"].append({
                        "field": field,
                        "coverage_pct": round(coverage_pct, 1),
                        "description": FDIC_FIELD_DESCRIPTIONS.get(field, {}).get("short", "Unknown")
                    })
                else:
                    analysis_report["available_series"].append({
                        "field": field,
                        "coverage_pct": round(coverage_pct, 1)
                    })

        return analysis_report
    def _optimize_df_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Reduces DataFrame memory usage and file size by downcasting numeric types
        and converting object columns with low cardinality to 'category'.
        """
        for col in df.columns:
            if df[col].dtype == 'float64':
                df[col] = pd.to_numeric(df[col], downcast='float')
            elif df[col].dtype == 'int64':
                df[col] = pd.to_numeric(df[col], downcast='integer')
            elif df[col].dtype == 'object':
                if len(df[col].unique()) / len(df[col]) < 0.5:
                    df[col] = df[col].astype('category')
        return df
    def _create_peer_composite(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Creates composite 'dummy' banks for each defined Peer Group.
        These allows us to plot 'Core Peer Avg' lines on charts easily.
        """
        logging.info("Generating historical composites for defined Peer Groups...")

        composites = []

        # Base CERT for composites (e.g. 90001, 90002...)
        base_dummy_cert = 90000

        numeric_cols = df.select_dtypes(include=np.number).columns.drop('CERT', errors='ignore')

        for group_key, group_info in PEER_GROUPS.items():
            # Filter for peers in this group
            group_certs = group_info['certs']
            peer_subset = df[df['CERT'].isin(group_certs)]

            if peer_subset.empty:
                continue

            # Calculate average for every quarter
            # We use a context manager to suppress grouping warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", FutureWarning)
                group_avg = peer_subset.groupby('REPDTE')[numeric_cols].mean().reset_index()

            # Assign Dummy Metadata
            dummy_cert = base_dummy_cert + group_info['display_order']
            group_avg['CERT'] = dummy_cert
            group_avg['NAME'] = f"AVG: {group_info['name']}"
            group_avg['HQ_STATE'] = 'AVG'

            composites.append(group_avg)
            logging.info(f"Created composite for {group_info['name']} (CERT {dummy_cert})")

        if composites:
            df = pd.concat([df] + composites, ignore_index=True)

        return df

    def run(self) -> Dict[str, Any]:
        logging.info("Starting dashboard generation...")
        fdic_df, _ = self.fdic_fetcher.fetch_all_banks()
        if fdic_df.empty: raise ValueError("No FDIC data retrieved.")

        all_certs = fdic_df['CERT'].unique().tolist()
        locations_df = get_bank_locations(all_certs)
        if 'NAME' in locations_df.columns:
            locations_df.drop(columns=['NAME'], inplace=True)
        fdic_df = pd.merge(fdic_df, locations_df, on='CERT', how='left')
        logging.info("Location data merged successfully.")

        fdic_analysis = self._analyze_fdic_data_availability(fdic_df)

        proc_df = self.processor.create_derived_metrics(fdic_df)
        proc_df_with_ttm = self.processor.calculate_ttm_metrics(proc_df)
        proc_df_with_peers = self._create_peer_composite(proc_df_with_ttm)

        peer_comp_df = self.analyzer.create_peer_comparison(proc_df_with_peers)
        snapshot_df = self.processor.create_latest_snapshot(proc_df_with_peers)
        avg_8q_all_metrics_df = self.processor.calculate_8q_averages(proc_df_with_peers)

        # === FRED DATA FETCHING - NO DUPLICATES ===
        logging.info("Fetching FRED macroeconomic data asynchronously...")
        start_time = time.perf_counter()

        # Define calculated series that should NOT be fetched
        CALCULATED_SERIES = ['SOFR3MTB3M']

        # Build list - EXCLUDING calculated series
        # Build list - EXCLUDING calculated series
        series_ids_to_fetch = []
        for category_dict in FRED_SERIES_TO_FETCH.values():
            for sid in category_dict.keys():
                if sid not in CALCULATED_SERIES:
                    series_ids_to_fetch.append(sid)

        logger.info(f"Fetching {len(series_ids_to_fetch)} series (excluding {CALCULATED_SERIES})")

        # Build descriptions DataFrame for ALL series (including calculated)
        # Use a completely unique variable name to avoid any collisions
        _temp_series_metadata_dict = {}
        for category, series_dict in FRED_SERIES_TO_FETCH.items():
            # Validate that series_dict is actually a dict of series
            if not isinstance(series_dict, dict):
                logger.error(f"Category '{category}' has invalid structure: {type(series_dict)}")
                continue

            for series_id, metadata in series_dict.items():
                # Skip if this looks like a metadata key, not a series ID
                if series_id in ['short', 'long', 'annualized', 'best_technical', 'technical_reason']:
                    logger.warning(f"Skipping metadata key '{series_id}' in category '{category}'")
                    continue

                if not isinstance(metadata, dict):
                    logger.error(f"Invalid metadata for {series_id} in category '{category}': {type(metadata)}")
                    continue

                _temp_series_metadata_dict[series_id] = {
                    'Series ID': series_id,
                    'Category': category,
                    **metadata
                }

        series_descriptions_df = pd.DataFrame.from_dict(_temp_series_metadata_dict, orient='index')
        logger.info(f"Created descriptions dataframe with {len(series_descriptions_df)} rows")

        # Handle event loop properly for Spyder/Jupyter
        try:
            loop = asyncio.get_running_loop()
            import nest_asyncio
            nest_asyncio.apply()
            fred_df, fred_desc_df, failed_fred_series, fred_metadata_df = asyncio.run(
                self.fred_fetcher.fetch_all_series_async(
                    series_ids=series_ids_to_fetch,
                    series_descriptions=series_descriptions_df
                )
            )
            # === PATCH S-RAW3a: pull stashed raw observations DF ===
            fred_obs_df = getattr(self.fred_fetcher, 'last_fred_obs_df', pd.DataFrame())
            # === END PATCH S-RAW3a ===

        except RuntimeError:
            fred_df, fred_desc_df, failed_fred_series, fred_metadata_df = asyncio.run(
                self.fred_fetcher.fetch_all_series_async(
                    series_ids=series_ids_to_fetch,
                    series_descriptions=series_descriptions_df
                )
            )
            # === PATCH S-RAW3a: pull stashed raw observations DF ===
            fred_obs_df = getattr(self.fred_fetcher, 'last_fred_obs_df', pd.DataFrame())
            # === END PATCH S-RAW3a ===

        end_time = time.perf_counter()
        logging.info(f"Asynchronous FRED data fetching completed in {end_time - start_time:.2f} seconds.")
        # Merge frequency information into descriptions
        # Merge frequency information into descriptions
        # === PATCH: enrich FRED metadata with DateBasis & QuarterOffset and apply overrides ===
        # Enrich FRED metadata with DateBasis / QuarterOffset and merge into descriptions
        # Defaults if metadata is missing
        if fred_metadata_df is None or fred_metadata_df.empty:
            fred_metadata_df = pd.DataFrame(columns=['Series ID','frequency','frequency_short','units','DateBasis','QuarterOffset'])
        else:
            fred_metadata_df = fred_metadata_df.copy()
            # Ensure the new columns exist with safe defaults
            if 'DateBasis' not in fred_metadata_df.columns:
                fred_metadata_df['DateBasis'] = 'period_start'
            else:
                fred_metadata_df['DateBasis'] = fred_metadata_df['DateBasis'].fillna('period_start')

            if 'QuarterOffset' not in fred_metadata_df.columns:
                fred_metadata_df['QuarterOffset'] = 0
            fred_metadata_df['QuarterOffset'] = pd.to_numeric(fred_metadata_df['QuarterOffset'], errors='coerce').fillna(0).astype(int)

        # Known publication-dated quarterlies that represent the PREVIOUS quarter
        _fred_overrides = {
            # Tightening CRE Lending Standards – publication month maps to prior quarter
            'DRCRELEXFACBS': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
            'DRBLACBS': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
            'DRTSCILM': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
            'DRTSCLCC': {
                'DateBasis': 'publication_date',
                'QuarterOffset': -1,  # Q2 publication refers to Q1 data
                'frequency': 'quarterly',
                'frequency_short': 'Q'
            },
        }

        # Apply overrides
        for _sid, _vals in _fred_overrides.items():
            _m = fred_metadata_df['Series ID'].astype(str).eq(_sid)
            for _k, _v in _vals.items():
                fred_metadata_df.loc[_m, _k] = _v

        # Merge the enriched metadata into descriptions
        fred_desc_df = pd.merge(
            fred_desc_df,
            fred_metadata_df[['Series ID','frequency','frequency_short','units','DateBasis','QuarterOffset']],
            on='Series ID',
            how='left'
        )


        # Set index and create calculated series
        if 'date' in fred_df.columns:
            fred_df.set_index('date', inplace=True)

        if 'SOFR' in fred_df.columns and 'TB3MS' in fred_df.columns:
            fred_df['SOFR'] = fred_df['SOFR'].ffill()
            fred_df['TB3MS'] = fred_df['TB3MS'].ffill()
            fred_df['SOFR3MTB3M'] = fred_df['SOFR'] - fred_df['TB3MS']
            if fred_desc_df is not None and fred_desc_df[fred_desc_df['Series ID'] == 'SOFR3MTB3M'].empty:
                new_row_data = {'Series ID': 'SOFR3MTB3M', 'Category': 'Middle Market, Healthcare, & Funding Indicators', 'short': 'SOFR vs T-Bill Spread', 'long': 'Calculated Spread: SOFR minus 3-Month T-Bill Rate'}
                new_row = pd.DataFrame([new_row_data])
                fred_desc_df = pd.concat([fred_desc_df, new_row], ignore_index=True)

        # Process technical indicators
        enhanced_analyzer = MacroTrendAnalyzer(self.config)
        processed_data = {}
        validation_reports = {}

        for category, series_in_category in FRED_SERIES_TO_FETCH.items():
            for series_id in series_in_category.keys():
                if series_id in CALCULATED_SERIES:
                    continue
                if series_id in fred_df.columns:
                    validation_reports[series_id] = enhanced_analyzer.validate_series_data(
                        series_id, fred_df[series_id]
                    )
                    if validation_reports[series_id]['status'] != 'Error':
                        result = enhanced_analyzer.calculate_technical_indicators(
                            fred_df[series_id], series_id
                        )
                        if not result.empty:
                            processed_data[series_id] = result

        # Process calculated series
        for series_id in CALCULATED_SERIES:
            if series_id in fred_df.columns:
                validation_reports[series_id] = enhanced_analyzer.validate_series_data(
                    series_id, fred_df[series_id]
                )
                if validation_reports[series_id]['status'] != 'Error':
                    result = enhanced_analyzer.calculate_technical_indicators(
                        fred_df[series_id], series_id
                    )
                    if not result.empty:
                        processed_data[series_id] = result

        validation_df = pd.DataFrame.from_dict(validation_reports, orient='index')
        powerbi_macro_df = enhanced_analyzer.generate_powerbi_output(processed_data)


        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f"{self.config.output_dir}/Bank_Performance_Dashboard_{ts}.xlsx"
        peer_comp_df = self._optimize_df_dtypes(peer_comp_df)
        snapshot_df = self._optimize_df_dtypes(snapshot_df)
        proc_df_with_peers = self._optimize_df_dtypes(proc_df_with_peers)
        avg_8q_all_metrics_df = self._optimize_df_dtypes(avg_8q_all_metrics_df)
        powerbi_macro_df = self._optimize_df_dtypes(powerbi_macro_df)
        # === FDIC-META: units & scaling for email tables ===
        # === FDIC-META (with Basis): tells consumer how to scale/format ===
        fdic_meta_df = pd.DataFrame([
            # Dollars in $000 → display $M (scale 1e-3)
            {"MetricCode":"ASSET", "Display":"Assets",                      "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},
            {"MetricCode":"LNLS",  "Display":"Gross Loans",                 "DisplayUnit":"$M", "Scale":1e-3, "Fmt":"currency_m", "Decimals":0, "Basis":"level"},

            # Ratios: mark as 'percent' (already %) OR 'fraction' (0–1) depending on your upstream
            # Given your HTML shows 8.86% (too big), your upstream is ALREADY percent → use Basis='percent' and Scale=1.0
            {"MetricCode":"CI_to_Capital_Risk",             "Display":"C&I to Capital Risk (%)",             "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"CRE_Concentration_Capital_Risk", "Display":"CRE Concentration Risk (%)",          "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"IDB_CRE_Growth_TTM",             "Display":"CRE Growth TTM (%)",                  "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"IDB_CRE_Growth_36M",             "Display":"CRE Growth 36M (%)",                  "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"percent"},
            {"MetricCode":"TTM_NCO_Rate",                  "Display":"TTM NCO Rate (%)",                    "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"NPL_to_Gross_Loans_Rate",       "Display":"NPL to Gross Loans (%)",              "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"Allowance_to_Gross_Loans_Rate", "Display":"Allowance to Gross Loans (%)",        "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"TTM_Past_Due_Rate",             "Display":"TTM Past Due Rate (%)",               "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"TTM_PD30_Rate",                 "Display":"TTM Past Due (30-90 Days) Rate (%)",  "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},
            {"MetricCode":"TTM_PD90_Rate",                 "Display":"TTM Past Due (90+ Days) Rate (%)",    "DisplayUnit":"%", "Scale":1.0, "Fmt":"percent", "Decimals":2, "Basis":"fraction"},

        ])
        # === END FDIC-META ===


        self.output_gen.write_excel_output(
            file_path=fname,
            Summary_Dashboard=peer_comp_df,
            Latest_Peer_Snapshot=snapshot_df,
            Averages_8Q_All_Metrics=avg_8q_all_metrics_df,
            FDIC_Metric_Descriptions=self.output_gen.fdic_desc_df,
            FDIC_Metadata=fdic_meta_df,                     # <<— NEW SHEET
            Macro_Analysis=powerbi_macro_df,
            FDIC_Data=proc_df_with_peers,
            FRED_Data=fred_df.reset_index(),
            FRED_Metadata=fred_metadata_df,
            FRED_Descriptions=fred_desc_df,
            Data_Validation_Report=validation_df,
        )


        # === ENHANCED DIAGNOSTIC: Inspect Critical FDIC Data for Subject Bank ===
        print("\n" + "="*80)
        print("CRITICAL FDIC DATA INSPECTION FOR YOUR BANK")
        print("="*80)

        try:
            # Get the latest data for your bank
            subject_bank_data = proc_df_with_peers[
                proc_df_with_peers['CERT'] == self.config.subject_bank_cert
            ].sort_values('REPDTE')

            if not subject_bank_data.empty:
                latest_record = subject_bank_data.iloc[-1]
                bank_name = latest_record.get('NAME', 'Unknown Bank')
                report_date = latest_record.get('REPDTE', 'Unknown Date')

                print(f"Bank: {bank_name} (CERT: {self.config.subject_bank_cert})")
                print(f"Latest Report Date: {report_date}")
                print("-" * 80)

                # Critical columns to inspect
                critical_columns = {
                    'ASSET': 'Total Assets',
                    'EQ': 'Total Equity Capital',
                    'LNLS': 'Gross Loans & Leases',
                    'LNCI': 'C&I Loans',
                    'LNRENROW': 'Owner-Occ CRE',
                    'LNRECONS': 'Construction & Development',
                    'LNRERES': '1-4 Family Residential',
                    'LNRENROT': 'Nonfarm CRE '
                }

                print("CRITICAL BALANCE SHEET ITEMS:")
                for col_code, description in critical_columns.items():
                    value = latest_record.get(col_code)
                    if pd.isna(value) or value is None:
                        print(f"  {col_code:12} ({description:30}): *** MISSING DATA ***")
                    elif value == 0:
                        print(f"  {col_code:12} ({description:30}): *** ZERO VALUE ***")
                    else:
                        # Format as currency in thousands
                        formatted_value = f"${value:,.0f}"
                        print(f"  {col_code:12} ({description:30}): {formatted_value}")

                # === NEW: Investigate C&I Loans specifically ===
                print("\n" + "="*60)
                print("C&I LOANS INVESTIGATION")
                print("="*60)

                # Check if LNCI exists in the raw data at all
                if 'LNCI' not in subject_bank_data.columns:
                    print("❌ LNCI column does not exist in the fetched data")
                    print("   This suggests FDIC API did not return this field")
                else:
                    print("✓ LNCI column exists in the data")

                    # Check LNCI values across all quarters
                    lnci_summary = subject_bank_data['LNCI'].describe()
                    print(f"   LNCI Data Summary across all quarters:")
                    print(f"     Count (non-null): {subject_bank_data['LNCI'].notna().sum()}")
                    print(f"     Count (null):     {subject_bank_data['LNCI'].isna().sum()}")
                    print(f"     Count (zero):     {(subject_bank_data['LNCI'] == 0).sum()}")
                    if subject_bank_data['LNCI'].notna().sum() > 0:
                        print(f"     Min value:        ${subject_bank_data['LNCI'].min():,.0f}")
                        print(f"     Max value:        ${subject_bank_data['LNCI'].max():,.0f}")
                        print(f"     Latest non-null:  ${subject_bank_data['LNCI'].dropna().iloc[-1] if len(subject_bank_data['LNCI'].dropna()) > 0 else 0:,.0f}")

                # Check alternative C&I loan fields that might be available
                print("\n   Checking for alternative C&I loan series:")
                ci_alternatives = {
                    'LNCI': 'Commercial and Industrial Loans',
                    'LNCIDOM': 'C&I Loans - Domestic',
                    'LNCINUM': 'C&I Loans - Number of Loans',
                    'LNCON': 'Consumer Loans (sometimes includes C&I)',
                    'BUSLOANS': 'Total Business Loans (broader category)'
                }

                for field, description in ci_alternatives.items():
                    if field in subject_bank_data.columns:
                        latest_val = latest_record.get(field)
                        if pd.notna(latest_val) and latest_val != 0:
                            print(f"   ✓ {field:10}: {description:35} = ${latest_val:,.0f}")
                        else:
                            print(f"   ⚠ {field:10}: {description:35} = Missing/Zero")
                    else:
                        print(f"   ❌ {field:10}: {description:35} = Not available")

                # Check what fields ARE available for this bank
                print(f"\n   Available loan fields for your bank:")
                loan_fields = [col for col in subject_bank_data.columns if col.startswith('LN') and latest_record.get(col, 0) > 0]
                for field in sorted(loan_fields):
                    value = latest_record.get(field, 0)
                    print(f"     {field}: ${value:,.0f}")

                # Check if this is a reporting threshold issue
                total_assets = latest_record.get('ASSET', 0)
                print(f"\n   Bank Size Analysis:")
                print(f"     Total Assets: ${total_assets:,.0f}")
                if total_assets < 300_000:  # Less than $300M
                    print("     ⚠ Small bank - may have different reporting requirements")
                elif total_assets < 1_000_000:  # Less than $1B
                    print("     ℹ Community bank - standard reporting applies")
                else:
                    print("     ℹ Large bank - full reporting requirements")

                # Calculate what C&I might be if it's combined with other categories
                total_loans = latest_record.get('LNLS', 0)
                known_loans = sum([
                    latest_record.get('LNRENROW', 0),  # CRE Non-Owner
                    latest_record.get('LNRECONS', 0),  # Construction
                    latest_record.get('LNRERES', 0),   # Residential
                    latest_record.get('LNRENROT', 0),  # CRE Owner-Occ
                    latest_record.get('LNRELOC', 0),   # HELOC
                    latest_record.get('LNCRCD', 0),    # Credit Cards
                    latest_record.get('LNAUTO', 0),    # Auto
                    latest_record.get('LNOTHER', 0),   # Other
                    latest_record.get('LS', 0),        # Leases
                ])

                implied_ci = total_loans - known_loans
                print(f"\n   Implied C&I Calculation:")
                print(f"     Total Loans:        ${total_loans:,.0f}")
                print(f"     Known Categories:   ${known_loans:,.0f}")
                print(f"     Implied C&I:        ${implied_ci:,.0f}")
                print(f"     Percentage of Total: {(implied_ci/total_loans)*100:.1f}%" if total_loans > 0 else "     Percentage: N/A")

                # Rest of your existing diagnostic code...
                print("\n" + "="*60)
                print("OTHER CALCULATED FIELDS")
                print("="*60)
                calc_fields = {
                    'Total_Capital': 'Total Capital (T1 + T2)',
                    'AOBS': 'Off-Balance Sheet Allowance',
                    'Total_ACL': 'Total Allowance for Credit Losses'
                }

                for col_code, description in calc_fields.items():
                    value = latest_record.get(col_code)
                    if pd.isna(value) or value is None:
                        print(f"  {col_code:15} ({description:30}): *** CALCULATION FAILED ***")
                    else:
                        formatted_value = f"${value:,.0f}"
                        print(f"  {col_code:15} ({description:30}): {formatted_value}")

            else:
                print(f"❌ NO DATA FOUND for your bank (CERT: {self.config.subject_bank_cert})")

        except Exception as e:
            print(f"❌ ERROR during diagnostic inspection: {e}")

        print("="*80)
        return {
            "output_file": fname,
            "powerbi_macro_df": powerbi_macro_df,
            "fred_desc_df": fred_desc_df,
            "failed_fred_series": failed_fred_series,
            "fdic_analysis": fdic_analysis,
        }



def load_config() -> DashboardConfig:
    """
    Loads configuration from environment variables with fallback to defaults.

    Create a .env file with:
        FRED_API_KEY=your_key_here
        SUBJECT_BANK_CERT=34221
        PEER_CERTS=26876,9396,18221,...

    Returns:
        DashboardConfig: Configuration object

    Raises:
        ValueError: If FRED_API_KEY is missing
    """
    import os
    from pathlib import Path

    # Try to load .env file (install python-dotenv: pip install python-dotenv)
    try:
        from dotenv import load_dotenv
        env_path = Path(__file__).parent / '.env'
        if env_path.exists():
            load_dotenv(env_path)
            logger.info(f"Loaded configuration from {env_path}")
        else:
            logger.warning(f".env file not found at {env_path}, using environment variables")
    except ImportError:
        logger.warning("python-dotenv not installed, reading from environment variables only")

    # Required
    fred_api_key = os.getenv('FRED_API_KEY')
    if not fred_api_key:
        raise ValueError(
            "FRED_API_KEY not found. Set it in .env file or environment:\n"
            "  export FRED_API_KEY='your_key_here'"
        )


    subject_cert = int(os.getenv('SUBJECT_BANK_CERT', '34221'))

    # UPDATED: Derive unique peer certs from the PEER_GROUPS dictionary
    # This ignores the .env PEER_CERTS list to ensure consistency with your logic
    peer_certs = get_all_peer_certs()

    # Remove subject bank from peer list if present to avoid fetching twice
    if subject_cert in peer_certs:
        peer_certs.remove(subject_cert)

    logger.info(f"Configuration loaded: Subject CERT={subject_cert}, {len(peer_certs)} unique peers from {len(PEER_GROUPS)} groups.")

    return DashboardConfig(
        fred_api_key=fred_api_key,
        subject_bank_cert=subject_cert,
        peer_bank_certs=peer_certs
    )


def main():
    run_results = {}
    try:
        config = load_config()
        dash = BankPerformanceDashboard(config)
        run_results = dash.run()
        print("\n" + "="*80)
        print("DASHBOARD GENERATION COMPLETE")
        print(f"Output file located at: {run_results.get('output_file', 'N/A')}")
        print("="*80)

        print("\n--- FDIC DATA AVAILABILITY REPORT ---")
        fdic_analysis = run_results.get('fdic_analysis', {})

        empty_series = fdic_analysis.get('empty_series', [])
        sparse_series = fdic_analysis.get('sparse_series', [])
        available_series = fdic_analysis.get('available_series', [])

        if empty_series:
            print(f"\n❌ EMPTY FDIC SERIES ({len(empty_series)} series with no data):")
            for series in empty_series:
                description = FDIC_FIELD_DESCRIPTIONS.get(series, {}).get("short", "Unknown")
                print(f"  - {series}: {description}")
        else:
            print("\n✅ All FDIC series returned data.")

        if sparse_series:
            print(f"\n⚠️  SPARSE FDIC SERIES ({len(sparse_series)} series with limited data):")
            for item in sparse_series:
                print(f"  - {item['field']}: {item['description']} ({item['coverage_pct']}% coverage)")

        print(f"\n✅ AVAILABLE FDIC SERIES: {len(available_series)} series with good data coverage")

        total_requested = len(FDIC_FIELDS_TO_FETCH)
        total_available = len(available_series)
        total_sparse = len(sparse_series)
        total_empty = len(empty_series)

        print(f"\nSUMMARY:")
        print(f"  Total FDIC series requested: {total_requested}")
        print(f"  Available (good coverage):   {total_available} ({(total_available/total_requested)*100:.1f}%)")
        print(f"  Sparse (limited coverage):   {total_sparse} ({(total_sparse/total_requested)*100:.1f}%)")
        print(f"  Empty (no data):             {total_empty} ({(total_empty/total_requested)*100:.1f}%)")

        print("\n--- POWER BI INTEGRATION ---")
        try:
            macro_df = run_results.get('powerbi_macro_df')
            desc_df = run_results.get('fred_desc_df')
            if macro_df is not None and not macro_df.empty and desc_df is not None:
                powerbi_generator = PowerBIGenerator()
                powerbi_generator.generate_powerbi_setup_file(
                    output_dir=config.output_dir,
                    powerbi_macro_df=macro_df,
                    fred_desc_df=desc_df
                )
                powerbi_setup_file = Path(config.output_dir) / "PowerBI_Setup.txt"
                print(f"[+] Power BI setup file generated successfully.")
                print(f"    -> Location: {powerbi_setup_file.resolve()}")
            else:
                print("[-] Power BI setup file was not generated because macro data was missing.")

        except Exception as e:
            logger.error(f"Failed to generate Power BI setup file: {e}", exc_info=True)
            print("[-] Power BI setup file was not generated due to an unexpected error.")

    except Exception as e:
        logger.error(f"Application failed to complete: {e}", exc_info=True)
        sys.exit(1)

    finally:
        print("\n--- DATA SERIES VERIFICATION ---")
        failed_series = run_results.get("failed_fred_series", [])
        if failed_series:
            print("\nWARNING: The following FRED series failed to fetch and are NOT in the output:")
            for series in failed_series:
                print(f"  - {series}")
        else:
            print("\nAll FRED series fetched successfully.")

        logging.shutdown()

if __name__ == '__main__':
    main()




