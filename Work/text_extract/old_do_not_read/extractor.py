"""
Split-RAG Document Extractor - Main Extraction Engine
Version: 1.0.0

This module processes PDF, DOCX, and XLSX files into structured Context Graph JSON.

Usage:
    python extractor.py --working_dir /path/to/working/directory
"""

from __future__ import annotations

# ============================================================================
# STANDARD LIBRARY IMPORTS (per CANON_001: always permitted)
# ============================================================================
import argparse
import csv
import hashlib
import io
import json
import logging
import os
import re
import sys
import time
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

# ============================================================================
# THIRD-PARTY IMPORTS (Tier 1 Approved per Section 1.1.1)
# ============================================================================
import numpy as np
import openpyxl
import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from PIL import Image, ImageFilter
from pydantic import ValidationError
from tqdm import tqdm

# Import our schema
from schema_v1 import (
    ChunkText,
    ContextGraph,
    ContextNode,
    EdgeType,
    FileRef,
    ImageData,
    Location,
    Provenance,
    TableData,
    utc_now_z,
    validate_graph,
)

# ============================================================================
# CONSTANTS
# ============================================================================

EXTRACTOR_VERSION = "1.0.0"
SCHEMA_VERSION = "1.0.0"
SUPPORTED_EXTENSIONS: Set[str] = {".pdf", ".docx", ".xlsx"}

DOCX_NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}

LOG_CODES = {
    "E001": "START_RUN",
    "E002": "END_RUN",
    "E003": "FILE_DISCOVERED",
    "E004": "FILE_SKIPPED_SHA256",
    "E005": "DOC_PARSE_PRIMARY_OK",
    "E006": "DOC_PARSE_PRIMARY_FAIL",
    "E007": "FALLBACK_USED",
    "E008": "NODE_VALIDATION_FAIL",
    "E009": "DOC_WRITE_OK",
    "E010": "DOC_WRITE_FAIL",
    "E011": "MANIFEST_UPDATE",
    "E012": "CONFIG_LOAD_FAIL",
}


# ============================================================================
# CUSTOM EXCEPTIONS
# ============================================================================


class ExtractorError(Exception):
    """Base exception for all extractor errors."""
    pass


class ConfigurationError(ExtractorError):
    """Raised when configuration is invalid or missing."""
    pass


class FileDiscoveryError(ExtractorError):
    """Raised when input directory is inaccessible."""
    pass


class ExtractionError(ExtractorError):
    """Raised when document extraction fails."""
    pass


class SchemaValidationError(ExtractorError):
    """Raised when output fails schema validation."""
    pass


class ManifestError(ExtractorError):
    """Raised when manifest operations fail."""
    pass


# ============================================================================
# SECTION 2.3: DETERMINISTIC ID GENERATION (SHA256)
# ============================================================================


def compute_file_sha256(file_path: Path, chunk_size: int = 8192) -> str:
    """Compute SHA256 hash of file contents using chunked reading."""
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            sha256_hash.update(chunk)
    return sha256_hash.hexdigest().lower()


def compute_node_id(file_sha256: str, node_type: str, content: str, index: int = 0) -> str:
    """Generate stable node ID from file + type + content + index."""
    hash_input = f"{file_sha256}|{node_type}|{content.strip()}|{index}"
    return hashlib.sha256(hash_input.encode("utf-8")).hexdigest().lower()


def compute_edge_id(from_id: str, to_id: str, relation: str) -> str:
    """Generate stable edge ID."""
    hash_input = f"{from_id}|{to_id}|{relation}"
    return hashlib.sha256(hash_input.encode("utf-8")).hexdigest().lower()


def compute_run_id(file_paths: List[Path], timestamp: str) -> str:
    """Generate run ID from processed files + timestamp."""
    hash_input = "|".join(sorted(str(p) for p in file_paths)) + f"|{timestamp}"
    return hashlib.sha256(hash_input.encode("utf-8")).hexdigest().lower()


def get_file_modified_utc(file_path: Path) -> str:
    """Get file modification time as ISO8601 UTC string ending in Z."""
    mtime = file_path.stat().st_mtime
    dt = datetime.utcfromtimestamp(mtime).replace(microsecond=0)
    return dt.isoformat() + "Z"


# ============================================================================
# SECTION 3.0: EXTRACTED ELEMENT DATA CLASS
# ============================================================================


@dataclass
class ExtractedElement:
    """Intermediate representation before conversion to ContextNode."""
    content_type: str  # "section" | "chunk" | "table" | "image"
    content: str
    page_number: Optional[int]  # 1-indexed, None for non-paginated
    section_title: str
    sheet_name: Optional[str] = None
    bbox: Optional[Tuple[float, float, float, float]] = None
    table_shape: Optional[Tuple[int, int]] = None  # (rows, cols)
    confidence: float = 0.95
    warnings: List[str] = field(default_factory=list)
    char_start: Optional[int] = None
    char_end: Optional[int] = None


# ============================================================================
# SECTION 3.2: FILE DISCOVERY
# ============================================================================


def discover_input_files(input_dir: Path) -> List[Path]:
    """
    Find all processable files in input directory.

    REQUIRED: Non-recursive (only top-level files).
    REQUIRED: Case-insensitive extension matching.
    REQUIRED: Skip hidden files (starting with .).
    REQUIRED: Skip temporary files (starting with ~).
    REQUIRED: Return sorted list for deterministic ordering.
    """
    files = []
    for item in input_dir.iterdir():
        if not item.is_file():
            continue
        if item.name.startswith(".") or item.name.startswith("~"):
            continue
        if item.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(item)

    return sorted(files, key=lambda p: p.name.lower())


# ============================================================================
# SECTION 3.3: MANIFEST MANAGEMENT
# ============================================================================


@dataclass
class ManifestEntry:
    sha256: str
    status: str  # "processed" | "failed" | "skipped" | "partial"
    processed_at: str  # ISO-8601
    output_file: Optional[str] = None
    error_message: Optional[str] = None


@dataclass
class Manifest:
    schema_version: str = "1.0.0"
    files: Dict[str, ManifestEntry] = field(default_factory=dict)


def load_manifest(manifest_path: Path) -> Manifest:
    """Load existing manifest or create new one."""
    if not manifest_path.exists():
        return Manifest()

    try:
        with open(manifest_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        files = {}
        for path, entry_data in data.get("files", {}).items():
            files[path] = ManifestEntry(**entry_data)

        return Manifest(schema_version=data.get("schema_version", "1.0.0"), files=files)
    except (json.JSONDecodeError, KeyError, TypeError):
        return Manifest()


def save_manifest(manifest: Manifest, manifest_path: Path) -> None:
    """Atomically save manifest to disk."""
    temp_path = manifest_path.with_suffix(".tmp")

    data = {
        "schema_version": manifest.schema_version,
        "files": {
            k: {
                "sha256": v.sha256,
                "status": v.status,
                "processed_at": v.processed_at,
                "output_file": v.output_file,
                "error_message": v.error_message,
            }
            for k, v in manifest.files.items()
        },
    }

    with open(temp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    temp_path.replace(manifest_path)


def should_process_file(
    file_path: Path, file_sha256: str, manifest: Manifest, config: dict
) -> bool:
    """Determine if file needs processing."""
    key = str(file_path.relative_to(file_path.parent.parent))

    if key not in manifest.files:
        return True

    entry = manifest.files[key]

    if entry.sha256 != file_sha256:
        return True

    if entry.status == "failed" and config.get("processing", {}).get("retry_failed", True):
        return True

    return False


# ============================================================================
# SECTION 3.7: TABLE TO CSV CONVERSION
# ============================================================================


def table_to_csv(table: List[List[Any]]) -> str:
    """Convert 2D table to CSV format."""
    if not table or not table[0]:
        return ""

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    for row in table:
        cleaned_row = [str(cell) if cell is not None else "" for cell in row]
        writer.writerow(cleaned_row)

    return output.getvalue()


def chunk_large_table(
    table_data: List[List[Any]], max_rows_per_chunk: int = 40
) -> List[List[List[Any]]]:
    """Split large tables into chunks, keeping header row in each."""
    if len(table_data) <= max_rows_per_chunk + 1:
        return [table_data]

    header = table_data[0]
    data_rows = table_data[1:]

    chunks = []
    for i in range(0, len(data_rows), max_rows_per_chunk):
        chunk_rows = data_rows[i : i + max_rows_per_chunk]
        chunks.append([header] + chunk_rows)

    return chunks


# ============================================================================
# SECTION 3.8: SEMANTIC CHUNKING ALGORITHM
# ============================================================================


def chunk_text_content(
    text: str, max_chars: int = 6000, overlap_chars: int = 200
) -> List[Tuple[str, int, int]]:
    """
    Split text into chunks respecting semantic boundaries.

    Returns:
        List of tuples: (chunk_text, start_char_index, end_char_index)
    """
    if len(text) <= max_chars:
        return [(text, 0, len(text))]

    chunks = []
    current_pos = 0

    while current_pos < len(text):
        chunk_end = min(current_pos + max_chars, len(text))

        if chunk_end < len(text):
            chunk_text = text[current_pos:chunk_end]

            # Try paragraph boundary
            split_pos = chunk_text.rfind("\n\n")
            if split_pos == -1 or split_pos < max_chars * 0.5:
                # Try sentence boundary
                sentence_match = None
                for match in re.finditer(r"[.!?]\s+", chunk_text):
                    if match.end() < len(chunk_text):
                        sentence_match = match

                if sentence_match and sentence_match.end() > max_chars * 0.5:
                    split_pos = sentence_match.end()
                else:
                    # Try word boundary
                    split_pos = chunk_text.rfind(" ")
                    if split_pos == -1 or split_pos < max_chars * 0.5:
                        split_pos = len(chunk_text)

            actual_end = current_pos + split_pos
        else:
            actual_end = chunk_end

        chunk_content = text[current_pos:actual_end].strip()
        if chunk_content:
            chunks.append((chunk_content, current_pos, actual_end))

        current_pos = max(actual_end - overlap_chars, current_pos + 1)

    return chunks


# ============================================================================
# SECTION 3.9: IMAGE EDGE DENSITY FILTER
# ============================================================================


def classify_image_by_edge_density(
    image_bytes: bytes, edge_threshold: int = 40, density_threshold: float = 0.10
) -> Tuple[str, float]:
    """
    Classify image as data-dense or decorative using edge detection.

    Returns:
        Tuple of (category: str, edge_density: float)
    """
    try:
        img = Image.open(io.BytesIO(image_bytes))

        if img.mode != "L":
            img = img.convert("L")

        edges = img.filter(ImageFilter.FIND_EDGES)

        pixels = list(edges.getdata())
        total_pixels = len(pixels)
        edge_pixels = sum(1 for p in pixels if p > edge_threshold)

        edge_density = edge_pixels / total_pixels if total_pixels > 0 else 0.0

        if edge_density >= 0.15:
            category = "chart"
        elif edge_density >= 0.10:
            category = "diagram"
        elif edge_density < 0.08:
            category = "photo"
        else:
            category = "unknown"

        return category, round(edge_density, 4)

    except Exception:
        return "unknown", 0.0


# ============================================================================
# SECTION 3.10: CONFIDENCE SCORING HEURISTICS
# ============================================================================


def calculate_confidence(
    content_type: str,
    extractor_used: str,
    extraction_warnings: List[str],
    table_shape: Optional[Tuple[int, int]] = None,
) -> float:
    """Calculate confidence score for extracted content."""
    base_scores = {"docling": 0.95, "pdfplumber": 0.85, "office_native": 0.90}
    confidence = base_scores.get(extractor_used, 0.80)

    if content_type == "section" and extractor_used == "pdfplumber":
        confidence -= 0.10

    warning_penalty = min(len(extraction_warnings) * 0.05, 0.15)
    confidence -= warning_penalty

    if content_type == "table" and table_shape:
        rows, cols = table_shape
        if cols == 0 or rows < 2:
            confidence -= 0.10

    return max(0.50, min(1.00, round(confidence, 2)))


# ============================================================================
# SECTION 3.5: PDFPLUMBER FALLBACK EXTRACTION
# ============================================================================


def is_header_line(line: str, rules: dict) -> bool:
    """Determine if a line is likely a section header."""
    header_rules = rules.get("header_detection", {})
    max_len = header_rules.get("max_len", 100)
    min_len = header_rules.get("min_len", 3)
    all_caps_ratio = header_rules.get("all_caps_ratio", 0.8)
    numbered_pattern = header_rules.get("numbered_heading_regex", r"^\d+\.[\d.]*\s+")
    additional_patterns = header_rules.get("additional_patterns", [])

    line = line.strip()

    if len(line) < min_len or len(line) > max_len:
        return False

    # Check ALL CAPS
    alpha_chars = [c for c in line if c.isalpha()]
    if alpha_chars:
        upper_ratio = sum(1 for c in alpha_chars if c.isupper()) / len(alpha_chars)
        if upper_ratio >= all_caps_ratio:
            return True

    # Check numbered heading
    if re.match(numbered_pattern, line):
        return True

    # Check additional patterns
    for pattern in additional_patterns:
        if re.match(pattern, line, re.IGNORECASE):
            return True

    return False


def should_skip_content(text: str, rules: dict) -> bool:
    """Check if content should be skipped based on filtering rules."""
    filtering_rules = rules.get("content_filtering", {})
    min_length = filtering_rules.get("min_text_length", 10)
    skip_patterns = filtering_rules.get("skip_patterns", [])

    text = text.strip()

    if len(text) < min_length:
        return True

    for pattern in skip_patterns:
        if re.match(pattern, text, re.IGNORECASE):
            return True

    return False


def extract_with_pdfplumber(
    file_path: Path, config: dict, rules: dict
) -> Tuple[List[ExtractedElement], int, str]:
    """Fallback PDF extraction using pdfplumber."""
    import pdfplumber

    elements = []
    page_count = 0
    current_section = "UNSPECIFIED"

    try:
        with pdfplumber.open(file_path) as pdf:
            page_count = len(pdf.pages)

            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                paragraph_lines = []

                for line in text.split("\n"):
                    line = line.strip()
                    if not line:
                        if paragraph_lines:
                            para_text = " ".join(paragraph_lines)
                            if not should_skip_content(para_text, rules):
                                elements.append(
                                    ExtractedElement(
                                        content_type="chunk",
                                        content=para_text,
                                        page_number=page_num,
                                        section_title=current_section,
                                        confidence=0.85,
                                    )
                                )
                            paragraph_lines = []
                        continue

                    if is_header_line(line, rules):
                        if paragraph_lines:
                            para_text = " ".join(paragraph_lines)
                            if not should_skip_content(para_text, rules):
                                elements.append(
                                    ExtractedElement(
                                        content_type="chunk",
                                        content=para_text,
                                        page_number=page_num,
                                        section_title=current_section,
                                        confidence=0.85,
                                    )
                                )
                            paragraph_lines = []

                        current_section = line
                        elements.append(
                            ExtractedElement(
                                content_type="section",
                                content=line,
                                page_number=page_num,
                                section_title=current_section,
                                confidence=0.75,
                            )
                        )
                    else:
                        paragraph_lines.append(line)

                if paragraph_lines:
                    para_text = " ".join(paragraph_lines)
                    if not should_skip_content(para_text, rules):
                        elements.append(
                            ExtractedElement(
                                content_type="chunk",
                                content=para_text,
                                page_number=page_num,
                                section_title=current_section,
                                confidence=0.85,
                            )
                        )

                # Extract tables
                tables = page.extract_tables()
                for table in tables:
                    if table and len(table) > 0:
                        csv_content = table_to_csv(table)
                        if csv_content:
                            elements.append(
                                ExtractedElement(
                                    content_type="table",
                                    content=csv_content,
                                    page_number=page_num,
                                    section_title=current_section,
                                    table_shape=(len(table), len(table[0]) if table[0] else 0),
                                    confidence=0.85,
                                )
                            )

        return elements, page_count, "pdfplumber"

    except Exception as e:
        raise ExtractionError(f"pdfplumber extraction failed: {str(e)}") from e


# ============================================================================
# SECTION 3.4: PRIMARY EXTRACTION - DOCLING
# ============================================================================


def extract_with_docling(
    file_path: Path, config: dict
) -> Tuple[List[ExtractedElement], int, str]:
    """Primary extraction using Docling library."""
    try:
        from docling.document_converter import DocumentConverter
    except ImportError as e:
        raise ExtractionError(f"Docling import failed: {e}") from e

    elements: List[ExtractedElement] = []
    current_section = "UNSPECIFIED"
    page_count = 0

    try:
        converter = DocumentConverter()
        result = converter.convert(str(file_path))

        if hasattr(result, "document") and hasattr(result.document, "pages"):
            page_count = len(result.document.pages) if result.document.pages else 0

        LABEL_TO_TYPE = {
            "title": "section",
            "section_header": "section",
            "heading": "section",
            "paragraph": "chunk",
            "text": "chunk",
            "list_item": "chunk",
            "table": "table",
            "caption": "chunk",
            "formula": "chunk",
            "code": "chunk",
            "footnote": "chunk",
        }

        if hasattr(result, "document") and hasattr(result.document, "iterate_items"):
            for item, level in result.document.iterate_items():
                label = (
                    getattr(item, "label", "text").lower()
                    if hasattr(item, "label")
                    else "text"
                )

                if label == "picture":
                    continue

                content_type = LABEL_TO_TYPE.get(label, "chunk")

                if label == "table":
                    if hasattr(item, "export_to_dataframe"):
                        df = item.export_to_dataframe()
                        table_data = [df.columns.tolist()] + df.values.tolist()
                        content = table_to_csv(table_data)
                        table_shape = (len(table_data), len(table_data[0]) if table_data else 0)
                    elif hasattr(item, "to_dataframe"):
                        df = item.to_dataframe()
                        table_data = [df.columns.tolist()] + df.values.tolist()
                        content = table_to_csv(table_data)
                        table_shape = (len(table_data), len(table_data[0]) if table_data else 0)
                    else:
                        content = str(item)
                        table_shape = (0, 0)
                else:
                    content = item.text if hasattr(item, "text") else str(item)
                    table_shape = None

                if not content or not content.strip():
                    continue

                page_number = 1
                if hasattr(item, "prov") and item.prov:
                    prov = item.prov[0] if isinstance(item.prov, list) else item.prov
                    if hasattr(prov, "page_no") and prov.page_no:
                        page_number = prov.page_no

                bbox = None
                if hasattr(item, "prov") and item.prov:
                    prov = item.prov[0] if isinstance(item.prov, list) else item.prov
                    if hasattr(prov, "bbox") and prov.bbox:
                        b = prov.bbox
                        bbox = (b.l, b.t, b.r, b.b) if hasattr(b, "l") else None

                if content_type == "section":
                    current_section = content.strip()

                elements.append(
                    ExtractedElement(
                        content_type=content_type,
                        content=content.strip(),
                        page_number=page_number,
                        section_title=current_section,
                        bbox=bbox,
                        table_shape=table_shape,
                        confidence=0.95,
                        warnings=[],
                    )
                )

        return elements, page_count, "docling"

    except Exception as e:
        raise ExtractionError(f"Docling extraction failed: {str(e)}") from e


# ============================================================================
# SECTION 3.6: OFFICE NATIVE EXTRACTION (DOCX/XLSX)
# ============================================================================


def extract_paragraph_text(para_element, ns: dict) -> str:
    """Extract text from a DOCX paragraph element."""
    texts = []
    for text_elem in para_element.iter(f"{{{ns['w']}}}t"):
        if text_elem.text:
            texts.append(text_elem.text)
    return "".join(texts)


def get_paragraph_style(para_element, ns: dict) -> str:
    """Get style name from a DOCX paragraph element."""
    pPr = para_element.find(f"{{{ns['w']}}}pPr", ns)
    if pPr is not None:
        pStyle = pPr.find(f"{{{ns['w']}}}pStyle", ns)
        if pStyle is not None:
            return pStyle.get(f"{{{ns['w']}}}val", "")
    return ""


def extract_table_from_docx_element(table_element, ns: dict) -> List[List[str]]:
    """Extract table data from a DOCX table element."""
    rows = []
    for tr in table_element.iter(f"{{{ns['w']}}}tr"):
        row = []
        for tc in tr.iter(f"{{{ns['w']}}}tc"):
            cell_texts = []
            for para in tc.iter(f"{{{ns['w']}}}p"):
                cell_texts.append(extract_paragraph_text(para, ns))
            row.append(" ".join(cell_texts))
        if row:
            rows.append(row)
    return rows


def extract_docx_native(
    file_path: Path, config: dict, rules: dict
) -> Tuple[List[ExtractedElement], int, str]:
    """Extract content from DOCX using zipfile + lxml."""
    elements = []
    current_section = "UNSPECIFIED"

    try:
        with zipfile.ZipFile(file_path, "r") as docx:
            with docx.open("word/document.xml") as doc_xml:
                tree = etree.parse(doc_xml)
                root = tree.getroot()

                ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                body = root.find(".//w:body", ns)

                if body is None:
                    return elements, 0, "office_native"

                for element in body:
                    tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

                    if tag == "p":
                        text = extract_paragraph_text(element, DOCX_NS)
                        style = get_paragraph_style(element, DOCX_NS)

                        if not text.strip():
                            continue

                        if style and style.lower().startswith("heading"):
                            current_section = text.strip()
                            elements.append(
                                ExtractedElement(
                                    content_type="section",
                                    content=text.strip(),
                                    page_number=None,
                                    section_title=current_section,
                                    confidence=0.95,
                                )
                            )
                        elif not should_skip_content(text, rules):
                            elements.append(
                                ExtractedElement(
                                    content_type="chunk",
                                    content=text.strip(),
                                    page_number=None,
                                    section_title=current_section,
                                    confidence=0.95,
                                )
                            )

                    elif tag == "tbl":
                        table_data = extract_table_from_docx_element(element, DOCX_NS)
                        if table_data:
                            csv_content = table_to_csv(table_data)
                            if csv_content:
                                elements.append(
                                    ExtractedElement(
                                        content_type="table",
                                        content=csv_content,
                                        page_number=None,
                                        section_title=current_section,
                                        table_shape=(
                                            len(table_data),
                                            len(table_data[0]) if table_data[0] else 0,
                                        ),
                                        confidence=0.95,
                                    )
                                )

        return elements, 0, "office_native"

    except Exception as e:
        raise ExtractionError(f"DOCX extraction failed: {str(e)}") from e


def extract_xlsx_native(
    file_path: Path, config: dict, rules: dict
) -> Tuple[List[ExtractedElement], int, str]:
    """Extract content from XLSX using openpyxl."""
    elements = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            current_section = f"Sheet: {sheet_name}"

            elements.append(
                ExtractedElement(
                    content_type="section",
                    content=current_section,
                    page_number=None,
                    section_title=current_section,
                    sheet_name=sheet_name,
                    confidence=1.0,
                )
            )

            data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append([str(cell) if cell is not None else "" for cell in row])

            if data and len(data) > 1:
                table_chunks = chunk_large_table(
                    data, config.get("chunking", {}).get("table_chunk_rows", 40)
                )

                for i, chunk_data in enumerate(table_chunks):
                    csv_content = table_to_csv(chunk_data)
                    if csv_content:
                        chunk_section = (
                            current_section
                            if len(table_chunks) == 1
                            else f"{current_section} (Part {i + 1})"
                        )
                        elements.append(
                            ExtractedElement(
                                content_type="table",
                                content=csv_content,
                                page_number=None,
                                section_title=chunk_section,
                                sheet_name=sheet_name,
                                table_shape=(len(data), len(data[0]) if data else 0),
                                confidence=0.95,
                            )
                        )

        wb.close()
        return elements, 0, "office_native"

    except Exception as e:
        raise ExtractionError(f"XLSX extraction failed: {str(e)}") from e


# ============================================================================
# CONFIGURATION LOADING (JSON format per CANON_003)
# ============================================================================


def load_config(working_dir: Path) -> dict:
    """Load configuration from config.json."""
    config_path = working_dir / "config.json"

    if not config_path.exists():
        raise ConfigurationError(f"Configuration file not found: {config_path}")

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        raise ConfigurationError(f"Invalid JSON in config file: {e}")


def load_rules(working_dir: Path) -> dict:
    """Load extraction rules from rules.json."""
    rules_path = working_dir / "rules.json"

    if not rules_path.exists():
        return {}

    try:
        with open(rules_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        raise ConfigurationError(f"Invalid JSON in rules file: {e}")


# ============================================================================
# LOGGING SETUP
# ============================================================================


def setup_logging(config: dict, working_dir: Path) -> logging.Logger:
    """Configure logging based on config settings."""
    log_config = config.get("logging", {})
    log_level = getattr(logging, log_config.get("level", "INFO").upper())
    log_format = log_config.get("format", "%(asctime)s | %(levelname)-8s | %(message)s")
    date_format = log_config.get("date_format", "%Y-%m-%d %H:%M:%S")

    log_dir = working_dir / config.get("paths", {}).get("log_dir", "logs")
    log_dir.mkdir(exist_ok=True)

    logger = logging.getLogger("split_rag_extractor")
    logger.setLevel(log_level)
    logger.handlers = []

    log_file = log_dir / "processing.log"
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(log_level)
    file_handler.setFormatter(logging.Formatter(log_format, date_format))
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(log_level)
    console_handler.setFormatter(logging.Formatter(log_format, date_format))
    logger.addHandler(console_handler)

    return logger


# ============================================================================
# OUTPUT FILE NAMING
# ============================================================================


def get_output_filename(input_path: Path) -> str:
    """Generate output filename from input path."""
    return f"{input_path.stem}.json"


# ============================================================================
# ELEMENT TO NODE CONVERSION
# ============================================================================


def build_file_ref(file_path: Path, working_dir: Path, file_sha256: str) -> dict:
    """Build FileRef dictionary for a file."""
    return {
        "file_path": str(file_path.relative_to(working_dir)).replace("\\", "/"),
        "file_name": file_path.name,
        "file_ext": file_path.suffix.lower().lstrip("."),
        "file_size_bytes": file_path.stat().st_size,
        "file_modified_utc": get_file_modified_utc(file_path),
        "file_sha256": file_sha256,
    }


def build_provenance(extractor_used: str, confidence: float, warnings: List[str]) -> dict:
    """Build Provenance dictionary."""
    return {
        "source_tier": "tier1_local",
        "extractor_version": EXTRACTOR_VERSION,
        "created_utc": utc_now_z(),
        "confidence": confidence,
        "warnings": warnings,
    }


def elements_to_nodes(
    elements: List[ExtractedElement],
    file_path: Path,
    working_dir: Path,
    file_sha256: str,
    extractor_used: str,
    config: dict,
) -> Tuple[List[dict], List[dict]]:
    """Convert ExtractedElements to ContextNode dictionaries and build edges."""
    nodes = []
    edges = []

    file_ref = build_file_ref(file_path, working_dir, file_sha256)
    chunking_config = config.get("chunking", {})
    max_chars = chunking_config.get("max_chars_per_chunk", 6000)
    overlap_chars = chunking_config.get("overlap_chars", 200)

    # Create document node
    doc_node_id = compute_node_id(file_sha256, "document", file_path.name, 0)
    doc_node = {
        "node_id": doc_node_id,
        "node_type": "document",
        "title": file_path.name,
        "parent_id": None,
        "file_ref": file_ref,
        "location": None,
        "provenance": build_provenance(extractor_used, 1.0, []),
        "text": None,
        "table": None,
        "image": None,
        "tags": [],
    }
    nodes.append(doc_node)

    # Track sections for parent linking
    current_section_id = doc_node_id
    section_index = 0
    chunk_index = 0

    for elem_idx, element in enumerate(elements):
        confidence = calculate_confidence(
            element.content_type,
            extractor_used,
            element.warnings,
            element.table_shape,
        )

        # Build location
        location = None
        if element.page_number or element.sheet_name or element.bbox:
            location = {}
            if element.page_number:
                location["page"] = element.page_number
            if element.sheet_name:
                location["sheet"] = element.sheet_name
            if element.bbox:
                location["bbox"] = list(element.bbox)

        if element.content_type == "section":
            section_index += 1
            node_id = compute_node_id(file_sha256, "section", element.content, section_index)
            node = {
                "node_id": node_id,
                "node_type": "section",
                "title": element.content,
                "parent_id": doc_node_id,
                "file_ref": file_ref,
                "location": location,
                "provenance": build_provenance(extractor_used, confidence, element.warnings),
                "text": None,
                "table": None,
                "image": None,
                "tags": [],
            }
            nodes.append(node)
            current_section_id = node_id

            # Add contains edge
            edge_id = compute_edge_id(doc_node_id, node_id, "contains")
            edges.append({
                "edge_id": edge_id,
                "from_id": doc_node_id,
                "to_id": node_id,
                "relation": "contains",
                "weight": 1.0,
                "metadata": {},
            })

        elif element.content_type == "chunk":
            # Apply chunking if needed
            if len(element.content) > max_chars:
                text_chunks = chunk_text_content(element.content, max_chars, overlap_chars)
                for chunk_text, char_start, char_end in text_chunks:
                    chunk_index += 1
                    node_id = compute_node_id(file_sha256, "chunk", chunk_text, chunk_index)
                    node = {
                        "node_id": node_id,
                        "node_type": "chunk",
                        "title": None,
                        "parent_id": current_section_id,
                        "file_ref": file_ref,
                        "location": location,
                        "provenance": build_provenance(extractor_used, confidence, element.warnings),
                        "text": {
                            "text": chunk_text,
                            "char_start": char_start,
                            "char_end": char_end,
                            "token_estimate": len(chunk_text) // 4,
                        },
                        "table": None,
                        "image": None,
                        "tags": [],
                    }
                    nodes.append(node)

                    edge_id = compute_edge_id(current_section_id, node_id, "contains")
                    edges.append({
                        "edge_id": edge_id,
                        "from_id": current_section_id,
                        "to_id": node_id,
                        "relation": "contains",
                        "weight": 1.0,
                        "metadata": {},
                    })
            else:
                chunk_index += 1
                node_id = compute_node_id(file_sha256, "chunk", element.content, chunk_index)
                node = {
                    "node_id": node_id,
                    "node_type": "chunk",
                    "title": None,
                    "parent_id": current_section_id,
                    "file_ref": file_ref,
                    "location": location,
                    "provenance": build_provenance(extractor_used, confidence, element.warnings),
                    "text": {
                        "text": element.content,
                        "char_start": 0,
                        "char_end": len(element.content),
                        "token_estimate": len(element.content) // 4,
                    },
                    "table": None,
                    "image": None,
                    "tags": [],
                }
                nodes.append(node)

                edge_id = compute_edge_id(current_section_id, node_id, "contains")
                edges.append({
                    "edge_id": edge_id,
                    "from_id": current_section_id,
                    "to_id": node_id,
                    "relation": "contains",
                    "weight": 1.0,
                    "metadata": {},
                })

        elif element.content_type == "table":
            chunk_index += 1
            node_id = compute_node_id(file_sha256, "table", element.content, chunk_index)
            node = {
                "node_id": node_id,
                "node_type": "table",
                "title": None,
                "parent_id": current_section_id,
                "file_ref": file_ref,
                "location": location,
                "provenance": build_provenance(extractor_used, confidence, element.warnings),
                "text": None,
                "table": {
                    "format": "csv",
                    "csv_text": element.content,
                },
                "image": None,
                "tags": [],
            }
            nodes.append(node)

            edge_id = compute_edge_id(current_section_id, node_id, "contains")
            edges.append({
                "edge_id": edge_id,
                "from_id": current_section_id,
                "to_id": node_id,
                "relation": "contains",
                "weight": 1.0,
                "metadata": {},
            })

    # Add "next" edges for sequential nodes
    chunk_nodes = [n for n in nodes if n["node_type"] in ("chunk", "table")]
    for i in range(len(chunk_nodes) - 1):
        edge_id = compute_edge_id(chunk_nodes[i]["node_id"], chunk_nodes[i + 1]["node_id"], "next")
        edges.append({
            "edge_id": edge_id,
            "from_id": chunk_nodes[i]["node_id"],
            "to_id": chunk_nodes[i + 1]["node_id"],
            "relation": "next",
            "weight": 1.0,
            "metadata": {},
        })

    return nodes, edges


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================


def main(working_dir: Path) -> int:
    """
    Main entry point for extraction.

    Returns:
        0 = success (all files processed or skipped)
        1 = partial success (some files failed)
        2 = total failure (no files processed)
    """
    start_time = time.time()

    # Step 1: Load configuration
    try:
        config = load_config(working_dir)
        rules = load_rules(working_dir)
    except ConfigurationError as e:
        print(f"ERROR: {e}")
        return 2

    # Step 2: Initialize logging
    logger = setup_logging(config, working_dir)
    logger.info(f"[{LOG_CODES['E001']}] Starting extraction run")

    # Step 3: Load/create manifest
    paths_config = config.get("paths", {})
    manifest_path = working_dir / paths_config.get("manifest_file", "output/manifest.json")
    manifest_path.parent.mkdir(exist_ok=True)
    manifest = load_manifest(manifest_path)

    # Step 4: Discover input files
    input_dir = working_dir / paths_config.get("input_dir", "input")
    output_dir = working_dir / paths_config.get("output_dir", "output")
    output_dir.mkdir(exist_ok=True)

    if not input_dir.exists():
        logger.error(f"Input directory not found: {input_dir}")
        return 2

    files = discover_input_files(input_dir)
    logger.info(f"[{LOG_CODES['E003']}] Discovered {len(files)} files")

    if not files:
        logger.warning("No processable files found in input directory")
        return 0

    # Step 5: Process each file
    processed_count = 0
    failed_count = 0
    skipped_count = 0

    max_files = config.get("processing", {}).get("max_files_per_run", 0)
    files_to_process = files[:max_files] if max_files > 0 else files

    run_timestamp = utc_now_z()
    run_id = compute_run_id(files_to_process, run_timestamp)

    for file_path in tqdm(files_to_process, desc="Processing files"):
        file_start_time = time.time()

        try:
            # Step 5a: Compute SHA256
            file_sha256 = compute_file_sha256(file_path)

            # Step 5b: Check manifest
            if not should_process_file(file_path, file_sha256, manifest, config):
                logger.info(f"[{LOG_CODES['E004']}] Skipping unchanged: {file_path.name}")
                skipped_count += 1
                continue

            # Step 5c-d: Extract content
            file_ext = file_path.suffix.lower()
            elements = []
            page_count = 0
            extractor_used = ""

            if file_ext == ".pdf":
                try:
                    elements, page_count, extractor_used = extract_with_docling(file_path, config)
                    logger.info(f"[{LOG_CODES['E005']}] Docling succeeded: {file_path.name}")
                except ExtractionError as e:
                    logger.warning(f"[{LOG_CODES['E006']}] Docling failed: {e}")
                    try:
                        elements, page_count, extractor_used = extract_with_pdfplumber(
                            file_path, config, rules
                        )
                        logger.info(f"[{LOG_CODES['E007']}] Fallback to pdfplumber: {file_path.name}")
                    except ExtractionError as e2:
                        raise ExtractionError(f"All extractors failed: {e2}") from e2

            elif file_ext == ".docx":
                elements, page_count, extractor_used = extract_docx_native(file_path, config, rules)
                logger.info(f"[{LOG_CODES['E005']}] DOCX native succeeded: {file_path.name}")

            elif file_ext == ".xlsx":
                elements, page_count, extractor_used = extract_xlsx_native(file_path, config, rules)
                logger.info(f"[{LOG_CODES['E005']}] XLSX native succeeded: {file_path.name}")

            else:
                raise ExtractionError(f"Unsupported file type: {file_ext}")

            # Step 5f-g: Convert to nodes and build edges
            nodes, edges = elements_to_nodes(
                elements, file_path, working_dir, file_sha256, extractor_used, config
            )

            # Step 5h: Build and validate output
            file_duration_ms = int((time.time() - file_start_time) * 1000)

            nodes_by_type = defaultdict(int)
            total_chars = 0
            for node in nodes:
                nodes_by_type[node["node_type"]] += 1
                if node.get("text"):
                    total_chars += len(node["text"]["text"])
                elif node.get("table"):
                    total_chars += len(node["table"]["csv_text"])

            output_data = {
                "schema_version": SCHEMA_VERSION,
                "run_id": run_id,
                "created_utc": run_timestamp,
                "nodes": nodes,
                "edges": edges,
                "stats": {
                    "total_nodes": len(nodes),
                    "nodes_by_type": dict(nodes_by_type),
                    "total_edges": len(edges),
                    "total_chars": total_chars,
                    "extraction_duration_ms": file_duration_ms,
                    "extractor_used": extractor_used,
                    "page_count": page_count,
                },
            }

            # Validate against schema
            validation_errors = 0
            try:
                validate_graph(output_data)
            except ValidationError as e:
                logger.error(f"[{LOG_CODES['E008']}] Validation failed: {file_path.name}: {e}")
                validation_errors = len(e.errors())
                output_data["stats"]["validation_errors_count"] = validation_errors

            # Step 5i: Write JSON output
            output_filename = get_output_filename(file_path)
            output_path = output_dir / output_filename

            try:
                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(output_data, f, indent=2, default=str)
                logger.info(f"[{LOG_CODES['E009']}] Wrote: {output_filename}")
            except IOError as e:
                logger.error(f"[{LOG_CODES['E010']}] Write failed: {output_filename}: {e}")
                raise ExtractionError(f"Failed to write output: {e}") from e

            # Step 5j: Update manifest
            source_path = str(file_path.relative_to(working_dir))
            status = "processed" if validation_errors == 0 else "partial"
            manifest.files[source_path] = ManifestEntry(
                sha256=file_sha256,
                status=status,
                processed_at=run_timestamp,
                output_file=output_filename,
                error_message=None,
            )

            processed_count += 1

        except Exception as e:
            logger.error(f"Failed to process {file_path.name}: {e}")
            source_path = str(file_path.relative_to(working_dir))
            manifest.files[source_path] = ManifestEntry(
                sha256=compute_file_sha256(file_path) if file_path.exists() else "",
                status="failed",
                processed_at=utc_now_z(),
                output_file=None,
                error_message=str(e),
            )
            failed_count += 1

    # Step 6: Write final manifest
    save_manifest(manifest, manifest_path)
    logger.info(f"[{LOG_CODES['E011']}] Manifest updated")

    # Step 7: Log summary statistics
    total_time = time.time() - start_time
    logger.info(f"[{LOG_CODES['E002']}] Extraction complete in {total_time:.2f}s")
    logger.info(f"  Processed: {processed_count}")
    logger.info(f"  Skipped: {skipped_count}")
    logger.info(f"  Failed: {failed_count}")

    if failed_count > 0 and processed_count == 0:
        return 2
    elif failed_count > 0:
        return 1
    else:
        return 0


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Split-RAG Document Extractor",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--working_dir",
        type=str,
        required=True,
        help="Path to working directory containing config.json and input/ folder",
    )

    args = parser.parse_args()
    working_dir = Path(args.working_dir).resolve()

    if not working_dir.exists():
        print(f"Error: Working directory does not exist: {working_dir}")
        sys.exit(2)

    exit_code = main(working_dir)
    sys.exit(exit_code)
